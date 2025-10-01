import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog, messagebox
import json
import os
import pandas as pd
import numpy as np
import sys
import io
from contextlib import redirect_stdout, redirect_stderr
from sklearn.model_selection import train_test_split, TimeSeriesSplit
from sklearn.ensemble import RandomForestClassifier, VotingClassifier
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
from sklearn.feature_selection import SelectKBest, f_classif
from sklearn.metrics import classification_report, accuracy_score
from sklearn.calibration import CalibratedClassifierCV
import warnings

warnings.filterwarnings("ignore")

try:
    from xgboost import XGBClassifier
    HAS_XGBOOST = True
except ImportError:
    HAS_XGBOOST = False
    print("XGBoost not available. Using RandomForest instead.")


class UFCPredictor:

    def __init__(self, use_xgboost=True):
        self.winner_model = None
        self.method_model = None
        self.label_encoders = {}
        self.use_xgboost = use_xgboost and HAS_XGBOOST
        self.df_train = None

    def calculate_streak(self, recent_wins, count_wins=True):
        """Calculate current win or loss streak"""
        if not recent_wins:
            return 0
        
        streak = 0
        target = 1 if count_wins else 0
        
        for result in reversed(recent_wins):
            if result == target:
                streak += 1
            else:
                break
        
        return streak

    def fix_data_leakage(self, df):
        """Recalculate comprehensive fighter statistics chronologically"""
        print("Fixing data leakage with comprehensive feature tracking...\n")

        import copy

        df["event_date"] = pd.to_datetime(df["event_date"])
        df = df.sort_values("event_date").reset_index(drop=True)

        fighter_stats = {}

        stats_to_track = {
            "wins": 0,
            "losses": 0,
            "draws": 0,
            "sig_str_total": 0,
            "sig_str_att_total": 0,
            "sig_str_absorbed_total": 0,
            "total_str_landed": 0,
            "total_str_att": 0,
            "total_str_absorbed": 0,
            "str_def_hits": 0,
            "str_def_att": 0,
            "td_total": 0,
            "td_att_total": 0,
            "td_def_success": 0,
            "td_def_att": 0,
            "sub_att_total": 0,
            "kd_total": 0,
            "kd_absorbed_total": 0,
            "fight_time_minutes": 0,
            "fight_count": 0,
            "ko_wins": 0,
            "sub_wins": 0,
            "dec_wins": 0,
            "ko_losses": 0,
            "sub_losses": 0,
            "recent_wins": [],
            "recent_finishes": [],
            "last_fight_date": None,
            "head_pct_sum": 0,
            "body_pct_sum": 0,
            "leg_pct_sum": 0,
            "location_fight_count": 0,
            "distance_pct_sum": 0,
            "clinch_pct_sum": 0,
            "ground_pct_sum": 0,
            "position_fight_count": 0,
        }

        for prefix in ["r", "b"]:
            for stat in [
                "wins", "losses", "draws", "win_loss_ratio",
                "pro_SLpM", "pro_sig_str_acc", "pro_SApM", "pro_str_def",
                "pro_total_str_pM", "pro_total_str_acc", "pro_total_str_absorbed_pM",
                "pro_td_avg", "pro_td_acc", "pro_td_def", "pro_sub_avg", "pro_kd_pM",
                "ko_rate", "sub_rate", "dec_rate", "recent_form",
                "head_pct", "body_pct", "leg_pct",
                "distance_pct", "clinch_pct", "ground_pct",
                "win_streak", "loss_streak", "last_5_wins",
                "days_since_last_fight", "recent_finish_rate", "durability",
                "fight_time_minutes"
            ]:
                df[f"{prefix}_{stat}_corrected"] = 0.0
        
        df['h2h_advantage'] = 0.0
        fighter_h2h = {}

        for idx, row in df.iterrows():
            if idx % 500 == 0:
                print(f"   Processing fight {idx}/{len(df)}...")

            r_fighter, b_fighter = row["r_fighter"], row["b_fighter"]
            
            # Head-to-head tracking
            h2h_key = (r_fighter, b_fighter)
            h2h_key_reverse = (b_fighter, r_fighter)
            
            if h2h_key in fighter_h2h:
                df.at[idx, 'h2h_advantage'] = fighter_h2h[h2h_key]
            elif h2h_key_reverse in fighter_h2h:
                df.at[idx, 'h2h_advantage'] = -fighter_h2h[h2h_key_reverse]

            if r_fighter not in fighter_stats:
                fighter_stats[r_fighter] = copy.deepcopy(stats_to_track)
            if b_fighter not in fighter_stats:
                fighter_stats[b_fighter] = copy.deepcopy(stats_to_track)

            for fighter, prefix in [(r_fighter, "r"), (b_fighter, "b")]:
                stats = fighter_stats[fighter]
                df.at[idx, f"{prefix}_wins_corrected"] = stats["wins"]
                df.at[idx, f"{prefix}_losses_corrected"] = stats["losses"]
                df.at[idx, f"{prefix}_draws_corrected"] = stats["draws"]
                df.at[idx, f"{prefix}_win_loss_ratio_corrected"] = stats["wins"] / max(
                    stats["losses"], 1
                )

                # Durability (inverse of finish losses)
                total_losses = stats["losses"]
                finish_losses = stats["ko_losses"] + stats["sub_losses"]
                df.at[idx, f"{prefix}_durability_corrected"] = (
                    1.0 / (1 + finish_losses) if total_losses > 0 else 1.0
                )

                total_fights = stats["wins"] + stats["losses"]
                if total_fights > 0:
                    df.at[idx, f"{prefix}_ko_rate_corrected"] = (
                        stats["ko_wins"] / total_fights
                    )
                    df.at[idx, f"{prefix}_sub_rate_corrected"] = (
                        stats["sub_wins"] / total_fights
                    )
                    df.at[idx, f"{prefix}_dec_rate_corrected"] = (
                        stats["dec_wins"] / total_fights
                    )

                # Recent finish rate
                if len(stats["recent_finishes"]) > 0:
                    df.at[idx, f"{prefix}_recent_finish_rate_corrected"] = (
                        sum(stats["recent_finishes"]) / len(stats["recent_finishes"])
                    )

                if len(stats["recent_wins"]) > 0:
                    df.at[idx, f"{prefix}_recent_form_corrected"] = sum(
                        stats["recent_wins"]
                    ) / len(stats["recent_wins"])
                
                # Momentum features
                df.at[idx, f"{prefix}_win_streak_corrected"] = self.calculate_streak(
                    stats["recent_wins"], True
                )
                df.at[idx, f"{prefix}_loss_streak_corrected"] = self.calculate_streak(
                    stats["recent_wins"], False
                )
                df.at[idx, f"{prefix}_last_5_wins_corrected"] = (
                    sum(stats["recent_wins"][-5:]) if len(stats["recent_wins"]) >= 5 else 0
                )

                # Days since last fight
                if stats["last_fight_date"]:
                    days_off = (row["event_date"] - stats["last_fight_date"]).days
                    df.at[idx, f"{prefix}_days_since_last_fight_corrected"] = days_off

                # Total fight time accumulated
                df.at[idx, f"{prefix}_fight_time_minutes_corrected"] = stats["fight_time_minutes"]

                if stats["fight_time_minutes"] > 0:
                    df.at[idx, f"{prefix}_pro_SLpM_corrected"] = (
                        stats["sig_str_total"] / stats["fight_time_minutes"]
                    )
                    df.at[idx, f"{prefix}_pro_SApM_corrected"] = (
                        stats["sig_str_absorbed_total"] / 
                        stats["fight_time_minutes"]
                    )
                    df.at[idx, f"{prefix}_pro_total_str_pM_corrected"] = (
                        stats["total_str_landed"] / stats["fight_time_minutes"]
                    )
                    df.at[idx, f"{prefix}_pro_total_str_absorbed_pM_corrected"] = (
                        stats["total_str_absorbed"] / 
                        stats["fight_time_minutes"]
                    )
                    df.at[idx, f"{prefix}_pro_td_avg_corrected"] = (
                        stats["td_total"] / stats["fight_time_minutes"]
                    ) * 15
                    df.at[idx, f"{prefix}_pro_sub_avg_corrected"] = (
                        stats["sub_att_total"] / stats["fight_time_minutes"]
                    ) * 15
                    df.at[idx, f"{prefix}_pro_kd_pM_corrected"] = (
                        stats["kd_total"] / stats["fight_time_minutes"]
                    )

                if stats["sig_str_att_total"] > 0:
                    df.at[idx, f"{prefix}_pro_sig_str_acc_corrected"] = (
                        stats["sig_str_total"] / stats["sig_str_att_total"]
                    )
                if stats["total_str_att"] > 0:
                    df.at[idx, f"{prefix}_pro_total_str_acc_corrected"] = (
                        stats["total_str_landed"] / stats["total_str_att"]
                    )
                if stats["str_def_att"] > 0:
                    df.at[idx, f"{prefix}_pro_str_def_corrected"] = (
                        stats["str_def_hits"] / stats["str_def_att"]
                    )
                if stats["td_att_total"] > 0:
                    df.at[idx, f"{prefix}_pro_td_acc_corrected"] = (
                        stats["td_total"] / stats["td_att_total"]
                    )
                if stats["td_def_att"] > 0:
                    df.at[idx, f"{prefix}_pro_td_def_corrected"] = (
                        stats["td_def_success"] / stats["td_def_att"]
                    )

                # Strike location percentages
                if stats["location_fight_count"] > 0:
                    df.at[idx, f"{prefix}_head_pct_corrected"] = (
                        stats["head_pct_sum"] / stats["location_fight_count"]
                    )
                    df.at[idx, f"{prefix}_body_pct_corrected"] = (
                        stats["body_pct_sum"] / stats["location_fight_count"]
                    )
                    df.at[idx, f"{prefix}_leg_pct_corrected"] = (
                        stats["leg_pct_sum"] / stats["location_fight_count"]
                    )

                if stats["position_fight_count"] > 0:
                    df.at[idx, f"{prefix}_distance_pct_corrected"] = (
                        stats["distance_pct_sum"] / 
                        stats["position_fight_count"]
                    )
                    df.at[idx, f"{prefix}_clinch_pct_corrected"] = (
                        stats["clinch_pct_sum"] / stats["position_fight_count"]
                    )
                    df.at[idx, f"{prefix}_ground_pct_corrected"] = (
                        stats["ground_pct_sum"] / stats["position_fight_count"]
                    )

            # Update stats after fight
            if pd.notna(row["winner"]):
                fight_time_min = (
                    row["total_fight_time_sec"] / 60
                    if pd.notna(row["total_fight_time_sec"])
                    else 0
                )
                method = str(row["method"]).lower()
                method_cat = (
                    "ko"
                    if "ko" in method or "tko" in method
                    else "sub"
                    if "sub" in method
                    else "dec"
                )
                is_finish = method_cat in ["ko", "sub"]

                if row["winner"] == "Red":
                    fighter_stats[r_fighter]["wins"] += 1
                    fighter_stats[b_fighter]["losses"] += 1
                    fighter_stats[r_fighter][f"{method_cat}_wins"] += 1
                    fighter_stats[r_fighter]["recent_wins"].append(1)
                    fighter_stats[b_fighter]["recent_wins"].append(0)
                    fighter_stats[r_fighter]["recent_finishes"].append(1 if is_finish else 0)
                    fighter_stats[b_fighter]["recent_finishes"].append(0)
                    if is_finish:
                        fighter_stats[b_fighter][f"{method_cat}_losses"] += 1
                    fighter_h2h[(r_fighter, b_fighter)] = fighter_h2h.get((r_fighter, b_fighter), 0) + 1
                elif row["winner"] == "Blue":
                    fighter_stats[b_fighter]["wins"] += 1
                    fighter_stats[r_fighter]["losses"] += 1
                    fighter_stats[b_fighter][f"{method_cat}_wins"] += 1
                    fighter_stats[b_fighter]["recent_wins"].append(1)
                    fighter_stats[r_fighter]["recent_wins"].append(0)
                    fighter_stats[b_fighter]["recent_finishes"].append(1 if is_finish else 0)
                    fighter_stats[r_fighter]["recent_finishes"].append(0)
                    if is_finish:
                        fighter_stats[r_fighter][f"{method_cat}_losses"] += 1
                    fighter_h2h[(r_fighter, b_fighter)] = fighter_h2h.get((r_fighter, b_fighter), 0) - 1

                for fighter in [r_fighter, b_fighter]:
                    if len(fighter_stats[fighter]["recent_wins"]) > 5:
                        fighter_stats[fighter]["recent_wins"] = fighter_stats[fighter][
                            "recent_wins"
                        ][-5:]
                    if len(fighter_stats[fighter]["recent_finishes"]) > 5:
                        fighter_stats[fighter]["recent_finishes"] = fighter_stats[fighter][
                            "recent_finishes"
                        ][-5:]
                    fighter_stats[fighter]["last_fight_date"] = row["event_date"]

                for fighter, f_prefix, opp_prefix in [
                    (r_fighter, "r", "b"),
                    (b_fighter, "b", "r"),
                ]:
                    for stat_pair in [
                        ("sig_str", "sig_str_total"),
                        ("str", "total_str_landed"),
                    ]:
                        if pd.notna(row[f"{f_prefix}_{stat_pair[0]}"]):
                            fighter_stats[fighter][stat_pair[1]] += row[
                                f"{f_prefix}_{stat_pair[0]}"
                            ]

                    if pd.notna(row[f"{opp_prefix}_sig_str"]):
                        fighter_stats[fighter]["sig_str_absorbed_total"] += row[
                            f"{opp_prefix}_sig_str"
                        ]
                    if pd.notna(row[f"{opp_prefix}_str"]):
                        fighter_stats[fighter]["total_str_absorbed"] += row[
                            f"{opp_prefix}_str"
                        ]

                    for att in ["sig_str_att", "str_att", "td_att", "sub_att"]:
                        col = f"{f_prefix}_{att}"
                        if pd.notna(row.get(col)):
                            key = (
                                att + "_total" if att != "str_att" else "total_str_att"
                            )
                            fighter_stats[fighter][key] += row[col]

                    if pd.notna(row[f"{f_prefix}_td"]):
                        fighter_stats[fighter]["td_total"] += row[f"{f_prefix}_td"]
                    
                    # Track knockdowns
                    if pd.notna(row.get(f"{f_prefix}_kd")):
                        fighter_stats[fighter]["kd_total"] += row[f"{f_prefix}_kd"]
                    if pd.notna(row.get(f"{opp_prefix}_kd")):
                        fighter_stats[fighter]["kd_absorbed_total"] += row[f"{opp_prefix}_kd"]

                    if pd.notna(row[f"{opp_prefix}_sig_str_att"]):
                        fighter_stats[fighter]["str_def_att"] += row[
                            f"{opp_prefix}_sig_str_att"
                        ]
                        if pd.notna(row[f"{opp_prefix}_sig_str"]):
                            fighter_stats[fighter]["str_def_hits"] += (
                                row[f"{opp_prefix}_sig_str_att"]
                                -row[f"{opp_prefix}_sig_str"]
                            )

                    if pd.notna(row[f"{opp_prefix}_td_att"]):
                        fighter_stats[fighter]["td_def_att"] += row[
                            f"{opp_prefix}_td_att"
                        ]
                        if pd.notna(row[f"{opp_prefix}_td"]):
                            fighter_stats[fighter]["td_def_success"] += (
                                row[f"{opp_prefix}_td_att"] - 
                                row[f"{opp_prefix}_td"]
                            )

                    # Strike location percentages
                    if (
                        pd.notna(row.get(f"{f_prefix}_head"))
                        or pd.notna(row.get(f"{f_prefix}_body"))
                        or pd.notna(row.get(f"{f_prefix}_leg"))
                    ):
                        if pd.notna(row.get(f"{f_prefix}_head")):
                            fighter_stats[fighter]["head_pct_sum"] += row[
                                f"{f_prefix}_head"
                            ]
                        if pd.notna(row.get(f"{f_prefix}_body")):
                            fighter_stats[fighter]["body_pct_sum"] += row[
                                f"{f_prefix}_body"
                            ]
                        if pd.notna(row.get(f"{f_prefix}_leg")):
                            fighter_stats[fighter]["leg_pct_sum"] += row[
                                f"{f_prefix}_leg"
                            ]
                        fighter_stats[fighter]["location_fight_count"] += 1

                    if (
                        pd.notna(row.get(f"{f_prefix}_distance"))
                        or pd.notna(row.get(f"{f_prefix}_clinch"))
                        or pd.notna(row.get(f"{f_prefix}_ground"))
                    ):
                        if pd.notna(row.get(f"{f_prefix}_distance")):
                            fighter_stats[fighter]["distance_pct_sum"] += row[
                                f"{f_prefix}_distance"
                            ]
                        if pd.notna(row.get(f"{f_prefix}_clinch")):
                            fighter_stats[fighter]["clinch_pct_sum"] += row[
                                f"{f_prefix}_clinch"
                            ]
                        if pd.notna(row.get(f"{f_prefix}_ground")):
                            fighter_stats[fighter]["ground_pct_sum"] += row[
                                f"{f_prefix}_ground"
                            ]
                        fighter_stats[fighter]["position_fight_count"] += 1

                    fighter_stats[fighter]["fight_time_minutes"] += fight_time_min
                    fighter_stats[fighter]["fight_count"] += 1

        diff_stats = [
            "wins", "losses", "draws", "win_loss_ratio",
            "pro_SLpM", "pro_sig_str_acc", "pro_SApM", "pro_str_def",
            "pro_total_str_pM", "pro_total_str_acc", "pro_total_str_absorbed_pM",
            "pro_td_avg", "pro_td_acc", "pro_td_def", "pro_sub_avg", "pro_kd_pM",
            "ko_rate", "sub_rate", "dec_rate", "recent_form",
            "head_pct", "body_pct", "leg_pct",
            "distance_pct", "clinch_pct", "ground_pct",
            "win_streak", "loss_streak", "last_5_wins",
            "days_since_last_fight", "recent_finish_rate", "durability",
            "fight_time_minutes"
        ]

        for stat in diff_stats:
            df[f"{stat}_diff_corrected"] = (
                df[f"r_{stat}_corrected"] - df[f"b_{stat}_corrected"]
            )

        return df

    def prepare_features(self, df):
        """Prepare enhanced features"""
        print("\nPreparing features...")

        df = df[df["winner"].isin(["Red", "Blue"])].copy()

        method_mapping = {
            "KO/TKO": "KO/TKO",
            "Submission": "Submission",
            "Decision - Unanimous": "Decision",
            "Decision - Split": "Decision",
            "Decision - Majority": "Decision",
            "TKO - Doctor's Stoppage": "KO/TKO",
            "Could Not Continue": "KO/TKO",
            "DQ": "Decision",
            "Overturned": "Decision",
        }

        df["method_simple"] = df["method"].map(
            method_mapping).fillna("Decision")
        df["winner_method_simple"] = df["winner"] + "_" + df["method_simple"]

        feature_columns = [
            "height_diff",
            "reach_diff",
            "weight_diff",
            "age_at_event_diff",
            "ape_index_diff",
            "wins_diff_corrected",
            "losses_diff_corrected",
            "win_loss_ratio_diff_corrected",
            "pro_SLpM_diff_corrected",
            "pro_sig_str_acc_diff_corrected",
            "pro_SApM_diff_corrected",
            "pro_str_def_diff_corrected",
            "pro_total_str_pM_diff_corrected",
            "pro_total_str_acc_diff_corrected",
            "pro_total_str_absorbed_pM_diff_corrected",
            "pro_td_avg_diff_corrected",
            "pro_td_acc_diff_corrected",
            "pro_td_def_diff_corrected",
            "pro_sub_avg_diff_corrected",
            "pro_kd_pM_diff_corrected",
            "ko_rate_diff_corrected",
            "sub_rate_diff_corrected",
            "dec_rate_diff_corrected",
            "recent_form_diff_corrected",
            "head_pct_diff_corrected",
            "body_pct_diff_corrected",
            "leg_pct_diff_corrected",
            "distance_pct_diff_corrected",
            "clinch_pct_diff_corrected",
            "ground_pct_diff_corrected",
            "win_streak_diff_corrected",
            "loss_streak_diff_corrected",
            "last_5_wins_diff_corrected",
            "days_since_last_fight_diff_corrected",
            "recent_finish_rate_diff_corrected",
            "durability_diff_corrected",
            "h2h_advantage",
            "total_rounds",
            "is_title_bout",
        ]

        # Enhanced derived features
        df["net_striking_advantage"] = (
            df["r_pro_SLpM_corrected"] - df["r_pro_SApM_corrected"]
        ) - (df["b_pro_SLpM_corrected"] - df["b_pro_SApM_corrected"])
        
        df["striking_efficiency"] = (
            df["r_pro_SLpM_corrected"] * df["r_pro_sig_str_acc_corrected"]
        ) - (df["b_pro_SLpM_corrected"] * df["b_pro_sig_str_acc_corrected"])
        
        df["defensive_striking"] = (
            df["r_pro_str_def_corrected"] - df["r_pro_SApM_corrected"]
        ) - (df["b_pro_str_def_corrected"] - df["b_pro_SApM_corrected"])
        
        df["grappling_control"] = (
            df["r_pro_td_avg_corrected"] * df["r_pro_td_acc_corrected"]
        ) - (df["b_pro_td_avg_corrected"] * df["b_pro_td_acc_corrected"])
        
        df["grappling_defense"] = (
            df["r_pro_td_def_corrected"] - df["r_pro_sub_avg_corrected"] / 5
        ) - (df["b_pro_td_def_corrected"] - df["b_pro_sub_avg_corrected"] / 5)
        
        df["offensive_output"] = (
            df["r_pro_SLpM_corrected"]
            +df["r_pro_td_avg_corrected"]
            +df["r_pro_sub_avg_corrected"]
        ) - (
            df["b_pro_SLpM_corrected"]
            +df["b_pro_td_avg_corrected"]
            +df["b_pro_sub_avg_corrected"]
        )
        
        df["defensive_composite"] = (
            (df["r_pro_str_def_corrected"] + df["r_pro_td_def_corrected"]) / 2
        ) - ((df["b_pro_str_def_corrected"] + df["b_pro_td_def_corrected"]) / 2)
        
        df["ko_specialist_gap"] = (
            df["r_ko_rate_corrected"] * df["r_pro_SLpM_corrected"]
        ) - (df["b_ko_rate_corrected"] * df["b_pro_SLpM_corrected"])
        
        df["submission_specialist_gap"] = (
            df["r_sub_rate_corrected"] * df["r_pro_sub_avg_corrected"]
        ) - (df["b_sub_rate_corrected"] * df["b_pro_sub_avg_corrected"])
        
        # Absolute fight counts and experience metrics
        df["r_total_fights"] = df["r_wins_corrected"] + df["r_losses_corrected"]
        df["b_total_fights"] = df["b_wins_corrected"] + df["b_losses_corrected"]
        
        df["experience_gap"] = (
            df["r_wins_corrected"] + df["r_losses_corrected"]
        ) - (df["b_wins_corrected"] + df["b_losses_corrected"])
        
        df["experience_ratio"] = df["r_total_fights"] / (df["b_total_fights"] + 1)
        
        # Cage time per fight (experience quality metric)
        df["r_avg_fight_time"] = df["r_fight_time_minutes_corrected"] / (df["r_total_fights"] + 1)
        df["b_avg_fight_time"] = df["b_fight_time_minutes_corrected"] / (df["b_total_fights"] + 1)
        df["avg_fight_time_diff"] = df["r_avg_fight_time"] - df["b_avg_fight_time"]
        
        df["skill_momentum"] = (
            df["pro_SLpM_diff_corrected"] * df["recent_form_diff_corrected"]
        )
        
        df["finish_threat"] = (
            df["r_ko_rate_corrected"] + df["r_sub_rate_corrected"]
        ) - (df["b_ko_rate_corrected"] + df["b_sub_rate_corrected"])
        
        df["momentum_advantage"] = (
            df["win_streak_diff_corrected"] - df["loss_streak_diff_corrected"]
        )

        # New inactivity penalty
        df["inactivity_penalty"] = np.where(
            df["days_since_last_fight_diff_corrected"] > 365,
            -1,
            np.where(df["days_since_last_fight_diff_corrected"] < -365, 1, 0)
        )

        # Pace differential
        df["pace_differential"] = (
            df["r_pro_SLpM_corrected"] + df["r_pro_td_avg_corrected"]
        ) - (
            df["b_pro_SLpM_corrected"] + df["b_pro_td_avg_corrected"]
        )

        feature_columns.extend(
            [
                "net_striking_advantage",
                "striking_efficiency",
                "defensive_striking",
                "grappling_control",
                "grappling_defense",
                "offensive_output",
                "defensive_composite",
                "ko_specialist_gap",
                "submission_specialist_gap",
                "experience_gap",
                "skill_momentum",
                "finish_threat",
                "momentum_advantage",
                "inactivity_penalty",
                "pace_differential",
                "r_total_fights",
                "b_total_fights",
                "experience_ratio",
                "avg_fight_time_diff",
            ]
        )

        if "r_stance" in df.columns and "b_stance" in df.columns:
            if "stance_encoder" not in self.label_encoders:
                self.label_encoders["stance_encoder"] = LabelEncoder()
                all_stances = pd.concat(
                    [df["r_stance"], df["b_stance"]]).unique()
                self.label_encoders["stance_encoder"].fit(all_stances)

            df["r_stance_encoded"] = self.label_encoders["stance_encoder"].transform(
                df["r_stance"].fillna("Orthodox")
            )
            df["b_stance_encoded"] = self.label_encoders["stance_encoder"].transform(
                df["b_stance"].fillna("Orthodox")
            )
            df["stance_diff"] = df["r_stance_encoded"] - df["b_stance_encoded"]
            feature_columns.append("stance_diff")

        for col in feature_columns:
            if col in df.columns:
                df[col] = df[col].fillna(0)

        df = df.replace([np.inf, -np.inf], [1e6, -1e6])

        print(f"Total features: {len(feature_columns)}")
        return df, feature_columns

    def train_models(self, df):
        """Train models with better regularization"""
        df, feature_columns = self.prepare_features(df)

        X = df[feature_columns]
        y_winner = (df["winner"] == "Red").astype(int)

        if "winner_method_encoder" not in self.label_encoders:
            self.label_encoders["winner_method_encoder"] = LabelEncoder()
        y_method = self.label_encoders["winner_method_encoder"].fit_transform(
            df["winner_method_simple"]
        )

        (
            X_train,
            X_test,
            y_winner_train,
            y_winner_test,
            y_method_train,
            y_method_test,
        ) = train_test_split(
            X, y_winner, y_method, test_size=0.2, random_state=42, stratify=y_winner
        )

        print("\n" + "=" * 70)
        print("TRAINING MODELS")
        print("=" * 70)

        numeric_transformer = Pipeline(
            [
                ("imputer", SimpleImputer(strategy="median")),
                ("scaler", StandardScaler()),
            ]
        )

        preprocessor = ColumnTransformer(
            [("num", numeric_transformer, feature_columns)]
        )

        if self.use_xgboost:
            print("\nUsing XGBoost Configuration...")
            scale_pos = len(y_winner[y_winner==0]) / max(len(y_winner[y_winner==1]), 1)
            
            # winner model
            self.winner_model = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectKBest(f_classif, k=min(55, len(feature_columns))),
                    ),
                    (
                        "classifier",
                        XGBClassifier(
                            n_estimators=400,
                            max_depth=6,
                            learning_rate=0.03,
                            subsample=0.75,
                            colsample_bytree=0.7,
                            colsample_bylevel=0.7,
                            reg_alpha=0.5,
                            reg_lambda=2.5,
                            min_child_weight=7,
                            gamma=0.4,
                            scale_pos_weight=scale_pos,
                            random_state=42,
                            eval_metric="logloss",
                        ),
                    ),
                ]
            )
            
            # Calibrate for better probability estimates
            self.winner_model = CalibratedClassifierCV(
                self.winner_model,
                method='isotonic',
                cv=3
            )
            
            # method model
            self.method_model = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectKBest(f_classif, k=min(50, len(feature_columns))),
                    ),
                    (
                        "classifier",
                        XGBClassifier(
                            n_estimators=400,
                            max_depth=7,
                            learning_rate=0.04,
                            subsample=0.8,
                            colsample_bytree=0.75,
                            reg_alpha=0.3,
                            reg_lambda=1.8,
                            min_child_weight=4,
                            gamma=0.25,
                            random_state=42,
                            objective="multi:softprob",
                        ),
                    ),
                ]
            )
        else:
            print("Using Random Forest Configuration...")
            self.winner_model = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectKBest(f_classif, k=min(50, len(feature_columns))),
                    ),
                    (
                        "classifier",
                        RandomForestClassifier(
                            n_estimators=400,
                            max_depth=18,
                            min_samples_split=12,
                            min_samples_leaf=6,
                            random_state=42,
                            n_jobs=-1,
                        ),
                    ),
                ]
            )
            
            self.method_model = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectKBest(f_classif, k=min(45, len(feature_columns))),
                    ),
                    (
                        "classifier",
                        RandomForestClassifier(
                            n_estimators=400,
                            max_depth=18,
                            min_samples_split=12,
                            min_samples_leaf=6,
                            random_state=42,
                            n_jobs=-1,
                        ),
                    ),
                ]
            )

        self.winner_model.fit(X_train, y_winner_train)
        self.method_model.fit(X_train, y_method_train)

        # Time-based cross-validation
        tscv = TimeSeriesSplit(n_splits=5)
        winner_cv_scores = []
        
        for train_idx, val_idx in tscv.split(X):
            X_fold_train, X_fold_val = X.iloc[train_idx], X.iloc[val_idx]
            y_fold_train, y_fold_val = y_winner.iloc[train_idx], y_winner.iloc[val_idx]
            
            # Use base model without calibration for CV
            if self.use_xgboost:
                fold_model = Pipeline([
                    ("preprocessor", preprocessor),
                    ("feature_selector", SelectKBest(f_classif, k=min(55, len(feature_columns)))),
                    ("classifier", XGBClassifier(
                        n_estimators=400,
                        max_depth=6,
                        learning_rate=0.03,
                        subsample=0.75,
                        colsample_bytree=0.7,
                        colsample_bylevel=0.7,
                        reg_alpha=0.5,
                        reg_lambda=2.5,
                        min_child_weight=7,
                        gamma=0.4,
                        scale_pos_weight=scale_pos,
                        random_state=42,
                        eval_metric="logloss",
                    ))
                ])
            
            fold_model.fit(X_fold_train, y_fold_train)
            score = fold_model.score(X_fold_val, y_fold_val)
            winner_cv_scores.append(score)

        print(
            f"\nTime-Based CV Accuracy: {np.mean(winner_cv_scores):.4f} "
            f"(±{np.std(winner_cv_scores):.4f})"
        )
        print(f"Test Set Accuracy: {self.winner_model.score(X_test, y_winner_test):.4f}\n")

        return feature_columns

    def get_fighter_latest_stats(self, fighter_name):
        """Get latest stats for fighter"""
        red_fights = self.df_train[
            self.df_train["r_fighter"] == fighter_name
        ].sort_values("event_date", ascending=False)
        blue_fights = self.df_train[
            self.df_train["b_fighter"] == fighter_name
        ].sort_values("event_date", ascending=False)

        if len(red_fights) > 0 and len(blue_fights) > 0:
            latest = (
                red_fights.iloc[0]
                if red_fights.iloc[0]["event_date"] > blue_fights.iloc[0]["event_date"]
                else blue_fights.iloc[0]
            )
            prefix = (
                "r"
                if red_fights.iloc[0]["event_date"] > blue_fights.iloc[0]["event_date"]
                else "b"
            )
        elif len(red_fights) > 0:
            latest, prefix = red_fights.iloc[0], "r"
        elif len(blue_fights) > 0:
            latest, prefix = blue_fights.iloc[0], "b"
        else:
            return None

        stats = {
            k: latest[f"{prefix}_{k}_corrected"]
            for k in [
                "wins", "losses", "draws", "win_loss_ratio",
                "pro_SLpM", "pro_sig_str_acc", "pro_SApM", "pro_str_def",
                "pro_total_str_pM", "pro_total_str_acc", "pro_total_str_absorbed_pM",
                "pro_td_avg", "pro_td_acc", "pro_td_def", "pro_sub_avg", "pro_kd_pM",
                "ko_rate", "sub_rate", "dec_rate", "recent_form",
                "head_pct", "body_pct", "leg_pct",
                "distance_pct", "clinch_pct", "ground_pct",
                "win_streak", "loss_streak", "last_5_wins",
                "days_since_last_fight", "recent_finish_rate", "durability",
                "fight_time_minutes"
            ]
        }

        stats.update(
            {
                k: latest[f"{prefix}_{k}"]
                for k in [
                    "height", "reach", "weight", "age_at_event", "stance", "ape_index"
                ]
            }
        )

        # Update with last fight result
        if pd.notna(latest["winner"]):
            fighter_won = (latest["winner"] == "Red" and prefix == "r") or (
                latest["winner"] == "Blue" and prefix == "b"
            )
            fighter_lost = (latest["winner"] == "Red" and prefix == "b") or (
                latest["winner"] == "Blue" and prefix == "r"
            )

            if fighter_won:
                stats["wins"] += 1
            elif fighter_lost:
                stats["losses"] += 1
            else:
                stats["draws"] += 1

            stats["win_loss_ratio"] = stats["wins"] / max(stats["losses"], 1)

            # Update finish rates
            total_fights = stats["wins"] + stats["losses"]
            if total_fights > 0:
                method = str(latest["method"]).lower()
                if "ko" in method or "tko" in method:
                    method_type = "ko"
                elif "sub" in method:
                    method_type = "sub"
                else:
                    method_type = "dec"

                prev_total = total_fights - 1
                if prev_total > 0:
                    prev_ko = int(stats["ko_rate"] * prev_total)
                    prev_sub = int(stats["sub_rate"] * prev_total)
                    prev_dec = int(stats["dec_rate"] * prev_total)
                else:
                    prev_ko = prev_sub = prev_dec = 0

                if fighter_won:
                    if method_type == "ko":
                        prev_ko += 1
                    elif method_type == "sub":
                        prev_sub += 1
                    else:
                        prev_dec += 1

                stats["ko_rate"] = prev_ko / total_fights
                stats["sub_rate"] = prev_sub / total_fights
                stats["dec_rate"] = prev_dec / total_fights

        return stats

    def prepare_upcoming_fight(self, fight, feature_columns):
        """Prepare upcoming fight with enhanced features"""
        r_stats = self.get_fighter_latest_stats(fight["red_fighter"])
        b_stats = self.get_fighter_latest_stats(fight["blue_fighter"])

        if not r_stats or not b_stats:
            return None, None, None

        fight_features = {
            "height_diff": (r_stats["height"] - b_stats["height"])
            if r_stats["height"] and b_stats["height"]
            else 0,
            "reach_diff": (r_stats["reach"] - b_stats["reach"])
            if r_stats["reach"] and b_stats["reach"]
            else 0,
            "weight_diff": (r_stats["weight"] - b_stats["weight"])
            if r_stats["weight"] and b_stats["weight"]
            else 0,
            "age_at_event_diff": (r_stats["age_at_event"] - b_stats["age_at_event"])
            if r_stats["age_at_event"] and b_stats["age_at_event"]
            else 0,
            "ape_index_diff": (r_stats["ape_index"] - b_stats["ape_index"])
            if r_stats["ape_index"] and b_stats["ape_index"]
            else 0,
            **{
                f"{k}_diff_corrected": r_stats[k] - b_stats[k]
                for k in [
                    "wins", "losses", "win_loss_ratio",
                    "pro_SLpM", "pro_sig_str_acc", "pro_SApM", "pro_str_def",
                    "pro_total_str_pM", "pro_total_str_acc", "pro_total_str_absorbed_pM",
                    "pro_td_avg", "pro_td_acc", "pro_td_def", "pro_sub_avg", "pro_kd_pM",
                    "ko_rate", "sub_rate", "dec_rate", "recent_form",
                    "head_pct", "body_pct", "leg_pct",
                    "distance_pct", "clinch_pct", "ground_pct",
                    "win_streak", "loss_streak", "last_5_wins",
                    "days_since_last_fight", "recent_finish_rate", "durability",
                    "fight_time_minutes"
                ]
            },
            "h2h_advantage": 0,
            "total_rounds": fight["total_rounds"],
            "is_title_bout": 1 if fight["total_rounds"] == 5 else 0,
            
            # ADD ALL INDIVIDUAL CORRECTED STATS NEEDED FOR METHOD CALCULATION
            "r_wins_corrected": r_stats["wins"],
            "b_wins_corrected": b_stats["wins"],
            "r_losses_corrected": r_stats["losses"],
            "b_losses_corrected": b_stats["losses"],
            "r_ko_rate_corrected": r_stats["ko_rate"],
            "b_ko_rate_corrected": b_stats["ko_rate"],
            "r_sub_rate_corrected": r_stats["sub_rate"],
            "b_sub_rate_corrected": b_stats["sub_rate"],
            "r_dec_rate_corrected": r_stats["dec_rate"],
            "b_dec_rate_corrected": b_stats["dec_rate"],
            "r_pro_str_def_corrected": r_stats["pro_str_def"],
            "b_pro_str_def_corrected": b_stats["pro_str_def"],
            "r_pro_td_def_corrected": r_stats["pro_td_def"],
            "b_pro_td_def_corrected": b_stats["pro_td_def"],
            "r_durability_corrected": r_stats["durability"],
            "b_durability_corrected": b_stats["durability"],
            "r_pro_SLpM_corrected": r_stats["pro_SLpM"],
            "b_pro_SLpM_corrected": b_stats["pro_SLpM"],
            "r_pro_sig_str_acc_corrected": r_stats["pro_sig_str_acc"],
            "b_pro_sig_str_acc_corrected": b_stats["pro_sig_str_acc"],
            "r_pro_SApM_corrected": r_stats["pro_SApM"],
            "b_pro_SApM_corrected": b_stats["pro_SApM"],
            "r_pro_td_avg_corrected": r_stats["pro_td_avg"],
            "b_pro_td_avg_corrected": b_stats["pro_td_avg"],
            "r_pro_td_acc_corrected": r_stats["pro_td_acc"],
            "b_pro_td_acc_corrected": b_stats["pro_td_acc"],
            "r_pro_sub_avg_corrected": r_stats["pro_sub_avg"],
            "b_pro_sub_avg_corrected": b_stats["pro_sub_avg"],
            "r_pro_kd_pM_corrected": r_stats["pro_kd_pM"],
            "b_pro_kd_pM_corrected": b_stats["pro_kd_pM"],
            "r_head_pct_corrected": r_stats["head_pct"],
            "b_head_pct_corrected": b_stats["head_pct"],
            "r_distance_pct_corrected": r_stats["distance_pct"],
            "b_distance_pct_corrected": b_stats["distance_pct"],
            "r_clinch_pct_corrected": r_stats["clinch_pct"],
            "b_clinch_pct_corrected": b_stats["clinch_pct"],
            "r_ground_pct_corrected": r_stats["ground_pct"],
            "b_ground_pct_corrected": b_stats["ground_pct"],
            "r_fight_time_minutes_corrected": r_stats["fight_time_minutes"],
            "b_fight_time_minutes_corrected": b_stats["fight_time_minutes"],
        }

        # Enhanced derived features
        r_total_fights = r_stats["wins"] + r_stats["losses"]
        b_total_fights = b_stats["wins"] + b_stats["losses"]
        
        r_avg_fight_time = r_stats["fight_time_minutes"] / (r_total_fights + 1)
        b_avg_fight_time = b_stats["fight_time_minutes"] / (b_total_fights + 1)
        
        fight_features.update(
            {
                "net_striking_advantage": (r_stats["pro_SLpM"] - r_stats["pro_SApM"])
                -(b_stats["pro_SLpM"] - b_stats["pro_SApM"]),
                "striking_efficiency": (
                    r_stats["pro_SLpM"] * r_stats["pro_sig_str_acc"]
                )
                -(b_stats["pro_SLpM"] * b_stats["pro_sig_str_acc"]),
                "defensive_striking": (r_stats["pro_str_def"] - r_stats["pro_SApM"])
                -(b_stats["pro_str_def"] - b_stats["pro_SApM"]),
                "grappling_control": (r_stats["pro_td_avg"] * r_stats["pro_td_acc"])
                -(b_stats["pro_td_avg"] * b_stats["pro_td_acc"]),
                "grappling_defense": (
                    r_stats["pro_td_def"] - r_stats["pro_sub_avg"] / 5
                )
                -(b_stats["pro_td_def"] - b_stats["pro_sub_avg"] / 5),
                "offensive_output": (
                    r_stats["pro_SLpM"] + r_stats["pro_td_avg"] + 
                    r_stats["pro_sub_avg"]
                )
                -(
                    b_stats["pro_SLpM"] + b_stats["pro_td_avg"] + 
                    b_stats["pro_sub_avg"]
                ),
                "defensive_composite": (
                    (r_stats["pro_str_def"] + r_stats["pro_td_def"]) / 2
                )
                -((b_stats["pro_str_def"] + b_stats["pro_td_def"]) / 2),
                "ko_specialist_gap": (r_stats["ko_rate"] * r_stats["pro_SLpM"])
                -(b_stats["ko_rate"] * b_stats["pro_SLpM"]),
                "submission_specialist_gap": (
                    r_stats["sub_rate"] * r_stats["pro_sub_avg"]
                )
                -(b_stats["sub_rate"] * b_stats["pro_sub_avg"]),
                "experience_gap": r_total_fights - b_total_fights,
                "skill_momentum": (
                    (r_stats["pro_SLpM"] - b_stats["pro_SLpM"]) * 
                    (r_stats["recent_form"] - b_stats["recent_form"])
                ),
                "finish_threat": (
                    r_stats["ko_rate"] + r_stats["sub_rate"]
                ) - (b_stats["ko_rate"] + b_stats["sub_rate"]),
                "momentum_advantage": (
                    (r_stats["win_streak"] - b_stats["win_streak"]) - 
                    (r_stats["loss_streak"] - b_stats["loss_streak"])
                ),
                "inactivity_penalty": (
                    -1 if r_stats["days_since_last_fight"] - b_stats["days_since_last_fight"] > 365
                    else 1 if b_stats["days_since_last_fight"] - r_stats["days_since_last_fight"] > 365
                    else 0
                ),
                "pace_differential": (
                    r_stats["pro_SLpM"] + r_stats["pro_td_avg"]
                ) - (
                    b_stats["pro_SLpM"] + b_stats["pro_td_avg"]
                ),
                "r_total_fights": r_total_fights,
                "b_total_fights": b_total_fights,
                "experience_ratio": r_total_fights / (b_total_fights + 1),
                "avg_fight_time_diff": r_avg_fight_time - b_avg_fight_time,
            }
        )

        if (
            "stance_encoder" in self.label_encoders
            and r_stats["stance"]
            and b_stats["stance"]
        ):
            r_enc = self.label_encoders["stance_encoder"].transform(
                [r_stats["stance"]]
            )[0]
            b_enc = self.label_encoders["stance_encoder"].transform(
                [b_stats["stance"]]
            )[0]
            fight_features["stance_diff"] = r_enc - b_enc
        else:
            fight_features["stance_diff"] = 0

        return fight_features, r_stats, b_stats

    def calculate_comprehensive_method_adjustments(self, fight_data, winner_prefix, loser_prefix):
        """Comprehensive method prediction using all available stats"""
        
        # Extract all relevant stats
        w_slpm = fight_data[f"{winner_prefix}_pro_SLpM_corrected"].values[0]
        w_sig_acc = fight_data[f"{winner_prefix}_pro_sig_str_acc_corrected"].values[0]
        w_td_avg = fight_data[f"{winner_prefix}_pro_td_avg_corrected"].values[0]
        w_td_acc = fight_data[f"{winner_prefix}_pro_td_acc_corrected"].values[0]
        w_sub_avg = fight_data[f"{winner_prefix}_pro_sub_avg_corrected"].values[0]
        w_ko_rate = fight_data[f"{winner_prefix}_ko_rate_corrected"].values[0]
        w_sub_rate = fight_data[f"{winner_prefix}_sub_rate_corrected"].values[0]
        w_dec_rate = fight_data[f"{winner_prefix}_dec_rate_corrected"].values[0]
        w_kd_rate = fight_data[f"{winner_prefix}_pro_kd_pM_corrected"].values[0]
        
        l_str_def = fight_data[f"{loser_prefix}_pro_str_def_corrected"].values[0]
        l_sapm = fight_data[f"{loser_prefix}_pro_SApM_corrected"].values[0]
        l_td_def = fight_data[f"{loser_prefix}_pro_td_def_corrected"].values[0]
        l_durability = fight_data[f"{loser_prefix}_durability_corrected"].values[0]
        
        w_head_pct = fight_data[f"{winner_prefix}_head_pct_corrected"].values[0]
        w_distance_pct = fight_data[f"{winner_prefix}_distance_pct_corrected"].values[0]
        w_clinch_pct = fight_data[f"{winner_prefix}_clinch_pct_corrected"].values[0]
        w_ground_pct = fight_data[f"{winner_prefix}_ground_pct_corrected"].values[0]
        
        l_distance_pct = fight_data[f"{loser_prefix}_distance_pct_corrected"].values[0]
        l_ground_pct = fight_data[f"{loser_prefix}_ground_pct_corrected"].values[0]
        
        total_rounds = fight_data["total_rounds"].values[0]
        
        # KO/TKO PROBABILITY
        ko_base = w_ko_rate
        
        striking_volume_factor = min(w_slpm / 6.0, 1.5)
        accuracy_factor = 1 + (w_sig_acc - 0.45) * 2
        accuracy_factor = max(0.5, min(accuracy_factor, 1.8))
        
        head_hunting_factor = 1 + (w_head_pct - 0.5) * 0.6
        head_hunting_factor = max(0.7, min(head_hunting_factor, 1.5))
        
        distance_factor = 1 + (w_distance_pct - 0.6) * 0.5
        distance_factor = max(0.8, min(distance_factor, 1.4))
        
        # Knockdown threat factor
        kd_threat_factor = 1 + min(w_kd_rate / 0.5, 1.2)
        
        opp_vulnerability = (
            (1 - l_str_def) * 0.4 +
            (l_sapm / 6.0) * 0.3 +
            (1 - l_durability) * 0.3
        )
        opp_vulnerability = min(opp_vulnerability, 1.0)
        
        power_differential = (w_slpm * w_sig_acc) - (l_sapm * (1 - l_str_def))
        power_differential_factor = 1 + max(0, power_differential / 5.0)
        power_differential_factor = min(power_differential_factor, 1.8)
        
        ko_prob = ko_base * (
            striking_volume_factor * 0.20 +
            accuracy_factor * 0.18 +
            head_hunting_factor * 0.12 +
            distance_factor * 0.08 +
            opp_vulnerability * 0.18 +
            power_differential_factor * 0.08 +
            kd_threat_factor * 0.16
        )
        
        if total_rounds == 5:
            ko_prob *= 1.15
        
        ko_prob = min(ko_prob, 0.85)
        
        # SUBMISSION PROBABILITY
        sub_base = w_sub_rate
        
        grappling_control = w_td_avg * w_td_acc
        control_factor = 1 + min(grappling_control / 2.0, 1.0)
        
        sub_attempt_factor = 1 + (w_sub_avg / 2.0)
        sub_attempt_factor = min(sub_attempt_factor, 2.0)
        
        ground_preference_factor = 1 + (w_ground_pct - 0.2) * 0.8
        ground_preference_factor = max(0.6, min(ground_preference_factor, 1.6))
        
        opp_grappling_weakness = (
            (1 - l_td_def) * 0.5 +
            (1 - l_durability) * 0.3 +
            l_ground_pct * 0.2
        )
        
        td_differential = (w_td_avg * w_td_acc) - (l_td_def * 2.0)
        td_factor = 1 + max(0, td_differential / 3.0)
        td_factor = min(td_factor, 1.7)
        
        sub_prob = sub_base * (
            control_factor * 0.25 +
            sub_attempt_factor * 0.20 +
            ground_preference_factor * 0.15 +
            opp_grappling_weakness * 0.25 +
            td_factor * 0.15
        )
        
        if total_rounds == 5:
            sub_prob *= 1.12
        
        sub_prob = min(sub_prob, 0.80)
        
        # DECISION PROBABILITY
        dec_base = w_dec_rate
        
        finish_threat = ko_prob + sub_prob
        
        defensive_composite = (l_str_def + l_td_def) / 2.0
        defensive_factor = 1 + defensive_composite * 0.5
        
        volume_no_finish = (w_slpm / 6.0) * (1 - w_ko_rate)
        volume_decision_factor = 1 + min(volume_no_finish, 0.6)
        
        style_variety = 1 - abs(w_distance_pct - 0.5) * 2
        variety_factor = 1 + style_variety * 0.3
        
        dec_prob = dec_base * (
            defensive_factor * 0.35 +
            volume_decision_factor * 0.25 +
            variety_factor * 0.20 +
            1.0 * 0.20
        )
        
        dec_prob *= (1 - finish_threat * 0.5)
        
        if total_rounds == 5:
            dec_prob *= 0.95
        
        # STYLE MATCHUP ADJUSTMENTS
        striker_advantage = (w_distance_pct - w_ground_pct) - (l_distance_pct - l_ground_pct)
        
        if striker_advantage > 0.3:
            ko_prob *= 1.15
            sub_prob *= 0.85
        elif striker_advantage < -0.3:
            sub_prob *= 1.15
            ko_prob *= 0.85
        
        pressure_differential = (w_slpm - l_sapm) - (l_str_def * 5)
        if pressure_differential > 2:
            ko_prob *= 1.10
            dec_prob *= 0.95
        
        # NORMALIZATION
        total = ko_prob + sub_prob + dec_prob
        if total > 0:
            ko_prob /= total
            sub_prob /= total
            dec_prob /= total
        
        return {
            'ko_prob': ko_prob,
            'sub_prob': sub_prob,
            'dec_prob': dec_prob,
            'finish_threat': ko_prob + sub_prob
        }

    def predict_fight(self, fight_data, feature_columns):
        """Enhanced fight prediction with comprehensive method logic"""
        X = fight_data[feature_columns]
        winner_proba = self.winner_model.predict_proba(X)[0]
        winner_pred = self.winner_model.predict(X)[0]
        winner_name = "Red" if winner_pred == 1 else "Blue"

        method_proba = self.method_model.predict_proba(X)[0]
        method_labels = self.label_encoders["winner_method_encoder"].classes_

        winner_prefix = "r" if winner_name == "Red" else "b"
        loser_prefix = "b" if winner_name == "Red" else "r"
        
        # Use comprehensive method calculation
        method_calcs = self.calculate_comprehensive_method_adjustments(
            fight_data, winner_prefix, loser_prefix
        )
        
        # Context-based weighting between model and comprehensive analysis
        total_rounds = fight_data["total_rounds"].values[0]
        if total_rounds == 5:
            model_weight = 0.40
            analysis_weight = 0.60
        else:
            model_weight = 0.45
            analysis_weight = 0.55
        
        # Build final probabilities
        final_method_probs = {}
        for i, label in enumerate(method_labels):
            if label.startswith(winner_name):
                method_type = label.split("_")[1]
                
                if method_type == "KO/TKO":
                    analysis_prob = method_calcs['ko_prob']
                elif method_type == "Submission":
                    analysis_prob = method_calcs['sub_prob']
                else:  # Decision
                    analysis_prob = method_calcs['dec_prob']
                
                # Blend model prediction with comprehensive analysis
                final_method_probs[label] = (
                    method_proba[i] * model_weight + analysis_prob * analysis_weight
                )

        # Normalize
        total = sum(final_method_probs.values())
        if total > 0:
            final_method_probs = {k: v / total for k, v in final_method_probs.items()}

        final_method = max(final_method_probs, key=final_method_probs.get)

        return {
            "winner": winner_name,
            "winner_confidence": winner_proba[winner_pred],
            "red_prob": winner_proba[1],
            "blue_prob": winner_proba[0],
            "method": final_method.split("_")[1],
            "method_probabilities": final_method_probs,
            "combined_method_prob": final_method_probs[final_method],
        }

    def predict_upcoming_fights(self, upcoming_fights, feature_columns):
        """Predict upcoming fights and return predictions + skipped fights"""
        predictions = []
        skipped_fights = []

        for fight in upcoming_fights:
            result = self.prepare_upcoming_fight(fight, feature_columns)
            if not result[0]:
                skipped_fights.append(f"{fight['red_fighter']} vs {fight['blue_fighter']}")
                continue

            fight_features, r_stats, b_stats = result
            fight_df = pd.DataFrame([fight_features])

            for col in feature_columns:
                if col not in fight_df.columns:
                    fight_df[col] = 0

            pred = self.predict_fight(fight_df, feature_columns)

            winner_prefix = pred['winner']
            ko_prob = pred['method_probabilities'].get(f"{winner_prefix}_KO/TKO", 0)
            sub_prob = pred['method_probabilities'].get(f"{winner_prefix}_Submission", 0)
            dec_prob = pred['method_probabilities'].get(f"{winner_prefix}_Decision", 0)

            predictions.append(
                {
                    "Red Fighter": fight["red_fighter"],
                    "Blue Fighter": fight["blue_fighter"],
                    "Weight Class": fight["weight_class"],
                    "Winner": fight["red_fighter"]
                    if pred["winner"] == "Red"
                    else fight["blue_fighter"],
                    "Win%": pred["winner_confidence"],
                    "Method": pred["method"],
                    "Method%": pred["combined_method_prob"],
                    "KO/TKO%": ko_prob,
                    "Submission%": sub_prob,
                    "Decision%": dec_prob,
                }
            )

        return predictions, skipped_fights

    def export_predictions_to_excel(self, predictions, filename="ufc_predictions_.xlsx"):
        """Export predictions to formatted Excel file"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        df = pd.DataFrame(predictions)
        df.to_excel(filename, index=False, engine="openpyxl")

        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        header_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        header_font = Font(bold=True, size=11)
        left_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = left_alignment
            cell.border = thin_border

        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.alignment = left_alignment
                cell.border = thin_border

        for row_idx in range(2, ws.max_row + 1):
            for col_idx in [5, 7, 8, 9, 10]:
                ws.cell(row=row_idx, column=col_idx).number_format = "0.00%"

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            header_length = 0

            for idx, cell in enumerate(column):
                try:
                    if cell.value is None:
                        cell_length = 0
                    elif isinstance(cell.value, (int, float)):
                        if cell.number_format == "0.00%":
                            cell_length = 7
                        else:
                            cell_length = len(str(cell.value))
                    else:
                        cell_length = len(str(cell.value))
                        if idx == 0:
                            header_length = cell_length
                    
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass

            if header_length > 0:
                adjusted_width = max(max_length + 2, int(header_length) + 2)
            else:
                adjusted_width = max_length + 2
            
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filename)
        print(f"\n Predictions exported to: {filename}")

        return filename



class UFCPredictorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("UFC Fight Predictor")
        self.root.geometry("1000x800")
        self.root.minsize(700, 550)
        
        # Variables
        self.data_file_path = tk.StringVar(value="fight_data.csv")
        self.output_file_path = tk.StringVar(value="UFC_predictions.xlsx")
        
        # Predictor will be initialized when running
        self.predictor = None
        
        # Create UI
        self.create_widgets()
        
    def create_widgets(self):
        # Title with UFC red color
        title_frame = tk.Frame(self.root, bg='#D20A0A')
        title_frame.pack(fill=tk.X)
        tk.Label(title_frame, text="UFC FIGHT PREDICTOR", 
                font=('Arial', 16, 'bold'), fg='white', bg='#D20A0A').pack(pady=(10, 8))
        
        # Data file selection
        file_frame = ttk.LabelFrame(self.root, text="Data File", padding="5")
        file_frame.pack(fill=tk.X, padx=10, pady=3)
        
        ttk.Entry(file_frame, textvariable=self.data_file_path, width=70).pack(side=tk.LEFT, padx=3)
        ttk.Button(file_frame, text="Browse", command=self.browse_data_file).pack(side=tk.LEFT)
        
        # Fight input section
        input_frame = ttk.LabelFrame(self.root, text="Enter Fights to Predict", padding="5")
        input_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=3)
        
        instructions = """Enter fights in CSV format (one per line):
Format: red_fighter,blue_fighter,weight_class,gender,total_rounds

Example:
Max Holloway,Dustin Poirier,Lightweight,Men,5
Ilia Topuria,Charles Oliveira,Lightweight,Men,5
Tatiana Suarez,Amanda Lemos,Strawweight,Women,3"""
        ttk.Label(input_frame, text=instructions, font=('Arial', 8), foreground="gray", justify=tk.LEFT).pack(anchor=tk.W, pady=(0, 3))
        
        self.fights_text = scrolledtext.ScrolledText(input_frame, height=8, width=95, wrap=tk.WORD)
        self.fights_text.pack(fill=tk.BOTH, expand=True, pady=3)
        
        
        # Output file
        output_frame = ttk.LabelFrame(self.root, text="Output File", padding="5")
        output_frame.pack(fill=tk.X, padx=10, pady=3)
        
        ttk.Entry(output_frame, textvariable=self.output_file_path, width=70).pack(side=tk.LEFT, padx=3)
        ttk.Button(output_frame, text="Browse", command=self.browse_output_file).pack(side=tk.LEFT)
        
        # Buttons
        button_frame = ttk.Frame(self.root, padding="5")
        button_frame.pack(fill=tk.X, padx=10)
        
        ttk.Button(button_frame, text="Load Sample", command=self.load_sample).pack(side=tk.LEFT, padx=3)
        ttk.Button(button_frame, text="Clear", command=self.clear_input).pack(side=tk.LEFT, padx=3)
        ttk.Button(button_frame, text="Generate Predictions", command=self.run_predictions).pack(side=tk.RIGHT, padx=3)
        
        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)
        
    def browse_data_file(self):
        filename = filedialog.askopenfilename(
            title="Select Fight Data CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filename:
            self.data_file_path.set(filename)
            
    def browse_output_file(self):
        filename = filedialog.asksaveasfilename(
            title="Save Predictions As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.output_file_path.set(filename)
            
    def load_sample(self):
        sample = '''Max Holloway,Dustin Poirier,Lightweight,Men,5
Ilia Topuria,Charles Oliveira,Lightweight,Men,5
Tatiana Suarez,Amanda Lemos,Strawweight,Women,3'''
        self.fights_text.delete('1.0', tk.END)
        self.fights_text.insert('1.0', sample)
        
    def clear_input(self):
        self.fights_text.delete('1.0', tk.END)
        
    def parse_fights_input(self, text):
        """Parse CSV input text to extract fight data"""
        text = text.strip()
        fights = []
        
        for line in text.split('\n'):
            line = line.strip()
            if not line:
                continue
                
            # Split by comma
            parts = [p.strip() for p in line.split(',')]
            
            if len(parts) != 5:
                raise ValueError(f"Invalid CSV format. Expected 5 fields, got {len(parts)} in line: {line}")
            
            try:
                fights.append({
                    'red_fighter': parts[0],
                    'blue_fighter': parts[1],
                    'weight_class': parts[2],
                    'gender': parts[3],
                    'total_rounds': int(parts[4])
                })
            except ValueError as e:
                raise ValueError(f"Error parsing line: {line}\nTotal rounds must be a number. {str(e)}")
        
        if not fights:
            raise ValueError("No valid fight data found. Please check the format.")
            
        return fights
        

        
    def run_predictions(self):
        try:
            self.status_var.set("Loading data and training models... Please wait...")
            self.root.update()
            
            # Get input
            fights_text = self.fights_text.get('1.0', tk.END)
            upcoming_fights = self.parse_fights_input(fights_text)
            
            if not upcoming_fights:
                raise ValueError("No fights entered")
                
            data_file = self.data_file_path.get()
            if not os.path.exists(data_file):
                raise FileNotFoundError(f"Data file not found: {data_file}")
            
            # Load data
            df = pd.read_csv(data_file)
            
            self.status_var.set(f"Loaded {len(df)} fights. Training models...")
            self.root.update()
            
            # Initialize predictor and train (suppress console output for EXE)
            self.predictor = UFCPredictor(use_xgboost=True)
            
            # Suppress print statements during training
            f = io.StringIO()
            with redirect_stdout(f):
                df = self.predictor.fix_data_leakage(df)
                self.predictor.df_train = df
                feature_columns = self.predictor.train_models(df)
            
            self.status_var.set("Generating predictions...")
            self.root.update()
            
            # Make predictions - now returns skipped fights too
            predictions, skipped_fights = self.predictor.predict_upcoming_fights(upcoming_fights, feature_columns)
            
            # Check if we have any predictions
            if not predictions:
                self.status_var.set("No predictions generated - all fights skipped")
                messagebox.showerror(
                    "No Predictions", 
                    "No predictions could be generated. All fights were skipped due to insufficient fighter data.\n\n"
                    "Please ensure the fighters exist in your training data."
                )
                return
            
            
            # Export to Excel
            output_file = self.output_file_path.get()
            
            # Suppress console output during export
            with redirect_stdout(io.StringIO()):
                self.predictor.export_predictions_to_excel(predictions, output_file)
            
            # Success message with skipped fights warning if applicable
            success_msg = f"Predictions generated successfully!\n\nSaved to: {output_file}\n\n"
            success_msg += f"📊 {len(predictions)} fight(s) predicted"
            
            if skipped_fights:
                success_msg += f"\n⚠️ {len(skipped_fights)} fight(s) skipped due to insufficient data"
                self.status_var.set(f"Success! {len(predictions)} predictions saved, {len(skipped_fights)} skipped")
                
                # Show detailed skipped fights
                skipped_msg = success_msg + "\n\n" + "Skipped Fights (insufficient fighter data):\n" + "\n".join(f"• {fight}" for fight in skipped_fights)
                messagebox.showwarning("Predictions Complete", skipped_msg)
            else:
                self.status_var.set(f"Success! All {len(predictions)} predictions saved to {output_file}")
                messagebox.showinfo("Success", success_msg)
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            self.status_var.set("Error occurred")
            messagebox.showerror("Error", f"An error occurred:\n\n{str(e)}\n\nDetails:\n{error_details}")


def main():
    root = tk.Tk()
    app = UFCPredictorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()