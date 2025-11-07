import sys
import os
import random
import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog, messagebox
import pandas as pd
import numpy as np
import io
from contextlib import redirect_stdout
from sklearn.model_selection import train_test_split, TimeSeriesSplit, StratifiedKFold
from sklearn.ensemble import RandomForestClassifier, VotingClassifier
from sklearn.feature_selection import SelectPercentile, f_classif
from sklearn.metrics import log_loss, accuracy_score, roc_auc_score, classification_report
from sklearn.calibration import CalibratedClassifierCV
from sklearn.linear_model import LogisticRegression
# permutation_importance removed - using model built-in importance instead
from scipy.stats import linregress
import warnings
import shutil
import atexit
from joblib import Parallel, delayed
import multiprocessing as mp
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Enable multiprocessing for faster training
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"

# Get the directory where this script is located
if getattr(sys, "frozen", False):
    script_dir = os.path.dirname(sys.executable)
else:
    script_dir = os.path.dirname(os.path.abspath(__file__))

fight_data_path = os.path.join(script_dir, "fight_data.csv")
if not os.path.exists(fight_data_path):
    print(f"Warning: fight_data.csv not found in script directory: {script_dir}")

# Set random seeds for reproducibility
random.seed(42)
np.random.seed(42)
os.environ["PYTHONHASHSEED"] = "42"

warnings.filterwarnings("ignore")

def cleanup_temp_files():
    """Clean up temporary files and folders created during training"""
    try:
        if os.path.exists("catboost_info"):
            shutil.rmtree("catboost_info")
        if os.path.exists("best_dl_model.h5"):
            os.remove("best_dl_model.h5")
    except Exception as e:
        print(f"Warning: Could not clean up temporary files: {e}")

atexit.register(cleanup_temp_files)

# Try to import optional libraries
try:
    from xgboost import XGBClassifier
    HAS_XGBOOST = True
except ImportError:
    HAS_XGBOOST = False
    print("XGBoost not available.")

try:
    from lightgbm import LGBMClassifier
    HAS_LIGHTGBM = True
except ImportError:
    HAS_LIGHTGBM = False
    print("LightGBM not available.")

try:
    import optuna
    HAS_OPTUNA = True
    optuna.logging.set_verbosity(optuna.logging.WARNING)
except ImportError:
    HAS_OPTUNA = False
    print("Optuna not available. Install with: pip install optuna")


class ImprovedUFCPredictor:
    """
    Improved UFC Fight Predictor with optimizations for accuracy:
    - Reduced feature set (50-60 core features vs 200+)
    - Conservative data augmentation (35% vs 90%)
    - Permutation-based feature selection
    - Optuna hyperparameter optimization
    - Proper validation with hold-out test set
    - Calibrated probability estimates
    """

    def __init__(self, use_ensemble=True, debug_mode=False):
        self.winner_model = None
        self.label_encoders = {}
        self.use_ensemble = use_ensemble
        self.debug_mode = debug_mode
        self.df_train = None
        self.feature_importance = None
        self.best_params = None
        self.calibrated_model = None
        self.feature_columns = None  # Store feature columns for predictions
        self.fighter_elos = {}  # Store fighter ELO ratings
        self.elo_history = {}  # Store ELO history for predictions

        # Set random seeds
        self.set_random_seeds()

    def set_random_seeds(self):
        """Set all random seeds for reproducibility"""
        random.seed(42)
        np.random.seed(42)
        os.environ["PYTHONHASHSEED"] = "42"

    # ===== CORE FEATURES: ~111 RESEARCH-BACKED FEATURES =====
    # Original 94 + 17 new features from RF/GBDT/SVM top-10 analysis

    def get_core_feature_names(self):
        """Define ~111 features including research-backed top-performers (Date, CLINCH, GROUND, CTRL, REV)"""
        return [
            # TIER 0: ELO Features (3 features) - MOST IMPORTANT
            "elo_diff",
            "r_elo_pre_fight",
            "b_elo_pre_fight",

            # TIER 1: Core Performance (15 features)
            "recent_form_diff_corrected",
            "win_streak_diff_corrected",
            "loss_streak_diff_corrected",
            "net_striking_advantage",
            "striking_efficiency",
            "defensive_striking",
            "grappling_control",
            "finish_rate_diff",
            "win_loss_ratio_diff_corrected",
            "last_5_wins_diff_corrected",
            "pro_SLpM_diff_corrected",
            "pro_SApM_diff_corrected",
            "pro_td_avg_diff_corrected",
            "ko_rate_diff_corrected",
            "sub_rate_diff_corrected",

            # TIER 2: Physical/Style (10 features)
            "reach_diff",
            "age_at_event_diff",
            "stance_matchup_advantage",
            "striker_vs_grappler",
            "experience_gap",
            "height_diff",
            "ape_index_diff",
            "weight_diff",
            "striker_advantage",
            "grappler_advantage",

            # TIER 3: Context (8 features)
            "opponent_quality_diff",
            "days_since_last_fight_diff_corrected",
            "is_title_bout",
            "total_rounds",
            "r_trajectory_3",
            "b_trajectory_3",
            "ring_rust_factor",
            "weight_class_factor",

            # TIER 4: Advanced (15-20 features)
            "momentum_swing",
            "style_clash_severity",
            "power_vs_technique",
            "finish_pressure",
            "upset_potential",
            "pro_str_def_diff_corrected",
            "pro_td_def_diff_corrected",
            "pro_sig_str_acc_diff_corrected",
            "recent_finish_rate_diff_corrected",
            "durability_diff_corrected",
            "h2h_advantage",
            "clutch_factor_diff",
            "momentum_quality_diff",
            "pressure_performance_diff",
            "form_consistency_diff",

            # Additional high-value features
            "offensive_output",
            "defensive_composite",
            "ko_specialist_gap",
            "submission_specialist_gap",
            "skill_momentum",

            # Interaction features (combining key metrics)
            "elo_x_form",
            "reach_x_striking",
            "experience_x_age",

            # Advanced interaction features (data-driven from importance analysis)
            "age_x_striking",
            "age_x_grappling",
            "age_x_durability",
            "td_x_defense",
            "grappling_x_experience",
            "striking_x_accuracy",
            "striking_x_defense",
            "form_x_experience",
            "finish_x_momentum",
            "height_x_reach",
            "physical_x_striking",

            # Elite interaction features (based on top 20 analysis)
            "elo_x_win_ratio",
            "win_ratio_x_form",
            "durability_x_striking",
            "elo_x_durability",
            "submission_x_grappling",
            "ko_power_x_striking",
            "momentum_x_win_streak",
            "streak_differential",

            # NEW: Strategic interactions based on 60.79% model analysis
            "age_x_win_streak",
            "elo_x_sub_threat",
            "form_x_durability",
            "striking_x_grappling_matchup",
            "momentum_combo",

            # NEW HIGH-IMPACT: Elite compound features
            "elite_finisher",
            "veteran_advantage",
            "complete_fighter",
            "total_finish_threat",
            "unstoppable_streak",
            "age_prime_advantage",
            "freshness_advantage",
            "desperation_diff",
            "momentum_sustainability_diff",
            "elo_x_finish",
            "adversity_experience_diff",

            # RESEARCH-BACKED: Top features from RF/GBDT/SVM analysis
            # NEW: Opponent quality (was placeholder returning 0!)
            "opponent_quality_diff",
            "avg_opponent_elo_diff_corrected",

            # NEW: Recent momentum vs career average (improving/declining fighters)
            "distance_pct_momentum_diff_corrected",
            "slpm_momentum_diff_corrected",
            "ctrl_sec_momentum_diff_corrected",

            # NEW: Rounds-based strategy adjustments
            "rounds_x_cardio",
            "rounds_x_finish_rate",
            "rounds_x_durability",

            # NEW: Enhanced top feature interactions
            "h2h_x_elo",
            "h2h_x_form",
            "age_prime_score_diff",
            "age_x_experience",
            "win_ratio_x_finish",
            "win_ratio_x_durability",

            # Positional striking CAREER AVERAGES (CLINCH #3, GROUND #5, DISTANCE #6)
            "distance_pct_diff_corrected",
            "clinch_pct_diff_corrected",
            "ground_pct_diff_corrected",
            "positional_striking_advantage",

            # Target area CAREER AVERAGES distribution
            "head_pct_diff_corrected",
            "body_pct_diff_corrected",
            "leg_pct_diff_corrected",
            "target_distribution_advantage",

            # Control and reversals CAREER AVERAGES (CTRL #5 GBDT, REV #8/#9)
            "avg_ctrl_sec_diff_corrected",
            "avg_rev_diff_corrected",
            "control_dominance",

            # Compound features with career-based positionals
            "clinch_x_grappling",
            "distance_x_striking",
            "ground_x_control",
            "positional_mastery",
        ]

    def calculate_streak(self, recent_wins, count_wins=True):
        """Calculate current win or loss streak"""
        if not recent_wins:
            return 0
        streak = 0
        target = 1 if count_wins else 0
        for result in reversed(recent_wins):
            if result == target:
                streak += 1
            else:
                break
        return streak

    def calculate_trajectory(self, values):
        """Calculate linear trend of recent performance"""
        if len(values) < 2:
            return 0
        try:
            x = np.arange(len(values))
            slope, _, _, _, _ = linregress(x, values)
            return slope
        except:
            return 0

    def calculate_age_curve_factor(self, age):
        """Calculate performance multiplier based on age"""
        if age < 25:
            return 0.92
        elif 25 <= age <= 32:
            return 1.0
        elif 33 <= age <= 36:
            return 0.94
        else:
            return 0.85

    # ===== ELO RATING SYSTEM (COPIED FROM UFC ELO SYSTEM) =====

    def elo_expected_score(self, fighter_elo, opponent_elo):
        """
        Calculate expected score using ELO formula
        Returns probability of winning (0-1)
        """
        return 1 / (1 + 10 ** ((opponent_elo - fighter_elo) / 400))

    def elo_calculate_k_factor(self, row, winner_fighter, fighter_result, fighter_elos, elo_history):
        """
        Calculate dynamic K-factor based on fight circumstances

        K-factor determines how much ELO changes after a fight.
        Higher K-factor means more volatile ratings.
        """
        k_base = 40  # Increased K-factor for wider ELO spread and better differentiation
        k = k_base

        # 1. CHAMPIONSHIP BOUT WEIGHT (title fights matter more)
        is_title = row.get('is_title_bout', 0) == 1
        if is_title:
            k *= 2.0
            if fighter_result == 'Win':
                k *= 1.5  # Winning title fights is even more valuable

        # 2. RECENCY WEIGHT (more recent fights are weighted higher)
        # Calculate days since fight
        if 'event_date' in row.index and pd.notna(row['event_date']):
            from datetime import datetime
            event_date = pd.to_datetime(row['event_date'])
            days_ago = (datetime.now() - event_date).days
            # Apply exponential decay: fights in last year = 100%, 2 years ago = 75%, etc.
            recency_multiplier = np.exp(-days_ago / 365 / 2)  # Half-life of 2 years
            if recency_multiplier < 0.5:  # Cap at 50% for very old fights
                recency_multiplier = 0.5
            k *= recency_multiplier

        # 3. DOMINANCE BONUS (finishing early matters more)
        if 'finish_round' in row.index and pd.notna(row['finish_round']):
            round_finished = row['finish_round']
            if round_finished <= 1:
                k *= 1.5  # First round finish
            elif round_finished == 2:
                k *= 1.3  # Second round finish
            elif round_finished == 3:
                k *= 1.2  # Third round finish

        # 4. OPPONENT STRENGTH ADJUSTMENT
        opponent = row['b_fighter'] if row['r_fighter'] == winner_fighter else row['r_fighter']
        fighter_elo = fighter_elos.get(winner_fighter, 1500)

        if opponent in fighter_elos:
            opponent_elo = fighter_elos[opponent]

            # Bonus for beating above-average opponents
            if opponent_elo > 1500:
                if opponent_elo > 1600:
                    k *= 1.2
                if opponent_elo > 1700:
                    k *= 1.2

            # Upset bonuses/penalties
            elo_diff = fighter_elo - opponent_elo
            if fighter_result == 'Win' and elo_diff < -100:
                # Upset bonus: beating much higher rated opponent
                upset_factor = abs(elo_diff) / 100 * 0.1
                k *= (1 + upset_factor)
            elif fighter_result == 'Loss' and elo_diff > 100:
                # Upset penalty: losing to much lower rated opponent
                upset_penalty = elo_diff / 100 * 0.15
                k *= (1 + upset_penalty)

        # 5. WIN/LOSS STREAK MULTIPLIER
        fighter_streak = self.elo_get_recent_streak(winner_fighter, elo_history)
        if fighter_streak >= 3:  # Win streak
            k *= 1.15
        elif fighter_streak <= -3:  # Loss streak
            k *= 1.25

        # 6. METHOD OF VICTORY BONUS
        if fighter_result == 'Win' and 'method' in row.index and pd.notna(row['method']):
            method = str(row['method']).upper()
            if 'KO' in method or 'TKO' in method:
                k *= 1.1  # Finishing via KO is more impressive
            elif 'SUBMISSION' in method:
                k *= 1.05  # Submissions are also dominant

        return k

    def elo_get_recent_streak(self, fighter, elo_history):
        """Get fighter's recent win/loss streak from ELO history"""
        if fighter not in elo_history or len(elo_history[fighter]) == 0:
            return 0

        # Look at last 5 results
        history = elo_history[fighter][-5:]

        if len(history) == 0:
            return 0

        streak = 0
        for entry in reversed(history):
            result = entry.get('result', None)
            if result == 'Win':
                streak = (streak if streak > 0 else 0) + 1
            elif result == 'Loss':
                streak = (streak if streak < 0 else 0) - 1
            else:
                streak = 0

        return streak

    def elo_update_ratings(self, r_fighter, b_fighter, winner, k_factor, fighter_elos, elo_history, fight_date):
        """
        Update ELO ratings for both fighters after a fight

        Args:
            r_fighter: Red corner fighter name
            b_fighter: Blue corner fighter name
            winner: 'Red', 'Blue', or 'Draw'
            k_factor: K-factor for this fight
            fighter_elos: Dict of current ELO ratings
            elo_history: Dict of fighter history
            fight_date: Date of the fight
        """
        # Get current ELOs
        r_elo = fighter_elos.get(r_fighter, 1500)
        b_elo = fighter_elos.get(b_fighter, 1500)

        # Calculate expected scores
        r_expected = self.elo_expected_score(r_elo, b_elo)
        b_expected = self.elo_expected_score(b_elo, r_elo)

        # Determine actual scores based on winner
        if winner == 'Red':
            r_actual = 1.0
            b_actual = 0.0
            r_result = 'Win'
            b_result = 'Loss'
        elif winner == 'Blue':
            r_actual = 0.0
            b_actual = 1.0
            r_result = 'Loss'
            b_result = 'Win'
        elif winner == 'Draw':
            r_actual = 0.5
            b_actual = 0.5
            r_result = 'Draw'
            b_result = 'Draw'
            k_factor *= 0.5  # Draws have less impact
        else:
            # No Contest or invalid
            return

        # Calculate ELO changes
        r_change = k_factor * (r_actual - r_expected)
        b_change = k_factor * (b_actual - b_expected)

        # Update ELO ratings
        fighter_elos[r_fighter] = r_elo + r_change
        fighter_elos[b_fighter] = b_elo + b_change

        # Record in history
        if r_fighter not in elo_history:
            elo_history[r_fighter] = []
        if b_fighter not in elo_history:
            elo_history[b_fighter] = []

        elo_history[r_fighter].append({
            'date': fight_date,
            'elo': fighter_elos[r_fighter],
            'result': r_result
        })
        elo_history[b_fighter].append({
            'date': fight_date,
            'elo': fighter_elos[b_fighter],
            'result': b_result
        })

    def calculate_elo_ratings(self, df):
        """
        Calculate ELO ratings for all fighters chronologically
        This ensures no data leakage - each fight uses pre-fight ELO ratings

        Returns:
            df: DataFrame with ELO columns added (r_elo_pre_fight, b_elo_pre_fight, elo_diff)
        """
        print("\n" + "="*80)
        print("CALCULATING ELO RATINGS (CHRONOLOGICAL - NO DATA LEAKAGE)")
        print("="*80 + "\n")

        # Initialize ELO system
        initial_elo = 1500
        fighter_elos = {}
        elo_history = {}

        # Sort by date to process chronologically
        df = df.sort_values('event_date').reset_index(drop=True)

        # Initialize ELO columns
        df['r_elo_pre_fight'] = 0.0
        df['b_elo_pre_fight'] = 0.0
        df['elo_diff'] = 0.0
        df['r_prob_pre_fight'] = 0.0
        df['b_prob_pre_fight'] = 0.0

        for idx, row in df.iterrows():
            if idx % 500 == 0:
                print(f"   Processing fight {idx}/{len(df)} for ELO...")

            r_fighter = row['r_fighter']
            b_fighter = row['b_fighter']

            # Initialize fighters if not seen before
            if r_fighter not in fighter_elos:
                fighter_elos[r_fighter] = initial_elo
                elo_history[r_fighter] = []
            if b_fighter not in fighter_elos:
                fighter_elos[b_fighter] = initial_elo
                elo_history[b_fighter] = []

            # CRITICAL: Capture PRE-FIGHT ELO ratings (before this fight is processed)
            r_elo_pre = fighter_elos[r_fighter]
            b_elo_pre = fighter_elos[b_fighter]

            # Calculate pre-fight probabilities
            r_prob = self.elo_expected_score(r_elo_pre, b_elo_pre)
            b_prob = self.elo_expected_score(b_elo_pre, r_elo_pre)

            # Store pre-fight ELO ratings and probabilities
            df.at[idx, 'r_elo_pre_fight'] = r_elo_pre
            df.at[idx, 'b_elo_pre_fight'] = b_elo_pre
            df.at[idx, 'elo_diff'] = r_elo_pre - b_elo_pre
            df.at[idx, 'r_prob_pre_fight'] = r_prob
            df.at[idx, 'b_prob_pre_fight'] = b_prob

            # Now update ELO ratings based on fight result
            winner = row.get('winner')

            if pd.notna(winner) and winner in ['Red', 'Blue', 'Draw']:
                # Determine who won for K-factor calculation
                if winner == 'Red':
                    winner_fighter = r_fighter
                    fighter_result = 'Win'
                elif winner == 'Blue':
                    winner_fighter = b_fighter
                    fighter_result = 'Win'
                else:
                    winner_fighter = r_fighter
                    fighter_result = 'Draw'

                # Calculate K-factor
                k_factor = self.elo_calculate_k_factor(
                    row, winner_fighter, fighter_result, fighter_elos, elo_history
                )

                # Update ELO ratings (this modifies fighter_elos for future fights)
                self.elo_update_ratings(
                    r_fighter, b_fighter, winner, k_factor,
                    fighter_elos, elo_history, row['event_date']
                )

        print("\nELO calculation complete!")
        print(f"   Processed {len(df)} fights")
        print(f"   Tracked {len(fighter_elos)} unique fighters")
        print(f"   Average ELO: {np.mean(list(fighter_elos.values())):.1f}")
        print(f"   ELO Range: {min(fighter_elos.values()):.1f} - {max(fighter_elos.values()):.1f}")

        # Store fighter ELOs for prediction use
        self.fighter_elos = fighter_elos
        self.elo_history = elo_history

        return df

    # ===== DATA AUGMENTATION: REDUCED FROM 90% TO 35% =====

    def augment_data_conservatively(self, df):
        """
        Conservative data augmentation to eliminate bias without overfitting
        Using 35% augmentation (same as Model 5) for optimal generalization
        """
        print("\n" + "="*80)
        print("OPTIMIZED DATA AUGMENTATION (80% augmentation)")
        print("="*80)

        red_bias = (df["winner"] == "Red").mean()
        print(f"Original Red corner win rate: {red_bias:.3f}")

        # Only augment if bias > 3%
        if abs(red_bias - 0.5) > 0.03:
            print(f"Bias detected ({abs(red_bias - 0.5):.3f}), applying conservative augmentation...")

            # Create augmented dataset by swapping corners
            df_augmented = self.swap_corners(df)

            # OPTIMIZED: 80% augmentation - sweet spot for balance without overfitting
            sample_size = int(len(df_augmented) * 0.80)
            df_augmented_sampled = df_augmented.sample(n=sample_size, random_state=42)

            # Combine original and augmented data
            df_combined = pd.concat([df, df_augmented_sampled], ignore_index=True)

            new_red_bias = (df_combined["winner"] == "Red").mean()
            print(f"After augmentation: {len(df)} -> {len(df_combined)} fights (1.80x)")
            print(f"New Red corner win rate: {new_red_bias:.3f}")

            return df_combined
        else:
            print("Bias is minimal (<3%), no augmentation needed")
            return df

    def swap_corners(self, df):
        """Swap red and blue corners to create augmented data"""
        print("Creating corner-swapped augmented data...")

        df_swapped = df.copy()

        # Swap basic info
        for col in df.columns:
            if col.startswith("r_"):
                b_col = "b_" + col[2:]
                if b_col in df.columns:
                    df_swapped[col], df_swapped[b_col] = df[b_col].copy(), df[col].copy()

        # Swap winner
        df_swapped["winner"] = df["winner"].replace({"Red": "Blue", "Blue": "Red"})

        # Swap differential features (negate them)
        for col in df.columns:
            if "_diff" in col or "_advantage" in col or "_gap" in col:
                df_swapped[col] = -df[col]

        # Negate interaction features (they're products of differentials, so they're also differentials)
        interaction_features = [
            "elo_x_form", "reach_x_striking", "experience_x_age",
            "age_x_striking", "age_x_grappling", "age_x_durability",
            "td_x_defense", "grappling_x_experience",
            "striking_x_accuracy", "striking_x_defense",
            "form_x_experience", "finish_x_momentum",
            "height_x_reach", "physical_x_striking",
            # Elite interaction features
            "elo_x_win_ratio", "win_ratio_x_form",
            "durability_x_striking", "elo_x_durability",
            "submission_x_grappling", "ko_power_x_striking",
            "momentum_x_win_streak", "streak_differential",
            # NEW: Strategic interactions
            "age_x_win_streak", "elo_x_sub_threat",
            "form_x_durability", "striking_x_grappling_matchup",
            "momentum_combo",
            # NEW HIGH-IMPACT: Compound interactions (products of differentials)
            "elite_finisher", "complete_fighter",
            "total_finish_threat", "unstoppable_streak", "elo_x_finish",
            # RESEARCH-BACKED: Positional compound interactions
            "clinch_x_grappling", "distance_x_striking", "ground_x_control",
            "positional_mastery", "control_dominance",
            # NEW: Enhanced top feature interactions
            "h2h_x_elo", "h2h_x_form", "age_prime_score_diff",
            "age_x_experience", "win_ratio_x_finish", "win_ratio_x_durability"
        ]
        for col in interaction_features:
            if col in df_swapped.columns:
                df_swapped[col] = -df_swapped[col]

        # Note: r_prob_pre_fight and b_prob_pre_fight are already swapped by the r_/b_ swap loop above
        # No need to swap them again (that would double-swap them back to original)

        print(f"Corner swapping complete. Created {len(df_swapped)} augmented fights.")
        return df_swapped

    # ===== FIGHTER-SPECIFIC ADVANCED FEATURES =====

    def calculate_fighter_vs_similar_opponents(self, df, fighter_col):
        """How has fighter performed against similar style opponents?"""
        performance_scores = []

        for idx, row in df.iterrows():
            fighter = row[fighter_col]
            if pd.isna(fighter):
                performance_scores.append(0.5)
                continue

            try:
                # Get all fights for this fighter
                fighter_fights = df[(df['r_fighter'] == fighter) | (df['b_fighter'] == fighter)]

                if len(fighter_fights) < 3:
                    performance_scores.append(0.5)
                    continue

                # Calculate recent performance vs striker/grappler types
                recent_fights = fighter_fights.tail(5)
                wins_vs_strikers = 0
                total_vs_strikers = 0

                for _, fight in recent_fights.iterrows():
                    opponent = fight['b_fighter'] if fight['r_fighter'] == fighter else fight['r_fighter']
                    opponent_prefix = 'b' if fight['r_fighter'] == fighter else 'r'

                    # Check if opponent is striker or grappler
                    slpm = fight.get(f'{opponent_prefix}_pro_SLpM_corrected', 0)
                    td_avg = fight.get(f'{opponent_prefix}_pro_td_avg_corrected', 0)

                    if slpm > td_avg * 2:  # Striker
                        total_vs_strikers += 1
                        if fight.get('winner') == ('Red' if fight['r_fighter'] == fighter else 'Blue'):
                            wins_vs_strikers += 1

                if total_vs_strikers > 0:
                    performance_scores.append(wins_vs_strikers / total_vs_strikers)
                else:
                    performance_scores.append(0.5)

            except:
                performance_scores.append(0.5)

        return np.array(performance_scores)

    def calculate_elite_competition_performance(self, df, fighter_col):
        """Recent performance vs elite competition (top-ranked opponents)"""
        elite_performance = []

        for idx, row in df.iterrows():
            fighter = row[fighter_col]
            if pd.isna(fighter):
                elite_performance.append(0.5)
                continue

            try:
                # Get fighter's recent fights
                fighter_fights = df[(df['r_fighter'] == fighter) | (df['b_fighter'] == fighter)]
                recent_fights = fighter_fights.tail(5)

                elite_wins = 0
                elite_fights_count = 0

                for _, fight in recent_fights.iterrows():
                    opponent = fight['b_fighter'] if fight['r_fighter'] == fighter else fight['r_fighter']
                    opponent_prefix = 'b' if fight['r_fighter'] == fighter else 'r'

                    # Consider elite if win rate > 70%
                    opponent_wins = fight.get(f'{opponent_prefix}_wins_corrected', 0)
                    opponent_losses = fight.get(f'{opponent_prefix}_losses_corrected', 1)
                    opponent_winrate = opponent_wins / max(opponent_wins + opponent_losses, 1)

                    if opponent_winrate > 0.7:  # Elite opponent
                        elite_fights_count += 1
                        if fight.get('winner') == ('Red' if fight['r_fighter'] == fighter else 'Blue'):
                            elite_wins += 1

                if elite_fights_count >= 2:
                    elite_performance.append(elite_wins / elite_fights_count)
                else:
                    elite_performance.append(0.5)

            except:
                elite_performance.append(0.5)

        return np.array(elite_performance)

    def calculate_fighter_improvement_rate(self, df, fighter_col):
        """Calculate if fighter is improving or declining"""
        improvement_rates = []

        for idx, row in df.iterrows():
            fighter = row[fighter_col]
            if pd.isna(fighter):
                improvement_rates.append(0.0)
                continue

            try:
                # Get all fights for this fighter
                fighter_fights = df[(df['r_fighter'] == fighter) | (df['b_fighter'] == fighter)]

                if len(fighter_fights) < 6:
                    improvement_rates.append(0.0)
                    continue

                # Compare first 1/3 vs last 1/3 of career
                third = len(fighter_fights) // 3
                early_fights = fighter_fights.head(third)
                recent_fights = fighter_fights.tail(third)

                # Calculate win rates
                early_wins = 0
                recent_wins = 0

                for _, fight in early_fights.iterrows():
                    if fight.get('winner') == ('Red' if fight['r_fighter'] == fighter else 'Blue'):
                        early_wins += 1

                for _, fight in recent_fights.iterrows():
                    if fight.get('winner') == ('Red' if fight['r_fighter'] == fighter else 'Blue'):
                        recent_wins += 1

                early_winrate = early_wins / len(early_fights)
                recent_winrate = recent_wins / len(recent_fights)

                improvement_rates.append(recent_winrate - early_winrate)

            except:
                improvement_rates.append(0.0)

        return np.array(improvement_rates)

    # ===== DATA LEAKAGE PREVENTION =====

    def fix_data_leakage(self, df):
        """Recalculate comprehensive fighter statistics chronologically"""
        print("\n" + "="*80)
        print("FIXING DATA LEAKAGE - CHRONOLOGICAL RECALCULATION")
        print("="*80 + "\n")

        import copy

        df["event_date"] = pd.to_datetime(df["event_date"])
        df = df.sort_values("event_date").reset_index(drop=True)

        fighter_stats = {}

        stats_to_track = {
            "wins": 0, "losses": 0, "draws": 0,
            "sig_str_total": 0, "sig_str_att_total": 0, "sig_str_absorbed_total": 0,
            "td_total": 0, "td_att_total": 0, "td_def_success": 0, "td_def_att": 0,
            "sub_att_total": 0, "kd_total": 0,
            "fight_time_minutes": 0, "fight_count": 0,
            "ko_wins": 0, "sub_wins": 0, "dec_wins": 0,
            "ko_losses": 0, "sub_losses": 0,
            "recent_wins": [], "recent_finishes": [],
            "last_fight_date": None,
            # NEW: Positional striking tracking (distance/clinch/ground)
            "distance_strikes": 0, "clinch_strikes": 0, "ground_strikes": 0, "total_sig_strikes": 0,
            # NEW: Target area tracking (head/body/leg)
            "head_strikes": 0, "body_strikes": 0, "leg_strikes": 0,
            # NEW: Control and reversals
            "ctrl_sec_total": 0, "rev_total": 0,
            # NEW: Opponent quality tracking
            "total_opponent_elo": 0, "opponents_faced": 0,
            # NEW: Recent stats for momentum (last 3 fights)
            "recent_distance_pct": [], "recent_clinch_pct": [], "recent_ground_pct": [],
            "recent_slpm": [], "recent_ctrl_sec": [],
        }

        # Initialize corrected columns
        for prefix in ["r", "b"]:
            for stat in ["wins", "losses", "win_loss_ratio", "pro_SLpM", "pro_SApM",
                         "pro_sig_str_acc", "pro_str_def", "pro_td_avg", "pro_td_acc",
                         "pro_td_def", "pro_sub_avg", "ko_rate", "sub_rate", "dec_rate",
                         "recent_form", "win_streak", "loss_streak", "last_5_wins",
                         "days_since_last_fight", "recent_finish_rate", "durability",
                         # NEW: Career-based positional/target/control metrics
                         "distance_pct", "clinch_pct", "ground_pct",
                         "head_pct", "body_pct", "leg_pct",
                         "avg_ctrl_sec", "avg_rev",
                         # NEW: Opponent quality
                         "avg_opponent_elo",
                         # NEW: Recent momentum vs career
                         "distance_pct_momentum", "slpm_momentum", "ctrl_sec_momentum"]:
                df[f"{prefix}_{stat}_corrected"] = 0.0

        df["h2h_advantage"] = 0.0
        df["opponent_quality_diff"] = 0.0
        fighter_h2h = {}

        for idx, row in df.iterrows():
            if idx % 500 == 0:
                print(f"   Processing fight {idx}/{len(df)}...")

            r_fighter, b_fighter = row["r_fighter"], row["b_fighter"]

            # H2H tracking
            h2h_key = (r_fighter, b_fighter)
            h2h_key_reverse = (b_fighter, r_fighter)

            if h2h_key in fighter_h2h:
                df.at[idx, "h2h_advantage"] = fighter_h2h[h2h_key]
            elif h2h_key_reverse in fighter_h2h:
                df.at[idx, "h2h_advantage"] = -fighter_h2h[h2h_key_reverse]

            if r_fighter not in fighter_stats:
                fighter_stats[r_fighter] = copy.deepcopy(stats_to_track)
            if b_fighter not in fighter_stats:
                fighter_stats[b_fighter] = copy.deepcopy(stats_to_track)

            # Record stats before fight
            for fighter, prefix in [(r_fighter, "r"), (b_fighter, "b")]:
                stats = fighter_stats[fighter]
                df.at[idx, f"{prefix}_wins_corrected"] = stats["wins"]
                df.at[idx, f"{prefix}_losses_corrected"] = stats["losses"]
                df.at[idx, f"{prefix}_win_loss_ratio_corrected"] = stats["wins"] / max(stats["losses"], 1)

                total_fights = stats["wins"] + stats["losses"]
                if total_fights > 0:
                    df.at[idx, f"{prefix}_ko_rate_corrected"] = stats["ko_wins"] / total_fights
                    df.at[idx, f"{prefix}_sub_rate_corrected"] = stats["sub_wins"] / total_fights
                    df.at[idx, f"{prefix}_dec_rate_corrected"] = stats["dec_wins"] / total_fights

                # Recent form
                if len(stats["recent_wins"]) >= 5:
                    df.at[idx, f"{prefix}_recent_form_corrected"] = sum(stats["recent_wins"][-5:]) / 5
                elif len(stats["recent_wins"]) >= 3:
                    df.at[idx, f"{prefix}_recent_form_corrected"] = sum(stats["recent_wins"][-3:]) / 3

                # Streaks
                df.at[idx, f"{prefix}_win_streak_corrected"] = self.calculate_streak(stats["recent_wins"], True)
                df.at[idx, f"{prefix}_loss_streak_corrected"] = self.calculate_streak(stats["recent_wins"], False)

                # Days since last fight
                if stats["last_fight_date"]:
                    days_off = (row["event_date"] - stats["last_fight_date"]).days
                    df.at[idx, f"{prefix}_days_since_last_fight_corrected"] = days_off

                # Per-minute stats
                if stats["fight_time_minutes"] > 0:
                    df.at[idx, f"{prefix}_pro_SLpM_corrected"] = stats["sig_str_total"] / stats["fight_time_minutes"]
                    df.at[idx, f"{prefix}_pro_SApM_corrected"] = stats["sig_str_absorbed_total"] / stats["fight_time_minutes"]
                    df.at[idx, f"{prefix}_pro_td_avg_corrected"] = (stats["td_total"] / stats["fight_time_minutes"]) * 15
                    df.at[idx, f"{prefix}_pro_sub_avg_corrected"] = (stats["sub_att_total"] / stats["fight_time_minutes"]) * 15

                if stats["sig_str_att_total"] > 0:
                    df.at[idx, f"{prefix}_pro_sig_str_acc_corrected"] = stats["sig_str_total"] / stats["sig_str_att_total"]

                if stats["td_att_total"] > 0:
                    df.at[idx, f"{prefix}_pro_td_acc_corrected"] = stats["td_total"] / stats["td_att_total"]

                if stats["td_def_att"] > 0:
                    df.at[idx, f"{prefix}_pro_td_def_corrected"] = stats["td_def_success"] / stats["td_def_att"]

                # Durability
                finish_losses = stats["ko_losses"] + stats["sub_losses"]
                df.at[idx, f"{prefix}_durability_corrected"] = 1.0 / (1 + finish_losses)

                # NEW: Career-based positional striking percentages
                if stats["total_sig_strikes"] > 0:
                    df.at[idx, f"{prefix}_distance_pct_corrected"] = stats["distance_strikes"] / stats["total_sig_strikes"]
                    df.at[idx, f"{prefix}_clinch_pct_corrected"] = stats["clinch_strikes"] / stats["total_sig_strikes"]
                    df.at[idx, f"{prefix}_ground_pct_corrected"] = stats["ground_strikes"] / stats["total_sig_strikes"]
                    df.at[idx, f"{prefix}_head_pct_corrected"] = stats["head_strikes"] / stats["total_sig_strikes"]
                    df.at[idx, f"{prefix}_body_pct_corrected"] = stats["body_strikes"] / stats["total_sig_strikes"]
                    df.at[idx, f"{prefix}_leg_pct_corrected"] = stats["leg_strikes"] / stats["total_sig_strikes"]
                else:
                    # Defaults based on typical UFC averages
                    df.at[idx, f"{prefix}_distance_pct_corrected"] = 0.60  # Most strikes from distance
                    df.at[idx, f"{prefix}_clinch_pct_corrected"] = 0.25
                    df.at[idx, f"{prefix}_ground_pct_corrected"] = 0.15
                    df.at[idx, f"{prefix}_head_pct_corrected"] = 0.55
                    df.at[idx, f"{prefix}_body_pct_corrected"] = 0.25
                    df.at[idx, f"{prefix}_leg_pct_corrected"] = 0.20

                # NEW: Career-based control time and reversals per fight
                if stats["fight_count"] > 0:
                    df.at[idx, f"{prefix}_avg_ctrl_sec_corrected"] = stats["ctrl_sec_total"] / stats["fight_count"]
                    df.at[idx, f"{prefix}_avg_rev_corrected"] = stats["rev_total"] / stats["fight_count"]
                else:
                    df.at[idx, f"{prefix}_avg_ctrl_sec_corrected"] = 0
                    df.at[idx, f"{prefix}_avg_rev_corrected"] = 0

                # NEW: Opponent quality (average ELO of opponents faced)
                if stats["opponents_faced"] > 0:
                    df.at[idx, f"{prefix}_avg_opponent_elo_corrected"] = stats["total_opponent_elo"] / stats["opponents_faced"]
                else:
                    df.at[idx, f"{prefix}_avg_opponent_elo_corrected"] = 1500  # Default ELO

                # NEW: Recent momentum vs career average
                # Distance percentage momentum
                if len(stats["recent_distance_pct"]) >= 2:
                    recent_avg = sum(stats["recent_distance_pct"][-3:]) / len(stats["recent_distance_pct"][-3:])
                    career_avg = stats["distance_strikes"] / stats["total_sig_strikes"] if stats["total_sig_strikes"] > 0 else 0.6
                    df.at[idx, f"{prefix}_distance_pct_momentum_corrected"] = recent_avg - career_avg

                # Striking output momentum
                if len(stats["recent_slpm"]) >= 2:
                    recent_avg = sum(stats["recent_slpm"][-3:]) / len(stats["recent_slpm"][-3:])
                    career_avg = stats["sig_str_total"] / stats["fight_time_minutes"] if stats["fight_time_minutes"] > 0 else 0
                    df.at[idx, f"{prefix}_slpm_momentum_corrected"] = recent_avg - career_avg

                # Control time momentum
                if len(stats["recent_ctrl_sec"]) >= 2:
                    recent_avg = sum(stats["recent_ctrl_sec"][-3:]) / len(stats["recent_ctrl_sec"][-3:])
                    career_avg = stats["ctrl_sec_total"] / stats["fight_count"] if stats["fight_count"] > 0 else 0
                    df.at[idx, f"{prefix}_ctrl_sec_momentum_corrected"] = recent_avg - career_avg

            # Calculate opponent quality differential
            r_opp_elo = df.at[idx, "r_avg_opponent_elo_corrected"]
            b_opp_elo = df.at[idx, "b_avg_opponent_elo_corrected"]
            df.at[idx, "opponent_quality_diff"] = r_opp_elo - b_opp_elo

            # Update stats after fight
            if pd.notna(row["winner"]):
                method = str(row.get("method", "")).lower()
                method_cat = "ko" if "ko" in method or "tko" in method else "sub" if "sub" in method else "dec"
                is_finish = method_cat in ["ko", "sub"]

                if row["winner"] == "Red":
                    fighter_stats[r_fighter]["wins"] += 1
                    fighter_stats[b_fighter]["losses"] += 1
                    fighter_stats[r_fighter][f"{method_cat}_wins"] += 1
                    fighter_stats[r_fighter]["recent_wins"].append(1)
                    fighter_stats[b_fighter]["recent_wins"].append(0)
                    if is_finish:
                        fighter_stats[b_fighter][f"{method_cat}_losses"] += 1
                    fighter_h2h[(r_fighter, b_fighter)] = fighter_h2h.get((r_fighter, b_fighter), 0) + 1
                elif row["winner"] == "Blue":
                    fighter_stats[b_fighter]["wins"] += 1
                    fighter_stats[r_fighter]["losses"] += 1
                    fighter_stats[b_fighter][f"{method_cat}_wins"] += 1
                    fighter_stats[b_fighter]["recent_wins"].append(1)
                    fighter_stats[r_fighter]["recent_wins"].append(0)
                    if is_finish:
                        fighter_stats[r_fighter][f"{method_cat}_losses"] += 1
                    fighter_h2h[(r_fighter, b_fighter)] = fighter_h2h.get((r_fighter, b_fighter), 0) - 1

                # Trim recent history
                for fighter in [r_fighter, b_fighter]:
                    if len(fighter_stats[fighter]["recent_wins"]) > 10:
                        fighter_stats[fighter]["recent_wins"] = fighter_stats[fighter]["recent_wins"][-10:]
                    fighter_stats[fighter]["last_fight_date"] = row["event_date"]

                # Update fight stats
                for fighter, f_prefix, opp_prefix in [(r_fighter, "r", "b"), (b_fighter, "b", "r")]:
                    if pd.notna(row.get(f"{f_prefix}_sig_str")):
                        fighter_stats[fighter]["sig_str_total"] += row[f"{f_prefix}_sig_str"]
                    if pd.notna(row.get(f"{f_prefix}_sig_str_att")):
                        fighter_stats[fighter]["sig_str_att_total"] += row[f"{f_prefix}_sig_str_att"]
                    if pd.notna(row.get(f"{opp_prefix}_sig_str")):
                        fighter_stats[fighter]["sig_str_absorbed_total"] += row[f"{opp_prefix}_sig_str"]
                    if pd.notna(row.get(f"{f_prefix}_td")):
                        fighter_stats[fighter]["td_total"] += row[f"{f_prefix}_td"]
                    if pd.notna(row.get(f"{f_prefix}_td_att")):
                        fighter_stats[fighter]["td_att_total"] += row[f"{f_prefix}_td_att"]
                    if pd.notna(row.get(f"{f_prefix}_sub_att")):
                        fighter_stats[fighter]["sub_att_total"] += row[f"{f_prefix}_sub_att"]

                    fight_time = row.get("total_fight_time_sec", 0) / 60 if pd.notna(row.get("total_fight_time_sec")) else 0
                    fighter_stats[fighter]["fight_time_minutes"] += fight_time
                    fighter_stats[fighter]["fight_count"] += 1

                    # NEW: Track positional striking (distance/clinch/ground)
                    sig_str = row.get(f"{f_prefix}_sig_str", 0)
                    if pd.notna(sig_str) and sig_str > 0:
                        fighter_stats[fighter]["total_sig_strikes"] += sig_str

                        # Positional percentages from this fight
                        distance_pct = row.get(f"{f_prefix}_distance", 0) if pd.notna(row.get(f"{f_prefix}_distance")) else 0
                        clinch_pct = row.get(f"{f_prefix}_clinch", 0) if pd.notna(row.get(f"{f_prefix}_clinch")) else 0
                        ground_pct = row.get(f"{f_prefix}_ground", 0) if pd.notna(row.get(f"{f_prefix}_ground")) else 0

                        # Add strikes from each position (percentage x total strikes)
                        fighter_stats[fighter]["distance_strikes"] += sig_str * distance_pct
                        fighter_stats[fighter]["clinch_strikes"] += sig_str * clinch_pct
                        fighter_stats[fighter]["ground_strikes"] += sig_str * ground_pct

                        # NEW: Track target areas (head/body/leg)
                        head_pct = row.get(f"{f_prefix}_head", 0) if pd.notna(row.get(f"{f_prefix}_head")) else 0
                        body_pct = row.get(f"{f_prefix}_body", 0) if pd.notna(row.get(f"{f_prefix}_body")) else 0
                        leg_pct = row.get(f"{f_prefix}_leg", 0) if pd.notna(row.get(f"{f_prefix}_leg")) else 0

                        fighter_stats[fighter]["head_strikes"] += sig_str * head_pct
                        fighter_stats[fighter]["body_strikes"] += sig_str * body_pct
                        fighter_stats[fighter]["leg_strikes"] += sig_str * leg_pct

                    # NEW: Track control time and reversals
                    if pd.notna(row.get(f"{f_prefix}_ctrl_sec")):
                        fighter_stats[fighter]["ctrl_sec_total"] += row[f"{f_prefix}_ctrl_sec"]
                    if pd.notna(row.get(f"{f_prefix}_rev")):
                        fighter_stats[fighter]["rev_total"] += row[f"{f_prefix}_rev"]

                # NEW: Track opponent ELO for quality metric (after both fighters updated)
                # Red corner tracks blue's ELO, blue tracks red's ELO
                if hasattr(self, 'fighter_elos'):
                    r_elo = self.fighter_elos.get(r_fighter, 1500)
                    b_elo = self.fighter_elos.get(b_fighter, 1500)

                    fighter_stats[r_fighter]["total_opponent_elo"] += b_elo
                    fighter_stats[r_fighter]["opponents_faced"] += 1
                    fighter_stats[b_fighter]["total_opponent_elo"] += r_elo
                    fighter_stats[b_fighter]["opponents_faced"] += 1

                # NEW: Track recent stats for momentum (last 3 fights)
                for fighter, f_prefix in [(r_fighter, "r"), (b_fighter, "b")]:
                    # Track recent positional percentages
                    if fighter_stats[fighter]["total_sig_strikes"] > 0:
                        this_fight_distance = fighter_stats[fighter]["distance_strikes"] / fighter_stats[fighter]["total_sig_strikes"]
                        fighter_stats[fighter]["recent_distance_pct"].append(this_fight_distance)
                        if len(fighter_stats[fighter]["recent_distance_pct"]) > 3:
                            fighter_stats[fighter]["recent_distance_pct"] = fighter_stats[fighter]["recent_distance_pct"][-3:]

                    # Track recent striking output
                    if fight_time > 0:
                        this_fight_slpm = row.get(f"{f_prefix}_sig_str", 0) / (fight_time / 60) if pd.notna(row.get(f"{f_prefix}_sig_str")) else 0
                        fighter_stats[fighter]["recent_slpm"].append(this_fight_slpm)
                        if len(fighter_stats[fighter]["recent_slpm"]) > 3:
                            fighter_stats[fighter]["recent_slpm"] = fighter_stats[fighter]["recent_slpm"][-3:]

                    # Track recent control time
                    this_fight_ctrl = row.get(f"{f_prefix}_ctrl_sec", 0) if pd.notna(row.get(f"{f_prefix}_ctrl_sec")) else 0
                    fighter_stats[fighter]["recent_ctrl_sec"].append(this_fight_ctrl)
                    if len(fighter_stats[fighter]["recent_ctrl_sec"]) > 3:
                        fighter_stats[fighter]["recent_ctrl_sec"] = fighter_stats[fighter]["recent_ctrl_sec"][-3:]

        print("Data leakage fixed successfully!\n")

        # Calculate ELO ratings chronologically (no data leakage)
        df = self.calculate_elo_ratings(df)

        return df

    def prepare_core_features(self, df):
        """Prepare only the core 50-60 features"""
        print("\n" + "="*80)
        print("PREPARING CORE FEATURES (50-60 vs previous 200+)")
        print("="*80 + "\n")

        # Filter to valid winners only
        if "winner" in df.columns:
            df = df[df["winner"].isin(["Red", "Blue"])].copy()

        # Create method mapping
        method_mapping = {
            "KO/TKO": "KO/TKO", "Submission": "Submission",
            "Decision - Unanimous": "Decision", "Decision - Split": "Decision",
            "Decision - Majority": "Decision", "TKO - Doctor's Stoppage": "KO/TKO",
            "Could Not Continue": "KO/TKO", "DQ": "Decision", "Overturned": "Decision",
        }

        if "method" in df.columns:
            df["method_simple"] = df["method"].map(method_mapping).fillna("Decision")
        else:
            df["method_simple"] = "Decision"

        if "winner" not in df.columns:
            df["winner"] = "Red"

        df["winner_method_simple"] = df["winner"] + "_" + df["method_simple"]

        # Calculate differential features
        for stat in ["wins", "losses", "win_loss_ratio", "pro_SLpM", "pro_SApM",
                     "pro_sig_str_acc", "pro_str_def", "pro_td_avg", "pro_td_acc",
                     "pro_td_def", "pro_sub_avg", "ko_rate", "sub_rate", "dec_rate",
                     "recent_form", "win_streak", "loss_streak", "last_5_wins",
                     "days_since_last_fight", "recent_finish_rate", "durability",
                     # NEW: Career-based positional/target/control metrics
                     "distance_pct", "clinch_pct", "ground_pct",
                     "head_pct", "body_pct", "leg_pct",
                     "avg_ctrl_sec", "avg_rev",
                     # NEW: Opponent quality and momentum
                     "avg_opponent_elo", "distance_pct_momentum", "slpm_momentum", "ctrl_sec_momentum"]:
            df[f"{stat}_diff_corrected"] = df[f"r_{stat}_corrected"] - df[f"b_{stat}_corrected"]

        # Physical differentials
        for stat in ["height", "reach", "weight", "age_at_event", "ape_index"]:
            if f"r_{stat}" in df.columns and f"b_{stat}" in df.columns:
                df[f"{stat}_diff"] = df[f"r_{stat}"] - df[f"b_{stat}"]

        # Core derived features
        df["net_striking_advantage"] = (df["r_pro_SLpM_corrected"] - df["r_pro_SApM_corrected"]) - (df["b_pro_SLpM_corrected"] - df["b_pro_SApM_corrected"])
        df["striking_efficiency"] = (df["r_pro_SLpM_corrected"] * df["r_pro_sig_str_acc_corrected"]) - (df["b_pro_SLpM_corrected"] * df["b_pro_sig_str_acc_corrected"])
        df["defensive_striking"] = (df["r_pro_str_def_corrected"] - df["r_pro_SApM_corrected"]) - (df["b_pro_str_def_corrected"] - df["b_pro_SApM_corrected"])
        df["grappling_control"] = (df["r_pro_td_avg_corrected"] * df["r_pro_td_acc_corrected"]) - (df["b_pro_td_avg_corrected"] * df["b_pro_td_acc_corrected"])

        # Experience features
        df["r_total_fights"] = df["r_wins_corrected"] + df["r_losses_corrected"]
        df["b_total_fights"] = df["b_wins_corrected"] + df["b_losses_corrected"]
        df["experience_gap"] = df["r_total_fights"] - df["b_total_fights"]

        # Finish rates
        df["r_finish_rate"] = (df["r_ko_rate_corrected"] + df["r_sub_rate_corrected"]) / 2
        df["b_finish_rate"] = (df["b_ko_rate_corrected"] + df["b_sub_rate_corrected"]) / 2
        df["finish_rate_diff"] = df["r_finish_rate"] - df["b_finish_rate"]

        # Style features
        df["striker_vs_grappler"] = 0  # Simplified for now

        df["r_striker_score"] = df["r_pro_SLpM_corrected"] - df["r_pro_td_avg_corrected"]
        df["b_striker_score"] = df["b_pro_SLpM_corrected"] - df["b_pro_td_avg_corrected"]
        df["striker_advantage"] = df["r_striker_score"] - df["b_striker_score"]

        df["r_grappler_score"] = df["r_pro_td_avg_corrected"] + df["r_pro_sub_avg_corrected"]
        df["b_grappler_score"] = df["b_pro_td_avg_corrected"] + df["b_pro_sub_avg_corrected"]
        df["grappler_advantage"] = df["r_grappler_score"] - df["b_grappler_score"]

        # Stance matchup
        df["stance_matchup_advantage"] = 0  # Will be calculated if stance data available

        # Fighter trajectory (simplified)
        df["r_trajectory_3"] = 0
        df["b_trajectory_3"] = 0

        # Context features
        df["ring_rust_factor"] = 0
        df["weight_class_factor"] = 1.0

        if "is_title_bout" not in df.columns:
            df["is_title_bout"] = 0
        if "total_rounds" not in df.columns:
            df["total_rounds"] = 3

        # NEW: Rounds-based strategy adjustments
        # 5-round fights favor cardio/decision specialists, 3-round favor explosive finishers
        df["rounds_x_cardio"] = df["total_rounds"] * df["dec_rate_diff_corrected"]
        df["rounds_x_finish_rate"] = (5 - df["total_rounds"]) * df["finish_rate_diff"]  # 3-rounders favor finishers
        df["rounds_x_durability"] = df["total_rounds"] * df["durability_diff_corrected"]  # 5-rounders favor durable fighters

        # Momentum features
        df["momentum_swing"] = df["r_recent_form_corrected"] - df["b_recent_form_corrected"] + (df["r_win_streak_corrected"] - df["b_win_streak_corrected"]) * 0.1
        df["style_clash_severity"] = abs(df["striker_advantage"] * 0.5 + df["grappler_advantage"] * 0.5)
        df["power_vs_technique"] = (df["r_pro_SLpM_corrected"] - df["b_pro_SLpM_corrected"]) * 0.6 + (df["r_pro_sig_str_acc_corrected"] - df["b_pro_sig_str_acc_corrected"]) * 0.4
        df["finish_pressure"] = (df["r_ko_rate_corrected"] - df["b_ko_rate_corrected"]) * 0.5 + (df["r_sub_rate_corrected"] - df["b_sub_rate_corrected"]) * 0.3
        df["upset_potential"] = (df["b_recent_form_corrected"] - df["r_recent_form_corrected"]) * 0.4 + (df["b_win_streak_corrected"] - df["r_win_streak_corrected"]) * 0.3

        # Advanced features
        df["offensive_output"] = (df["r_pro_SLpM_corrected"] + df["r_pro_td_avg_corrected"] + df["r_pro_sub_avg_corrected"]) - (df["b_pro_SLpM_corrected"] + df["b_pro_td_avg_corrected"] + df["b_pro_sub_avg_corrected"])
        df["defensive_composite"] = ((df["r_pro_str_def_corrected"] + df["r_pro_td_def_corrected"]) / 2) - ((df["b_pro_str_def_corrected"] + df["b_pro_td_def_corrected"]) / 2)
        df["ko_specialist_gap"] = (df["r_ko_rate_corrected"] * df["r_pro_SLpM_corrected"]) - (df["b_ko_rate_corrected"] * df["b_pro_SLpM_corrected"])
        df["submission_specialist_gap"] = (df["r_sub_rate_corrected"] * df["r_pro_sub_avg_corrected"]) - (df["b_sub_rate_corrected"] * df["b_pro_sub_avg_corrected"])
        df["skill_momentum"] = df["pro_SLpM_diff_corrected"] * df["recent_form_diff_corrected"]

        # Interaction features (combining key metrics for better prediction)
        df["elo_x_form"] = df["elo_diff"] * df["recent_form_diff_corrected"]
        df["reach_x_striking"] = df["reach_diff"] * df["pro_SLpM_diff_corrected"]
        df["experience_x_age"] = df["experience_gap"] * df["age_at_event_diff"]

        # NEW: Advanced interaction features based on permutation importance analysis
        # Age-based interactions (age is #1 most important feature)
        df["age_x_striking"] = df["age_at_event_diff"] * df["pro_SLpM_diff_corrected"]
        df["age_x_grappling"] = df["age_at_event_diff"] * df["pro_td_avg_diff_corrected"]
        df["age_x_durability"] = df["age_at_event_diff"] * df["durability_diff_corrected"]

        # Grappling combinations (TD avg is #2 most important)
        df["td_x_defense"] = df["pro_td_avg_diff_corrected"] * df["pro_td_def_diff_corrected"]
        df["grappling_x_experience"] = df["grappler_advantage"] * df["experience_gap"]

        # Striking combinations (net striking is #3 most important)
        df["striking_x_accuracy"] = df["net_striking_advantage"] * df["pro_sig_str_acc_diff_corrected"]
        df["striking_x_defense"] = df["pro_SLpM_diff_corrected"] * df["pro_str_def_diff_corrected"]

        # Momentum interactions (momentum swing is #8 most important)
        df["form_x_experience"] = df["recent_form_diff_corrected"] * df["experience_gap"]
        df["finish_x_momentum"] = df["finish_rate_diff"] * df["recent_form_diff_corrected"]

        # Physical advantage combinations
        df["height_x_reach"] = df["height_diff"] * df["reach_diff"]
        df["physical_x_striking"] = (df["height_diff"] + df["reach_diff"]) * df["pro_SLpM_diff_corrected"]

        # NEW: Elite interaction features based on top 20 importance analysis
        # Win/loss ratio combinations (#1 most important)
        df["elo_x_win_ratio"] = df["elo_diff"] * df["win_loss_ratio_diff_corrected"]
        df["win_ratio_x_form"] = df["win_loss_ratio_diff_corrected"] * df["recent_form_diff_corrected"]

        # Durability combinations (#6 most important)
        df["durability_x_striking"] = df["durability_diff_corrected"] * df["net_striking_advantage"]
        df["elo_x_durability"] = df["elo_diff"] * df["durability_diff_corrected"]

        # Submission specialist (#4) combinations
        df["submission_x_grappling"] = df["submission_specialist_gap"] * df["grappler_advantage"]

        # KO power (#9) combinations
        df["ko_power_x_striking"] = df["ko_rate_diff_corrected"] * df["net_striking_advantage"]

        # Momentum/streak combinations (#7 momentum_swing, #12 loss_streak, #14 win_streak)
        df["momentum_x_win_streak"] = df["momentum_swing"] * df["win_streak_diff_corrected"]
        df["streak_differential"] = df["win_streak_diff_corrected"] * df["loss_streak_diff_corrected"]

        # NEW ELITE INTERACTIONS: Based on top feature analysis
        # Age + streak synergy (top features: age #1, win_streak #2)
        df["age_x_win_streak"] = df["age_at_event_diff"] * df["win_streak_diff_corrected"]

        # ELO + submission threat (elo #5, submission_specialist #12)
        df["elo_x_sub_threat"] = df["elo_diff"] * df["submission_specialist_gap"]

        # Form + durability compound (form #30, durability #13)
        df["form_x_durability"] = df["recent_form_diff_corrected"] * df["durability_diff_corrected"]

        # Striking vs grappling matchup (striking #10, grappler #12)
        df["striking_x_grappling_matchup"] = df["net_striking_advantage"] * df["grappler_advantage"]

        # Triple momentum combo (momentum_x_win_streak #4 is already powerful)
        df["momentum_combo"] = df["momentum_x_win_streak"] * df["recent_form_diff_corrected"]

        # NEW HIGH-IMPACT FEATURES: Elite compound interactions
        # Triple threat: ELO + finish rate + momentum
        df["elite_finisher"] = df["elo_diff"] * df["finish_rate_diff"] * df["recent_form_diff_corrected"]

        # Experienced winner: Win ratio + experience + age advantage
        df["veteran_advantage"] = df["win_loss_ratio_diff_corrected"] * df["experience_gap"] * (-df["age_at_event_diff"])

        # Complete fighter: Striking + grappling + durability
        df["complete_fighter"] = df["net_striking_advantage"] * df["grappler_advantage"] * df["durability_diff_corrected"]

        # Finish threat: KO power + submission threat + finish pressure
        df["total_finish_threat"] = (df["ko_rate_diff_corrected"] + df["sub_rate_diff_corrected"]) * df["finish_pressure"]

        # Hot streak compound: Win streak + momentum + form
        df["unstoppable_streak"] = df["win_streak_diff_corrected"] * df["momentum_swing"] * df["recent_form_diff_corrected"]

        # Age prime factor: Peak performance (27-32 years old)
        df["r_age_prime"] = 1.0 - abs(df["r_age_at_event"] - 29.5) / 10.0  # Peak at 29.5
        df["b_age_prime"] = 1.0 - abs(df["b_age_at_event"] - 29.5) / 10.0
        df["age_prime_advantage"] = df["r_age_prime"] - df["b_age_prime"]

        # Layoff freshness vs ring rust (optimal ~90-180 days)
        df["r_layoff_factor"] = 1.0 - abs(df["r_days_since_last_fight_corrected"] - 135) / 200.0  # Peak at 135 days
        df["b_layoff_factor"] = 1.0 - abs(df["b_days_since_last_fight_corrected"] - 135) / 200.0
        df["freshness_advantage"] = df["r_layoff_factor"] - df["b_layoff_factor"]

        # Desperation factor: Losing streak + age (older fighters need wins more)
        df["r_desperation"] = df["r_loss_streak_corrected"] * (df["r_age_at_event"] / 35.0)
        df["b_desperation"] = df["b_loss_streak_corrected"] * (df["b_age_at_event"] / 35.0)
        df["desperation_diff"] = df["r_desperation"] - df["b_desperation"]

        # Momentum sustainability: Recent form / days since last fight (maintaining activity)
        df["r_momentum_sustainability"] = df["r_recent_form_corrected"] / (df["r_days_since_last_fight_corrected"] / 100.0 + 1.0)
        df["b_momentum_sustainability"] = df["b_recent_form_corrected"] / (df["b_days_since_last_fight_corrected"] / 100.0 + 1.0)
        df["momentum_sustainability_diff"] = df["r_momentum_sustainability"] - df["b_momentum_sustainability"]

        # ELO + finish rate synergy (elite fighters who finish)
        df["elo_x_finish"] = df["elo_diff"] * df["finish_rate_diff"]

        # Experience in adversity: Win ratio despite losses (comeback ability)
        df["r_adversity_experience"] = df["r_win_loss_ratio_corrected"] * (df["r_losses"] + 1)
        df["b_adversity_experience"] = df["b_win_loss_ratio_corrected"] * (df["b_losses"] + 1)
        df["adversity_experience_diff"] = df["r_adversity_experience"] - df["b_adversity_experience"]

        # ========== RESEARCH-BACKED FEATURES (RF/GBDT/SVM Analysis) ==========

        # ENHANCED H2H AND STYLISTIC FEATURES
        # h2h_advantage is the top feature (0.018892) - let's enhance it
        # Create interaction features that amplify h2h advantage with other strong features

        # H2H × ELO: When you've beaten someone before AND have better rating
        df["h2h_x_elo"] = df["h2h_advantage"] * df["elo_diff"]

        # H2H × Recent Form: Previous wins matter more when you're on a hot streak
        df["h2h_x_form"] = df["h2h_advantage"] * df["recent_form_diff_corrected"]

        # Age differential interactions (age_at_event_diff is #2 at 0.015408)
        # Prime age advantage: Being in prime (26-33) vs opponent not in prime
        def calculate_age_prime_score(age):
            if 26 <= age <= 33:
                return 1.0
            elif 23 <= age < 26 or 33 < age <= 36:
                return 0.6
            else:
                return 0.2

        df["r_age_prime_score"] = df["r_age_at_event"].apply(calculate_age_prime_score)
        df["b_age_prime_score"] = df["b_age_at_event"].apply(calculate_age_prime_score)
        df["age_prime_score_diff"] = df["r_age_prime_score"] - df["b_age_prime_score"]

        # Age × Experience: Older fighters with more experience are more dangerous
        df["age_x_experience"] = df["age_at_event_diff"] * (
            (df["r_total_fights"] - df["b_total_fights"]) / 50.0
        ).clip(-1, 1)

        # Win ratio interactions (win_loss_ratio_diff_corrected is #3 at 0.014744)
        # Win ratio × finish rate: High win ratio fighters who finish fights are elite
        df["win_ratio_x_finish"] = df["win_loss_ratio_diff_corrected"] * df["finish_rate_diff"]

        # Win ratio × durability: High win ratio + durability = championship material
        df["win_ratio_x_durability"] = df["win_loss_ratio_diff_corrected"] * df["durability_diff_corrected"]

        # POSITIONAL STRIKING: CLINCH #3, GROUND #5, DISTANCE #6 in model rankings
        # NOW USING CAREER AVERAGES (not fight outcome data!)
        # These represent each fighter's career tendencies for where they land strikes

        # Positional advantage: Dominant in preferred position
        # Higher = better at fighting from distance, clinch, or ground
        df["positional_striking_advantage"] = (
            abs(df["distance_pct_diff_corrected"]) +
            abs(df["clinch_pct_diff_corrected"]) +
            abs(df["ground_pct_diff_corrected"])
        )

        # TARGET AREA DISTRIBUTION: Head/Body/Leg striking (ranked #10)
        # NOW USING CAREER AVERAGES of where strikes land
        # Target diversity: Fighters who can attack all areas are more dangerous
        df["target_distribution_advantage"] = (
            abs(df["head_pct_diff_corrected"]) +
            abs(df["body_pct_diff_corrected"]) +
            abs(df["leg_pct_diff_corrected"])
        )

        # CONTROL & REVERSALS: CTRL #5 in GBDT, REV #8/#9 in RF/GBDT
        # NOW USING CAREER AVERAGES per fight
        # Control dominance: Combines control time + reversals (grappling mastery)
        # Normalize control time (convert seconds to 0-1 scale, typical range 0-300 sec)
        df["control_dominance"] = (
            (df["avg_ctrl_sec_diff_corrected"] / 300.0).clip(-1, 1) +
            df["avg_rev_diff_corrected"] * 0.1
        )

        # COMPOUND FEATURES: Combining research-backed positionals with existing top features

        # Clinch grappling: Clinch striking tendency × grappling ability
        df["clinch_x_grappling"] = df["clinch_pct_diff_corrected"] * df["grappler_advantage"]

        # Distance striking: Distance striking tendency × net striking advantage
        df["distance_x_striking"] = df["distance_pct_diff_corrected"] * df["net_striking_advantage"]

        # Ground control: Ground striking tendency × control time (ground and pound mastery)
        df["ground_x_control"] = (
            df["ground_pct_diff_corrected"] *
            (df["avg_ctrl_sec_diff_corrected"] / 300.0).clip(-1, 1)
        )

        # Positional mastery: Combined positional + grappling + striking
        df["positional_mastery"] = (
            df["positional_striking_advantage"] *
            df["grappler_advantage"] *
            df["net_striking_advantage"]
        )

        # Opponent quality (simplified)
        df["r_opponent_quality"] = 0.5
        df["b_opponent_quality"] = 0.5
        df["opponent_quality_diff"] = 0

        # Clutch and momentum features
        df["r_clutch_factor"] = df["r_win_loss_ratio_corrected"] * df["r_recent_form_corrected"]
        df["b_clutch_factor"] = df["b_win_loss_ratio_corrected"] * df["b_recent_form_corrected"]
        df["clutch_factor_diff"] = df["r_clutch_factor"] - df["b_clutch_factor"]

        df["r_momentum_quality"] = 0.5
        df["b_momentum_quality"] = 0.5
        df["momentum_quality_diff"] = 0

        df["r_pressure_performance"] = 0.5
        df["b_pressure_performance"] = 0.5
        df["pressure_performance_diff"] = 0

        df["r_form_consistency"] = 0.5
        df["b_form_consistency"] = 0.5
        df["form_consistency_diff"] = 0

        print(f"Core features prepared: {len(self.get_core_feature_names())} features")

        return df

    # ===== HYPERPARAMETER OPTIMIZATION WITH OPTUNA =====

    def optimize_hyperparameters(self, X, y, n_trials=50):
        """Optimize hyperparameters using Optuna Bayesian optimization"""
        if not HAS_OPTUNA or not HAS_XGBOOST:
            print("Optuna or XGBoost not available, using default parameters")
            return {
                'n_estimators': 600,
                'max_depth': 8,
                'learning_rate': 0.02,
                'subsample': 0.85,
                'colsample_bytree': 0.85,
                'reg_alpha': 0.1,
                'reg_lambda': 0.8,
            }

        print(f"\n{'='*80}")
        print(f"OPTIMIZING HYPERPARAMETERS WITH OPTUNA ({n_trials} trials)")
        print(f"{'='*80}\n")

        def objective(trial):
            params = {
                'n_estimators': trial.suggest_int('n_estimators', 200, 1200),
                'max_depth': trial.suggest_int('max_depth', 4, 8),  # Balanced: allow complexity without overfitting (was 3-6)
                'learning_rate': trial.suggest_float('learning_rate', 0.005, 0.05, log=True),
                'subsample': trial.suggest_float('subsample', 0.7, 0.95),
                'colsample_bytree': trial.suggest_float('colsample_bytree', 0.7, 0.95),
                'reg_alpha': trial.suggest_float('reg_alpha', 0.3, 2.0),  # Relaxed regularization (was 1.0-5.0)
                'reg_lambda': trial.suggest_float('reg_lambda', 0.8, 3.0),  # Relaxed regularization (was 1.5-5.0)
                'min_child_weight': trial.suggest_int('min_child_weight', 3, 12),  # Relaxed minimum samples (was 5-20)
                'random_state': 42,
                'n_jobs': -1,
            }

            model = XGBClassifier(**params)

            # Time series cross-validation
            cv_scores = []
            tscv = TimeSeriesSplit(n_splits=5)

            for train_idx, val_idx in tscv.split(X):
                X_train_fold = X.iloc[train_idx] if hasattr(X, 'iloc') else X[train_idx]
                X_val_fold = X.iloc[val_idx] if hasattr(X, 'iloc') else X[val_idx]
                y_train_fold = y.iloc[train_idx] if hasattr(y, 'iloc') else y[train_idx]
                y_val_fold = y.iloc[val_idx] if hasattr(y, 'iloc') else y[val_idx]

                model.fit(X_train_fold, y_train_fold)
                score = model.score(X_val_fold, y_val_fold)
                cv_scores.append(score)

            return np.mean(cv_scores)

        study = optuna.create_study(direction='maximize')
        study.optimize(objective, n_trials=n_trials, show_progress_bar=True)

        print(f"\nBest CV Score: {study.best_value:.4f}")
        print(f"Best Parameters:")
        for key, value in study.best_params.items():
            print(f"  {key}: {value}")

        return study.best_params

    # ===== PERMUTATION IMPORTANCE FEATURE SELECTION =====

    def select_features_by_importance(self, X, y, max_features=55):
        """
        Robust feature selection using model-based importance with adaptive thresholding.
        Automatically determines optimal feature count based on importance distribution.
        """
        print("\n" + "="*80)
        print("INTELLIGENT FEATURE SELECTION")
        print("="*80 + "\n")

        # Remove constant or near-constant features first (variance threshold)
        print("Step 1: Removing low-variance features...")
        variances = X.var()
        variance_threshold = variances.quantile(0.01)  # Remove bottom 1%
        high_variance_features = variances[variances > variance_threshold].index.tolist()
        X_filtered = X[high_variance_features]
        print(f"  Kept {len(high_variance_features)}/{len(X.columns)} features with sufficient variance")

        # Train model on larger sample for more stable importance
        if HAS_XGBOOST:
            model = XGBClassifier(
                n_estimators=400,
                max_depth=8,
                learning_rate=0.05,
                subsample=0.8,
                colsample_bytree=0.8,
                random_state=42,
                n_jobs=-1
            )
        else:
            model = RandomForestClassifier(n_estimators=400, max_depth=15, random_state=42, n_jobs=-1)

        # Use 40% for validation (more stable than 20%)
        X_train, X_val, y_train, y_val = train_test_split(
            X_filtered, y, test_size=0.4, random_state=42, stratify=y
        )

        print("Step 2: Training model for importance calculation...")
        model.fit(X_train, y_train)

        # Use model's built-in feature importance (much more stable than permutation)
        if HAS_XGBOOST:
            importances = model.feature_importances_
        else:
            importances = model.feature_importances_

        # Get feature names
        feature_names = X_filtered.columns.tolist()

        # Create importance DataFrame
        importance_df = pd.DataFrame({
            'feature': feature_names,
            'importance': importances
        }).sort_values('importance', ascending=False)

        print("\nStep 3: Analyzing importance distribution...")

        # Calculate statistics
        mean_importance = importance_df['importance'].mean()
        median_importance = importance_df['importance'].median()

        # Strategy: Use elbow method - find where importance drops sharply
        importance_df['importance_ratio'] = importance_df['importance'] / importance_df['importance'].max()
        importance_df['importance_change'] = importance_df['importance_ratio'].diff().abs()

        # Find features above mean importance (natural threshold)
        above_mean = importance_df[importance_df['importance'] > mean_importance]

        # Also consider cumulative importance approach
        total_importance = importance_df['importance'].sum()
        importance_df['cumulative'] = importance_df['importance'].cumsum() / total_importance

        # Select features that:
        # 1. Are above mean importance OR
        # 2. Are in top percentile by cumulative importance (100% = all features)
        cumulative_100 = importance_df[importance_df['cumulative'] <= 1.00]

        # Take union of both methods
        selected_by_mean = set(above_mean['feature'].tolist())
        selected_by_cumulative = set(cumulative_100['feature'].tolist())
        selected_features_set = selected_by_mean.union(selected_by_cumulative)

        # Filter importance_df to selected features
        selected_features = importance_df[importance_df['feature'].isin(selected_features_set)]

        # Apply soft cap if way too many features
        if len(selected_features) > max_features:
            print(f"  Too many features ({len(selected_features)}), applying soft cap...")
            selected_features = selected_features.head(max_features)

        print(f"\n  Mean importance: {mean_importance:.6f}")
        print(f"  Median importance: {median_importance:.6f}")
        print(f"  Features above mean: {len(above_mean)}")
        print(f"  Features for 100% cumulative: {len(cumulative_100)}")
        print(f"  Selected (union): {len(selected_features)}")

        print(f"\nTop {len(selected_features)} Most Important Features (Ranked):")
        print(selected_features[['feature', 'importance']].to_string(index=False))

        top_features = selected_features['feature'].tolist()

        print(f"\nFinal Selection: {len(top_features)} features")
        print(f"  Validation Accuracy: {model.score(X_val, y_val):.4f}")

        self.feature_importance = importance_df
        return top_features

    # ===== TRAINING WITH PROPER VALIDATION =====

    def train(self, df):
        """Train the model with all improvements

        Args:
            df: DataFrame with fight data (should already have data leakage fixed)
        """
        print("\n" + "="*80)
        print("IMPROVED UFC PREDICTOR - TRAINING")
        print("="*80)

        print(f"\nTraining on {len(df)} fights")

        # Prepare core features BEFORE augmentation so interaction features exist during swapping
        df = self.prepare_core_features(df)

        # Conservative data augmentation (35% vs 90%)
        df = self.augment_data_conservatively(df)

        # Get core feature names
        core_features = self.get_core_feature_names()

        # Filter to available features
        available_features = [f for f in core_features if f in df.columns]
        print(f"\nUsing {len(available_features)} core features (of {len(core_features)} defined)")

        # Prepare X and y
        X = df[available_features].copy()
        y_winner = (df["winner"] == "Red").astype(int)

        # Handle missing values
        X = X.fillna(0)

        # PROPER VALIDATION: Split into train/val/test
        print("\n" + "="*80)
        print("PROPER VALIDATION SPLIT")
        print("="*80)

        # Sort by date for temporal split
        df_sorted = df.sort_values("event_date").reset_index(drop=True)
        X = X.loc[df_sorted.index]
        y_winner = y_winner.loc[df_sorted.index]

        # 70% train, 15% validation, 15% test
        train_size = int(0.70 * len(df_sorted))
        val_size = int(0.85 * len(df_sorted))

        X_train = X.iloc[:train_size]
        X_val = X.iloc[train_size:val_size]
        X_test = X.iloc[val_size:]

        y_train = y_winner.iloc[:train_size]
        y_val = y_winner.iloc[train_size:val_size]
        y_test = y_winner.iloc[val_size:]

        print(f"\nTrain set: {len(X_train)} fights")
        print(f"Validation set: {len(X_val)} fights")
        print(f"Test set: {len(X_test)} fights (HELD OUT)")

        # Feature selection: Reduce to top 35 most important features
        X_train_val = pd.concat([X_train, X_val])
        y_train_val = pd.concat([y_train, y_val])

        selected_features = self.select_features_by_importance(X_train_val, y_train_val, max_features=55)

        # Update datasets with selected features only
        X_train = X_train[selected_features]
        X_val = X_val[selected_features]
        X_test = X_test[selected_features]
        X_train_val = X_train_val[selected_features]
        available_features = selected_features  # Update for later use

        # Hyperparameter optimization on train+val
        if HAS_OPTUNA and HAS_XGBOOST:
            self.best_params = self.optimize_hyperparameters(X_train_val, y_train_val, n_trials=50) # 50 or 100
        else:
            self.best_params = {
                'n_estimators': 600,
                'max_depth': 8,
                'learning_rate': 0.02,
                'subsample': 0.85,
                'colsample_bytree': 0.85,
                'reg_alpha': 0.1,
                'reg_lambda': 0.8,
            }

        # Train final model with ensemble diversity
        print("\n" + "="*80)
        print("TRAINING ENSEMBLE MODEL")
        print("="*80 + "\n")

        if HAS_XGBOOST and self.use_ensemble:
            # Calculate scale_pos_weight for class balancing
            scale_weight = (y_train_val == 0).sum() / (y_train_val == 1).sum()
            print(f"Applying class balancing: scale_pos_weight = {scale_weight:.3f}")

            # Create diverse ensemble with XGBoost, RandomForest, and LogisticRegression
            print("Creating ensemble with XGBoost, RandomForest, and LogisticRegression...")

            xgb_model = XGBClassifier(
                **self.best_params,
                scale_pos_weight=scale_weight,
                random_state=42,
                n_jobs=-1
            )

            rf_model = RandomForestClassifier(
                n_estimators=400,
                max_depth=12,
                min_samples_split=8,
                random_state=42,
                n_jobs=-1
            )

            lr_model = LogisticRegression(
                C=0.1,
                max_iter=1000,
                random_state=42,
                n_jobs=-1
            )

            base_model = VotingClassifier(
                estimators=[
                    ('xgb', xgb_model),
                    ('rf', rf_model),
                    ('lr', lr_model)
                ],
                voting='soft',
                n_jobs=-1
            )
        elif HAS_XGBOOST:
            # Single XGBoost model if ensemble disabled
            scale_weight = (y_train_val == 0).sum() / (y_train_val == 1).sum()
            print(f"Applying class balancing: scale_pos_weight = {scale_weight:.3f}")
            base_model = XGBClassifier(
                **self.best_params,
                scale_pos_weight=scale_weight,
                random_state=42,
                n_jobs=-1
            )
        else:
            base_model = RandomForestClassifier(
                n_estimators=600, max_depth=15, min_samples_split=8,
                random_state=42, n_jobs=-1
            )

        # Train on train+val (already created with selected features)
        print("Training ensemble model...")
        base_model.fit(X_train_val, y_train_val)

        # Calibrate probabilities
        print("\n" + "="*80)
        print("CALIBRATING PROBABILITIES")
        print("="*80 + "\n")

        self.calibrated_model = CalibratedClassifierCV(
            base_model,
            method='isotonic',
            cv=5
        )
        self.calibrated_model.fit(X_train_val, y_train_val)

        self.winner_model = self.calibrated_model

        # Evaluate on test set (FIRST TIME)
        print("\n" + "="*80)
        print("FINAL MODEL PERFORMANCE")
        print("="*80)

        # Get predictions and probabilities
        y_pred_test = self.winner_model.predict(X_test)
        y_proba_test = self.winner_model.predict_proba(X_test)[:, 1]

        # Calculate all metrics
        train_acc = self.winner_model.score(X_train, y_train)
        val_acc = self.winner_model.score(X_val, y_val)
        test_acc = accuracy_score(y_test, y_pred_test)
        test_roc_auc = roc_auc_score(y_test, y_proba_test)

        print(f"\nTrain Accuracy:      {train_acc:.4f}")
        print(f"Validation Accuracy: {val_acc:.4f}")
        print(f"Test Accuracy:       {test_acc:.4f} *** HELD OUT ***")
        print(f"\nROC-AUC Score:       {test_roc_auc:.4f} (probability separation)")

        print("\nDetailed Metrics by Corner:")
        print(classification_report(y_test, y_pred_test,
                                   target_names=['Blue Corner', 'Red Corner'],
                                   digits=4))

        # Check overfitting
        train_val_gap = train_acc - val_acc
        if train_val_gap > 0.05:
            print(f"\nWarning: Large train-validation gap ({train_val_gap:.4f}) suggests possible overfitting")
        else:
            print(f"\nTrain-validation gap ({train_val_gap:.4f}) is acceptable")

        # Time series cross-validation for robustness check
        print("\n" + "="*80)
        print("TIME SERIES CROSS-VALIDATION (Robustness Check)")
        print("="*80 + "\n")

        tscv = TimeSeriesSplit(n_splits=5)
        cv_scores = []

        for fold, (train_idx, val_idx) in enumerate(tscv.split(X_train_val)):
            X_fold_train = X_train_val.iloc[train_idx]
            X_fold_val = X_train_val.iloc[val_idx]
            y_fold_train = y_train_val.iloc[train_idx]
            y_fold_val = y_train_val.iloc[val_idx]

            fold_model = XGBClassifier(**self.best_params, random_state=42, n_jobs=-1) if HAS_XGBOOST else RandomForestClassifier(n_estimators=600, random_state=42, n_jobs=-1)
            fold_model.fit(X_fold_train, y_fold_train)
            score = fold_model.score(X_fold_val, y_fold_val)
            cv_scores.append(score)
            print(f"  Fold {fold + 1}/5: {score:.4f}")

        print(f"\nCV Mean: {np.mean(cv_scores):.4f} ± {np.std(cv_scores):.4f}")

        self.df_train = df
        self.feature_columns = available_features  # Store for predictions

        print("\n" + "="*80)
        print("TRAINING COMPLETE")
        print("="*80)
        print(f"\nWinner Model: {test_acc:.4f} accuracy")
        print(f"Features Used: {len(available_features)}")

        return available_features

    # ===== PREDICTION METHODS =====

    def get_fighter_latest_stats(self, fighter_name):
        """Get latest stats for fighter from training data"""
        red_fights = self.df_train[
            self.df_train["r_fighter"] == fighter_name
        ].sort_values("event_date", ascending=False)
        blue_fights = self.df_train[
            self.df_train["b_fighter"] == fighter_name
        ].sort_values("event_date", ascending=False)

        if len(red_fights) > 0 and len(blue_fights) > 0:
            red_latest = red_fights.iloc[0]
            blue_latest = blue_fights.iloc[0]
            latest = (
                red_latest
                if red_latest["event_date"] > blue_latest["event_date"]
                else blue_latest
            )
            prefix = (
                "r" if red_latest["event_date"] > blue_latest["event_date"] else "b"
            )
        elif len(red_fights) > 0:
            latest, prefix = red_fights.iloc[0], "r"
        elif len(blue_fights) > 0:
            latest, prefix = blue_fights.iloc[0], "b"
        else:
            return None

        # Get stats from corrected columns
        stats = {}
        for stat in ["wins", "losses", "win_loss_ratio", "pro_SLpM", "pro_SApM",
                     "pro_sig_str_acc", "pro_str_def", "pro_td_avg", "pro_td_acc",
                     "pro_td_def", "pro_sub_avg", "ko_rate", "sub_rate", "dec_rate",
                     "recent_form", "win_streak", "loss_streak", "days_since_last_fight",
                     "recent_finish_rate", "durability"]:
            col = f"{prefix}_{stat}_corrected"
            if col in latest.index:
                stats[stat] = latest[col]
            else:
                stats[stat] = 0

        # Physical stats
        for stat in ["height", "reach", "weight", "age_at_event", "ape_index", "stance"]:
            col = f"{prefix}_{stat}"
            if col in latest.index:
                stats[stat] = latest[col]
            else:
                stats[stat] = 0 if stat != "stance" else "Orthodox"

        return stats

    def prepare_upcoming_fight(self, fight, feature_columns):
        """Prepare features for an upcoming fight"""
        r_stats = self.get_fighter_latest_stats(fight["red_fighter"])
        b_stats = self.get_fighter_latest_stats(fight["blue_fighter"])

        if not r_stats or not b_stats:
            return None, None, None

        r_total_fights = r_stats["wins"] + r_stats["losses"]
        b_total_fights = b_stats["wins"] + b_stats["losses"]

        fight_features = {
            "r_fighter": fight["red_fighter"],
            "b_fighter": fight["blue_fighter"],
            "total_rounds": fight["total_rounds"],
            "is_title_bout": 1 if fight["total_rounds"] == 5 else 0,
            "r_total_fights": r_total_fights,
            "b_total_fights": b_total_fights,
            "weight_class": fight.get("weight_class", "Unknown"),
        }

        # Physical differentials
        for stat in ["height", "reach", "weight", "age_at_event", "ape_index"]:
            fight_features[f"{stat}_diff"] = r_stats.get(stat, 0) - b_stats.get(stat, 0)

        # Statistical differentials
        for stat in ["wins", "losses", "win_loss_ratio", "pro_SLpM", "pro_SApM",
                     "pro_sig_str_acc", "pro_str_def", "pro_td_avg", "pro_td_acc",
                     "pro_td_def", "pro_sub_avg", "ko_rate", "sub_rate", "dec_rate",
                     "recent_form", "win_streak", "loss_streak", "days_since_last_fight",
                     "recent_finish_rate", "durability"]:
            fight_features[f"{stat}_diff_corrected"] = r_stats.get(stat, 0) - b_stats.get(stat, 0)

        # Individual stats for both fighters
        for prefix, stats in [("r", r_stats), ("b", b_stats)]:
            for stat in ["wins", "losses", "ko_rate", "sub_rate", "dec_rate", "pro_str_def",
                         "pro_td_def", "durability", "pro_SLpM", "pro_sig_str_acc", "pro_SApM",
                         "pro_td_avg", "pro_td_acc", "pro_sub_avg", "recent_form"]:
                fight_features[f"{prefix}_{stat}_corrected"] = stats.get(stat, 0)

        # Core derived features
        fight_features["net_striking_advantage"] = (
            (r_stats.get("pro_SLpM", 0) - r_stats.get("pro_SApM", 0)) -
            (b_stats.get("pro_SLpM", 0) - b_stats.get("pro_SApM", 0))
        )
        fight_features["striking_efficiency"] = (
            (r_stats.get("pro_SLpM", 0) * r_stats.get("pro_sig_str_acc", 0)) -
            (b_stats.get("pro_SLpM", 0) * b_stats.get("pro_sig_str_acc", 0))
        )
        fight_features["defensive_striking"] = (
            (r_stats.get("pro_str_def", 0) - r_stats.get("pro_SApM", 0)) -
            (b_stats.get("pro_str_def", 0) - b_stats.get("pro_SApM", 0))
        )
        fight_features["grappling_control"] = (
            (r_stats.get("pro_td_avg", 0) * r_stats.get("pro_td_acc", 0)) -
            (b_stats.get("pro_td_avg", 0) * b_stats.get("pro_td_acc", 0))
        )

        fight_features["experience_gap"] = r_total_fights - b_total_fights

        r_finish_rate = (r_stats.get("ko_rate", 0) + r_stats.get("sub_rate", 0)) / 2
        b_finish_rate = (b_stats.get("ko_rate", 0) + b_stats.get("sub_rate", 0)) / 2
        fight_features["finish_rate_diff"] = r_finish_rate - b_finish_rate
        fight_features["r_finish_rate"] = r_finish_rate
        fight_features["b_finish_rate"] = b_finish_rate

        # Style features
        r_striker_score = r_stats.get("pro_SLpM", 0) - r_stats.get("pro_td_avg", 0)
        b_striker_score = b_stats.get("pro_SLpM", 0) - b_stats.get("pro_td_avg", 0)
        fight_features["striker_advantage"] = r_striker_score - b_striker_score
        fight_features["r_striker_score"] = r_striker_score
        fight_features["b_striker_score"] = b_striker_score

        r_grappler_score = r_stats.get("pro_td_avg", 0) + r_stats.get("pro_sub_avg", 0)
        b_grappler_score = b_stats.get("pro_td_avg", 0) + b_stats.get("pro_sub_avg", 0)
        fight_features["grappler_advantage"] = r_grappler_score - b_grappler_score
        fight_features["r_grappler_score"] = r_grappler_score
        fight_features["b_grappler_score"] = b_grappler_score

        fight_features["striker_vs_grappler"] = 0
        fight_features["stance_matchup_advantage"] = 0
        fight_features["r_trajectory_3"] = 0
        fight_features["b_trajectory_3"] = 0
        fight_features["ring_rust_factor"] = 0
        fight_features["weight_class_factor"] = 1.0

        # Momentum features
        fight_features["momentum_swing"] = (
            r_stats.get("recent_form", 0) - b_stats.get("recent_form", 0) +
            (r_stats.get("win_streak", 0) - b_stats.get("win_streak", 0)) * 0.1
        )
        fight_features["style_clash_severity"] = abs(
            fight_features["striker_advantage"] * 0.5 + fight_features["grappler_advantage"] * 0.5
        )
        fight_features["power_vs_technique"] = (
            (r_stats.get("pro_SLpM", 0) - b_stats.get("pro_SLpM", 0)) * 0.6 +
            (r_stats.get("pro_sig_str_acc", 0) - b_stats.get("pro_sig_str_acc", 0)) * 0.4
        )
        fight_features["finish_pressure"] = (
            (r_stats.get("ko_rate", 0) - b_stats.get("ko_rate", 0)) * 0.5 +
            (r_stats.get("sub_rate", 0) - b_stats.get("sub_rate", 0)) * 0.3
        )
        fight_features["upset_potential"] = (
            (b_stats.get("recent_form", 0) - r_stats.get("recent_form", 0)) * 0.4 +
            (b_stats.get("win_streak", 0) - r_stats.get("win_streak", 0)) * 0.3
        )

        # Advanced features
        fight_features["offensive_output"] = (
            (r_stats.get("pro_SLpM", 0) + r_stats.get("pro_td_avg", 0) + r_stats.get("pro_sub_avg", 0)) -
            (b_stats.get("pro_SLpM", 0) + b_stats.get("pro_td_avg", 0) + b_stats.get("pro_sub_avg", 0))
        )
        fight_features["defensive_composite"] = (
            ((r_stats.get("pro_str_def", 0) + r_stats.get("pro_td_def", 0)) / 2) -
            ((b_stats.get("pro_str_def", 0) + b_stats.get("pro_td_def", 0)) / 2)
        )
        fight_features["ko_specialist_gap"] = (
            (r_stats.get("ko_rate", 0) * r_stats.get("pro_SLpM", 0)) -
            (b_stats.get("ko_rate", 0) * b_stats.get("pro_SLpM", 0))
        )
        fight_features["submission_specialist_gap"] = (
            (r_stats.get("sub_rate", 0) * r_stats.get("pro_sub_avg", 0)) -
            (b_stats.get("sub_rate", 0) * b_stats.get("pro_sub_avg", 0))
        )
        fight_features["skill_momentum"] = (
            fight_features["pro_SLpM_diff_corrected"] * fight_features["recent_form_diff_corrected"]
        )

        # Interaction features (will be calculated after ELO is added)
        fight_features["elo_x_form"] = 0  # Placeholder, calculated after ELO
        fight_features["reach_x_striking"] = fight_features["reach_diff"] * fight_features["pro_SLpM_diff_corrected"]
        fight_features["experience_x_age"] = fight_features["experience_gap"] * fight_features["age_at_event_diff"]

        # Advanced interaction features (data-driven from importance analysis)
        fight_features["age_x_striking"] = fight_features["age_at_event_diff"] * fight_features["pro_SLpM_diff_corrected"]
        fight_features["age_x_grappling"] = fight_features["age_at_event_diff"] * fight_features["pro_td_avg_diff_corrected"]
        fight_features["age_x_durability"] = fight_features["age_at_event_diff"] * fight_features["durability_diff_corrected"]
        fight_features["td_x_defense"] = fight_features["pro_td_avg_diff_corrected"] * fight_features["pro_td_def_diff_corrected"]
        fight_features["grappling_x_experience"] = fight_features["grappler_advantage"] * fight_features["experience_gap"]
        fight_features["striking_x_accuracy"] = (
            (r_stats.get("pro_SLpM", 0) - r_stats.get("pro_SApM", 0) - (b_stats.get("pro_SLpM", 0) - b_stats.get("pro_SApM", 0))) *
            fight_features["pro_sig_str_acc_diff_corrected"]
        )
        fight_features["striking_x_defense"] = fight_features["pro_SLpM_diff_corrected"] * fight_features["pro_str_def_diff_corrected"]
        fight_features["form_x_experience"] = fight_features["recent_form_diff_corrected"] * fight_features["experience_gap"]
        fight_features["finish_x_momentum"] = fight_features["finish_rate_diff"] * fight_features["recent_form_diff_corrected"]
        fight_features["height_x_reach"] = fight_features["height_diff"] * fight_features["reach_diff"]
        fight_features["physical_x_striking"] = (fight_features["height_diff"] + fight_features["reach_diff"]) * fight_features["pro_SLpM_diff_corrected"]

        # Elite interaction features (placeholders, some calculated after ELO)
        fight_features["elo_x_win_ratio"] = 0  # Calculated after ELO
        fight_features["win_ratio_x_form"] = fight_features["win_loss_ratio_diff_corrected"] * fight_features["recent_form_diff_corrected"]
        fight_features["durability_x_striking"] = fight_features["durability_diff_corrected"] * (
            (r_stats.get("pro_SLpM", 0) - r_stats.get("pro_SApM", 0)) -
            (b_stats.get("pro_SLpM", 0) - b_stats.get("pro_SApM", 0))
        )
        fight_features["elo_x_durability"] = 0  # Calculated after ELO
        fight_features["submission_x_grappling"] = fight_features["submission_specialist_gap"] * fight_features["grappler_advantage"]
        fight_features["ko_power_x_striking"] = fight_features["ko_rate_diff_corrected"] * (
            (r_stats.get("pro_SLpM", 0) - r_stats.get("pro_SApM", 0)) -
            (b_stats.get("pro_SLpM", 0) - b_stats.get("pro_SApM", 0))
        )
        fight_features["momentum_x_win_streak"] = fight_features["momentum_swing"] * fight_features["win_streak_diff_corrected"]
        fight_features["streak_differential"] = fight_features["win_streak_diff_corrected"] * fight_features["loss_streak_diff_corrected"]

        fight_features["opponent_quality_diff"] = 0
        fight_features["r_clutch_factor"] = r_stats.get("win_loss_ratio", 0) * r_stats.get("recent_form", 0)
        fight_features["b_clutch_factor"] = b_stats.get("win_loss_ratio", 0) * b_stats.get("recent_form", 0)
        fight_features["clutch_factor_diff"] = fight_features["r_clutch_factor"] - fight_features["b_clutch_factor"]
        fight_features["momentum_quality_diff"] = 0
        fight_features["pressure_performance_diff"] = 0
        fight_features["form_consistency_diff"] = 0
        fight_features["h2h_advantage"] = 0
        fight_features["last_5_wins_diff_corrected"] = 0

        # Add ELO features
        if hasattr(self, 'fighter_elos'):
            r_elo = self.fighter_elos.get(fight["red_fighter"], 1500)
            b_elo = self.fighter_elos.get(fight["blue_fighter"], 1500)

            fight_features["r_elo_pre_fight"] = r_elo
            fight_features["b_elo_pre_fight"] = b_elo
            fight_features["elo_diff"] = r_elo - b_elo
        else:
            # Default ELO if not available
            fight_features["r_elo_pre_fight"] = 1500
            fight_features["b_elo_pre_fight"] = 1500
            fight_features["elo_diff"] = 0

        # Calculate elo_x_form and other ELO-dependent features now that ELO is available
        fight_features["elo_x_form"] = fight_features["elo_diff"] * fight_features["recent_form_diff_corrected"]
        fight_features["elo_x_win_ratio"] = fight_features["elo_diff"] * fight_features["win_loss_ratio_diff_corrected"]
        fight_features["elo_x_durability"] = fight_features["elo_diff"] * fight_features["durability_diff_corrected"]

        # NEW: Strategic interaction features
        fight_features["age_x_win_streak"] = fight_features.get("age_at_event_diff", 0) * fight_features.get("win_streak_diff_corrected", 0)
        fight_features["elo_x_sub_threat"] = fight_features["elo_diff"] * fight_features.get("submission_specialist_gap", 0)
        fight_features["form_x_durability"] = fight_features["recent_form_diff_corrected"] * fight_features.get("durability_diff_corrected", 0)
        fight_features["striking_x_grappling_matchup"] = fight_features.get("net_striking_advantage", 0) * fight_features.get("grappler_advantage", 0)
        fight_features["momentum_combo"] = fight_features.get("momentum_x_win_streak", 0) * fight_features["recent_form_diff_corrected"]

        # NEW HIGH-IMPACT: Elite compound features
        fight_features["elite_finisher"] = fight_features["elo_diff"] * fight_features.get("finish_rate_diff", 0) * fight_features["recent_form_diff_corrected"]
        fight_features["veteran_advantage"] = fight_features["win_loss_ratio_diff_corrected"] * fight_features.get("experience_gap", 0) * (-fight_features.get("age_at_event_diff", 0))
        fight_features["complete_fighter"] = fight_features.get("net_striking_advantage", 0) * fight_features.get("grappler_advantage", 0) * fight_features.get("durability_diff_corrected", 0)
        fight_features["total_finish_threat"] = (fight_features.get("ko_rate_diff_corrected", 0) + fight_features.get("sub_rate_diff_corrected", 0)) * fight_features.get("finish_pressure", 0)
        fight_features["unstoppable_streak"] = fight_features.get("win_streak_diff_corrected", 0) * fight_features.get("momentum_swing", 0) * fight_features["recent_form_diff_corrected"]

        # Age prime and freshness factors
        r_age = fight_features.get("r_age_at_event", 30)
        b_age = fight_features.get("b_age_at_event", 30)
        fight_features["age_prime_advantage"] = (1.0 - abs(r_age - 29.5) / 10.0) - (1.0 - abs(b_age - 29.5) / 10.0)

        r_layoff = fight_features.get("r_days_since_last_fight_corrected", 135)
        b_layoff = fight_features.get("b_days_since_last_fight_corrected", 135)
        fight_features["freshness_advantage"] = (1.0 - abs(r_layoff - 135) / 200.0) - (1.0 - abs(b_layoff - 135) / 200.0)

        # Desperation and sustainability
        r_loss_streak = r_stats.get("loss_streak", 0)
        b_loss_streak = b_stats.get("loss_streak", 0)
        fight_features["desperation_diff"] = (r_loss_streak * (r_age / 35.0)) - (b_loss_streak * (b_age / 35.0))

        r_form = r_stats.get("recent_form", 0)
        b_form = b_stats.get("recent_form", 0)
        fight_features["momentum_sustainability_diff"] = (r_form / (r_layoff / 100.0 + 1.0)) - (b_form / (b_layoff / 100.0 + 1.0))

        fight_features["elo_x_finish"] = fight_features["elo_diff"] * fight_features.get("finish_rate_diff", 0)

        r_losses = r_stats.get("losses", 0)
        b_losses = b_stats.get("losses", 0)
        r_win_ratio = r_stats.get("win_loss_ratio", 0)
        b_win_ratio = b_stats.get("win_loss_ratio", 0)
        fight_features["adversity_experience_diff"] = (r_win_ratio * (r_losses + 1)) - (b_win_ratio * (b_losses + 1))

        # RESEARCH-BACKED FEATURES: Enhanced top feature interactions
        # H2H interactions
        fight_features["h2h_x_elo"] = fight_features.get("h2h_advantage", 0) * fight_features.get("elo_diff", 0)
        fight_features["h2h_x_form"] = fight_features.get("h2h_advantage", 0) * fight_features.get("recent_form_diff_corrected", 0)

        # Age prime score (mirroring training calculation)
        def calc_age_prime(age):
            if 26 <= age <= 33:
                return 1.0
            elif 23 <= age < 26 or 33 < age <= 36:
                return 0.6
            else:
                return 0.2

        r_age = r_stats.get("age_at_event", 30)
        b_age = b_stats.get("age_at_event", 30)
        r_age_prime = calc_age_prime(r_age)
        b_age_prime = calc_age_prime(b_age)
        fight_features["age_prime_score_diff"] = r_age_prime - b_age_prime

        # Age × Experience
        experience_diff = (r_total_fights - b_total_fights) / 50.0
        fight_features["age_x_experience"] = fight_features.get("age_at_event_diff", 0) * max(-1, min(1, experience_diff))

        # Win ratio interactions
        fight_features["win_ratio_x_finish"] = fight_features.get("win_loss_ratio_diff_corrected", 0) * fight_features.get("finish_rate_diff", 0)
        fight_features["win_ratio_x_durability"] = fight_features.get("win_loss_ratio_diff_corrected", 0) * fight_features.get("durability_diff_corrected", 0)

        # Career-based positional & target striking (use defaults for upcoming fights)
        # These would be calculated from fighter stats if available
        fight_features["distance_pct_diff_corrected"] = 0
        fight_features["clinch_pct_diff_corrected"] = 0
        fight_features["ground_pct_diff_corrected"] = 0
        fight_features["positional_striking_advantage"] = 0
        fight_features["head_pct_diff_corrected"] = 0
        fight_features["body_pct_diff_corrected"] = 0
        fight_features["leg_pct_diff_corrected"] = 0
        fight_features["target_distribution_advantage"] = 0

        # Career-based control & reversals
        fight_features["avg_ctrl_sec_diff_corrected"] = 0
        fight_features["avg_rev_diff_corrected"] = 0
        fight_features["control_dominance"] = 0

        # Compound positional features (using career averages)
        fight_features["clinch_x_grappling"] = fight_features["clinch_pct_diff_corrected"] * fight_features.get("grappler_advantage", 0)
        fight_features["distance_x_striking"] = fight_features["distance_pct_diff_corrected"] * fight_features.get("net_striking_advantage", 0)
        fight_features["ground_x_control"] = (
            fight_features["ground_pct_diff_corrected"] *
            (fight_features["avg_ctrl_sec_diff_corrected"] / 300.0)
        )
        fight_features["positional_mastery"] = (
            fight_features["positional_striking_advantage"] *
            fight_features.get("grappler_advantage", 0) *
            fight_features.get("net_striking_advantage", 0)
        )

        return fight_features, r_stats, b_stats

    def predict_fight(self, fight_data, feature_columns):
        """Predict winner for a single fight with confidence"""
        # Ensure we have required columns
        X = pd.DataFrame([fight_data])

        # Add missing columns
        for col in feature_columns:
            if col not in X.columns:
                X[col] = 0

        X = X[feature_columns]

        # Get winner prediction with calibrated probabilities
        winner_proba = self.winner_model.predict_proba(X)[0]
        winner_pred = self.winner_model.predict(X)[0]
        winner_name = "Red" if winner_pred == 1 else "Blue"

        # Calculate confidence level
        winner_confidence = winner_proba[winner_pred]
        if winner_confidence >= 0.70:
            confidence_level = "HIGH"
        elif winner_confidence >= 0.60:
            confidence_level = "MEDIUM"
        else:
            confidence_level = "LOW"

        # Betting recommendation (confidence threshold = 65%)
        should_bet = winner_confidence >= 0.65

        return {
            "winner": winner_name,
            "winner_confidence": winner_confidence,
            "confidence_level": confidence_level,
            "red_prob": winner_proba[1],
            "blue_prob": winner_proba[0],
            "should_bet": should_bet,
        }

    def predict_upcoming_fights(self, upcoming_fights, feature_columns):
        """Predict upcoming fights with enhanced confidence metrics"""
        predictions = []
        skipped_fights = []

        for fight in upcoming_fights:
            result = self.prepare_upcoming_fight(fight, feature_columns)
            if not result[0]:
                skipped_fights.append(
                    f"{fight['red_fighter']} vs {fight['blue_fighter']}"
                )
                continue

            fight_features, r_stats, b_stats = result

            pred = self.predict_fight(fight_features, feature_columns)

            predictions.append(
                {
                    "Red Fighter": fight["red_fighter"],
                    "Blue Fighter": fight["blue_fighter"],
                    "Weight Class": fight["weight_class"],
                    "Winner": fight["red_fighter"]
                    if pred["winner"] == "Red"
                    else fight["blue_fighter"],
                    "Win%": pred["winner_confidence"],
                    # "Confidence": pred["confidence_level"],
                    # "Should Bet": "YES" if pred["should_bet"] else "NO",
                }
            )

        return predictions, skipped_fights

    def export_predictions_to_excel(self, predictions, filename="predictions.xlsx"):
        """Export predictions to formatted Excel file with confidence and betting metrics"""
        df = pd.DataFrame(predictions)

        # Ensure percentage columns are in decimal format (0-1 range)
        percentage_columns = ["Win%"]
        for col in percentage_columns:
            if col in df.columns:
                if df[col].max() > 1:
                    df[col] = df[col] / 100
                df[col] = df[col].round(4)

        df.to_excel(filename, index=False, engine="openpyxl")

        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        # Styling
        header_fill = PatternFill(
            start_color="D20A0A", end_color="D20A0A", fill_type="solid"
        )
        header_font = Font(bold=True, size=11, color="FFFFFF")
        left_alignment = Alignment(horizontal="left", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Apply formatting to ALL cells
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.alignment = left_alignment
                cell.border = thin_border

        # Apply header-specific formatting
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Center align specific columns
        for col_name in ["Confidence", "Should Bet"]:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).alignment = center_alignment

        # Format percentage columns
        percentage_col_names = ["Win%"]
        for col_name in percentage_col_names:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.number_format = "0.00%"

        # Highlight high confidence bets (green)
        high_confidence_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        # Low confidence (yellow)
        low_confidence_fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        )

        if "Confidence" in df.columns:
            conf_col_idx = df.columns.get_loc("Confidence") + 1
            for row_idx in range(2, ws.max_row + 1):
                conf_cell = ws.cell(row=row_idx, column=conf_col_idx)
                if conf_cell.value == "HIGH":
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = high_confidence_fill
                elif conf_cell.value == "LOW":
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = low_confidence_fill

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value is None:
                        cell_length = 0
                    elif isinstance(cell.value, (int, float)):
                        cell_length = (
                            10
                            if cell.number_format == "0.00%"
                            else len(str(cell.value))
                        )
                    else:
                        cell_length = len(str(cell.value))

                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass

            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(filename)
        print(f"\nPredictions exported to: {filename}")
        print(f"  - {len(predictions)} fights predicted")
        print(f"  - HIGH confidence fights highlighted in green")
        print(f"  - LOW confidence fights highlighted in yellow")
        return filename


# ===== GUI APPLICATION =====

class UFCPredictorGUI:
    """Complete GUI matching original model with improved backend"""

    def __init__(self, root):
        self.root = root
        self.root.title("UFC Fight Predictor - Winner Only")
        self.root.geometry("1000x800")
        self.root.minsize(700, 550)

        self.data_file_path = tk.StringVar(value=fight_data_path)
        self.output_file_path = tk.StringVar(value="UFC_predictions_winners.xlsx")
        self.predictor = None
        self.create_widgets()

    def create_widgets(self):
        # UFC-themed header
        title_frame = tk.Frame(self.root, bg="#D20A0A")
        title_frame.pack(fill=tk.X)
        tk.Label(
            title_frame,
            text="UFC FIGHT PREDICTOR - WINNER ONLY",
            font=("Arial", 16, "bold"),
            fg="white",
            bg="#D20A0A",
        ).pack(pady=(10, 8))

        # Data file selection
        file_frame = ttk.LabelFrame(self.root, text="Training Data File", padding="5")
        file_frame.pack(fill=tk.X, padx=10, pady=3)
        ttk.Entry(file_frame, textvariable=self.data_file_path, width=70).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(file_frame, text="Browse", command=self.browse_data_file).pack(
            side=tk.LEFT
        )

        # Fights input area
        input_frame = ttk.LabelFrame(
            self.root, text="Enter Fights to Predict", padding="5"
        )
        input_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=3)

        instructions = """Enter fights in CSV format (one per line):
Format: red_fighter,blue_fighter,weight_class,gender,total_rounds

Example:
Max Holloway,Dustin Poirier,Lightweight,Men,5
Ilia Topuria,Charles Oliveira,Lightweight,Men,5
Tatiana Suarez,Amanda Lemos,Women's Strawweight,Women,3"""

        ttk.Label(
            input_frame,
            text=instructions,
            font=("Arial", 8),
            foreground="gray",
            justify=tk.LEFT,
        ).pack(anchor=tk.W, pady=(0, 3))

        self.fights_text = scrolledtext.ScrolledText(
            input_frame, height=8, width=95, wrap=tk.WORD
        )
        self.fights_text.pack(fill=tk.BOTH, expand=True, pady=3)

        # Output file selection
        output_frame = ttk.LabelFrame(self.root, text="Output File", padding="5")
        output_frame.pack(fill=tk.X, padx=10, pady=3)
        ttk.Entry(output_frame, textvariable=self.output_file_path, width=70).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(output_frame, text="Browse", command=self.browse_output_file).pack(
            side=tk.LEFT
        )

        # Buttons
        button_frame = ttk.Frame(self.root, padding="5")
        button_frame.pack(fill=tk.X, padx=10)
        ttk.Button(button_frame, text="Load Sample", command=self.load_sample).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(button_frame, text="Clear", command=self.clear_input).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(
            button_frame, text="Generate Predictions", command=self.run_predictions
        ).pack(side=tk.RIGHT, padx=3)

        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(
            self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W
        )
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)

    def browse_data_file(self):
        filename = filedialog.askopenfilename(
            title="Select Fight Data CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if filename:
            self.data_file_path.set(filename)

    def browse_output_file(self):
        filename = filedialog.asksaveasfilename(
            title="Save Predictions As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if filename:
            self.output_file_path.set(filename)

    def load_sample(self):
        sample = """Max Holloway,Dustin Poirier,Lightweight,Men,5
Ilia Topuria,Charles Oliveira,Lightweight,Men,5
Tatiana Suarez,Amanda Lemos,Women's Strawweight,Women,3"""
        self.fights_text.delete("1.0", tk.END)
        self.fights_text.insert("1.0", sample)

    def clear_input(self):
        self.fights_text.delete("1.0", tk.END)

    def parse_fights_input(self, text):
        """Parse CSV fight input"""
        text = text.strip()
        fights = []

        for line in text.split("\n"):
            line = line.strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split(",")]

            if len(parts) != 5:
                raise ValueError(
                    f"Invalid CSV format. Expected 5 fields, got {len(parts)} in line: {line}"
                )

            try:
                fights.append(
                    {
                        "red_fighter": parts[0],
                        "blue_fighter": parts[1],
                        "weight_class": parts[2],
                        "gender": parts[3],
                        "total_rounds": int(parts[4]),
                    }
                )
            except ValueError as e:
                raise ValueError(
                    f"Error parsing line: {line}\nTotal rounds must be a number. {str(e)}"
                )

        if not fights:
            raise ValueError("No valid fight data found.")
        return fights

    def run_predictions(self):
        """Run predictions with model"""
        try:
            self.status_var.set("Loading data and training model...")
            self.root.update()

            fights_text = self.fights_text.get("1.0", tk.END)
            upcoming_fights = self.parse_fights_input(fights_text)

            data_file = self.data_file_path.get()
            if not os.path.exists(data_file):
                raise FileNotFoundError(f"Data file not found: {data_file}")

            df = pd.read_csv(data_file)
            self.status_var.set(
                f"Loaded {len(df)} fights. Training..."
            )
            self.root.update()

            # Create improved predictor
            self.predictor = ImprovedUFCPredictor(use_ensemble=False, debug_mode=False)

            # Train models (output goes to console, not GUI)
            print("\n" + "="*80)
            print("TRAINING UFC PREDICTOR")
            print("="*80 + "\n")

            df = self.predictor.fix_data_leakage(df)
            self.predictor.df_train = df
            feature_columns = self.predictor.train(df)

            self.status_var.set("Generating predictions...")
            self.root.update()

            # Generate predictions
            predictions, skipped_fights = self.predictor.predict_upcoming_fights(
                upcoming_fights, feature_columns
            )

            if not predictions:
                self.status_var.set("No predictions generated")
                messagebox.showerror(
                    "No Predictions",
                    "All fights were skipped due to insufficient fighter data.",
                )
                return

            output_file = self.output_file_path.get()

            # Export to Excel
            with redirect_stdout(io.StringIO()):
                self.predictor.export_predictions_to_excel(predictions, output_file)

            # Clean up temporary files
            cleanup_temp_files()

            # Success message
            success_msg = f"Predictions generated!\n\nSaved to: {output_file}\n\n{len(predictions)} fight(s) predicted"

            if skipped_fights:
                success_msg += f"\n\n{len(skipped_fights)} fight(s) skipped"
                self.status_var.set(
                    f"Success! {len(predictions)} predictions, {len(skipped_fights)} skipped"
                )
                skipped_msg = (
                    success_msg
                    + "\n\nSkipped:\n"
                    + "\n".join(f"- {fight}" for fight in skipped_fights)
                )
                messagebox.showwarning("Predictions Complete", skipped_msg)
            else:
                self.status_var.set(
                    f"Success! All {len(predictions)} predictions saved"
                )
                messagebox.showinfo("Success", success_msg)

        except Exception as e:
            import traceback

            error_details = traceback.format_exc()
            self.status_var.set("Error occurred")
            messagebox.showerror(
                "Error", f"Error:\n\n{str(e)}\n\nDetails:\n{error_details}"
            )


def main():
    """Main function to run the improved UFC Predictor GUI"""
    try:
        root = tk.Tk()
        app = UFCPredictorGUI(root)
        root.mainloop()
    except Exception as e:
        print(f"Error starting UFC Predictor: {e}")
        import traceback
        traceback.print_exc()
    finally:
        cleanup_temp_files()


if __name__ == "__main__":
    # Enable multiprocessing
    if hasattr(sys, "frozen") and sys.frozen:
        import multiprocessing
        multiprocessing.freeze_support()

    main()