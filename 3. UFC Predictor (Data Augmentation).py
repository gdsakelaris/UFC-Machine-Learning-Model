import sys
import os
import random
import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog, messagebox
import pandas as pd
import numpy as np
import io
from contextlib import redirect_stdout
from sklearn.model_selection import train_test_split, TimeSeriesSplit
from sklearn.ensemble import (
    RandomForestClassifier,
    VotingClassifier,
    StackingClassifier,
)
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
from sklearn.feature_selection import SelectPercentile, f_classif
from sklearn.metrics import log_loss
from sklearn.calibration import CalibratedClassifierCV
from sklearn.linear_model import LogisticRegression
from sklearn.neural_network import MLPClassifier
from scipy.stats import linregress
import warnings
import shutil
import atexit
from joblib import Parallel, delayed
import multiprocessing as mp

# Enable multiprocessing for faster training
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["VECLIB_MAXIMUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"
os.environ["NUMEXPR_MAX_THREADS"] = "1"
os.environ["XGBOOST_DISABLE_MULTIPROCESSING"] = "1"

# Enable multiprocessing for our custom parallel operations
os.environ["JOBLIB_MULTIPROCESSING"] = "1"
os.environ["LOKY_MAX_WORKERS"] = "12"

# Note: multiprocessing.freeze_support() will be called later in the main execution block

# Get the directory where this script is located
# Handle both regular Python execution and PyInstaller executable
if getattr(sys, "frozen", False):
    # Running as PyInstaller executable
    script_dir = os.path.dirname(sys.executable)
else:
    # Running as regular Python script
    script_dir = os.path.dirname(os.path.abspath(__file__))

# Check for fight_data.csv in the same directory as the script/executable
fight_data_path = os.path.join(script_dir, "fight_data.csv")
if not os.path.exists(fight_data_path):
    print(f"Warning: fight_data.csv not found in script directory: {script_dir}")


# Set random seeds for reproducibility
random.seed(42)
np.random.seed(42)

# Additional deterministic settings
os.environ["PYTHONHASHSEED"] = "42"
os.environ["TF_DETERMINISTIC_OPS"] = "1"
os.environ["TF_CUDNN_DETERMINISTIC"] = "1"


warnings.filterwarnings("ignore")


def cleanup_temp_files():
    """Clean up temporary files and folders created during training"""
    try:
        # Remove catboost_info folder if it exists
        if os.path.exists("catboost_info"):
            shutil.rmtree("catboost_info")
            print("Cleaned up catboost_info folder")

        # Remove best_dl_model.h5 file if it exists
        if os.path.exists("best_dl_model.h5"):
            os.remove("best_dl_model.h5")
            print("Cleaned up best_dl_model.h5 file")

    except Exception as e:
        print(f"Warning: Could not clean up temporary files: {e}")


# Register cleanup function to run on exit
atexit.register(cleanup_temp_files)

try:
    from xgboost import XGBClassifier

    HAS_XGBOOST = True
except ImportError:
    HAS_XGBOOST = False
    print("XGBoost not available.")

try:
    from lightgbm import LGBMClassifier

    HAS_LIGHTGBM = True
except ImportError:
    HAS_LIGHTGBM = False
    print("LightGBM not available. Install with: pip install lightgbm")

try:
    from catboost import CatBoostClassifier

    HAS_CATBOOST = True
except ImportError:
    HAS_CATBOOST = False
    print("CatBoost not available. Install with: pip install catboost")

try:
    import tensorflow as tf
    from tensorflow import keras

    HAS_TENSORFLOW = True
except ImportError:
    HAS_TENSORFLOW = False
    print("TensorFlow not available. Install with: pip install tensorflow")


class AdvancedUFCPredictor:
    def __init__(
        self,
        use_ensemble=True,
        use_neural_net=False,
        use_deep_learning=False,
        debug_mode=False,
    ):
        self.winner_model = None
        self.method_model = None  # Single method prediction model
        self.deep_learning_model = None  # TensorFlow model
        self.label_encoders = {}
        self.use_ensemble = use_ensemble
        self.use_neural_net = use_neural_net
        self.use_deep_learning = use_deep_learning and HAS_TENSORFLOW
        self.debug_mode = debug_mode
        self.df_train = None
        self.fighter_style_cache = {}
        self.fighter_encoder = None  # For fighter embeddings
        self.num_fighters = 0

        # MEMORY OPTIMIZATION: Feature caching for 2x speedup
        self.feature_cache = {}
        self.preprocessor_cache = None
        self.feature_columns_cache = None

        # EARLY STOPPING: Performance tracking
        self.early_stopping_patience = 10
        self.best_score = 0
        self.patience_counter = 0

        # Set all random seeds for maximum reproducibility
        self.set_random_seeds()

    def set_random_seeds(self):
        """Set all random seeds for maximum reproducibility while maintaining accuracy"""
        # Python random
        random.seed(42)

        # NumPy random
        np.random.seed(42)

        # Set environment variables for deterministic operations
        os.environ["PYTHONHASHSEED"] = "42"
        os.environ["TF_DETERMINISTIC_OPS"] = "1"
        os.environ["TF_CUDNN_DETERMINISTIC"] = "1"

        # XGBoost specific random seed settings
        os.environ["XGBOOST_DISABLE_MULTIPROCESSING"] = "1"

        # GPU device management to prevent conflicts
        os.environ["CUDA_VISIBLE_DEVICES"] = "0"

        # Clear GPU memory to prevent conflicts
        try:
            import gc

            gc.collect()
        except:
            pass

        # Set TensorFlow random seed if available
        if HAS_TENSORFLOW:
            try:
                tf.random.set_seed(42)
                # Enable deterministic operations for consistency
                tf.config.experimental.enable_op_determinism()
            except:
                pass

    def calculate_streak(self, recent_wins, count_wins=True):
        """Calculate current win or loss streak"""
        if not recent_wins:
            return 0

        streak = 0
        target = 1 if count_wins else 0

        for result in reversed(recent_wins):
            if result == target:
                streak += 1
            else:
                break

        return streak

    def calculate_trajectory(self, values):
        """Calculate linear trend of recent performance"""
        if len(values) < 2:
            return 0
        try:
            x = np.arange(len(values))
            slope, _, _, _, _ = linregress(x, values)
            return slope
        except:
            return 0

    def classify_fighter_style(self, stats):
        """Classify fighter as striker, grappler, or balanced"""
        striker_score = (
            stats.get("pro_SLpM_corrected", 0) * 0.4
            + stats.get("distance_pct_corrected", 0) * 0.3
            + stats.get("head_pct_corrected", 0) * 0.3
            - stats.get("ground_pct_corrected", 0) * 0.4
        )

        grappler_score = (
            stats.get("pro_td_avg_corrected", 0) * 0.4
            + stats.get("pro_sub_avg_corrected", 0) * 0.3
            + stats.get("ground_pct_corrected", 0) * 0.3
            - stats.get("distance_pct_corrected", 0) * 0.2
        )

        if striker_score > 0.3 and striker_score > grappler_score:
            return "striker"
        elif grappler_score > 0.3 and grappler_score > striker_score:
            return "grappler"
        else:
            return "balanced"

    def calculate_age_curve_factor(self, age):
        """Calculate performance multiplier based on age"""
        if age < 25:
            return 0.92  # Still developing
        elif 25 <= age <= 32:
            return 1.0  # Prime years
        elif 33 <= age <= 36:
            return 0.94  # Slight decline
        else:
            return 0.85  # Significant decline

    def is_fighter_at_peak(self, wins, losses, age, recent_form):
        """Determine if fighter is at career peak"""
        total_fights = wins + losses
        experience_factor = min(total_fights / 20, 1.0)  # Peaks around 20 fights
        age_factor = self.calculate_age_curve_factor(age)
        form_factor = recent_form

        peak_score = experience_factor * 0.4 + age_factor * 0.3 + form_factor * 0.3
        return peak_score

    # ===== ADVANCED TEMPORAL FEATURE METHODS =====
    
    def calculate_fighter_trajectory(self, df, fighter_col, window):
        """Calculate fighter performance trajectory over last N fights"""
        trajectories = []
        for idx, row in df.iterrows():
            fighter = row[fighter_col]
            if pd.isna(fighter):
                trajectories.append(0.0)
                continue
                
            try:
                # Get all fights for this fighter up to current fight
                fighter_fights = df[df[fighter_col] == fighter].copy()
                if 'event_date' in fighter_fights.columns:
                    fighter_fights = fighter_fights.sort_values('event_date')
                
                if len(fighter_fights) < 2:
                    trajectories.append(0.0)
                    continue
                    
                # Get last N fights (excluding current)
                recent_fights = fighter_fights.iloc[:-1].tail(window)
            except (KeyError, IndexError):
                trajectories.append(0.0)
                continue
            
            if len(recent_fights) < 2:
                trajectories.append(0.0)
                continue
                
            # Calculate trajectory as slope of win rate over time
            win_rates = []
            for i, fight in recent_fights.iterrows():
                if 'winner' in fight and 'r_fighter' in fight and 'b_fighter' in fight:
                    if fight['r_fighter'] == fighter:
                        win_rates.append(1.0 if fight['winner'] == 'Red' else 0.0)
                    elif fight['b_fighter'] == fighter:
                        win_rates.append(1.0 if fight['winner'] == 'Blue' else 0.0)
                    else:
                        win_rates.append(0.5)  # Neutral if fighter not found
                else:
                    win_rates.append(0.5)
            
            if len(win_rates) >= 2:
                # Calculate linear regression slope
                x = np.arange(len(win_rates))
                try:
                    slope, _, _, _, _ = linregress(x, win_rates)
                    trajectories.append(slope)
                except (ValueError, ZeroDivisionError):
                    trajectories.append(0.0)
            else:
                trajectories.append(0.0)
                
        return np.array(trajectories)

    def calculate_opponent_quality(self, df, fighter_col):
        """Calculate average opponent quality for each fighter"""
        qualities = []
        for idx, row in df.iterrows():
            fighter = row[fighter_col]
            if pd.isna(fighter):
                qualities.append(0.0)
                continue
                
            try:
                # Get all fights for this fighter up to current fight
                fighter_fights = df[df[fighter_col] == fighter].copy()
                if 'event_date' in fighter_fights.columns:
                    fighter_fights = fighter_fights.sort_values('event_date')
                
                if len(fighter_fights) < 2:
                    qualities.append(0.0)
                    continue
                    
                # Get last 5 fights (excluding current)
                recent_fights = fighter_fights.iloc[:-1].tail(5)
            except (KeyError, IndexError):
                qualities.append(0.0)
                continue
            
            opponent_qualities = []
            for i, fight in recent_fights.iterrows():
                if 'r_fighter' in fight and 'b_fighter' in fight:
                    if fight['r_fighter'] == fighter:
                        opponent = fight['b_fighter']
                    else:
                        opponent = fight['r_fighter']
                    
                    # Calculate opponent's win rate
                    opponent_fights = df[(df['r_fighter'] == opponent) | (df['b_fighter'] == opponent)]
                    if len(opponent_fights) > 0:
                        wins = 0
                        for _, opp_fight in opponent_fights.iterrows():
                            if 'winner' in opp_fight:
                                if (opp_fight['r_fighter'] == opponent and opp_fight['winner'] == 'Red') or \
                                   (opp_fight['b_fighter'] == opponent and opp_fight['winner'] == 'Blue'):
                                    wins += 1
                        win_rate = wins / len(opponent_fights) if len(opponent_fights) > 0 else 0.5
                        opponent_qualities.append(win_rate)
                    else:
                        opponent_qualities.append(0.5)
            
            if opponent_qualities:
                qualities.append(np.mean(opponent_qualities))
            else:
                qualities.append(0.5)
                
        return np.array(qualities)

    def calculate_fight_frequency(self, df, fighter_col):
        """Calculate fight frequency (fights per year) for each fighter"""
        frequencies = []
        for idx, row in df.iterrows():
            fighter = row[fighter_col]
            if pd.isna(fighter):
                frequencies.append(0.0)
                continue
                
            try:
                # Get all fights for this fighter up to current fight
                fighter_fights = df[df[fighter_col] == fighter].copy()
                if 'event_date' in fighter_fights.columns:
                    fighter_fights = fighter_fights.sort_values('event_date')
                    fighter_fights = fighter_fights.iloc[:-1]  # Exclude current fight
            except (KeyError, IndexError):
                frequencies.append(0.0)
                continue
                
                if len(fighter_fights) >= 2:
                    # Calculate time span
                    try:
                        first_fight = pd.to_datetime(fighter_fights['event_date'].iloc[0])
                        last_fight = pd.to_datetime(fighter_fights['event_date'].iloc[-1])
                        time_span_years = (last_fight - first_fight).days / 365.25
                        
                        if time_span_years > 0:
                            frequency = len(fighter_fights) / time_span_years
                            frequencies.append(frequency)
                        else:
                            frequencies.append(0.0)
                    except (ValueError, TypeError, ZeroDivisionError):
                        frequencies.append(0.0)
                else:
                    frequencies.append(0.0)
            else:
                frequencies.append(0.0)
                
        return np.array(frequencies)

    def calculate_career_stage(self, df, fighter_col):
        """Calculate career stage: 0=rookie, 1=developing, 2=prime, 3=declining, 4=veteran"""
        stages = []
        for idx, row in df.iterrows():
            fighter = row[fighter_col]
            if pd.isna(fighter):
                stages.append(2.0)  # Default to prime
                continue
                
            try:
                # Get all fights for this fighter up to current fight
                fighter_fights = df[df[fighter_col] == fighter].copy()
                if 'event_date' in fighter_fights.columns:
                    fighter_fights = fighter_fights.sort_values('event_date')
                
                total_fights = len(fighter_fights) - 1  # Exclude current fight
                age = row.get('r_age_at_event' if fighter_col == 'r_fighter' else 'b_age_at_event', 25)
            except (KeyError, IndexError):
                stages.append(2.0)  # Default to prime
                continue
            
            if total_fights <= 3:
                stage = 0  # Rookie
            elif total_fights <= 8:
                stage = 1  # Developing
            elif total_fights <= 20 and age <= 32:
                stage = 2  # Prime
            elif total_fights <= 30 and age <= 35:
                stage = 3  # Declining
            else:
                stage = 4  # Veteran
                
            stages.append(float(stage))
            
        return np.array(stages)

    # ===== CONTEXTUAL FEATURE METHODS =====
    
    def calculate_weight_class_factor(self, df):
        """Calculate weight class specific factors"""
        factors = []
        for idx, row in df.iterrows():
            weight_class = row.get('weight_class', 'Unknown')
            
            # Different weight classes have different fighting dynamics
            if 'Heavyweight' in str(weight_class):
                factor = 1.2  # More KO power, less technical
            elif 'Light Heavyweight' in str(weight_class):
                factor = 1.1  # Good balance of power and technique
            elif 'Middleweight' in str(weight_class):
                factor = 1.0  # Baseline
            elif 'Welterweight' in str(weight_class):
                factor = 0.95  # More technical, less power
            elif 'Lightweight' in str(weight_class):
                factor = 0.9  # Very technical, speed focused
            elif 'Featherweight' in str(weight_class):
                factor = 0.85  # High pace, technical
            elif 'Bantamweight' in str(weight_class):
                factor = 0.8  # Very high pace, technical
            elif 'Flyweight' in str(weight_class):
                factor = 0.75  # Highest pace, most technical
            else:
                factor = 1.0  # Default
                
            factors.append(factor)
            
        return np.array(factors)

    def calculate_location_advantage(self, df):
        """Calculate location-based advantage"""
        advantages = []
        for idx, row in df.iterrows():
            location = row.get('event_location', 'Unknown')
            
            # Simple location advantage calculation
            # In reality, this would be more complex based on fighter origins
            if 'Las Vegas' in str(location) or 'Nevada' in str(location):
                advantage = 0.0  # Neutral venue
            elif 'New York' in str(location) or 'New Jersey' in str(location):
                advantage = 0.0  # Neutral venue
            else:
                advantage = 0.0  # Default neutral
                
            advantages.append(advantage)
            
        return np.array(advantages)

    def calculate_event_significance(self, df):
        """Calculate event significance factor"""
        significances = []
        for idx, row in df.iterrows():
            is_title = row.get('is_title_bout', 0)
            total_rounds = row.get('total_rounds', 3)
            event_name = row.get('event_name', '')
            
            # Title fights are more significant
            title_factor = 1.5 if is_title else 1.0
            
            # 5-round fights are more significant
            rounds_factor = 1.2 if total_rounds == 5 else 1.0
            
            # Main event factor (simplified)
            main_event_factor = 1.1 if 'main event' in str(event_name).lower() else 1.0
            
            significance = title_factor * rounds_factor * main_event_factor
            significances.append(significance)
            
        return np.array(significances)

    def calculate_referee_factor(self, df):
        """Calculate referee-specific factors"""
        factors = []
        for idx, row in df.iterrows():
            referee = row.get('referee', 'Unknown')
            
            # Simple referee factor (in reality, would be based on historical data)
            # Some refs are known for stopping fights early, others let them go longer
            if 'Herb' in str(referee):
                factor = 1.0  # Neutral
            elif 'Mazzagatti' in str(referee):
                factor = 0.9  # Known for late stoppages
            else:
                factor = 1.0  # Default neutral
                
            factors.append(factor)
            
        return np.array(factors)

    # ===== ADVANCED STATISTICAL FEATURE METHODS =====
    
    def calculate_momentum_quality(self, df, fighter_col):
        """Calculate momentum with quality weighting"""
        momentums = []
        for idx, row in df.iterrows():
            fighter = row[fighter_col]
            if pd.isna(fighter):
                momentums.append(0.0)
                continue
                
            try:
                # Get recent fights for this fighter
                fighter_fights = df[df[fighter_col] == fighter].copy()
                if 'event_date' in fighter_fights.columns:
                    fighter_fights = fighter_fights.sort_values('event_date')
                recent_fights = fighter_fights.iloc[:-1].tail(5)  # Last 5 fights
            except (KeyError, IndexError):
                momentums.append(0.0)
                continue
            
            if len(recent_fights) == 0:
                momentums.append(0.0)
                continue
                
            weighted_wins = 0
            total_weight = 0
            
            for i, fight in recent_fights.iterrows():
                # Weight more recent fights higher
                weight = 1.0 / (len(recent_fights) - i)
                
                # Check if fighter won
                won = False
                if 'winner' in fight and 'r_fighter' in fight and 'b_fighter' in fight:
                    if fight['r_fighter'] == fighter and fight['winner'] == 'Red':
                        won = True
                    elif fight['b_fighter'] == fighter and fight['winner'] == 'Blue':
                        won = True
                
                # Add opponent quality weighting
                opponent_quality = 0.5  # Default
                if 'r_fighter' in fight and 'b_fighter' in fight:
                    opponent = fight['b_fighter'] if fight['r_fighter'] == fighter else fight['r_fighter']
                    opponent_fights = df[(df['r_fighter'] == opponent) | (df['b_fighter'] == opponent)]
                    if len(opponent_fights) > 0:
                        wins = 0
                        for _, opp_fight in opponent_fights.iterrows():
                            if 'winner' in opp_fight:
                                if (opp_fight['r_fighter'] == opponent and opp_fight['winner'] == 'Red') or \
                                   (opp_fight['b_fighter'] == opponent and opp_fight['winner'] == 'Blue'):
                                    wins += 1
                        opponent_quality = wins / len(opponent_fights) if len(opponent_fights) > 0 else 0.5
                
                # Weight by opponent quality
                weight *= (0.5 + opponent_quality)
                
                if won:
                    weighted_wins += weight
                total_weight += weight
            
            momentum = weighted_wins / total_weight if total_weight > 0 else 0.0
            momentums.append(momentum)
            
        return np.array(momentums)

    def calculate_pressure_performance(self, df, fighter_col):
        """Calculate performance under pressure (title fights, main events)"""
        performances = []
        for idx, row in df.iterrows():
            fighter = row[fighter_col]
            if pd.isna(fighter):
                performances.append(0.5)
                continue
                
            try:
                # Get high-pressure fights for this fighter
                fighter_fights = df[df[fighter_col] == fighter].copy()
                if 'event_date' in fighter_fights.columns:
                    fighter_fights = fighter_fights.sort_values('event_date')
                fighter_fights = fighter_fights.iloc[:-1]  # Exclude current fight
            except (KeyError, IndexError):
                performances.append(0.5)
                continue
            
            # Filter for high-pressure fights
            high_pressure_fights = fighter_fights[
                (fighter_fights.get('is_title_bout', 0) == 1) |
                (fighter_fights.get('total_rounds', 3) == 5)
            ]
            
            if len(high_pressure_fights) == 0:
                performances.append(0.5)  # Default neutral
                continue
                
            wins = 0
            for _, fight in high_pressure_fights.iterrows():
                if 'winner' in fight and 'r_fighter' in fight and 'b_fighter' in fight:
                    if (fight['r_fighter'] == fighter and fight['winner'] == 'Red') or \
                       (fight['b_fighter'] == fighter and fight['winner'] == 'Blue'):
                        wins += 1
            
            performance = wins / len(high_pressure_fights) if len(high_pressure_fights) > 0 else 0.5
            performances.append(performance)
            
        return np.array(performances)

    def calculate_style_matchup(self, df):
        """Calculate style matchup advantage"""
        advantages = []
        for idx, row in df.iterrows():
            r_stance = row.get('r_stance', 'Orthodox')
            b_stance = row.get('b_stance', 'Orthodox')
            
            # Stance matchup analysis
            if r_stance == 'Southpaw' and b_stance == 'Orthodox':
                advantage = 0.1  # Southpaw advantage
            elif r_stance == 'Orthodox' and b_stance == 'Southpaw':
                advantage = -0.1  # Southpaw disadvantage
            elif r_stance == 'Switch' and b_stance != 'Switch':
                advantage = 0.05  # Switch advantage
            elif b_stance == 'Switch' and r_stance != 'Switch':
                advantage = -0.05  # Switch disadvantage
            else:
                advantage = 0.0  # Neutral
                
            advantages.append(advantage)
            
        return np.array(advantages)

    def calculate_physical_advantage_composite(self, df):
        """Calculate composite physical advantage"""
        advantages = []
        for idx, row in df.iterrows():
            height_diff = row.get('height_diff', 0)
            reach_diff = row.get('reach_diff', 0)
            weight_diff = row.get('weight_diff', 0)
            age_diff = row.get('age_at_event_diff', 0)
            
            # Normalize and weight each factor
            height_adv = np.tanh(height_diff / 3.0) * 0.3  # Height advantage
            reach_adv = np.tanh(reach_diff / 3.0) * 0.4   # Reach advantage
            weight_adv = np.tanh(weight_diff / 10.0) * 0.2  # Weight advantage
            age_adv = np.tanh(-age_diff / 5.0) * 0.1      # Age advantage (younger is better)
            
            composite = height_adv + reach_adv + weight_adv + age_adv
            advantages.append(composite)
            
        return np.array(advantages)

    def calculate_fight_iq_differential(self, df):
        """Calculate fight IQ differential based on technical metrics"""
        diffs = []
        for idx, row in df.iterrows():
            # Technical metrics that indicate fight IQ
            r_acc = row.get('r_pro_sig_str_acc_corrected', 0.5)
            b_acc = row.get('b_pro_sig_str_acc_corrected', 0.5)
            r_def = row.get('r_pro_str_def_corrected', 0.5)
            b_def = row.get('b_pro_str_def_corrected', 0.5)
            r_td_acc = row.get('r_pro_td_acc_corrected', 0.5)
            b_td_acc = row.get('b_pro_td_acc_corrected', 0.5)
            
            # Calculate fight IQ score
            r_iq = (r_acc + r_def + r_td_acc) / 3.0
            b_iq = (b_acc + b_def + b_td_acc) / 3.0
            
            diff = r_iq - b_iq
            diffs.append(diff)
            
        return np.array(diffs)

    def calculate_mental_toughness_differential(self, df):
        """Calculate mental toughness differential"""
        diffs = []
        for idx, row in df.iterrows():
            # Factors that indicate mental toughness
            r_wins = row.get('r_wins_corrected', 0)
            b_wins = row.get('b_wins_corrected', 0)
            r_losses = row.get('r_losses_corrected', 0)
            b_losses = row.get('b_losses_corrected', 0)
            
            # Calculate comeback ability (wins after losses)
            r_comeback = r_wins / (r_losses + 1) if r_losses > 0 else (r_wins if r_wins > 0 else 0.5)
            b_comeback = b_wins / (b_losses + 1) if b_losses > 0 else (b_wins if b_wins > 0 else 0.5)
            
            diff = r_comeback - b_comeback
            diffs.append(diff)
            
        return np.array(diffs)

    def calculate_injury_resistance_differential(self, df):
        """Calculate injury resistance differential"""
        diffs = []
        for idx, row in df.iterrows():
            # Factors that indicate injury resistance
            r_fights = row.get('r_total_fights', 0)
            b_fights = row.get('b_total_fights', 0)
            r_avg_time = row.get('r_avg_fight_time', 0)
            b_avg_time = row.get('b_avg_fight_time', 0)
            
            # More fights and longer average fight times indicate durability
            r_avg_time_safe = r_avg_time if r_avg_time > 0 else 7.5  # Default to 7.5 minutes
            b_avg_time_safe = b_avg_time if b_avg_time > 0 else 7.5  # Default to 7.5 minutes
            r_durability = r_fights * (r_avg_time_safe / 15.0)  # Normalize by 15 minutes
            b_durability = b_fights * (b_avg_time_safe / 15.0)
            
            diff = r_durability - b_durability
            diffs.append(diff)
            
        return np.array(diffs)

    def calculate_weight_cut_impact_differential(self, df):
        """Calculate weight cut impact differential"""
        diffs = []
        for idx, row in df.iterrows():
            # Weight class and weight difference can indicate cut difficulty
            weight_class = row.get('weight_class', 'Unknown')
            weight_diff = row.get('weight_diff', 0)
            
            # Heavier fighters in lower weight classes may have harder cuts
            if 'Flyweight' in str(weight_class):
                cut_factor = 1.2
            elif 'Bantamweight' in str(weight_class):
                cut_factor = 1.1
            elif 'Featherweight' in str(weight_class):
                cut_factor = 1.0
            elif 'Lightweight' in str(weight_class):
                cut_factor = 0.9
            elif 'Welterweight' in str(weight_class):
                cut_factor = 0.8
            else:
                cut_factor = 1.0
                
            # Positive weight difference means red corner is heavier (harder cut)
            impact = weight_diff * cut_factor * 0.01  # Scale down
            diffs.append(impact)
            
        return np.array(diffs)

    # ===== HYPERPARAMETER OPTIMIZATION METHODS =====
    
    def optimize_hyperparameters(self, X, y, model_type='xgb', n_trials=50):
        """Optimize hyperparameters using Bayesian optimization"""
        print(f"Optimizing {model_type} hyperparameters with {n_trials} trials...")
        
        if model_type == 'xgb' and HAS_XGBOOST:
            return self._optimize_xgb_hyperparameters(X, y, n_trials)
        elif model_type == 'lgbm' and HAS_LIGHTGBM:
            return self._optimize_lgbm_hyperparameters(X, y, n_trials)
        elif model_type == 'catboost' and HAS_CATBOOST:
            return self._optimize_catboost_hyperparameters(X, y, n_trials)
        else:
            print(f"Model type {model_type} not available, using default parameters")
            return {}

    def _optimize_xgb_hyperparameters(self, X, y, n_trials):
        """Optimize XGBoost hyperparameters using simple grid search"""
        best_params = {}
        best_score = 0
        
        # Define parameter grid
        param_grid = {
            'n_estimators': [800, 1000, 1200],
            'max_depth': [8, 10, 12],
            'learning_rate': [0.008, 0.01, 0.015],
            'subsample': [0.85, 0.9, 0.95],
            'colsample_bytree': [0.85, 0.9, 0.95],
            'reg_alpha': [0.03, 0.05, 0.08],
            'reg_lambda': [0.4, 0.5, 0.6],
            'min_child_weight': [1, 2, 3],
            'gamma': [0.03, 0.05, 0.08]
        }
        
        # Use TimeSeriesSplit for validation
        tscv = TimeSeriesSplit(n_splits=3)
        
        for i in range(min(n_trials, 27)):  # Limit to reasonable number
            # Sample parameters
            params = {}
            for key, values in param_grid.items():
                params[key] = np.random.choice(values)
            
            # Add fixed parameters
            params.update({
                'n_jobs': -1,
                'random_state': 42,
                'eval_metric': 'logloss',
                'tree_method': 'hist',
                'device': 'cpu',
                'seed': 42,
                'enable_categorical': True
            })
            
            # Cross-validation
            scores = []
            for train_idx, val_idx in tscv.split(X):
                X_train, X_val = X.iloc[train_idx], X.iloc[val_idx]
                y_train, y_val = y[train_idx], y[val_idx]
                
                model = XGBClassifier(**params)
                model.fit(X_train, y_train)
                y_pred = model.predict_proba(X_val)[:, 1]
                score = log_loss(y_val, y_pred)
                scores.append(score)
            
            avg_score = np.mean(scores)
            if avg_score > best_score:
                best_score = avg_score
                best_params = params.copy()
                print(f"New best score: {best_score:.4f}")
        
        print(f"Best XGBoost parameters: {best_params}")
        return best_params

    def _optimize_lgbm_hyperparameters(self, X, y, n_trials):
        """Optimize LightGBM hyperparameters"""
        best_params = {}
        best_score = 0
        
        param_grid = {
            'n_estimators': [300, 400, 500],
            'max_depth': [6, 7, 8],
            'learning_rate': [0.025, 0.03, 0.035],
            'num_leaves': [30, 40, 50],
            'subsample': [0.75, 0.8, 0.85],
            'colsample_bytree': [0.75, 0.8, 0.85],
            'reg_alpha': [0.08, 0.1, 0.12],
            'reg_lambda': [0.7, 0.8, 0.9]
        }
        
        tscv = TimeSeriesSplit(n_splits=3)
        
        for i in range(min(n_trials, 27)):
            params = {}
            for key, values in param_grid.items():
                params[key] = np.random.choice(values)
            
            params.update({
                'device': 'cpu',
                'random_state': 42,
                'verbose': -1
            })
            
            scores = []
            for train_idx, val_idx in tscv.split(X):
                X_train, X_val = X.iloc[train_idx], X.iloc[val_idx]
                y_train, y_val = y[train_idx], y[val_idx]
                
                model = LGBMClassifier(**params)
                model.fit(X_train, y_train)
                y_pred = model.predict_proba(X_val)[:, 1]
                score = log_loss(y_val, y_pred)
                scores.append(score)
            
            avg_score = np.mean(scores)
            if avg_score > best_score:
                best_score = avg_score
                best_params = params.copy()
                print(f"New best LGBM score: {best_score:.4f}")
        
        print(f"Best LightGBM parameters: {best_params}")
        return best_params

    def _optimize_catboost_hyperparameters(self, X, y, n_trials):
        """Optimize CatBoost hyperparameters"""
        best_params = {}
        best_score = 0
        
        param_grid = {
            'iterations': [300, 400, 500],
            'depth': [6, 7, 8],
            'learning_rate': [0.025, 0.03, 0.035],
            'l2_leaf_reg': [0.7, 0.8, 0.9]
        }
        
        tscv = TimeSeriesSplit(n_splits=3)
        
        for i in range(min(n_trials, 27)):
            params = {}
            for key, values in param_grid.items():
                params[key] = np.random.choice(values)
            
            params.update({
                'task_type': 'CPU',
                'random_state': 42,
                'verbose': 0
            })
            
            scores = []
            for train_idx, val_idx in tscv.split(X):
                X_train, X_val = X.iloc[train_idx], X.iloc[val_idx]
                y_train, y_val = y[train_idx], y[val_idx]
                
                model = CatBoostClassifier(**params)
                model.fit(X_train, y_train)
                y_pred = model.predict_proba(X_val)[:, 1]
                score = log_loss(y_val, y_pred)
                scores.append(score)
            
            avg_score = np.mean(scores)
            if avg_score > best_score:
                best_score = avg_score
                best_params = params.copy()
                print(f"New best CatBoost score: {best_score:.4f}")
        
        print(f"Best CatBoost parameters: {best_params}")
        return best_params

    # ===== ENHANCED ENSEMBLE METHODS =====
    
    def create_dynamic_ensemble(self, base_models, X, y):
        """Create dynamic ensemble with confidence-based weighting"""
        print("Creating dynamic ensemble with confidence-based weighting...")
        
        # Train base models
        trained_models = {}
        for name, model in base_models:
            print(f"Training {name}...")
            model.fit(X, y)
            trained_models[name] = model
        
        # Create meta-features for dynamic weighting
        meta_features = self._create_meta_features(X, trained_models)
        
        # Train dynamic weighting model
        dynamic_weights = self._train_dynamic_weights(meta_features, y)
        
        return trained_models, dynamic_weights

    def _create_meta_features(self, X, trained_models):
        """Create meta-features for dynamic weighting"""
        meta_features = []
        
        for name, model in trained_models.items():
            # Get predictions and confidence
            pred_proba = model.predict_proba(X)
            confidence = np.max(pred_proba, axis=1)
            
            # Get feature importance (if available)
            if hasattr(model, 'feature_importances_'):
                importance = model.feature_importances_
            else:
                importance = np.zeros(X.shape[1])
            
            meta_features.extend([confidence, importance])
        
        return np.column_stack(meta_features)

    def _train_dynamic_weights(self, meta_features, y):
        """Train dynamic weighting model"""
        # Simple linear model for dynamic weighting
        from sklearn.linear_model import Ridge
        
        # Create target for weighting (accuracy on validation set)
        tscv = TimeSeriesSplit(n_splits=3)
        weights_target = []
        
        for train_idx, val_idx in tscv.split(meta_features):
            X_train, X_val = meta_features[train_idx], meta_features[val_idx]
            y_train, y_val = y[train_idx], y[val_idx]
            
            # Calculate accuracy for each model
            val_acc = []
            for i in range(0, meta_features.shape[1], 2):  # Every other feature is confidence
                confidence = X_val[:, i]
                # Simple accuracy proxy based on confidence
                acc = np.mean(confidence)
                val_acc.append(acc)
            
            weights_target.extend(val_acc)
        
        # Train Ridge regression for dynamic weighting
        weight_model = Ridge(alpha=0.1)
        weight_model.fit(meta_features, weights_target)
        
        return weight_model

    def predict_with_dynamic_ensemble(self, X, trained_models, dynamic_weights):
        """Make predictions with dynamic ensemble"""
        predictions = []
        confidences = []
        
        for name, model in trained_models.items():
            pred_proba = model.predict_proba(X)
            pred = model.predict(X)
            confidence = np.max(pred_proba, axis=1)
            
            predictions.append(pred_proba)
            confidences.append(confidence)
        
        # Create meta-features for dynamic weighting
        meta_features = self._create_meta_features(X, trained_models)
        
        # Get dynamic weights
        weights = dynamic_weights.predict(meta_features)
        weights = np.maximum(weights, 0)  # Ensure non-negative
        weights = weights / np.sum(weights, axis=1, keepdims=True)  # Normalize
        
        # Weighted ensemble prediction
        final_pred = np.zeros_like(predictions[0])
        for i, pred in enumerate(predictions):
            final_pred += weights[:, i:i+1] * pred
        
        return final_pred

    def build_deep_learning_model(self, num_features, num_fighters):
        """Build enhanced TensorFlow deep learning model with fighter embeddings and method prediction"""
        if not HAS_TENSORFLOW:
            return None

        # Ensure deterministic operations for consistency
        self.set_random_seeds()

        # Fighter embedding input
        fighter_r_input = keras.Input(shape=(1,), name="fighter_r")
        fighter_b_input = keras.Input(shape=(1,), name="fighter_b")

        # Enhanced embedding layer with larger dimensions
        fighter_embedding = keras.layers.Embedding(
            input_dim=num_fighters,
            output_dim=min(64, num_fighters // 4),
            name="fighter_embedding",
        )

        fighter_r_embed = keras.layers.Flatten()(fighter_embedding(fighter_r_input))
        fighter_b_embed = keras.layers.Flatten()(fighter_embedding(fighter_b_input))

        # Statistical features input with enhanced processing
        stats_input = keras.Input(shape=(num_features,), name="stats")
        stats_dense = keras.layers.Dense(256, activation="relu")(stats_input)
        stats_dense = keras.layers.BatchNormalization()(stats_dense)
        stats_dense = keras.layers.Dropout(0.3)(stats_dense)

        stats_dense = keras.layers.Dense(128, activation="relu")(stats_dense)
        stats_dense = keras.layers.Dropout(0.2)(stats_dense)

        # Combine all inputs
        combined = keras.layers.Concatenate()(
            [fighter_r_embed, fighter_b_embed, stats_dense]
        )

        # Enhanced shared layers
        x = keras.layers.Dense(512, activation="relu")(combined)
        x = keras.layers.BatchNormalization()(x)
        x = keras.layers.Dropout(0.3)(x)

        x = keras.layers.Dense(256, activation="relu")(x)
        x = keras.layers.BatchNormalization()(x)
        x = keras.layers.Dropout(0.3)(x)

        x = keras.layers.Dense(128, activation="relu")(x)
        x = keras.layers.Dropout(0.2)(x)

        # Winner prediction branch
        winner_branch = keras.layers.Dense(64, activation="relu")(x)
        winner_branch = keras.layers.Dropout(0.1)(winner_branch)
        winner_output = keras.layers.Dense(1, activation="sigmoid", name="winner")(
            winner_branch
        )

        # Enhanced method prediction branch
        method_branch = keras.layers.Dense(128, activation="relu")(x)
        method_branch = keras.layers.BatchNormalization()(method_branch)
        method_branch = keras.layers.Dropout(0.2)(method_branch)

        # Specialized method prediction layers
        method_branch = keras.layers.Dense(64, activation="relu")(method_branch)
        method_branch = keras.layers.Dropout(0.1)(method_branch)

        # Enhanced method prediction with residual connections
        method_residual = keras.layers.Dense(64, activation="relu")(method_branch)
        method_residual = keras.layers.Dropout(0.1)(method_residual)
        method_branch = keras.layers.Add()([method_branch, method_residual])

        # Additional method-specific processing
        method_branch = keras.layers.Dense(128, activation="relu")(method_branch)
        method_branch = keras.layers.BatchNormalization()(method_branch)
        method_branch = keras.layers.Dropout(0.2)(method_branch)

        # Method output (6 classes: Red_KO/TKO, Red_Submission, Red_Decision, Blue_KO/TKO, Blue_Submission, Blue_Decision)
        method_output = keras.layers.Dense(6, activation="softmax", name="method")(
            method_branch
        )

        model = keras.Model(
            inputs=[fighter_r_input, fighter_b_input, stats_input],
            outputs=[winner_output, method_output],
        )

        # Enhanced optimizer with fixed learning rate
        model.compile(
            optimizer=keras.optimizers.Adam(learning_rate=0.001),
            loss={
                "winner": "binary_crossentropy",
                "method": "categorical_crossentropy",
            },
            loss_weights={
                "winner": 1.0,
                "method": 1.5,  # Increased weight for method prediction
            },
            metrics={"winner": "accuracy", "method": "accuracy"},
        )

        return model

    def fix_data_leakage(self, df):
        """Recalculate comprehensive fighter statistics chronologically"""
        print("Fixing data leakage with advanced feature tracking...\n")

        import copy

        df["event_date"] = pd.to_datetime(df["event_date"])
        df = df.sort_values("event_date").reset_index(drop=True)

        fighter_stats = {}

        stats_to_track = {
            "wins": 0,
            "losses": 0,
            "draws": 0,
            "sig_str_total": 0,
            "sig_str_att_total": 0,
            "sig_str_absorbed_total": 0,
            "total_str_landed": 0,
            "total_str_att": 0,
            "total_str_absorbed": 0,
            "str_def_hits": 0,
            "str_def_att": 0,
            "td_total": 0,
            "td_att_total": 0,
            "td_def_success": 0,
            "td_def_att": 0,
            "sub_att_total": 0,
            "kd_total": 0,
            "kd_absorbed_total": 0,
            "fight_time_minutes": 0,
            "fight_count": 0,
            "ko_wins": 0,
            "sub_wins": 0,
            "dec_wins": 0,
            "ko_losses": 0,
            "sub_losses": 0,
            "recent_wins": [],
            "recent_finishes": [],
            "last_fight_date": None,
            "head_pct_sum": 0,
            "body_pct_sum": 0,
            "leg_pct_sum": 0,
            "location_fight_count": 0,
            "distance_pct_sum": 0,
            "clinch_pct_sum": 0,
            "ground_pct_sum": 0,
            "position_fight_count": 0,
            # Trajectory tracking
            "slpm_history": [],
            "td_avg_history": [],
            "win_history": [],
        }

        for prefix in ["r", "b"]:
            for stat in [
                "wins",
                "losses",
                "draws",
                "win_loss_ratio",
                "pro_SLpM",
                "pro_sig_str_acc",
                "pro_SApM",
                "pro_str_def",
                "pro_total_str_pM",
                "pro_total_str_acc",
                "pro_total_str_absorbed_pM",
                "pro_td_avg",
                "pro_td_acc",
                "pro_td_def",
                "pro_sub_avg",
                "pro_kd_pM",
                "ko_rate",
                "sub_rate",
                "dec_rate",
                "recent_form",
                "head_pct",
                "body_pct",
                "leg_pct",
                "distance_pct",
                "clinch_pct",
                "ground_pct",
                "win_streak",
                "loss_streak",
                "last_5_wins",
                "days_since_last_fight",
                "recent_finish_rate",
                "durability",
                "fight_time_minutes",
                "slpm_trend",
                "td_avg_trend",
                "age_adjusted_performance",
                "peak_indicator",
            ]:
                df[f"{prefix}_{stat}_corrected"] = 0.0

        df["h2h_advantage"] = 0.0
        fighter_h2h = {}

        for idx, row in df.iterrows():
            if idx % 500 == 0:
                print(f"   Processing fight {idx}/{len(df)}...")

            r_fighter, b_fighter = row["r_fighter"], row["b_fighter"]

            # Head-to-head tracking
            h2h_key = (r_fighter, b_fighter)
            h2h_key_reverse = (b_fighter, r_fighter)

            if h2h_key in fighter_h2h:
                df.at[idx, "h2h_advantage"] = fighter_h2h[h2h_key]
            elif h2h_key_reverse in fighter_h2h:
                df.at[idx, "h2h_advantage"] = -fighter_h2h[h2h_key_reverse]

            if r_fighter not in fighter_stats:
                fighter_stats[r_fighter] = copy.deepcopy(stats_to_track)
            if b_fighter not in fighter_stats:
                fighter_stats[b_fighter] = copy.deepcopy(stats_to_track)

            for fighter, prefix in [(r_fighter, "r"), (b_fighter, "b")]:
                stats = fighter_stats[fighter]
                df.at[idx, f"{prefix}_wins_corrected"] = stats["wins"]
                df.at[idx, f"{prefix}_losses_corrected"] = stats["losses"]
                df.at[idx, f"{prefix}_draws_corrected"] = stats["draws"]
                df.at[idx, f"{prefix}_win_loss_ratio_corrected"] = stats["wins"] / max(
                    stats["losses"], 1
                )

                # Durability
                total_losses = stats["losses"]
                finish_losses = stats["ko_losses"] + stats["sub_losses"]
                df.at[idx, f"{prefix}_durability_corrected"] = (
                    1.0 / (1 + finish_losses) if total_losses > 0 else 1.0
                )

                total_fights = stats["wins"] + stats["losses"]
                if total_fights > 0:
                    df.at[idx, f"{prefix}_ko_rate_corrected"] = (
                        stats["ko_wins"] / total_fights
                    )
                    df.at[idx, f"{prefix}_sub_rate_corrected"] = (
                        stats["sub_wins"] / total_fights
                    )
                    df.at[idx, f"{prefix}_dec_rate_corrected"] = (
                        stats["dec_wins"] / total_fights
                    )

                # Recent finish rate
                if len(stats["recent_finishes"]) >= 10:
                    df.at[idx, f"{prefix}_recent_finish_rate_corrected"] = (
                        sum(stats["recent_finishes"][-10:]) / 10
                    )
                elif len(stats["recent_finishes"]) >= 8:
                    df.at[idx, f"{prefix}_recent_finish_rate_corrected"] = (
                        sum(stats["recent_finishes"][-8:]) / 8
                    )
                elif len(stats["recent_finishes"]) >= 6:
                    df.at[idx, f"{prefix}_recent_finish_rate_corrected"] = (
                        sum(stats["recent_finishes"][-6:]) / 6
                    )
                elif len(stats["recent_finishes"]) >= 4:
                    df.at[idx, f"{prefix}_recent_finish_rate_corrected"] = (
                        sum(stats["recent_finishes"][-4:]) / 4
                    )
                elif len(stats["recent_finishes"]) >= 2:
                    df.at[idx, f"{prefix}_recent_finish_rate_corrected"] = (
                        sum(stats["recent_finishes"][-2:]) / 2
                    )

                # Recent form
                if len(stats["recent_wins"]) >= 10:
                    df.at[idx, f"{prefix}_recent_form_corrected"] = (
                        sum(stats["recent_wins"][-10:]) / 10
                    )
                elif len(stats["recent_wins"]) >= 8:
                    df.at[idx, f"{prefix}_recent_form_corrected"] = (
                        sum(stats["recent_wins"][-8:]) / 8
                    )
                elif len(stats["recent_wins"]) >= 6:
                    df.at[idx, f"{prefix}_recent_form_corrected"] = (
                        sum(stats["recent_wins"][-6:]) / 6
                    )
                elif len(stats["recent_wins"]) >= 4:
                    df.at[idx, f"{prefix}_recent_form_corrected"] = (
                        sum(stats["recent_wins"][-4:]) / 4
                    )
                elif len(stats["recent_wins"]) >= 2:
                    df.at[idx, f"{prefix}_recent_form_corrected"] = (
                        sum(stats["recent_wins"][-2:]) / 2
                    )

                # Momentum features
                df.at[idx, f"{prefix}_win_streak_corrected"] = self.calculate_streak(
                    stats["recent_wins"], True
                )
                df.at[idx, f"{prefix}_loss_streak_corrected"] = self.calculate_streak(
                    stats["recent_wins"], False
                )

                if len(stats["recent_wins"]) >= 10:
                    df.at[idx, f"{prefix}_last_5_wins_corrected"] = sum(
                        stats["recent_wins"][-10:]
                    )
                elif len(stats["recent_wins"]) >= 8:
                    df.at[idx, f"{prefix}_last_5_wins_corrected"] = sum(
                        stats["recent_wins"][-8:]
                    )
                elif len(stats["recent_wins"]) >= 6:
                    df.at[idx, f"{prefix}_last_5_wins_corrected"] = sum(
                        stats["recent_wins"][-6:]
                    )
                elif len(stats["recent_wins"]) >= 4:
                    df.at[idx, f"{prefix}_last_5_wins_corrected"] = sum(
                        stats["recent_wins"][-4:]
                    )
                elif len(stats["recent_wins"]) >= 2:
                    df.at[idx, f"{prefix}_last_5_wins_corrected"] = sum(
                        stats["recent_wins"][-2:]
                    )

                # Days since last fight
                if stats["last_fight_date"]:
                    days_off = (row["event_date"] - stats["last_fight_date"]).days
                    df.at[idx, f"{prefix}_days_since_last_fight_corrected"] = days_off

                # Fight time
                df.at[idx, f"{prefix}_fight_time_minutes_corrected"] = stats[
                    "fight_time_minutes"
                ]

                if stats["fight_time_minutes"] > 0:
                    df.at[idx, f"{prefix}_pro_SLpM_corrected"] = (
                        stats["sig_str_total"] / stats["fight_time_minutes"]
                    )
                    df.at[idx, f"{prefix}_pro_SApM_corrected"] = (
                        stats["sig_str_absorbed_total"] / stats["fight_time_minutes"]
                    )
                    df.at[idx, f"{prefix}_pro_total_str_pM_corrected"] = (
                        stats["total_str_landed"] / stats["fight_time_minutes"]
                    )
                    df.at[idx, f"{prefix}_pro_total_str_absorbed_pM_corrected"] = (
                        stats["total_str_absorbed"] / stats["fight_time_minutes"]
                    )
                    df.at[idx, f"{prefix}_pro_td_avg_corrected"] = (
                        stats["td_total"] / stats["fight_time_minutes"]
                    ) * 15
                    df.at[idx, f"{prefix}_pro_sub_avg_corrected"] = (
                        stats["sub_att_total"] / stats["fight_time_minutes"]
                    ) * 15
                    df.at[idx, f"{prefix}_pro_kd_pM_corrected"] = (
                        stats["kd_total"] / stats["fight_time_minutes"]
                    )

                if stats["sig_str_att_total"] > 0:
                    df.at[idx, f"{prefix}_pro_sig_str_acc_corrected"] = (
                        stats["sig_str_total"] / stats["sig_str_att_total"]
                    )
                if stats["total_str_att"] > 0:
                    df.at[idx, f"{prefix}_pro_total_str_acc_corrected"] = (
                        stats["total_str_landed"] / stats["total_str_att"]
                    )
                if stats["str_def_att"] > 0:
                    df.at[idx, f"{prefix}_pro_str_def_corrected"] = (
                        stats["str_def_hits"] / stats["str_def_att"]
                    )
                if stats["td_att_total"] > 0:
                    df.at[idx, f"{prefix}_pro_td_acc_corrected"] = (
                        stats["td_total"] / stats["td_att_total"]
                    )
                if stats["td_def_att"] > 0:
                    df.at[idx, f"{prefix}_pro_td_def_corrected"] = (
                        stats["td_def_success"] / stats["td_def_att"]
                    )

                # Strike location percentages
                if stats["location_fight_count"] > 0:
                    df.at[idx, f"{prefix}_head_pct_corrected"] = (
                        stats["head_pct_sum"] / stats["location_fight_count"]
                    )
                    df.at[idx, f"{prefix}_body_pct_corrected"] = (
                        stats["body_pct_sum"] / stats["location_fight_count"]
                    )
                    df.at[idx, f"{prefix}_leg_pct_corrected"] = (
                        stats["leg_pct_sum"] / stats["location_fight_count"]
                    )

                if stats["position_fight_count"] > 0:
                    df.at[idx, f"{prefix}_distance_pct_corrected"] = (
                        stats["distance_pct_sum"] / stats["position_fight_count"]
                    )
                    df.at[idx, f"{prefix}_clinch_pct_corrected"] = (
                        stats["clinch_pct_sum"] / stats["position_fight_count"]
                    )
                    df.at[idx, f"{prefix}_ground_pct_corrected"] = (
                        stats["ground_pct_sum"] / stats["position_fight_count"]
                    )

                # TRAJECTORY FEATURES
                df.at[idx, f"{prefix}_slpm_trend_corrected"] = (
                    self.calculate_trajectory(
                        stats["slpm_history"][-5:]
                        if len(stats["slpm_history"]) >= 5
                        else stats["slpm_history"]
                    )
                )
                df.at[idx, f"{prefix}_td_avg_trend_corrected"] = (
                    self.calculate_trajectory(
                        stats["td_avg_history"][-5:]
                        if len(stats["td_avg_history"]) >= 5
                        else stats["td_avg_history"]
                    )
                )

                # Age-adjusted performance
                age = (
                    row[f"{prefix}_age_at_event"]
                    if pd.notna(row.get(f"{prefix}_age_at_event"))
                    else 30
                )
                age_curve = self.calculate_age_curve_factor(age)
                recent_form = df.at[idx, f"{prefix}_recent_form_corrected"]
                df.at[idx, f"{prefix}_age_adjusted_performance_corrected"] = (
                    recent_form * age_curve
                )

                # Peak indicator
                df.at[idx, f"{prefix}_peak_indicator_corrected"] = (
                    self.is_fighter_at_peak(
                        stats["wins"], stats["losses"], age, recent_form
                    )
                )

            # Update stats after fight
            if pd.notna(row["winner"]):
                fight_time_min = (
                    row["total_fight_time_sec"] / 60
                    if pd.notna(row["total_fight_time_sec"])
                    else 0
                )
                method = str(row["method"]).lower()
                method_cat = (
                    "ko"
                    if "ko" in method or "tko" in method
                    else "sub"
                    if "sub" in method
                    else "dec"
                )
                is_finish = method_cat in ["ko", "sub"]

                if row["winner"] == "Red":
                    fighter_stats[r_fighter]["wins"] += 1
                    fighter_stats[b_fighter]["losses"] += 1
                    fighter_stats[r_fighter][f"{method_cat}_wins"] += 1
                    fighter_stats[r_fighter]["recent_wins"].append(1)
                    fighter_stats[b_fighter]["recent_wins"].append(0)
                    fighter_stats[r_fighter]["recent_finishes"].append(
                        1 if is_finish else 0
                    )
                    fighter_stats[b_fighter]["recent_finishes"].append(0)
                    if is_finish:
                        fighter_stats[b_fighter][f"{method_cat}_losses"] += 1
                    fighter_h2h[(r_fighter, b_fighter)] = (
                        fighter_h2h.get((r_fighter, b_fighter), 0) + 1
                    )
                elif row["winner"] == "Blue":
                    fighter_stats[b_fighter]["wins"] += 1
                    fighter_stats[r_fighter]["losses"] += 1
                    fighter_stats[b_fighter][f"{method_cat}_wins"] += 1
                    fighter_stats[b_fighter]["recent_wins"].append(1)
                    fighter_stats[r_fighter]["recent_wins"].append(0)
                    fighter_stats[b_fighter]["recent_finishes"].append(
                        1 if is_finish else 0
                    )
                    fighter_stats[r_fighter]["recent_finishes"].append(0)
                    if is_finish:
                        fighter_stats[r_fighter][f"{method_cat}_losses"] += 1
                    fighter_h2h[(r_fighter, b_fighter)] = (
                        fighter_h2h.get((r_fighter, b_fighter), 0) - 1
                    )

                for fighter in [r_fighter, b_fighter]:
                    if len(fighter_stats[fighter]["recent_wins"]) > 10:
                        fighter_stats[fighter]["recent_wins"] = fighter_stats[fighter][
                            "recent_wins"
                        ][-10:]
                    if len(fighter_stats[fighter]["recent_finishes"]) > 10:
                        fighter_stats[fighter]["recent_finishes"] = fighter_stats[
                            fighter
                        ]["recent_finishes"][-10:]
                    fighter_stats[fighter]["last_fight_date"] = row["event_date"]

                for fighter, f_prefix, opp_prefix in [
                    (r_fighter, "r", "b"),
                    (b_fighter, "b", "r"),
                ]:
                    for stat_pair in [
                        ("sig_str", "sig_str_total"),
                        ("str", "total_str_landed"),
                    ]:
                        if pd.notna(row[f"{f_prefix}_{stat_pair[0]}"]):
                            fighter_stats[fighter][stat_pair[1]] += row[
                                f"{f_prefix}_{stat_pair[0]}"
                            ]

                    if pd.notna(row[f"{opp_prefix}_sig_str"]):
                        fighter_stats[fighter]["sig_str_absorbed_total"] += row[
                            f"{opp_prefix}_sig_str"
                        ]
                    if pd.notna(row[f"{opp_prefix}_str"]):
                        fighter_stats[fighter]["total_str_absorbed"] += row[
                            f"{opp_prefix}_str"
                        ]

                    for att in ["sig_str_att", "str_att", "td_att", "sub_att"]:
                        col = f"{f_prefix}_{att}"
                        if pd.notna(row.get(col)):
                            key = (
                                att + "_total" if att != "str_att" else "total_str_att"
                            )
                            fighter_stats[fighter][key] += row[col]

                    if pd.notna(row[f"{f_prefix}_td"]):
                        fighter_stats[fighter]["td_total"] += row[f"{f_prefix}_td"]

                    if pd.notna(row.get(f"{f_prefix}_kd")):
                        fighter_stats[fighter]["kd_total"] += row[f"{f_prefix}_kd"]
                    if pd.notna(row.get(f"{opp_prefix}_kd")):
                        fighter_stats[fighter]["kd_absorbed_total"] += row[
                            f"{opp_prefix}_kd"
                        ]

                    if pd.notna(row[f"{opp_prefix}_sig_str_att"]):
                        fighter_stats[fighter]["str_def_att"] += row[
                            f"{opp_prefix}_sig_str_att"
                        ]
                        if pd.notna(row[f"{opp_prefix}_sig_str"]):
                            fighter_stats[fighter]["str_def_hits"] += (
                                row[f"{opp_prefix}_sig_str_att"]
                                - row[f"{opp_prefix}_sig_str"]
                            )

                    if pd.notna(row[f"{opp_prefix}_td_att"]):
                        fighter_stats[fighter]["td_def_att"] += row[
                            f"{opp_prefix}_td_att"
                        ]
                        if pd.notna(row[f"{opp_prefix}_td"]):
                            fighter_stats[fighter]["td_def_success"] += (
                                row[f"{opp_prefix}_td_att"] - row[f"{opp_prefix}_td"]
                            )

                    # Strike location
                    if (
                        pd.notna(row.get(f"{f_prefix}_head"))
                        or pd.notna(row.get(f"{f_prefix}_body"))
                        or pd.notna(row.get(f"{f_prefix}_leg"))
                    ):
                        if pd.notna(row.get(f"{f_prefix}_head")):
                            fighter_stats[fighter]["head_pct_sum"] += row[
                                f"{f_prefix}_head"
                            ]
                        if pd.notna(row.get(f"{f_prefix}_body")):
                            fighter_stats[fighter]["body_pct_sum"] += row[
                                f"{f_prefix}_body"
                            ]
                        if pd.notna(row.get(f"{f_prefix}_leg")):
                            fighter_stats[fighter]["leg_pct_sum"] += row[
                                f"{f_prefix}_leg"
                            ]
                        fighter_stats[fighter]["location_fight_count"] += 1

                    if (
                        pd.notna(row.get(f"{f_prefix}_distance"))
                        or pd.notna(row.get(f"{f_prefix}_clinch"))
                        or pd.notna(row.get(f"{f_prefix}_ground"))
                    ):
                        if pd.notna(row.get(f"{f_prefix}_distance")):
                            fighter_stats[fighter]["distance_pct_sum"] += row[
                                f"{f_prefix}_distance"
                            ]
                        if pd.notna(row.get(f"{f_prefix}_clinch")):
                            fighter_stats[fighter]["clinch_pct_sum"] += row[
                                f"{f_prefix}_clinch"
                            ]
                        if pd.notna(row.get(f"{f_prefix}_ground")):
                            fighter_stats[fighter]["ground_pct_sum"] += row[
                                f"{f_prefix}_ground"
                            ]
                        fighter_stats[fighter]["position_fight_count"] += 1

                    fighter_stats[fighter]["fight_time_minutes"] += fight_time_min
                    fighter_stats[fighter]["fight_count"] += 1

                    # Track trajectory
                    current_slpm = df.at[idx, f"{f_prefix}_pro_SLpM_corrected"]
                    current_td_avg = df.at[idx, f"{f_prefix}_pro_td_avg_corrected"]
                    fighter_stats[fighter]["slpm_history"].append(current_slpm)
                    fighter_stats[fighter]["td_avg_history"].append(current_td_avg)

                    if len(fighter_stats[fighter]["slpm_history"]) > 10:
                        fighter_stats[fighter]["slpm_history"] = fighter_stats[fighter][
                            "slpm_history"
                        ][-10:]
                    if len(fighter_stats[fighter]["td_avg_history"]) > 10:
                        fighter_stats[fighter]["td_avg_history"] = fighter_stats[
                            fighter
                        ]["td_avg_history"][-10:]

        diff_stats = [
            "wins",
            "losses",
            "draws",
            "win_loss_ratio",
            "pro_SLpM",
            "pro_sig_str_acc",
            "pro_SApM",
            "pro_str_def",
            "pro_total_str_pM",
            "pro_total_str_acc",
            "pro_total_str_absorbed_pM",
            "pro_td_avg",
            "pro_td_acc",
            "pro_td_def",
            "pro_sub_avg",
            "pro_kd_pM",
            "ko_rate",
            "sub_rate",
            "dec_rate",
            "recent_form",
            "head_pct",
            "body_pct",
            "leg_pct",
            "distance_pct",
            "clinch_pct",
            "ground_pct",
            "win_streak",
            "loss_streak",
            "last_5_wins",
            "days_since_last_fight",
            "recent_finish_rate",
            "durability",
            "fight_time_minutes",
            "slpm_trend",
            "td_avg_trend",
            "age_adjusted_performance",
            "peak_indicator",
        ]

        for stat in diff_stats:
            df[f"{stat}_diff_corrected"] = (
                df[f"r_{stat}_corrected"] - df[f"b_{stat}_corrected"]
            )

        return df

    def augment_data_with_corner_swapping(self, df):
        """Augment data by swapping red and blue corners to eliminate bias - OPTIMIZED FOR SPEED"""
        print(
            "Augmenting data with SMART corner swapping to eliminate red corner bias..."
        )

        # Add neutral features that don't favor either corner
        df = self.create_neutral_features(df)

        # Calculate original bias
        if "winner" in df.columns:
            red_count = (df["winner"] == "Red").sum()
            blue_count = (df["winner"] == "Blue").sum()
            total_count = red_count + blue_count
        else:
            red_count = blue_count = total_count = 0

        if total_count > 0:
            red_bias = red_count / total_count
            print(f"Original bias - Red: {red_bias:.3f}, Blue: {1 - red_bias:.3f}")

        # ENHANCED AUGMENTATION: More aggressive for data augmentation model
        bias_threshold = 0.02  # Even lower threshold for more augmentation
        if abs(red_bias - 0.5) > bias_threshold:
            print(
                f"Significant bias detected ({abs(red_bias - 0.5):.3f}), applying enhanced augmentation..."
            )

            # Create augmented dataset by swapping corners
            df_augmented = self.swap_corners(df)

            # ENHANCED SAMPLING: Use 90% of augmented data for better balance
            sample_size = int(len(df_augmented) * 0.90)
            df_augmented_sampled = df_augmented.sample(n=sample_size, random_state=42)

            # Additional augmentation for close fights to improve method prediction
            # Check if winner_method_simple column exists first
            if "winner_method_simple" in df.columns:
                # Check if winner_confidence column exists, if not use a default approach
                if "winner_confidence" in df.columns:
                    close_fights = df[
                        (df["winner_method_simple"].str.contains("Decision"))
                        & (df["winner_confidence"] < 0.7)  # Low confidence decisions
                    ].copy()
                else:
                    # If no confidence column, just use decision fights
                    close_fights = df[
                        df["winner_method_simple"].str.contains("Decision")
                    ].copy()
            else:
                # If no winner_method_simple column, skip close fight augmentation
                close_fights = pd.DataFrame()
            if len(close_fights) > 0:
                # Add 25% of close fights as additional augmentation
                close_sample_size = min(len(close_fights) // 2, len(df) // 4)
                if close_sample_size > 0:  # Ensure we have something to sample
                    close_subset = close_fights.sample(
                        n=close_sample_size, random_state=42
                    )
                    close_swapped = self.swap_corners(close_subset)
                    df_augmented_sampled = pd.concat(
                        [df_augmented_sampled, close_swapped], ignore_index=True
                    )

            # Combine original and sampled augmented data
            df_combined = pd.concat([df, df_augmented_sampled], ignore_index=True)

            print(
                f"Enhanced augmentation: {len(df)} -> {len(df_combined)} (1.75x size)"
            )
        else:
            print(
                "Bias is minimal, applying light augmentation for method prediction..."
            )
            # Light augmentation even with minimal bias for method prediction
            df_augmented = self.swap_corners(df)
            sample_size = len(df_augmented) // 4  # 25% augmentation
            df_augmented_sampled = df_augmented.sample(n=sample_size, random_state=42)
            df_combined = pd.concat([df, df_augmented_sampled], ignore_index=True)
            print(f"Light augmentation: {len(df)} -> {len(df_combined)} (1.25x size)")

        # Shuffle the combined dataset to mix original and swapped fights
        df_combined = df_combined.sample(frac=1, random_state=42).reset_index(drop=True)

        # Calculate new bias after augmentation
        new_red_count = (df_combined["winner"] == "Red").sum()
        new_blue_count = (df_combined["winner"] == "Blue").sum()
        new_total_count = new_red_count + new_blue_count

        if new_total_count > 0:
            new_red_bias = new_red_count / new_total_count
            print(
                f"After smart augmentation - Red: {new_red_bias:.3f}, Blue: {1 - new_red_bias:.3f}"
            )

        # Set balanced class weights for the augmented dataset
        self.class_weight = {
            0: 0.5,
            1: 0.5,
        }  # Balanced weights since bias is eliminated

        return df_combined

    def swap_corners(self, df):
        """Swap red and blue corners to create augmented data"""
        print("Creating corner-swapped augmented data...")

        # Create a copy of the dataframe
        df_swapped = df.copy()

        # Swap fighter names (both r_fighter/b_fighter and r_name/b_name)
        df_swapped["r_fighter"], df_swapped["b_fighter"] = (
            df_swapped["b_fighter"],
            df_swapped["r_fighter"],
        )
        if "r_name" in df_swapped.columns and "b_name" in df_swapped.columns:
            df_swapped["r_name"], df_swapped["b_name"] = (
                df_swapped["b_name"],
                df_swapped["r_name"],
            )

        # Swap winner (Red becomes Blue, Blue becomes Red)
        df_swapped["winner"] = df_swapped["winner"].map({"Red": "Blue", "Blue": "Red"})

        # Swap all red/blue prefixed columns (excluding fighter names which are handled manually)
        red_columns = [
            col
            for col in df.columns
            if col.startswith("r_") and col not in ["r_fighter", "r_name"]
        ]
        blue_columns = [
            col
            for col in df.columns
            if col.startswith("b_") and col not in ["b_fighter", "b_name"]
        ]

        # Create mapping for swapping (only map r_ columns to avoid duplicates)
        swap_mapping = {}
        for r_col in red_columns:
            b_col = r_col.replace("r_", "b_")
            if b_col in blue_columns:
                swap_mapping[r_col] = b_col

        # Perform the swap
        for r_col, b_col in swap_mapping.items():
            if r_col in df_swapped.columns and b_col in df_swapped.columns:
                # Store original values
                r_val = df_swapped[r_col].copy()
                b_val = df_swapped[b_col].copy()
                # Perform the swap
                df_swapped[r_col] = b_val
                df_swapped[b_col] = r_val

        # Swap diff columns (r_stat - b_stat becomes b_stat - r_stat, so multiply by -1)
        diff_columns = [
            col
            for col in df.columns
            if col.endswith("_diff_corrected") or col.endswith("_diff")
        ]
        for col in diff_columns:
            if col in df_swapped.columns:
                df_swapped[col] = -df_swapped[col]

        # Handle special cases for derived features that need manual adjustment
        # These features are calculated as Red - Blue, so they need to be negated
        derived_features_to_negate = [
            "net_striking_advantage",
            "striking_efficiency",
            "defensive_striking",
            "grappling_control",
            "grappling_defense",
            "offensive_output",
            "defensive_composite",
            "ko_specialist_gap",
            "submission_specialist_gap",
            "experience_gap",
            "skill_momentum",
            "finish_threat",
            "momentum_advantage",
            "pace_differential",
            "avg_fight_time_diff",
            "quality_experience_gap",
            "championship_exp_diff",
            "adversity_exp_diff",
            "experience_skill_interaction",
            "veteran_edge",
            "novice_vulnerability",
            "striker_advantage",
            "grappler_advantage",
            "effective_reach_advantage",
            "ko_specialist_matchup",
            "sub_specialist_matchup",
            "durability_advantage",
            "kd_resistance_advantage",
            "striker_vs_grappler",
            "distance_ground_preference",
            "clinch_advantage",
            "head_hunting_advantage",
            "vulnerability_advantage",
            "age_experience_interaction",
            "power_technique_advantage",
            "grappling_threat_advantage",
            "recent_ko_trend_advantage",
            "ko_susceptibility_advantage",
            "sub_opportunity_advantage",
            "championship_pressure_advantage",
            "opponent_quality_advantage",
            "technical_striker_advantage",
            "clinch_effectiveness_advantage",
            "momentum_velocity",
            "championship_impact",
            "opponent_quality_momentum",
            "finishing_pressure_stress",
            "ring_rust_vs_momentum",
            "weight_class_adaptation",
            "stance_versatility_impact",
            "power_vs_technique",
            "cardio_advantage",
            "finish_pressure",
            "opponent_quality_gap",
            "recent_opponent_strength",
            "upset_potential",
        ]

        for feature in derived_features_to_negate:
            if feature in df_swapped.columns:
                df_swapped[feature] = -df_swapped[feature]

        # Handle ratio columns (r_stat / b_stat becomes b_stat / r_stat)
        ratio_columns = [col for col in df.columns if "ratio" in col.lower()]
        for col in ratio_columns:
            if col in df_swapped.columns:
                # Avoid division by zero
                df_swapped[col] = np.where(
                    df_swapped[col] != 0, 1 / df_swapped[col], df_swapped[col]
                )

        # Handle advantage/better columns that need special logic
        # Skip features that were already handled in derived_features_to_negate
        advantage_columns = [
            col
            for col in df.columns
            if col.endswith("_advantage") or col.endswith("_better")
        ]
        for col in advantage_columns:
            if col in df_swapped.columns and col not in derived_features_to_negate:
                if col.endswith("_better"):
                    # Better columns: 1 for red, -1 for blue -> swap to -1 for red, 1 for blue
                    df_swapped[col] = -df_swapped[col]
                else:
                    # Advantage columns: positive for red advantage -> negative for blue advantage
                    df_swapped[col] = -df_swapped[col]

        # Handle matchup-specific features
        if "orthodox_southpaw_matchup" in df_swapped.columns:
            df_swapped["orthodox_southpaw_matchup"] = -df_swapped[
                "orthodox_southpaw_matchup"
            ]

        if "stance_diff" in df_swapped.columns:
            df_swapped["stance_diff"] = -df_swapped["stance_diff"]

        # Handle inactivity penalty (logic needs to be reversed)
        if "inactivity_penalty" in df_swapped.columns:
            df_swapped["inactivity_penalty"] = -df_swapped["inactivity_penalty"]

        print(f"Corner swapping complete. Created {len(df_swapped)} augmented fights.")
        return df_swapped

    def create_neutral_features(self, df):
        """Create neutral features that don't favor either corner"""

        # Create total fights columns if they don't exist
        if "r_total_fights" not in df.columns:
            if "r_wins_corrected" in df.columns and "r_losses_corrected" in df.columns:
                df["r_total_fights"] = df["r_wins_corrected"] + df["r_losses_corrected"]
            else:
                df["r_total_fights"] = 0
        if "b_total_fights" not in df.columns:
            if "b_wins_corrected" in df.columns and "b_losses_corrected" in df.columns:
                df["b_total_fights"] = df["b_wins_corrected"] + df["b_losses_corrected"]
            else:
                df["b_total_fights"] = 0

        # Key stats for neutral comparison
        neutral_stats = [
            "pro_SLpM",
            "pro_td_avg",
            "wins",
            "losses",
            "pro_sig_str_acc",
            "pro_str_def",
            "ko_rate",
            "sub_rate",
            "recent_form",
        ]

        for stat in neutral_stats:
            r_col = f"r_{stat}_corrected"
            b_col = f"b_{stat}_corrected"

            if r_col in df.columns and b_col in df.columns:
                # Absolute advantage (magnitude of difference)
                df[f"{stat}_advantage"] = np.abs(df[r_col] - df[b_col])

                # Which fighter is better (numeric encoding: 1 for red, -1 for blue)
                df[f"{stat}_better"] = np.where(df[r_col] > df[b_col], 1, -1)

        # Experience mismatch (neutral)
        df["experience_mismatch"] = np.abs(df["r_total_fights"] - df["b_total_fights"])

        # Enhanced style matchup features for data augmentation
        if (
            "r_distance_pct_corrected" in df.columns
            and "b_ground_pct_corrected" in df.columns
        ):
            df["striker_vs_grappler"] = np.where(
                (df["r_distance_pct_corrected"] > df["b_ground_pct_corrected"])
                & (df["b_ground_pct_corrected"] > df["r_distance_pct_corrected"]),
                1,
                0,
            )

            # Additional style features for better method prediction
            df["power_vs_technique"] = np.where(
                (df["r_pro_kd_pM_corrected"] > df["r_pro_sig_str_acc_corrected"] * 0.1)
                & (
                    df["b_pro_kd_pM_corrected"]
                    < df["b_pro_sig_str_acc_corrected"] * 0.1
                ),
                1,
                np.where(
                    (
                        df["b_pro_kd_pM_corrected"]
                        > df["b_pro_sig_str_acc_corrected"] * 0.1
                    )
                    & (
                        df["r_pro_kd_pM_corrected"]
                        < df["r_pro_sig_str_acc_corrected"] * 0.1
                    ),
                    -1,
                    0,
                ),
            )

            df["volume_vs_accuracy"] = np.where(
                (df["r_pro_SLpM_corrected"] > df["r_pro_sig_str_acc_corrected"] * 10)
                & (df["b_pro_SLpM_corrected"] < df["b_pro_sig_str_acc_corrected"] * 10),
                1,
                np.where(
                    (
                        df["b_pro_SLpM_corrected"]
                        > df["b_pro_sig_str_acc_corrected"] * 10
                    )
                    & (
                        df["r_pro_SLpM_corrected"]
                        < df["r_pro_sig_str_acc_corrected"] * 10
                    ),
                    -1,
                    0,
                ),
            )

        # Method prediction specific features
        if "r_ko_rate_corrected" in df.columns and "b_ko_rate_corrected" in df.columns:
            df["ko_rate_differential"] = (
                df["r_ko_rate_corrected"] - df["b_ko_rate_corrected"]
            )
        if (
            "r_sub_rate_corrected" in df.columns
            and "b_sub_rate_corrected" in df.columns
        ):
            df["sub_rate_differential"] = (
                df["r_sub_rate_corrected"] - df["b_sub_rate_corrected"]
            )
        if (
            "r_dec_rate_corrected" in df.columns
            and "b_dec_rate_corrected" in df.columns
        ):
            df["dec_rate_differential"] = (
                df["r_dec_rate_corrected"] - df["b_dec_rate_corrected"]
            )
        else:
            df["striker_vs_grappler"] = 0

        return df

    def prepare_features(self, df):
        """Prepare enhanced features with all advanced metrics and caching"""
        # FEATURE CACHING: Check if features already computed
        cache_key = f"features_{len(df)}_{hash(str(df.columns.tolist()))}"
        if cache_key in self.feature_cache:
            print("Using cached features for 2x speedup...")
            return self.feature_cache[cache_key]

        # Ensure consistent random state for feature preparation
        self.set_random_seeds()

        # Check for and remove duplicate columns
        if df.columns.duplicated().any():
            print("Warning: Found duplicate columns, removing duplicates...")
            df = df.loc[:, ~df.columns.duplicated()]

        print("\nPreparing advanced features...")

        # Check if winner column exists before filtering
        if "winner" in df.columns:
            df = df[df["winner"].isin(["Red", "Blue"])].copy()
        else:
            print("Warning: No 'winner' column found, skipping winner filtering")

        method_mapping = {
            "KO/TKO": "KO/TKO",
            "Submission": "Submission",
            "Decision - Unanimous": "Decision",
            "Decision - Split": "Decision",
            "Decision - Majority": "Decision",
            "TKO - Doctor's Stoppage": "KO/TKO",
            "Could Not Continue": "KO/TKO",
            "DQ": "Decision",
            "Overturned": "Decision",
        }

        # Check if method column exists, if not create a default
        if "method" in df.columns:
            df["method_simple"] = df["method"].map(method_mapping).fillna("Decision")
        else:
            # If no method column, create default method based on other available columns
            df["method_simple"] = "Decision"  # Default to Decision

        # Check if winner column exists, if not create a default
        if "winner" not in df.columns:
            # Try to infer winner from other columns or create default
            if "r_fighter" in df.columns and "b_fighter" in df.columns:
                # If we have fighter names but no winner, create a default
                df["winner"] = "Red"  # Default to Red
            else:
                # If no fighter info, create a default winner column
                df["winner"] = "Red"

        df["winner_method_simple"] = df["winner"] + "_" + df["method_simple"]

        feature_columns = [
            "height_diff",
            "reach_diff",
            "weight_diff",
            "age_at_event_diff",
            "ape_index_diff",
            "wins_diff_corrected",
            "losses_diff_corrected",
            "win_loss_ratio_diff_corrected",
            "pro_SLpM_diff_corrected",
            "pro_sig_str_acc_diff_corrected",
            "pro_SApM_diff_corrected",
            "pro_str_def_diff_corrected",
            "pro_total_str_pM_diff_corrected",
            "pro_total_str_acc_diff_corrected",
            "pro_total_str_absorbed_pM_diff_corrected",
            "pro_td_avg_diff_corrected",
            "pro_td_acc_diff_corrected",
            "pro_td_def_diff_corrected",
            "pro_sub_avg_diff_corrected",
            "pro_kd_pM_diff_corrected",
            "ko_rate_diff_corrected",
            "sub_rate_diff_corrected",
            "dec_rate_diff_corrected",
            "recent_form_diff_corrected",
            "head_pct_diff_corrected",
            "body_pct_diff_corrected",
            "leg_pct_diff_corrected",
            "distance_pct_diff_corrected",
            "clinch_pct_diff_corrected",
            "ground_pct_diff_corrected",
            "win_streak_diff_corrected",
            "loss_streak_diff_corrected",
            "last_5_wins_diff_corrected",
            "days_since_last_fight_diff_corrected",
            "recent_finish_rate_diff_corrected",
            "durability_diff_corrected",
            "slpm_trend_diff_corrected",
            "td_avg_trend_diff_corrected",
            "age_adjusted_performance_diff_corrected",
            "peak_indicator_diff_corrected",
            "h2h_advantage",
            "total_rounds",
            "is_title_bout",
            # Add neutral features
            "striker_vs_grappler",
            "experience_mismatch",
            "r_total_fights",
            "b_total_fights",
            # NEW HIGH-IMPACT FEATURES (only include features that are actually created)
            "ape_index_advantage",
            "stance_matchup_advantage",
            "stance_versatility_advantage",
            "weight_class_factor",
            "ring_rust_factor",
            "championship_pressure_advantage",  # This is the actual feature name
            "momentum_swing",
            "style_clash_severity",
            "power_vs_technique",
            "cardio_advantage",
            "finish_pressure",
            "opponent_quality_gap",
            "recent_opponent_strength",
            "upset_potential",
            # NEW HIGH-IMPACT WINNER PREDICTION FEATURES
            "recent_opponent_quality_diff",
            "finish_rate_diff",
            "stamina_factor_diff",
            "pressure_performance_diff",
            "comeback_ability_diff",
            "clutch_factor_diff",
            "recent_trend_diff",
            "momentum_shift_diff",
            "form_consistency_diff",
            "upset_history_diff",
        ]

        # Add neutral advantage features
        neutral_stats = [
            "pro_SLpM",
            "pro_td_avg",
            "wins",
            "losses",
            "pro_sig_str_acc",
            "pro_str_def",
            "ko_rate",
            "sub_rate",
            "recent_form",
        ]

        for stat in neutral_stats:
            new_features = [f"{stat}_advantage", f"{stat}_better"]
            for feature in new_features:
                if feature not in feature_columns:
                    feature_columns.append(feature)

        # ===== ADVANCED TEMPORAL FEATURES =====
        print("Adding advanced temporal features...")
        
        # Fighter trajectory analysis (last 3, 5, 10 fights)
        try:
            for window in [3, 5, 10]:
                df[f"r_trajectory_{window}"] = self.calculate_fighter_trajectory(df, "r_fighter", window)
                df[f"b_trajectory_{window}"] = self.calculate_fighter_trajectory(df, "b_fighter", window)
                df[f"trajectory_diff_{window}"] = df[f"r_trajectory_{window}"] - df[f"b_trajectory_{window}"]
                feature_columns.extend([f"r_trajectory_{window}", f"b_trajectory_{window}", f"trajectory_diff_{window}"])
        except Exception as e:
            print(f"Warning: Error in trajectory features: {e}")
            # Add default values
            for window in [3, 5, 10]:
                df[f"r_trajectory_{window}"] = 0.0
                df[f"b_trajectory_{window}"] = 0.0
                df[f"trajectory_diff_{window}"] = 0.0
                feature_columns.extend([f"r_trajectory_{window}", f"b_trajectory_{window}", f"trajectory_diff_{window}"])

        # Opponent quality progression
        try:
            df["r_opponent_quality"] = self.calculate_opponent_quality(df, "r_fighter")
            df["b_opponent_quality"] = self.calculate_opponent_quality(df, "b_fighter")
            df["opponent_quality_diff"] = df["r_opponent_quality"] - df["b_opponent_quality"]
            feature_columns.extend(["r_opponent_quality", "b_opponent_quality", "opponent_quality_diff"])
        except Exception as e:
            print(f"Warning: Error in opponent quality features: {e}")
            df["r_opponent_quality"] = 0.5
            df["b_opponent_quality"] = 0.5
            df["opponent_quality_diff"] = 0.0
            feature_columns.extend(["r_opponent_quality", "b_opponent_quality", "opponent_quality_diff"])

        # Fight frequency impact (ring rust vs overtraining)
        try:
            df["r_fight_frequency"] = self.calculate_fight_frequency(df, "r_fighter")
            df["b_fight_frequency"] = self.calculate_fight_frequency(df, "b_fighter")
            df["fight_frequency_diff"] = df["r_fight_frequency"] - df["b_fight_frequency"]
            feature_columns.extend(["r_fight_frequency", "b_fight_frequency", "fight_frequency_diff"])
        except Exception as e:
            print(f"Warning: Error in fight frequency features: {e}")
            df["r_fight_frequency"] = 0.0
            df["b_fight_frequency"] = 0.0
            df["fight_frequency_diff"] = 0.0
            feature_columns.extend(["r_fight_frequency", "b_fight_frequency", "fight_frequency_diff"])

        # Career stage modeling
        try:
            df["r_career_stage"] = self.calculate_career_stage(df, "r_fighter")
            df["b_career_stage"] = self.calculate_career_stage(df, "b_fighter")
            df["career_stage_diff"] = df["r_career_stage"] - df["b_career_stage"]
            feature_columns.extend(["r_career_stage", "b_career_stage", "career_stage_diff"])
        except Exception as e:
            print(f"Warning: Error in career stage features: {e}")
            df["r_career_stage"] = 2.0
            df["b_career_stage"] = 2.0
            df["career_stage_diff"] = 0.0
            feature_columns.extend(["r_career_stage", "b_career_stage", "career_stage_diff"])

        # ===== CONTEXTUAL FEATURES =====
        print("Adding contextual features...")
        
        # Weight class dynamics
        df["weight_class_factor"] = self.calculate_weight_class_factor(df)
        feature_columns.append("weight_class_factor")

        # Fight location impact
        df["location_advantage"] = self.calculate_location_advantage(df)
        feature_columns.append("location_advantage")

        # Event significance
        df["event_significance"] = self.calculate_event_significance(df)
        feature_columns.append("event_significance")

        # Referee tendencies (if available)
        if "referee" in df.columns:
            df["referee_factor"] = self.calculate_referee_factor(df)
            feature_columns.append("referee_factor")

        # ===== ADVANCED STATISTICAL FEATURES =====
        print("Adding advanced statistical features...")
        
        # Momentum indicators with quality weighting
        try:
            df["r_momentum_quality"] = self.calculate_momentum_quality(df, "r_fighter")
            df["b_momentum_quality"] = self.calculate_momentum_quality(df, "b_fighter")
            df["momentum_quality_diff"] = df["r_momentum_quality"] - df["b_momentum_quality"]
            feature_columns.extend(["r_momentum_quality", "b_momentum_quality", "momentum_quality_diff"])
        except Exception as e:
            print(f"Warning: Error in momentum features: {e}")
            df["r_momentum_quality"] = 0.5
            df["b_momentum_quality"] = 0.5
            df["momentum_quality_diff"] = 0.0
            feature_columns.extend(["r_momentum_quality", "b_momentum_quality", "momentum_quality_diff"])

        # Pressure performance
        try:
            df["r_pressure_performance"] = self.calculate_pressure_performance(df, "r_fighter")
            df["b_pressure_performance"] = self.calculate_pressure_performance(df, "b_fighter")
            df["pressure_performance_diff"] = df["r_pressure_performance"] - df["b_pressure_performance"]
            feature_columns.extend(["r_pressure_performance", "b_pressure_performance", "pressure_performance_diff"])
        except Exception as e:
            print(f"Warning: Error in pressure performance features: {e}")
            df["r_pressure_performance"] = 0.5
            df["b_pressure_performance"] = 0.5
            df["pressure_performance_diff"] = 0.0
            feature_columns.extend(["r_pressure_performance", "b_pressure_performance", "pressure_performance_diff"])

        # Style matchup analysis
        try:
            df["style_matchup_advantage"] = self.calculate_style_matchup(df)
            feature_columns.append("style_matchup_advantage")
        except Exception as e:
            print(f"Warning: Error in style matchup features: {e}")
            df["style_matchup_advantage"] = 0.0
            feature_columns.append("style_matchup_advantage")

        # Physical advantage composite
        try:
            df["physical_advantage_composite"] = self.calculate_physical_advantage_composite(df)
            feature_columns.append("physical_advantage_composite")
        except Exception as e:
            print(f"Warning: Error in physical advantage features: {e}")
            df["physical_advantage_composite"] = 0.0
            feature_columns.append("physical_advantage_composite")

        # Fight IQ differential
        try:
            df["fight_iq_diff"] = self.calculate_fight_iq_differential(df)
            feature_columns.append("fight_iq_diff")
        except Exception as e:
            print(f"Warning: Error in fight IQ features: {e}")
            df["fight_iq_diff"] = 0.0
            feature_columns.append("fight_iq_diff")

        # Mental toughness differential
        try:
            df["mental_toughness_diff"] = self.calculate_mental_toughness_differential(df)
            feature_columns.append("mental_toughness_diff")
        except Exception as e:
            print(f"Warning: Error in mental toughness features: {e}")
            df["mental_toughness_diff"] = 0.0
            feature_columns.append("mental_toughness_diff")

        # Injury resistance and durability
        try:
            df["injury_resistance_diff"] = self.calculate_injury_resistance_differential(df)
            feature_columns.append("injury_resistance_diff")
        except Exception as e:
            print(f"Warning: Error in injury resistance features: {e}")
            df["injury_resistance_diff"] = 0.0
            feature_columns.append("injury_resistance_diff")

        # Weight cut impact
        try:
            df["weight_cut_impact_diff"] = self.calculate_weight_cut_impact_differential(df)
            feature_columns.append("weight_cut_impact_diff")
        except Exception as e:
            print(f"Warning: Error in weight cut impact features: {e}")
            df["weight_cut_impact_diff"] = 0.0
            feature_columns.append("weight_cut_impact_diff")

        # Core derived features
        df["net_striking_advantage"] = (
            df["r_pro_SLpM_corrected"] - df["r_pro_SApM_corrected"]
        ) - (df["b_pro_SLpM_corrected"] - df["b_pro_SApM_corrected"])

        df["striking_efficiency"] = (
            df["r_pro_SLpM_corrected"] * df["r_pro_sig_str_acc_corrected"]
        ) - (df["b_pro_SLpM_corrected"] * df["b_pro_sig_str_acc_corrected"])

        df["defensive_striking"] = (
            df["r_pro_str_def_corrected"] - df["r_pro_SApM_corrected"]
        ) - (df["b_pro_str_def_corrected"] - df["b_pro_SApM_corrected"])

        df["grappling_control"] = (
            df["r_pro_td_avg_corrected"] * df["r_pro_td_acc_corrected"]
        ) - (df["b_pro_td_avg_corrected"] * df["b_pro_td_acc_corrected"])

        df["grappling_defense"] = (
            df["r_pro_td_def_corrected"] - df["r_pro_sub_avg_corrected"] / 5
        ) - (df["b_pro_td_def_corrected"] - df["b_pro_sub_avg_corrected"] / 5)

        df["offensive_output"] = (
            df["r_pro_SLpM_corrected"]
            + df["r_pro_td_avg_corrected"]
            + df["r_pro_sub_avg_corrected"]
        ) - (
            df["b_pro_SLpM_corrected"]
            + df["b_pro_td_avg_corrected"]
            + df["b_pro_sub_avg_corrected"]
        )

        df["defensive_composite"] = (
            (df["r_pro_str_def_corrected"] + df["r_pro_td_def_corrected"]) / 2
        ) - ((df["b_pro_str_def_corrected"] + df["b_pro_td_def_corrected"]) / 2)

        df["ko_specialist_gap"] = (
            df["r_ko_rate_corrected"] * df["r_pro_SLpM_corrected"]
        ) - (df["b_ko_rate_corrected"] * df["b_pro_SLpM_corrected"])

        df["submission_specialist_gap"] = (
            df["r_sub_rate_corrected"] * df["r_pro_sub_avg_corrected"]
        ) - (df["b_sub_rate_corrected"] * df["b_pro_sub_avg_corrected"])

        # Experience features
        df["r_total_fights"] = df["r_wins_corrected"] + df["r_losses_corrected"]
        df["b_total_fights"] = df["b_wins_corrected"] + df["b_losses_corrected"]
        df["experience_gap"] = df["r_total_fights"] - df["b_total_fights"]
        df["experience_ratio"] = df["r_total_fights"] / (df["b_total_fights"] + 1)

        df["r_avg_fight_time"] = df["r_fight_time_minutes_corrected"] / (
            df["r_total_fights"] + 1
        )
        df["b_avg_fight_time"] = df["b_fight_time_minutes_corrected"] / (
            df["b_total_fights"] + 1
        )
        df["avg_fight_time_diff"] = df["r_avg_fight_time"] - df["b_avg_fight_time"]

        df["skill_momentum"] = (
            df["pro_SLpM_diff_corrected"] * df["recent_form_diff_corrected"]
        )
        df["finish_threat"] = (
            df["r_ko_rate_corrected"] + df["r_sub_rate_corrected"]
        ) - (df["b_ko_rate_corrected"] + df["b_sub_rate_corrected"])

        df["momentum_advantage"] = (
            df["win_streak_diff_corrected"] - df["loss_streak_diff_corrected"]
        )

        df["inactivity_penalty"] = np.where(
            df["days_since_last_fight_diff_corrected"] > 365,
            -1,
            np.where(df["days_since_last_fight_diff_corrected"] < -365, 1, 0),
        )

        df["pace_differential"] = (
            df["r_pro_SLpM_corrected"] + df["r_pro_td_avg_corrected"]
        ) - (df["b_pro_SLpM_corrected"] + df["b_pro_td_avg_corrected"])

        # Enhanced experience features
        df["r_elite_fight_ratio"] = np.where(
            df["r_total_fights"] > 0, df["r_wins_corrected"] / df["r_total_fights"], 0
        )
        df["b_elite_fight_ratio"] = np.where(
            df["b_total_fights"] > 0, df["b_wins_corrected"] / df["b_total_fights"], 0
        )

        df["quality_experience_gap"] = (
            df["r_total_fights"] * df["r_elite_fight_ratio"]
            - df["b_total_fights"] * df["b_elite_fight_ratio"]
        )

        df["r_championship_experience"] = np.where(
            df["r_total_fights"] > 0,
            (df["r_fight_time_minutes_corrected"] / df["r_total_fights"]) > 12,
            0,
        ).astype(int)
        df["b_championship_experience"] = np.where(
            df["b_total_fights"] > 0,
            (df["b_fight_time_minutes_corrected"] / df["b_total_fights"]) > 12,
            0,
        ).astype(int)
        df["championship_exp_diff"] = (
            df["r_championship_experience"] - df["b_championship_experience"]
        )

        df["r_adversity_experience"] = df["r_losses_corrected"] * (
            df["r_win_loss_ratio_corrected"] - 1
        )
        df["b_adversity_experience"] = df["b_losses_corrected"] * (
            df["b_win_loss_ratio_corrected"] - 1
        )
        df["adversity_exp_diff"] = (
            df["r_adversity_experience"] - df["b_adversity_experience"]
        )

        df["experience_skill_interaction"] = (
            df["experience_gap"] * df["pro_SLpM_diff_corrected"] / 50
        )

        df["veteran_edge"] = np.where(
            np.abs(df["experience_gap"]) > 5,
            df["experience_gap"] * df["recent_form_diff_corrected"],
            0,
        )

        df["novice_vulnerability"] = np.where(
            (df["r_total_fights"] < 10) | (df["b_total_fights"] < 10),
            -np.abs(df["experience_gap"]) * 0.5,
            0,
        )

        # Experience gap historical win rate
        bins = [-np.inf, -10, -5, 0, 5, 10, np.inf]
        bucket_indices = np.digitize(df["experience_gap"], bins[1:-1])
        bucket_labels = [
            "huge_blue",
            "mod_blue",
            "slight_blue",
            "slight_red",
            "mod_red",
            "huge_red",
        ]
        df["exp_gap_bucket_temp"] = [bucket_labels[i] for i in bucket_indices]

        exp_gap_win_rate = (
            df.groupby("exp_gap_bucket_temp")["winner"]
            .apply(lambda x: (x == "Red").mean())
            .to_dict()
        )

        df["exp_gap_historical_win_rate"] = (
            df["exp_gap_bucket_temp"].map(exp_gap_win_rate).fillna(0.5)
        )
        df.drop("exp_gap_bucket_temp", axis=1, inplace=True)

        # STYLE MATCHUP FEATURES
        df["r_striker_score"] = (
            df["r_distance_pct_corrected"] * 1.2
            + df["r_head_pct_corrected"] * 0.8
            - df["r_ground_pct_corrected"] * 1.0
        )
        df["b_striker_score"] = (
            df["b_distance_pct_corrected"] * 1.2
            + df["b_head_pct_corrected"] * 0.8
            - df["b_ground_pct_corrected"] * 1.0
        )
        df["striker_advantage"] = df["r_striker_score"] - df["b_striker_score"]

        df["r_grappler_score"] = (
            df["r_pro_td_avg_corrected"] * 0.4
            + df["r_pro_sub_avg_corrected"] * 0.3
            + df["r_ground_pct_corrected"] * 0.3
        )
        df["b_grappler_score"] = (
            df["b_pro_td_avg_corrected"] * 0.4
            + df["b_pro_sub_avg_corrected"] * 0.3
            + df["b_ground_pct_corrected"] * 0.3
        )
        df["grappler_advantage"] = df["r_grappler_score"] - df["b_grappler_score"]

        # Reach advantage in striking context
        df["effective_reach_advantage"] = np.where(
            df["striker_advantage"] > 0.2, df["reach_diff"] * 1.5, df["reach_diff"]
        )

        # Stance matchup
        if "r_stance" in df.columns and "b_stance" in df.columns:
            if "stance_encoder" not in self.label_encoders:
                self.label_encoders["stance_encoder"] = LabelEncoder()
                all_stances = pd.concat([df["r_stance"], df["b_stance"]]).unique()
                self.label_encoders["stance_encoder"].fit(all_stances)

            df["r_stance_encoded"] = self.label_encoders["stance_encoder"].transform(
                df["r_stance"].fillna("Orthodox")
            )
            df["b_stance_encoded"] = self.label_encoders["stance_encoder"].transform(
                df["b_stance"].fillna("Orthodox")
            )
            df["stance_diff"] = df["r_stance_encoded"] - df["b_stance_encoded"]

            # Orthodox vs Southpaw advantage
            df["orthodox_southpaw_matchup"] = np.where(
                (df["r_stance"] == "Orthodox") & (df["b_stance"] == "Southpaw"),
                0.05,
                np.where(
                    (df["r_stance"] == "Southpaw") & (df["b_stance"] == "Orthodox"),
                    -0.05,
                    0,
                ),
            )

            feature_columns.extend(["stance_diff", "orthodox_southpaw_matchup"])

        # ENHANCED FEATURE ENGINEERING FOR METHOD PREDICTION

        # Fighter-specific finishing pattern analysis
        df["r_ko_specialist"] = np.where(
            (df["r_ko_rate_corrected"] > 0.4) & (df["r_pro_SLpM_corrected"] > 4.0), 1, 0
        )
        df["b_ko_specialist"] = np.where(
            (df["b_ko_rate_corrected"] > 0.4) & (df["b_pro_SLpM_corrected"] > 4.0), 1, 0
        )
        df["ko_specialist_matchup"] = df["r_ko_specialist"] - df["b_ko_specialist"]

        df["r_sub_specialist"] = np.where(
            (df["r_sub_rate_corrected"] > 0.3) & (df["r_pro_td_avg_corrected"] > 2.0),
            1,
            0,
        )
        df["b_sub_specialist"] = np.where(
            (df["b_sub_rate_corrected"] > 0.3) & (df["b_pro_td_avg_corrected"] > 2.0),
            1,
            0,
        )
        df["sub_specialist_matchup"] = df["r_sub_specialist"] - df["b_sub_specialist"]

        # Durability and chin analysis
        df["r_durability_score"] = (
            df["r_durability_corrected"] * 0.4
            + (1 - df["r_pro_SApM_corrected"] / 6.0) * 0.3
            + (df["r_pro_str_def_corrected"]) * 0.3
        )
        df["b_durability_score"] = (
            df["b_durability_corrected"] * 0.4
            + (1 - df["b_pro_SApM_corrected"] / 6.0) * 0.3
            + (df["b_pro_str_def_corrected"]) * 0.3
        )
        df["durability_advantage"] = df["r_durability_score"] - df["b_durability_score"]

        # Knockdown resistance analysis
        df["r_kd_resistance"] = 1 - (df["r_pro_SApM_corrected"] / 6.0) * (
            1 - df["r_pro_str_def_corrected"]
        )
        df["b_kd_resistance"] = 1 - (df["b_pro_SApM_corrected"] / 6.0) * (
            1 - df["b_pro_str_def_corrected"]
        )
        df["kd_resistance_advantage"] = df["r_kd_resistance"] - df["b_kd_resistance"]

        # Style matchup analysis for method prediction
        df["striker_vs_grappler"] = np.where(
            (df["striker_advantage"] > 0.3) & (df["grappler_advantage"] < -0.3),
            1,
            np.where(
                (df["striker_advantage"] < -0.3) & (df["grappler_advantage"] > 0.3),
                -1,
                0,
            ),
        )

        # Distance vs ground fighting preference
        df["distance_ground_preference"] = (
            df["r_distance_pct_corrected"] - df["r_ground_pct_corrected"]
        ) - (df["b_distance_pct_corrected"] - df["b_ground_pct_corrected"])

        # Clinch fighting advantage
        df["clinch_advantage"] = (
            df["r_clinch_pct_corrected"] * df["r_pro_SLpM_corrected"]
        ) - (df["b_clinch_pct_corrected"] * df["b_pro_SLpM_corrected"])

        # Head hunting vs body work preference
        df["head_hunting_advantage"] = (
            df["r_head_pct_corrected"] * df["r_pro_SLpM_corrected"]
        ) - (df["b_head_pct_corrected"] * df["b_pro_SLpM_corrected"])

        # Opponent vulnerability scoring
        df["r_opponent_vulnerability"] = (
            (1 - df["b_pro_str_def_corrected"]) * 0.3
            + (df["b_pro_SApM_corrected"] / 6.0) * 0.2
            + (1 - df["b_durability_corrected"]) * 0.3
            + (df["b_pro_kd_pM_corrected"] / 0.5) * 0.2
        )
        df["b_opponent_vulnerability"] = (
            (1 - df["r_pro_str_def_corrected"]) * 0.3
            + (df["r_pro_SApM_corrected"] / 6.0) * 0.2
            + (1 - df["r_durability_corrected"]) * 0.3
            + (df["r_pro_kd_pM_corrected"] / 0.5) * 0.2
        )
        df["vulnerability_advantage"] = (
            df["r_opponent_vulnerability"] - df["b_opponent_vulnerability"]
        )

        # Contextual fight factors
        df["title_fight_ko_factor"] = np.where(df["is_title_bout"] == 1, 1.2, 1.0)
        df["five_round_advantage"] = np.where(df["total_rounds"] == 5, 1.15, 1.0)

        # Recent form and momentum analysis
        df["r_momentum_score"] = (
            df["r_recent_form_corrected"] * 0.4
            + df["r_win_streak_corrected"] / 10.0 * 0.3
            + df["r_recent_finish_rate_corrected"] * 0.3
        )
        df["b_momentum_score"] = (
            df["b_recent_form_corrected"] * 0.4
            + df["b_win_streak_corrected"] / 10.0 * 0.3
            + df["b_recent_finish_rate_corrected"] * 0.3
        )
        df["momentum_advantage"] = df["r_momentum_score"] - df["b_momentum_score"]

        # Age and experience interaction
        df["age_experience_interaction"] = (
            df["age_at_event_diff"] * df["experience_gap"] / 100.0
        )

        # Power vs technique analysis
        df["r_power_technique_ratio"] = (
            df["r_pro_SLpM_corrected"] * df["r_pro_sig_str_acc_corrected"]
        ) / (df["r_pro_SApM_corrected"] + 0.1)
        df["b_power_technique_ratio"] = (
            df["b_pro_SLpM_corrected"] * df["b_pro_sig_str_acc_corrected"]
        ) / (df["b_pro_SApM_corrected"] + 0.1)
        df["power_technique_advantage"] = (
            df["r_power_technique_ratio"] - df["b_power_technique_ratio"]
        )

        # Grappling control vs submission threat
        df["r_grappling_threat"] = (
            df["r_pro_td_avg_corrected"] * df["r_pro_td_acc_corrected"] * 0.6
            + df["r_pro_sub_avg_corrected"] * 0.4
        )
        df["b_grappling_threat"] = (
            df["b_pro_td_avg_corrected"] * df["b_pro_td_acc_corrected"] * 0.6
            + df["b_pro_sub_avg_corrected"] * 0.4
        )
        df["grappling_threat_advantage"] = (
            df["r_grappling_threat"] - df["b_grappling_threat"]
        )

        # ============================================================================
        # ADVANCED FEATURE ENGINEERING FOR HIGHER ACCURACY
        # ============================================================================

        # Fighter-specific finishing patterns over time
        df["r_recent_ko_trend"] = np.where(
            df["r_recent_finish_rate_corrected"] > df["r_ko_rate_corrected"], 1, 0
        )
        df["b_recent_ko_trend"] = np.where(
            df["b_recent_finish_rate_corrected"] > df["b_ko_rate_corrected"], 1, 0
        )
        df["recent_ko_trend_advantage"] = (
            df["r_recent_ko_trend"] - df["b_recent_ko_trend"]
        )

        # Opponent-specific finishing success rates
        df["r_opponent_ko_susceptibility"] = (
            df["b_pro_SApM_corrected"]
            * (1 - df["b_pro_str_def_corrected"])
            * (1 - df["b_durability_corrected"])
        )
        df["b_opponent_ko_susceptibility"] = (
            df["r_pro_SApM_corrected"]
            * (1 - df["r_pro_str_def_corrected"])
            * (1 - df["r_durability_corrected"])
        )
        df["ko_susceptibility_advantage"] = (
            df["r_opponent_ko_susceptibility"] - df["b_opponent_ko_susceptibility"]
        )

        # Submission opportunity analysis
        df["r_sub_opportunity_score"] = (
            df["r_pro_td_avg_corrected"]
            * df["r_pro_td_acc_corrected"]
            * df["r_sub_rate_corrected"]
            * (1 - df["b_pro_td_def_corrected"])
        )
        df["b_sub_opportunity_score"] = (
            df["b_pro_td_avg_corrected"]
            * df["b_pro_td_acc_corrected"]
            * df["b_sub_rate_corrected"]
            * (1 - df["r_pro_td_def_corrected"])
        )
        df["sub_opportunity_advantage"] = (
            df["r_sub_opportunity_score"] - df["b_sub_opportunity_score"]
        )

        # Fight pace and cardio analysis
        df["r_cardio_factor"] = (
            df["r_avg_fight_time"]
            * df["r_pro_SLpM_corrected"]
            / (df["r_pro_SApM_corrected"] + 0.1)
        )
        df["b_cardio_factor"] = (
            df["b_avg_fight_time"]
            * df["b_pro_SLpM_corrected"]
            / (df["b_pro_SApM_corrected"] + 0.1)
        )
        # cardio_advantage will be created later in the new features section

        # Weight class specific adjustments
        df["weight_class_ko_factor"] = np.where(
            df["weight_class"].isin(["Heavyweight", "Light Heavyweight"]),
            1.3,
            np.where(
                df["weight_class"].isin(["Middleweight", "Welterweight"]), 1.1, 1.0
            ),
        )

        # Championship experience and pressure
        df["r_championship_pressure"] = (
            df["r_championship_experience"]
            * df["is_title_bout"]
            * df["r_recent_form_corrected"]
        )
        df["b_championship_pressure"] = (
            df["b_championship_experience"]
            * df["is_title_bout"]
            * df["b_recent_form_corrected"]
        )
        df["championship_pressure_advantage"] = (
            df["r_championship_pressure"] - df["b_championship_pressure"]
        )

        # Recent opponent quality analysis (using available features)
        df["r_opponent_quality"] = (
            df["r_recent_form_corrected"] * 0.6
            + df["r_recent_finish_rate_corrected"] * 0.4
        )
        df["b_opponent_quality"] = (
            df["b_recent_form_corrected"] * 0.6
            + df["b_recent_finish_rate_corrected"] * 0.4
        )
        df["opponent_quality_advantage"] = (
            df["r_opponent_quality"] - df["b_opponent_quality"]
        )

        # Fight ending probability by round
        df["early_finish_probability"] = (
            (df["r_ko_rate_corrected"] + df["r_sub_rate_corrected"])
            * (df["b_ko_rate_corrected"] + df["b_sub_rate_corrected"])
            * 0.5
        )

        # Technical striking vs brawling analysis
        df["r_technical_striker"] = np.where(
            (df["r_pro_sig_str_acc_corrected"] > 0.5)
            & (df["r_pro_SLpM_corrected"] > 3.0),
            1,
            0,
        )
        df["b_technical_striker"] = np.where(
            (df["b_pro_sig_str_acc_corrected"] > 0.5)
            & (df["b_pro_SLpM_corrected"] > 3.0),
            1,
            0,
        )
        df["technical_striker_advantage"] = (
            df["r_technical_striker"] - df["b_technical_striker"]
        )

        # Clinch and dirty boxing effectiveness
        df["r_clinch_effectiveness"] = (
            df["r_clinch_pct_corrected"]
            * df["r_pro_SLpM_corrected"]
            * df["r_pro_sig_str_acc_corrected"]
        )
        df["b_clinch_effectiveness"] = (
            df["b_clinch_pct_corrected"]
            * df["b_pro_SLpM_corrected"]
            * df["b_pro_sig_str_acc_corrected"]
        )
        df["clinch_effectiveness_advantage"] = (
            df["r_clinch_effectiveness"] - df["b_clinch_effectiveness"]
        )

        # ============================================================================
        # NEW HIGH-IMPACT FEATURES FOR BETTER ACCURACY
        # ============================================================================

        # APE INDEX ADVANTAGE (Reach - Height) - Critical for striking range
        df["ape_index_advantage"] = df["ape_index_diff"]

        # ENHANCED MOMENTUM FEATURES
        df["momentum_velocity"] = (
            df["r_recent_form_corrected"] - df["b_recent_form_corrected"]
        ) * (df["r_win_streak_corrected"] - df["b_win_streak_corrected"])

        # FIGHTER STYLE MATCHUP DEPTH
        df["style_matchup_depth"] = (
            abs(df["striker_advantage"])
            + abs(df["grappler_advantage"])
            + abs(df["r_pro_SLpM_corrected"] - df["b_pro_SLpM_corrected"]) * 0.3
        )

        # CHAMPIONSHIP EXPERIENCE IMPACT
        df["championship_impact"] = (
            df["championship_exp_diff"]
            * df["is_title_bout"]
            * (df["r_recent_form_corrected"] - df["b_recent_form_corrected"])
        )

        # OPPONENT QUALITY MOMENTUM
        df["opponent_quality_momentum"] = (
            df["r_opponent_quality"] - df["b_opponent_quality"]
        ) * (df["r_recent_form_corrected"] - df["b_recent_form_corrected"])

        # FINISHING PRESSURE UNDER STRESS
        df["finishing_pressure_stress"] = (
            df["r_ko_rate_corrected"] + df["r_sub_rate_corrected"]
        ) * (
            df["r_recent_form_corrected"] * 0.6 + df["r_win_streak_corrected"] * 0.4
        ) - (df["b_ko_rate_corrected"] + df["b_sub_rate_corrected"]) * (
            df["b_recent_form_corrected"] * 0.6 + df["b_win_streak_corrected"] * 0.4
        )

        # RING RUST FACTOR - Time since last fight (moved up to be created first)
        df["ring_rust_factor"] = np.where(
            df["days_since_last_fight_diff_corrected"] > 0,
            -0.1
            * (df["days_since_last_fight_diff_corrected"] / 365),  # Red corner rust
            np.where(
                df["days_since_last_fight_diff_corrected"] < 0,
                0.1
                * (
                    abs(df["days_since_last_fight_diff_corrected"]) / 365
                ),  # Blue corner rust
                0,
            ),
        )

        # RING RUST VS MOMENTUM
        df["ring_rust_vs_momentum"] = df["ring_rust_factor"] * (
            df["r_recent_form_corrected"] - df["b_recent_form_corrected"]
        )

        # WEIGHT CLASS FACTOR - Different dynamics in each division (moved up to be created first)
        weight_class_factors = {
            "Heavyweight": 1.2,
            "Light Heavyweight": 1.1,
            "Middleweight": 1.0,
            "Welterweight": 0.95,
            "Lightweight": 0.9,
            "Featherweight": 0.85,
            "Bantamweight": 0.8,
            "Flyweight": 0.75,
            "Women's Bantamweight": 0.8,
            "Women's Flyweight": 0.75,
            "Women's Strawweight": 0.7,
        }
        df["weight_class_factor"] = (
            df["weight_class"].map(weight_class_factors).fillna(1.0)
        )

        # WEIGHT CLASS ADAPTATION
        df["weight_class_adaptation"] = df["weight_class_factor"] * (
            df["r_pro_SLpM_corrected"] - df["b_pro_SLpM_corrected"]
        )

        # STYLE CLASH SEVERITY - How different the fighting styles are (moved up to be created first)
        df["style_clash_severity"] = abs(
            df["striker_vs_grappler"] * 0.5
            + (df["r_pro_SLpM_corrected"] - df["b_pro_SLpM_corrected"]) * 0.3
            + (df["r_pro_td_avg_corrected"] - df["b_pro_td_avg_corrected"]) * 0.2
        )

        # ENHANCED STANCE MATCHUP ADVANTAGE (including Switch stance) - moved up to be created first
        if "r_stance" in df.columns and "b_stance" in df.columns:
            # Create comprehensive stance matchup matrix
            def calculate_stance_advantage(r_stance, b_stance):
                # Based on real-world fight data analysis
                if r_stance == "Orthodox" and b_stance == "Southpaw":
                    return -0.06  # Southpaw has 6% advantage (52% vs 46%)
                elif r_stance == "Southpaw" and b_stance == "Orthodox":
                    return 0.06  # Southpaw has 6% advantage (52% vs 46%)
                elif r_stance == "Switch" and b_stance == "Orthodox":
                    return 0.15  # Switch has 15% advantage (57% vs 42%)
                elif r_stance == "Orthodox" and b_stance == "Switch":
                    return -0.15  # Switch has 15% advantage (57% vs 42%)
                elif r_stance == "Switch" and b_stance == "Southpaw":
                    return 0.13  # Switch has 13% advantage (56% vs 43%)
                elif r_stance == "Southpaw" and b_stance == "Switch":
                    return -0.13  # Switch has 13% advantage (56% vs 43%)
                elif r_stance == "Switch" and b_stance == "Switch":
                    return 0  # Neutral (no data available)
                else:
                    return 0  # Default neutral

            df["stance_matchup_advantage"] = df.apply(
                lambda row: calculate_stance_advantage(
                    row["r_stance"], row["b_stance"]
                ),
                axis=1,
            )

            # Add stance versatility factor (Switch fighters are more versatile based on real data)
            # Switch fighters show 15% advantage vs Orthodox and 13% advantage vs Southpaw
            df["stance_versatility_advantage"] = np.where(
                df["r_stance"] == "Switch",
                0.14,  # Average of 15% and 13% advantages
                np.where(df["b_stance"] == "Switch", -0.14, 0),
            )
        else:
            df["stance_matchup_advantage"] = 0
            df["stance_versatility_advantage"] = 0

        # STANCE VERSATILITY IMPACT
        df["stance_versatility_impact"] = (
            df["stance_versatility_advantage"] * df["style_clash_severity"]
        )

        # WEIGHT CLASS FACTOR already created above

        # RING RUST FACTOR already created above

        # CHAMPIONSHIP PRESSURE - Title fight experience and pressure
        df["championship_pressure"] = (
            df["championship_exp_diff"]
            * df["is_title_bout"]
            * (df["r_recent_form_corrected"] - df["b_recent_form_corrected"])
        )

        # MOMENTUM SWING - Recent performance trends
        df["momentum_swing"] = (
            df["r_recent_form_corrected"]
            - df["b_recent_form_corrected"]
            + (df["r_win_streak_corrected"] - df["b_win_streak_corrected"]) * 0.1
        )

        # STYLE CLASH SEVERITY already created above

        # POWER VS TECHNIQUE - Striking power vs technical striking
        df["power_vs_technique"] = (
            df["r_pro_SLpM_corrected"] - df["b_pro_SLpM_corrected"]
        ) * 0.6 + (
            df["r_pro_sig_str_acc_corrected"] - df["b_pro_sig_str_acc_corrected"]
        ) * 0.4

        # CARDIO ADVANTAGE - Endurance and pace
        df["cardio_advantage"] = (
            df["r_avg_fight_time"]
            - df["b_avg_fight_time"]
            + (df["r_pro_SLpM_corrected"] - df["b_pro_SLpM_corrected"]) * 0.1
        )

        # FINISH PRESSURE - Finishing ability under pressure
        df["finish_pressure"] = (
            (df["r_ko_rate_corrected"] - df["b_ko_rate_corrected"]) * 0.5
            + (df["r_sub_rate_corrected"] - df["b_sub_rate_corrected"]) * 0.3
            + (
                df["r_recent_finish_rate_corrected"]
                - df["b_recent_finish_rate_corrected"]
            )
            * 0.2
        )

        # Create recent opponent strength columns if they don't exist
        if "r_recent_opponent_strength" not in df.columns:
            df["r_recent_opponent_strength"] = df[
                "r_opponent_quality"
            ]  # Use opponent quality as proxy
        if "b_recent_opponent_strength" not in df.columns:
            df["b_recent_opponent_strength"] = df[
                "b_opponent_quality"
            ]  # Use opponent quality as proxy

        # OPPONENT QUALITY GAP - Quality of recent opponents
        df["opponent_quality_gap"] = (
            df["r_opponent_quality"]
            - df["b_opponent_quality"]
            + (df["r_recent_opponent_strength"] - df["b_recent_opponent_strength"])
            * 0.5
        )

        # RECENT OPPONENT STRENGTH - Strength of recent competition
        df["recent_opponent_strength"] = (
            df["r_recent_opponent_strength"] - df["b_recent_opponent_strength"]
        )

        # UPSET POTENTIAL - Factors that could lead to an upset
        df["upset_potential"] = (
            (df["b_recent_form_corrected"] - df["r_recent_form_corrected"]) * 0.4
            + (df["b_win_streak_corrected"] - df["r_win_streak_corrected"]) * 0.3
            + (df["b_age_at_event"] - df["r_age_at_event"]) * 0.1
            + (df["b_total_fights"] - df["r_total_fights"]) * 0.2
        )

        # ============================================================================
        # NEW HIGH-IMPACT WINNER PREDICTION FEATURES
        # ============================================================================

        # RECENT OPPONENT QUALITY - Quality of recent opponents
        df["r_recent_opponent_quality"] = (
            df["r_recent_opponent_strength"] * 0.7 + df["r_wins_corrected"] * 0.3
        )
        df["b_recent_opponent_quality"] = (
            df["b_recent_opponent_strength"] * 0.7 + df["b_wins_corrected"] * 0.3
        )
        df["recent_opponent_quality_diff"] = (
            df["r_recent_opponent_quality"] - df["b_recent_opponent_quality"]
        )

        # FINISH RATE - Recent finishing ability
        df["r_finish_rate"] = (
            df["r_ko_rate_corrected"] + df["r_sub_rate_corrected"]
        ) / 2
        df["b_finish_rate"] = (
            df["b_ko_rate_corrected"] + df["b_sub_rate_corrected"]
        ) / 2
        df["finish_rate_diff"] = df["r_finish_rate"] - df["b_finish_rate"]

        # STAMINA FACTOR - Based on fight duration patterns
        df["r_stamina_factor"] = np.where(
            df["r_avg_fight_time"] > 12,
            1.2,  # Championship rounds experience
            np.where(df["r_avg_fight_time"] > 8, 1.0, 0.8),  # Regular vs short fights
        )
        df["b_stamina_factor"] = np.where(
            df["b_avg_fight_time"] > 12,
            1.2,
            np.where(df["b_avg_fight_time"] > 8, 1.0, 0.8),
        )
        df["stamina_factor_diff"] = df["r_stamina_factor"] - df["b_stamina_factor"]

        # PRESSURE PERFORMANCE - Performance in high-stakes fights
        df["r_pressure_performance"] = np.where(
            df["is_title_bout"] == 1,
            df["r_recent_form_corrected"] * 1.2,
            df["r_recent_form_corrected"],
        )
        df["b_pressure_performance"] = np.where(
            df["is_title_bout"] == 1,
            df["b_recent_form_corrected"] * 1.2,
            df["b_recent_form_corrected"],
        )
        df["pressure_performance_diff"] = (
            df["r_pressure_performance"] - df["b_pressure_performance"]
        )

        # COMEBACK ABILITY - Ability to recover from adversity
        df["r_comeback_ability"] = np.where(
            df["r_loss_streak_corrected"] > 0,
            df["r_recent_form_corrected"] * 0.8,  # Penalty for recent losses
            df["r_recent_form_corrected"] * 1.1,  # Bonus for avoiding losses
        )
        df["b_comeback_ability"] = np.where(
            df["b_loss_streak_corrected"] > 0,
            df["b_recent_form_corrected"] * 0.8,
            df["b_recent_form_corrected"] * 1.1,
        )
        df["comeback_ability_diff"] = (
            df["r_comeback_ability"] - df["b_comeback_ability"]
        )

        # CLUTCH FACTOR - Performance in close fights
        df["r_clutch_factor"] = (
            df["r_win_loss_ratio_corrected"] * df["r_recent_form_corrected"]
        )
        df["b_clutch_factor"] = (
            df["b_win_loss_ratio_corrected"] * df["b_recent_form_corrected"]
        )
        df["clutch_factor_diff"] = df["r_clutch_factor"] - df["b_clutch_factor"]

        # RECENT TREND - Performance trajectory over last 3-5 fights
        df["r_recent_trend"] = (
            df["r_recent_form_corrected"] * df["r_win_streak_corrected"]
        )
        df["b_recent_trend"] = (
            df["b_recent_form_corrected"] * df["b_win_streak_corrected"]
        )
        df["recent_trend_diff"] = df["r_recent_trend"] - df["b_recent_trend"]

        # MOMENTUM SHIFT - Change in performance momentum
        df["r_momentum_shift"] = (
            df["r_win_streak_corrected"] - df["r_loss_streak_corrected"]
        )
        df["b_momentum_shift"] = (
            df["b_win_streak_corrected"] - df["b_loss_streak_corrected"]
        )
        df["momentum_shift_diff"] = df["r_momentum_shift"] - df["b_momentum_shift"]

        # FORM CONSISTENCY - How consistent recent performance is
        df["r_form_consistency"] = (
            1.0 - np.abs(df["r_recent_form_corrected"] - 0.5) * 2
        )  # Closer to 0.5 = more consistent
        df["b_form_consistency"] = 1.0 - np.abs(df["b_recent_form_corrected"] - 0.5) * 2
        df["form_consistency_diff"] = (
            df["r_form_consistency"] - df["b_form_consistency"]
        )

        # UPSET HISTORY - History of causing/being upset
        df["r_upset_history"] = np.where(
            df["r_recent_form_corrected"] > 0.7,
            1.2,  # Strong favorite
            np.where(
                df["r_recent_form_corrected"] < 0.3, 0.8, 1.0
            ),  # Underdog or neutral
        )
        df["b_upset_history"] = np.where(
            df["b_recent_form_corrected"] > 0.7,
            1.2,
            np.where(df["b_recent_form_corrected"] < 0.3, 0.8, 1.0),
        )
        df["upset_history_diff"] = df["r_upset_history"] - df["b_upset_history"]

        # Add all new features to column list
        feature_columns.extend(
            [
                "net_striking_advantage",
                "striking_efficiency",
                "defensive_striking",
                "grappling_control",
                "grappling_defense",
                "offensive_output",
                "defensive_composite",
                "ko_specialist_gap",
                "submission_specialist_gap",
                "experience_gap",
                "skill_momentum",
                "finish_threat",
                "momentum_advantage",
                "inactivity_penalty",
                "pace_differential",
                "r_total_fights",
                "b_total_fights",
                "experience_ratio",
                "avg_fight_time_diff",
                "quality_experience_gap",
                "championship_exp_diff",
                "adversity_exp_diff",
                "experience_skill_interaction",
                "veteran_edge",
                "novice_vulnerability",
                "exp_gap_historical_win_rate",
                "striker_advantage",
                "grappler_advantage",
                "effective_reach_advantage",
                "ko_specialist_matchup",
                "sub_specialist_matchup",
                "durability_advantage",
                "kd_resistance_advantage",
                "striker_vs_grappler",
                "distance_ground_preference",
                "clinch_advantage",
                "head_hunting_advantage",
                "vulnerability_advantage",
                "title_fight_ko_factor",
                "five_round_advantage",
                "age_experience_interaction",
                "power_technique_advantage",
                "grappling_threat_advantage",
                "recent_ko_trend_advantage",
                "ko_susceptibility_advantage",
                "sub_opportunity_advantage",
                "cardio_advantage",
                "weight_class_ko_factor",
                "championship_pressure_advantage",
                "opponent_quality_advantage",
                "early_finish_probability",
                "technical_striker_advantage",
                "clinch_effectiveness_advantage",
                # NEW HIGH-IMPACT FEATURES
                "ape_index_advantage",
                "stance_matchup_advantage",
                "stance_versatility_advantage",
                "weight_class_factor",
                "ring_rust_factor",
                "championship_pressure",
                "momentum_swing",
                "style_clash_severity",
                "power_vs_technique",
                "cardio_advantage",  # potentially causes error
                "finish_pressure",
                "opponent_quality_gap",
                "recent_opponent_strength",
                "upset_potential",
                # ENHANCED ACCURACY FEATURES
                "momentum_velocity",
                "style_matchup_depth",
                "championship_impact",
                "opponent_quality_momentum",
                "finishing_pressure_stress",
                "ring_rust_vs_momentum",
                "weight_class_adaptation",
                "stance_versatility_impact",
            ]
        )

        # ============================================================================
        # NEW FEATURES FOR ENHANCED PREDICTION ACCURACY
        # ============================================================================
        
        # WINNER PREDICTION FEATURES (10 new features)
        
        # 1. Clinch Control Advantage - effectiveness in clinch range
        if all(col in df.columns for col in ["r_clinch_pct_corrected", "b_clinch_pct_corrected", "r_pro_SLpM_corrected", "b_pro_SLpM_corrected"]):
            df["clinch_control_advantage"] = (df["r_clinch_pct_corrected"] * df["r_pro_SLpM_corrected"]) - (df["b_clinch_pct_corrected"] * df["b_pro_SLpM_corrected"])
        
        # 2. Ground Control Advantage - ground control effectiveness
        if all(col in df.columns for col in ["r_ground_pct_corrected", "b_ground_pct_corrected", "r_pro_td_avg_corrected", "b_pro_td_avg_corrected"]):
            df["ground_control_advantage"] = (df["r_ground_pct_corrected"] * df["r_pro_td_avg_corrected"]) - (df["b_ground_pct_corrected"] * df["b_pro_td_avg_corrected"])
        
        # 3. Striking Volume Advantage - volume × accuracy composite
        if all(col in df.columns for col in ["r_pro_SLpM_corrected", "b_pro_SLpM_corrected", "r_pro_sig_str_acc_corrected", "b_pro_sig_str_acc_corrected"]):
            df["striking_volume_advantage"] = (df["r_pro_SLpM_corrected"] * df["r_pro_sig_str_acc_corrected"]) - (df["b_pro_SLpM_corrected"] * df["b_pro_sig_str_acc_corrected"])
        
        # 4. Defensive Advantage - overall defensive capability
        if all(col in df.columns for col in ["r_pro_str_def_corrected", "b_pro_str_def_corrected", "r_pro_SApM_corrected", "b_pro_SApM_corrected"]):
            df["defensive_advantage"] = (df["r_pro_str_def_corrected"] - df["r_pro_SApM_corrected"]) - (df["b_pro_str_def_corrected"] - df["b_pro_SApM_corrected"])
        
        # 5. Finish Avoidance Advantage - ability to avoid being finished
        if all(col in df.columns for col in ["r_ko_losses_corrected", "b_ko_losses_corrected", "r_sub_losses_corrected", "b_sub_losses_corrected"]):
            df["finish_avoidance_advantage"] = (1 / (1 + df["r_ko_losses_corrected"] + df["r_sub_losses_corrected"])) - (1 / (1 + df["b_ko_losses_corrected"] + df["b_sub_losses_corrected"]))
        
        # 6. Pace Control Advantage - fight pace control ability
        if all(col in df.columns for col in ["r_avg_fight_time_corrected", "b_avg_fight_time_corrected", "r_pro_SLpM_corrected", "b_pro_SLpM_corrected"]):
            df["pace_control_advantage"] = (df["r_avg_fight_time_corrected"] - df["b_avg_fight_time_corrected"]) * (df["r_pro_SLpM_corrected"] - df["b_pro_SLpM_corrected"])
        
        # 7. Experience Quality Advantage - quality-adjusted experience
        if all(col in df.columns for col in ["r_total_fights_corrected", "b_total_fights_corrected", "r_win_loss_ratio_corrected", "b_win_loss_ratio_corrected"]):
            df["experience_quality_advantage"] = (df["r_total_fights_corrected"] * df["r_win_loss_ratio_corrected"]) - (df["b_total_fights_corrected"] * df["b_win_loss_ratio_corrected"])
        
        # 8. Momentum Consistency Advantage - consistent momentum advantage
        if all(col in df.columns for col in ["r_recent_form_corrected", "b_recent_form_corrected", "r_win_streak_corrected", "b_win_streak_corrected"]):
            df["momentum_consistency_advantage"] = (df["r_recent_form_corrected"] * df["r_win_streak_corrected"]) - (df["b_recent_form_corrected"] * df["b_win_streak_corrected"])
        
        # 9. Striking Defense Advantage - striking defense effectiveness
        if all(col in df.columns for col in ["r_pro_str_def_corrected", "b_pro_str_def_corrected", "r_pro_SApM_corrected", "b_pro_SApM_corrected"]):
            df["striking_defense_advantage"] = (df["r_pro_str_def_corrected"] * (1 - df["r_pro_SApM_corrected"]/10)) - (df["b_pro_str_def_corrected"] * (1 - df["b_pro_SApM_corrected"]/10))
        
        # 10. Grappling Offense Advantage - grappling offensive effectiveness
        if all(col in df.columns for col in ["r_pro_td_avg_corrected", "b_pro_td_avg_corrected", "r_pro_td_acc_corrected", "b_pro_td_acc_corrected"]):
            df["grappling_offense_advantage"] = (df["r_pro_td_avg_corrected"] * df["r_pro_td_acc_corrected"]) - (df["b_pro_td_avg_corrected"] * df["b_pro_td_acc_corrected"])
        
        # METHOD PREDICTION FEATURES (5 new features)
        
        # 1. KO Power Advantage - KO power composite (rate × knockdowns)
        if all(col in df.columns for col in ["r_ko_rate_corrected", "b_ko_rate_corrected", "r_pro_kd_pM_corrected", "b_pro_kd_pM_corrected"]):
            df["ko_power_advantage"] = (df["r_ko_rate_corrected"] * df["r_pro_kd_pM_corrected"]) - (df["b_ko_rate_corrected"] * df["b_pro_kd_pM_corrected"])
        
        # 2. Submission Threat Advantage - submission threat composite
        if all(col in df.columns for col in ["r_sub_rate_corrected", "b_sub_rate_corrected", "r_pro_sub_avg_corrected", "b_pro_sub_avg_corrected"]):
            df["submission_threat_advantage"] = (df["r_sub_rate_corrected"] * df["r_pro_sub_avg_corrected"]) - (df["b_sub_rate_corrected"] * df["b_pro_sub_avg_corrected"])
        
        # 3. Decision Tendency Advantage - decision likelihood composite
        if all(col in df.columns for col in ["r_dec_rate_corrected", "b_dec_rate_corrected", "r_avg_fight_time_corrected", "b_avg_fight_time_corrected"]):
            df["decision_tendency_advantage"] = (df["r_dec_rate_corrected"] * df["r_avg_fight_time_corrected"]) - (df["b_dec_rate_corrected"] * df["b_avg_fight_time_corrected"])
        
        # 4. Early Finish Advantage - early finish likelihood
        if all(col in df.columns for col in ["r_ko_rate_corrected", "b_ko_rate_corrected", "r_sub_rate_corrected", "b_sub_rate_corrected"]):
            df["early_finish_advantage"] = (df["r_ko_rate_corrected"] + df["r_sub_rate_corrected"]) - (df["b_ko_rate_corrected"] + df["b_sub_rate_corrected"])
        
        # 5. Durability Advantage - knockdown resistance
        if all(col in df.columns for col in ["r_ko_losses_corrected", "b_ko_losses_corrected"]):
            df["durability_advantage"] = (1 / (1 + df["r_ko_losses_corrected"])) - (1 / (1 + df["b_ko_losses_corrected"]))
        
        # ============================================================================
        # SOPHISTICATED WINNER PREDICTION FEATURES (12 advanced features)
        # ============================================================================
        
        # 1. Fight IQ Composite Advantage - defensive intelligence
        if all(col in df.columns for col in ["r_pro_str_def_corrected", "b_pro_str_def_corrected", "r_pro_td_def_corrected", "b_pro_td_def_corrected", "r_pro_sig_str_acc_corrected", "b_pro_sig_str_acc_corrected", "r_pro_SApM_corrected", "b_pro_SApM_corrected"]):
            df["fight_iq_composite_advantage"] = (
                (df["r_pro_str_def_corrected"] * df["r_pro_td_def_corrected"] * df["r_pro_sig_str_acc_corrected"]) / 
                (df["r_pro_SApM_corrected"] + 0.1)
            ) - (
                (df["b_pro_str_def_corrected"] * df["b_pro_td_def_corrected"] * df["b_pro_sig_str_acc_corrected"]) / 
                (df["b_pro_SApM_corrected"] + 0.1)
            )
        
        # 2. Championship Pressure Performance - big fight experience
        if all(col in df.columns for col in ["r_championship_pressure_advantage", "b_championship_pressure_advantage", "r_clutch_factor_diff", "b_clutch_factor_diff", "r_recent_form_diff_corrected", "b_recent_form_diff_corrected", "r_days_since_last_fight_diff_corrected", "b_days_since_last_fight_diff_corrected"]):
            df["championship_pressure_performance"] = (
                (df["r_championship_pressure_advantage"] * df["r_clutch_factor_diff"] * df["r_recent_form_diff_corrected"]) / 
                (df["r_days_since_last_fight_diff_corrected"] + 30)
            ) - (
                (df["b_championship_pressure_advantage"] * df["b_clutch_factor_diff"] * df["b_recent_form_diff_corrected"]) / 
                (df["b_days_since_last_fight_diff_corrected"] + 30)
            )
        
        # 3. Momentum Velocity Advantage - hot streak dynamics
        if all(col in df.columns for col in ["r_win_streak_diff_corrected", "b_win_streak_diff_corrected", "r_recent_form_diff_corrected", "b_recent_form_diff_corrected", "r_finish_rate_diff", "b_finish_rate_diff", "r_loss_streak_diff_corrected", "b_loss_streak_diff_corrected"]):
            df["momentum_velocity_advantage"] = (
                (df["r_win_streak_diff_corrected"] * df["r_recent_form_diff_corrected"] * df["r_finish_rate_diff"]) / 
                (abs(df["r_loss_streak_diff_corrected"]) + 1)
            ) - (
                (df["b_win_streak_diff_corrected"] * df["b_recent_form_diff_corrected"] * df["b_finish_rate_diff"]) / 
                (abs(df["b_loss_streak_diff_corrected"]) + 1)
            )
        
        # 4. Style Matchup Depth Score - complex style interactions
        if all(col in df.columns for col in ["striker_vs_grappler", "power_vs_technique", "stance_matchup_advantage", "style_clash_severity", "experience_gap"]):
            df["style_matchup_depth_score"] = (
                (df["striker_vs_grappler"] * df["power_vs_technique"] * df["stance_matchup_advantage"]) / 
                (df["style_clash_severity"] + 0.1)
            ) * (df["experience_gap"] / 10)
        
        # 5. Durability vs Power Ratio - survival vs knockout power
        if all(col in df.columns for col in ["r_durability_score", "b_durability_score", "r_pro_str_def_corrected", "b_pro_str_def_corrected", "r_ko_rate_corrected", "b_ko_rate_corrected", "r_pro_kd_pM_corrected", "b_pro_kd_pM_corrected"]):
            df["durability_vs_power_ratio"] = (
                (df["r_durability_score"] * df["r_pro_str_def_corrected"]) / 
                (df["b_ko_rate_corrected"] * df["b_pro_kd_pM_corrected"] + 0.1)
            ) - (
                (df["b_durability_score"] * df["b_pro_str_def_corrected"]) / 
                (df["r_ko_rate_corrected"] * df["r_pro_kd_pM_corrected"] + 0.1)
            )
        
        # 6. Pace Control Mastery - fight duration and efficiency
        if all(col in df.columns for col in ["r_avg_fight_time", "b_avg_fight_time", "r_pro_SLpM_corrected", "b_pro_SLpM_corrected", "r_pro_SApM_corrected", "b_pro_SApM_corrected", "r_pro_str_def_corrected", "b_pro_str_def_corrected"]):
            df["pace_control_mastery"] = (
                (df["r_avg_fight_time"] / 5) * (df["r_pro_SLpM_corrected"] - df["r_pro_SApM_corrected"]) * 
                (df["r_pro_str_def_corrected"] / (df["r_pro_SApM_corrected"] + 0.1))
            ) - (
                (df["b_avg_fight_time"] / 5) * (df["b_pro_SLpM_corrected"] - df["b_pro_SApM_corrected"]) * 
                (df["b_pro_str_def_corrected"] / (df["b_pro_SApM_corrected"] + 0.1))
            )
        
        # 7. Grappling Transition Mastery - ground control effectiveness
        if all(col in df.columns for col in ["r_pro_td_avg_corrected", "b_pro_td_avg_corrected", "r_pro_td_acc_corrected", "b_pro_td_acc_corrected", "r_pro_sub_avg_corrected", "b_pro_sub_avg_corrected", "r_pro_td_def_corrected", "b_pro_td_def_corrected", "r_ground_pct_corrected", "b_ground_pct_corrected", "r_clinch_pct_corrected", "b_clinch_pct_corrected"]):
            df["grappling_transition_mastery"] = (
                (df["r_pro_td_avg_corrected"] * df["r_pro_td_acc_corrected"] * df["r_pro_sub_avg_corrected"]) / 
                (df["r_pro_td_def_corrected"] + 0.1)
            ) * (df["r_ground_pct_corrected"] / (df["r_clinch_pct_corrected"] + 0.1)) - (
                (df["b_pro_td_avg_corrected"] * df["b_pro_td_acc_corrected"] * df["b_pro_sub_avg_corrected"]) / 
                (df["b_pro_td_def_corrected"] + 0.1)
            ) * (df["b_ground_pct_corrected"] / (df["b_clinch_pct_corrected"] + 0.1))
        
        # 8. Upset Potential vs Experience - underdog dynamics
        if all(col in df.columns for col in ["r_upset_history", "b_upset_history", "r_recent_opponent_quality_diff", "b_recent_opponent_quality_diff", "r_total_fights", "b_total_fights", "r_age_at_event", "b_age_at_event"]):
            df["upset_potential_vs_experience"] = (
                (df["r_upset_history"] * df["r_recent_opponent_quality_diff"]) / 
                (df["b_total_fights"] + 1)
            ) * (df["r_age_at_event"] / (df["b_age_at_event"] + 1)) - (
                (df["b_upset_history"] * df["b_recent_opponent_quality_diff"]) / 
                (df["r_total_fights"] + 1)
            ) * (df["b_age_at_event"] / (df["r_age_at_event"] + 1))
        
        # 9. Physical Dominance Composite - size and athleticism
        if all(col in df.columns for col in ["r_height", "b_height", "r_reach", "b_reach", "r_weight", "b_weight", "r_ape_index", "b_ape_index", "r_age_at_event", "b_age_at_event"]):
            df["physical_dominance_composite"] = (
                ((df["r_height"] * df["r_reach"] * df["r_weight"]) / 1000) * 
                (df["r_ape_index"] / 10) * (df["r_age_at_event"] / 30)
            ) - (
                ((df["b_height"] * df["b_reach"] * df["b_weight"]) / 1000) * 
                (df["b_ape_index"] / 10) * (df["b_age_at_event"] / 30)
            )
        
        # 10. Technical Striking Mastery - precision and targeting
        if all(col in df.columns for col in ["r_pro_sig_str_acc_corrected", "b_pro_sig_str_acc_corrected", "r_pro_str_def_corrected", "b_pro_str_def_corrected", "r_pro_SLpM_corrected", "b_pro_SLpM_corrected", "r_pro_SApM_corrected", "b_pro_SApM_corrected", "r_head_pct_corrected", "b_head_pct_corrected", "r_body_pct_corrected", "b_body_pct_corrected", "r_leg_pct_corrected", "b_leg_pct_corrected"]):
            df["technical_striking_mastery"] = (
                (df["r_pro_sig_str_acc_corrected"] * df["r_pro_str_def_corrected"] * df["r_pro_SLpM_corrected"]) / 
                (df["r_pro_SApM_corrected"] + 0.1)
            ) * (df["r_head_pct_corrected"] / (df["r_body_pct_corrected"] + df["r_leg_pct_corrected"] + 0.1)) - (
                (df["b_pro_sig_str_acc_corrected"] * df["b_pro_str_def_corrected"] * df["b_pro_SLpM_corrected"]) / 
                (df["b_pro_SApM_corrected"] + 0.1)
            ) * (df["b_head_pct_corrected"] / (df["b_body_pct_corrected"] + df["b_leg_pct_corrected"] + 0.1))
        
        # 11. Clutch Performance Under Pressure - mental toughness
        if all(col in df.columns for col in ["r_clutch_factor", "b_clutch_factor", "r_championship_pressure", "b_championship_pressure", "r_recent_form_corrected", "b_recent_form_corrected", "r_days_since_last_fight_diff_corrected", "b_days_since_last_fight_diff_corrected", "r_win_streak_corrected", "b_win_streak_corrected", "r_loss_streak_corrected", "b_loss_streak_corrected"]):
            df["clutch_performance_under_pressure"] = (
                (df["r_clutch_factor"] * df["r_championship_pressure"] * df["r_recent_form_corrected"]) / 
                (df["r_days_since_last_fight_diff_corrected"] + 30)
            ) * (df["r_win_streak_corrected"] / (abs(df["r_loss_streak_corrected"]) + 1)) - (
                (df["b_clutch_factor"] * df["b_championship_pressure"] * df["b_recent_form_corrected"]) / 
                (df["b_days_since_last_fight_diff_corrected"] + 30)
            ) * (df["b_win_streak_corrected"] / (abs(df["b_loss_streak_corrected"]) + 1))
        
        # 12. Veteran Wisdom vs Youth Advantage - experience vs athleticism
        if all(col in df.columns for col in ["r_total_fights", "b_total_fights", "r_age_at_event", "b_age_at_event", "r_pro_str_def_corrected", "b_pro_str_def_corrected", "r_pro_SApM_corrected", "b_pro_SApM_corrected", "r_championship_experience", "b_championship_experience", "r_wins_corrected", "b_wins_corrected"]):
            df["veteran_wisdom_vs_youth"] = (
                (df["r_total_fights"] * df["r_age_at_event"] * df["r_pro_str_def_corrected"]) / 
                (df["r_pro_SApM_corrected"] + 0.1)
            ) * (df["r_championship_experience"] / (df["r_wins_corrected"] + 1)) - (
                (df["b_total_fights"] * df["b_age_at_event"] * df["b_pro_str_def_corrected"]) / 
                (df["b_pro_SApM_corrected"] + 0.1)
            ) * (df["b_championship_experience"] / (df["b_wins_corrected"] + 1))
        
        # Add new features to feature_columns list
        new_winner_features = [
            "clinch_control_advantage",
            "ground_control_advantage", 
            "striking_volume_advantage",
            "defensive_advantage",
            "finish_avoidance_advantage",
            "pace_control_advantage",
            "experience_quality_advantage",
            "momentum_consistency_advantage",
            "striking_defense_advantage",
            "grappling_offense_advantage",
            # SOPHISTICATED WINNER FEATURES (12 advanced features)
            "fight_iq_composite_advantage",
            "championship_pressure_performance",
            "momentum_velocity_advantage",
            "style_matchup_depth_score",
            "durability_vs_power_ratio",
            "pace_control_mastery",
            "grappling_transition_mastery",
            "upset_potential_vs_experience",
            "physical_dominance_composite",
            "technical_striking_mastery",
            "clutch_performance_under_pressure",
            "veteran_wisdom_vs_youth"
        ]
        
        new_method_features = [
            "ko_power_advantage",
            "submission_threat_advantage", 
            "decision_tendency_advantage",
            "early_finish_advantage",
            "durability_advantage"
        ]
        
        # Add to feature_columns
        for feature in new_winner_features + new_method_features:
            if feature not in feature_columns:
                feature_columns.append(feature)

        # ============================================================================
        # ULTRA-SOPHISTICATED WINNER PREDICTION FEATURES (10 advanced features)
        # ============================================================================
        
        # 1. Championship Clutch Factor Composite
        if all(col in df.columns for col in ["r_championship_experience", "b_championship_experience", "r_clutch_factor", "b_clutch_factor", "r_recent_form_corrected", "b_recent_form_corrected", "r_days_since_last_fight_diff_corrected", "b_days_since_last_fight_diff_corrected", "r_win_streak_corrected", "b_win_streak_corrected", "r_loss_streak_corrected", "b_loss_streak_corrected"]):
            df["championship_clutch_composite"] = (
                (df["r_championship_experience"] * df["r_clutch_factor"] * df["r_recent_form_corrected"]) / 
                (df["r_days_since_last_fight_diff_corrected"] + 30)
            ) * (df["r_win_streak_corrected"] / (abs(df["r_loss_streak_corrected"]) + 1)) - (
                (df["b_championship_experience"] * df["b_clutch_factor"] * df["b_recent_form_corrected"]) / 
                (df["b_days_since_last_fight_diff_corrected"] + 30)
            ) * (df["b_win_streak_corrected"] / (abs(df["b_loss_streak_corrected"]) + 1))
        
        # 2. Advanced Style Clash Resolution Matrix
        if all(col in df.columns for col in ["striker_vs_grappler", "power_vs_technique", "stance_matchup_advantage", "style_clash_severity", "experience_gap", "r_total_fights", "b_total_fights"]):
            df["style_clash_resolution_matrix"] = (
                (df["striker_vs_grappler"] * df["power_vs_technique"] * df["stance_matchup_advantage"]) / 
                (df["style_clash_severity"] + 0.1)
            ) * (df["experience_gap"] / 10) * (df["r_total_fights"] / (df["b_total_fights"] + 1))
        
        # 3. Momentum Velocity with Pressure Response
        if all(col in df.columns for col in ["r_win_streak_diff_corrected", "b_win_streak_diff_corrected", "r_recent_form_diff_corrected", "b_recent_form_diff_corrected", "r_finish_rate_diff", "b_finish_rate_diff", "r_loss_streak_diff_corrected", "b_loss_streak_diff_corrected", "r_championship_pressure", "b_championship_pressure", "r_days_since_last_fight_diff_corrected", "b_days_since_last_fight_diff_corrected"]):
            df["momentum_pressure_velocity"] = (
                (df["r_win_streak_diff_corrected"] * df["r_recent_form_diff_corrected"] * df["r_finish_rate_diff"]) / 
                (abs(df["r_loss_streak_diff_corrected"]) + 1)
            ) * (df["r_championship_pressure"] / (df["r_days_since_last_fight_diff_corrected"] + 30)) - (
                (df["b_win_streak_diff_corrected"] * df["b_recent_form_diff_corrected"] * df["b_finish_rate_diff"]) / 
                (abs(df["b_loss_streak_diff_corrected"]) + 1)
            ) * (df["b_championship_pressure"] / (df["b_days_since_last_fight_diff_corrected"] + 30))
        
        # 4. Technical Mastery vs Physical Dominance Ratio
        if all(col in df.columns for col in ["r_pro_sig_str_acc_corrected", "b_pro_sig_str_acc_corrected", "r_pro_str_def_corrected", "b_pro_str_def_corrected", "r_pro_SLpM_corrected", "b_pro_SLpM_corrected", "r_pro_SApM_corrected", "b_pro_SApM_corrected", "r_head_pct_corrected", "b_head_pct_corrected", "r_body_pct_corrected", "b_body_pct_corrected", "r_leg_pct_corrected", "b_leg_pct_corrected", "r_height", "b_height", "r_reach", "b_reach", "r_weight", "b_weight", "r_ape_index", "b_ape_index"]):
            df["technical_vs_physical_dominance"] = (
                (df["r_pro_sig_str_acc_corrected"] * df["r_pro_str_def_corrected"] * df["r_pro_SLpM_corrected"]) / 
                (df["r_pro_SApM_corrected"] + 0.1)
            ) * (df["r_head_pct_corrected"] / (df["r_body_pct_corrected"] + df["r_leg_pct_corrected"] + 0.1)) / (
                ((df["r_height"] * df["r_reach"] * df["r_weight"]) / 1000) * (df["r_ape_index"] / 10)
            ) - (
                (df["b_pro_sig_str_acc_corrected"] * df["b_pro_str_def_corrected"] * df["b_pro_SLpM_corrected"]) / 
                (df["b_pro_SApM_corrected"] + 0.1)
            ) * (df["b_head_pct_corrected"] / (df["b_body_pct_corrected"] + df["b_leg_pct_corrected"] + 0.1)) / (
                ((df["b_height"] * df["b_reach"] * df["b_weight"]) / 1000) * (df["b_ape_index"] / 10)
            )
        
        # 5. Advanced Grappling Control Matrix
        if all(col in df.columns for col in ["r_pro_td_avg_corrected", "b_pro_td_avg_corrected", "r_pro_td_acc_corrected", "b_pro_td_acc_corrected", "r_pro_sub_avg_corrected", "b_pro_sub_avg_corrected", "r_pro_td_def_corrected", "b_pro_td_def_corrected", "r_ground_pct_corrected", "b_ground_pct_corrected", "r_clinch_pct_corrected", "b_clinch_pct_corrected", "r_pro_ctrl_sec_corrected", "b_pro_ctrl_sec_corrected", "r_avg_fight_time", "b_avg_fight_time"]):
            df["grappling_control_matrix"] = (
                (df["r_pro_td_avg_corrected"] * df["r_pro_td_acc_corrected"] * df["r_pro_sub_avg_corrected"]) / 
                (df["r_pro_td_def_corrected"] + 0.1)
            ) * (df["r_ground_pct_corrected"] / (df["r_clinch_pct_corrected"] + 0.1)) * (
                df["r_pro_ctrl_sec_corrected"] / (df["r_avg_fight_time"] + 1)
            ) - (
                (df["b_pro_td_avg_corrected"] * df["b_pro_td_acc_corrected"] * df["b_pro_sub_avg_corrected"]) / 
                (df["b_pro_td_def_corrected"] + 0.1)
            ) * (df["b_ground_pct_corrected"] / (df["b_clinch_pct_corrected"] + 0.1)) * (
                df["b_pro_ctrl_sec_corrected"] / (df["b_avg_fight_time"] + 1)
            )
        
        # 6. Upset Potential vs Experience Matrix
        if all(col in df.columns for col in ["r_upset_history", "b_upset_history", "r_recent_opponent_quality_diff", "b_recent_opponent_quality_diff", "r_total_fights", "b_total_fights", "r_age_at_event", "b_age_at_event", "r_recent_form_corrected", "b_recent_form_corrected"]):
            df["upset_experience_matrix"] = (
                (df["r_upset_history"] * df["r_recent_opponent_quality_diff"]) / 
                (df["b_total_fights"] + 1)
            ) * (df["r_age_at_event"] / (df["b_age_at_event"] + 1)) * (
                df["r_recent_form_corrected"] / (df["b_recent_form_corrected"] + 0.1)
            ) - (
                (df["b_upset_history"] * df["b_recent_opponent_quality_diff"]) / 
                (df["r_total_fights"] + 1)
            ) * (df["b_age_at_event"] / (df["r_age_at_event"] + 1)) * (
                df["b_recent_form_corrected"] / (df["r_recent_form_corrected"] + 0.1)
            )
        
        # 7. Advanced Fight IQ Composite
        if all(col in df.columns for col in ["r_pro_str_def_corrected", "b_pro_str_def_corrected", "r_pro_td_def_corrected", "b_pro_td_def_corrected", "r_pro_sig_str_acc_corrected", "b_pro_sig_str_acc_corrected", "r_pro_SApM_corrected", "b_pro_SApM_corrected", "r_total_fights", "b_total_fights", "r_age_at_event", "b_age_at_event"]):
            df["advanced_fight_iq_composite"] = (
                (df["r_pro_str_def_corrected"] * df["r_pro_td_def_corrected"] * df["r_pro_sig_str_acc_corrected"]) / 
                (df["r_pro_SApM_corrected"] + 0.1)
            ) * (df["r_pro_str_def_corrected"] / (df["r_pro_SApM_corrected"] + 0.1)) * (
                df["r_total_fights"] / (df["r_age_at_event"] + 1)
            ) - (
                (df["b_pro_str_def_corrected"] * df["b_pro_td_def_corrected"] * df["b_pro_sig_str_acc_corrected"]) / 
                (df["b_pro_SApM_corrected"] + 0.1)
            ) * (df["b_pro_str_def_corrected"] / (df["b_pro_SApM_corrected"] + 0.1)) * (
                df["b_total_fights"] / (df["b_age_at_event"] + 1)
            )
        
        # 8. Durability vs Power Advanced Ratio
        if all(col in df.columns for col in ["r_durability_score", "b_durability_score", "r_pro_str_def_corrected", "b_pro_str_def_corrected", "r_ko_rate_corrected", "b_ko_rate_corrected", "r_pro_kd_pM_corrected", "b_pro_kd_pM_corrected", "r_avg_fight_time", "b_avg_fight_time"]):
            df["durability_power_advanced_ratio"] = (
                (df["r_durability_score"] * df["r_pro_str_def_corrected"]) / 
                (df["b_ko_rate_corrected"] * df["b_pro_kd_pM_corrected"] + 0.1)
            ) * (df["r_avg_fight_time"] / 5) - (
                (df["b_durability_score"] * df["b_pro_str_def_corrected"]) / 
                (df["r_ko_rate_corrected"] * df["r_pro_kd_pM_corrected"] + 0.1)
            ) * (df["b_avg_fight_time"] / 5)
        
        # 9. Pace Control Mastery with Momentum
        if all(col in df.columns for col in ["r_avg_fight_time", "b_avg_fight_time", "r_pro_SLpM_corrected", "b_pro_SLpM_corrected", "r_pro_SApM_corrected", "b_pro_SApM_corrected", "r_pro_str_def_corrected", "b_pro_str_def_corrected", "r_win_streak_corrected", "b_win_streak_corrected", "r_loss_streak_corrected", "b_loss_streak_corrected"]):
            df["pace_control_momentum_mastery"] = (
                (df["r_avg_fight_time"] / 5) * (df["r_pro_SLpM_corrected"] - df["r_pro_SApM_corrected"]) * 
                (df["r_pro_str_def_corrected"] / (df["r_pro_SApM_corrected"] + 0.1))
            ) * (df["r_win_streak_corrected"] / (abs(df["r_loss_streak_corrected"]) + 1)) - (
                (df["b_avg_fight_time"] / 5) * (df["b_pro_SLpM_corrected"] - df["b_pro_SApM_corrected"]) * 
                (df["b_pro_str_def_corrected"] / (df["b_pro_SApM_corrected"] + 0.1))
            ) * (df["b_win_streak_corrected"] / (abs(df["b_loss_streak_corrected"]) + 1))
        
        # 10. Veteran Wisdom vs Youth Advanced
        if all(col in df.columns for col in ["r_total_fights", "b_total_fights", "r_age_at_event", "b_age_at_event", "r_pro_str_def_corrected", "b_pro_str_def_corrected", "r_pro_SApM_corrected", "b_pro_SApM_corrected", "r_championship_experience", "b_championship_experience", "r_wins_corrected", "b_wins_corrected", "r_recent_form_corrected", "b_recent_form_corrected"]):
            df["veteran_youth_advanced"] = (
                (df["r_total_fights"] * df["r_age_at_event"] * df["r_pro_str_def_corrected"]) / 
                (df["r_pro_SApM_corrected"] + 0.1)
            ) * (df["r_championship_experience"] / (df["r_wins_corrected"] + 1)) * (
                df["r_recent_form_corrected"] / (df["b_recent_form_corrected"] + 0.1)
            ) - (
                (df["b_total_fights"] * df["b_age_at_event"] * df["b_pro_str_def_corrected"]) / 
                (df["b_pro_SApM_corrected"] + 0.1)
            ) * (df["b_championship_experience"] / (df["b_wins_corrected"] + 1)) * (
                df["b_recent_form_corrected"] / (df["r_recent_form_corrected"] + 0.1)
            )
        
        # ============================================================================
        # ULTRA-SOPHISTICATED METHOD PREDICTION FEATURES (5 advanced features)
        # ============================================================================
        
        # 1. Advanced KO Power Composite
        if all(col in df.columns for col in ["r_ko_rate_corrected", "b_ko_rate_corrected", "r_pro_kd_pM_corrected", "b_pro_kd_pM_corrected", "r_pro_SLpM_corrected", "b_pro_SLpM_corrected", "r_pro_SApM_corrected", "b_pro_SApM_corrected", "r_head_pct_corrected", "b_head_pct_corrected", "r_body_pct_corrected", "b_body_pct_corrected", "r_leg_pct_corrected", "b_leg_pct_corrected", "r_win_streak_corrected", "b_win_streak_corrected", "r_loss_streak_corrected", "b_loss_streak_corrected"]):
            df["advanced_ko_power_composite"] = (
                (df["r_ko_rate_corrected"] * df["r_pro_kd_pM_corrected"] * df["r_pro_SLpM_corrected"]) / 
                (df["r_pro_SApM_corrected"] + 0.1)
            ) * (df["r_head_pct_corrected"] / (df["r_body_pct_corrected"] + df["r_leg_pct_corrected"] + 0.1)) * (
                df["r_win_streak_corrected"] / (abs(df["r_loss_streak_corrected"]) + 1)
            ) - (
                (df["b_ko_rate_corrected"] * df["b_pro_kd_pM_corrected"] * df["b_pro_SLpM_corrected"]) / 
                (df["b_pro_SApM_corrected"] + 0.1)
            ) * (df["b_head_pct_corrected"] / (df["b_body_pct_corrected"] + df["b_leg_pct_corrected"] + 0.1)) * (
                df["b_win_streak_corrected"] / (abs(df["b_loss_streak_corrected"]) + 1)
            )
        
        # 2. Submission Threat Advanced Matrix
        if all(col in df.columns for col in ["r_sub_rate_corrected", "b_sub_rate_corrected", "r_pro_sub_avg_corrected", "b_pro_sub_avg_corrected", "r_pro_td_def_corrected", "b_pro_td_def_corrected", "r_ground_pct_corrected", "b_ground_pct_corrected", "r_clinch_pct_corrected", "b_clinch_pct_corrected", "r_pro_td_avg_corrected", "b_pro_td_avg_corrected", "r_pro_td_acc_corrected", "b_pro_td_acc_corrected"]):
            df["submission_threat_advanced_matrix"] = (
                (df["r_sub_rate_corrected"] * df["r_pro_sub_avg_corrected"]) / 
                (df["r_pro_td_def_corrected"] + 0.1)
            ) * (df["r_ground_pct_corrected"] / (df["r_clinch_pct_corrected"] + 0.1)) * (
                df["r_pro_td_avg_corrected"] * df["r_pro_td_acc_corrected"]
            ) - (
                (df["b_sub_rate_corrected"] * df["b_pro_sub_avg_corrected"]) / 
                (df["b_pro_td_def_corrected"] + 0.1)
            ) * (df["b_ground_pct_corrected"] / (df["b_clinch_pct_corrected"] + 0.1)) * (
                df["b_pro_td_avg_corrected"] * df["b_pro_td_acc_corrected"]
            )
        
        # 3. Decision Tendency Advanced Composite
        if all(col in df.columns for col in ["r_dec_rate_corrected", "b_dec_rate_corrected", "r_avg_fight_time", "b_avg_fight_time", "r_ko_rate_corrected", "b_ko_rate_corrected", "r_sub_rate_corrected", "b_sub_rate_corrected", "r_pro_str_def_corrected", "b_pro_str_def_corrected", "r_pro_SApM_corrected", "b_pro_SApM_corrected", "r_durability_score", "b_durability_score", "r_ko_losses_corrected", "b_ko_losses_corrected"]):
            df["decision_tendency_advanced_composite"] = (
                (df["r_dec_rate_corrected"] * df["r_avg_fight_time"]) / 
                (df["r_ko_rate_corrected"] + df["r_sub_rate_corrected"] + 0.1)
            ) * (df["r_pro_str_def_corrected"] / (df["r_pro_SApM_corrected"] + 0.1)) * (
                df["r_durability_score"] / (df["r_ko_losses_corrected"] + 1)
            ) - (
                (df["b_dec_rate_corrected"] * df["b_avg_fight_time"]) / 
                (df["b_ko_rate_corrected"] + df["b_sub_rate_corrected"] + 0.1)
            ) * (df["b_pro_str_def_corrected"] / (df["b_pro_SApM_corrected"] + 0.1)) * (
                df["b_durability_score"] / (df["b_ko_losses_corrected"] + 1)
            )
        
        # 4. Early Finish Specialist Advantage
        if all(col in df.columns for col in ["r_ko_rate_corrected", "b_ko_rate_corrected", "r_sub_rate_corrected", "b_sub_rate_corrected", "r_avg_fight_time", "b_avg_fight_time", "r_pro_kd_pM_corrected", "b_pro_kd_pM_corrected", "r_pro_sub_avg_corrected", "b_pro_sub_avg_corrected", "r_recent_form_corrected", "b_recent_form_corrected", "r_days_since_last_fight_diff_corrected", "b_days_since_last_fight_diff_corrected"]):
            df["early_finish_specialist_advantage"] = (
                (df["r_ko_rate_corrected"] + df["r_sub_rate_corrected"]) / 
                (df["r_avg_fight_time"] + 1)
            ) * (df["r_pro_kd_pM_corrected"] * df["r_pro_sub_avg_corrected"]) * (
                df["r_recent_form_corrected"] / (df["r_days_since_last_fight_diff_corrected"] + 30)
            ) - (
                (df["b_ko_rate_corrected"] + df["b_sub_rate_corrected"]) / 
                (df["b_avg_fight_time"] + 1)
            ) * (df["b_pro_kd_pM_corrected"] * df["b_pro_sub_avg_corrected"]) * (
                df["b_recent_form_corrected"] / (df["b_days_since_last_fight_diff_corrected"] + 30)
            )
        
        # 5. Method Versatility vs Specialization
        if all(col in df.columns for col in ["r_ko_rate_corrected", "b_ko_rate_corrected", "r_sub_rate_corrected", "b_sub_rate_corrected", "r_dec_rate_corrected", "b_dec_rate_corrected", "r_pro_kd_pM_corrected", "b_pro_kd_pM_corrected", "r_pro_sub_avg_corrected", "b_pro_sub_avg_corrected", "r_total_fights", "b_total_fights", "r_wins_corrected", "b_wins_corrected"]):
            df["method_versatility_vs_specialization"] = (
                (df["r_ko_rate_corrected"] * df["r_sub_rate_corrected"] * df["r_dec_rate_corrected"]) / 
                (df["r_pro_kd_pM_corrected"] * df["r_pro_sub_avg_corrected"] + 0.1)
            ) * (df["r_total_fights"] / (df["r_wins_corrected"] + 1)) - (
                (df["b_ko_rate_corrected"] * df["b_sub_rate_corrected"] * df["b_dec_rate_corrected"]) / 
                (df["b_pro_kd_pM_corrected"] * df["b_pro_sub_avg_corrected"] + 0.1)
            ) * (df["b_total_fights"] / (df["b_wins_corrected"] + 1))
        
        # Add new ultra-sophisticated features to feature_columns
        ultra_sophisticated_winner_features = [
            "championship_clutch_composite",
            "style_clash_resolution_matrix",
            "momentum_pressure_velocity",
            "technical_vs_physical_dominance",
            "grappling_control_matrix",
            "upset_experience_matrix",
            "advanced_fight_iq_composite",
            "durability_power_advanced_ratio",
            "pace_control_momentum_mastery",
            "veteran_youth_advanced"
        ]
        
        ultra_sophisticated_method_features = [
            "advanced_ko_power_composite",
            "submission_threat_advanced_matrix",
            "decision_tendency_advanced_composite",
            "early_finish_specialist_advantage",
            "method_versatility_vs_specialization"
        ]
        
        # Add to feature_columns
        for feature in ultra_sophisticated_winner_features + ultra_sophisticated_method_features:
            if feature not in feature_columns:
                feature_columns.append(feature)

        # Filter feature columns to only include those that exist and are unique
        available_features = []
        for col in feature_columns:
            if col in df.columns and col not in available_features:
                available_features.append(col)
                df[col] = df[col].fillna(0)

        df = df.replace([np.inf, -np.inf], [1e6, -1e6])

        print(
            f"Total features: {len(available_features)} (filtered from {len(feature_columns)} requested)"
        )

        # FEATURE CACHING: Store computed features for future use
        self.feature_cache[cache_key] = (df, available_features)
        print("Features cached for future use...")

        return df, available_features

    def train_models(self, df):
        """Train stacked ensemble with specialized method models and deep learning"""
        # Ensure consistent random state before training
        self.set_random_seeds()

        # Augment data with corner swapping to eliminate red corner bias
        df = self.augment_data_with_corner_swapping(df)

        df, feature_columns = self.prepare_features(df)

        X = df[feature_columns]

        # Use standard target encoding (class weights will handle bias)
        if "winner" in df.columns:
            y_winner = (df["winner"] == "Red").astype(int)
        else:
            # If no winner column, create default labels
            y_winner = np.zeros(len(df), dtype=int)

        if "winner_method_encoder" not in self.label_encoders:
            self.label_encoders["winner_method_encoder"] = LabelEncoder()

        # Check if winner_method_simple column exists
        if "winner_method_simple" in df.columns:
            y_method = self.label_encoders["winner_method_encoder"].fit_transform(
                df["winner_method_simple"]
            )
        else:
            # If no winner_method_simple column, create default method labels
            y_method = np.zeros(len(df), dtype=int)

        # ===== ENHANCED VALIDATION STRATEGY =====
        print("\n📊 IMPLEMENTING ENHANCED VALIDATION STRATEGY...")
        
        # Use TimeSeriesSplit for temporal validation
        tscv = TimeSeriesSplit(n_splits=5)
        
        # For final train/test split, use stratified split but with time awareness
        try:
            if 'event_date' in df.columns:
                # Sort by event date for temporal split
                df_sorted = df.sort_values('event_date')
                # Reset index to avoid index issues
                df_sorted = df_sorted.reset_index(drop=True)
                X_sorted = X.loc[df_sorted.index].reset_index(drop=True)
                y_winner_sorted = y_winner[df_sorted.index]
                y_method_sorted = y_method[df_sorted.index]
                
                # Use 80% for training, 20% for testing (temporal split)
                split_idx = int(0.8 * len(X_sorted))
                X_train = X_sorted.iloc[:split_idx]
                X_test = X_sorted.iloc[split_idx:]
                y_winner_train = y_winner_sorted[:split_idx]
                y_winner_test = y_winner_sorted[split_idx:]
                y_method_train = y_method_sorted[:split_idx]
                y_method_test = y_method_sorted[split_idx:]
            else:
                # Fallback to stratified split if no date column
                (
                    X_train,
                    X_test,
                    y_winner_train,
                    y_winner_test,
                    y_method_train,
                    y_method_test,
                ) = train_test_split(
                    X, y_winner, y_method, test_size=0.2, random_state=42, stratify=y_winner
                )
        except (KeyError, IndexError) as e:
            print(f"Warning: Error in temporal split, using standard split: {e}")
            # Fallback to standard stratified split
            (
                X_train,
                X_test,
                y_winner_train,
                y_winner_test,
                y_method_train,
                y_method_test,
            ) = train_test_split(
                X, y_winner, y_method, test_size=0.2, random_state=42, stratify=y_winner
            )

        print("\n" + "=" * 80)
        print("TRAINING ADVANCED STACKED ENSEMBLE MODEL")
        print("=" * 80)

        numeric_transformer = Pipeline(
            [
                ("imputer", SimpleImputer(strategy="median")),
                ("scaler", StandardScaler()),
            ]
        )

        preprocessor = ColumnTransformer(
            [("num", numeric_transformer, feature_columns)]
        )

        # ===== HYPERPARAMETER OPTIMIZATION =====
        print("\n🔧 OPTIMIZING HYPERPARAMETERS...")
        
        # Optimize XGBoost parameters
        if HAS_XGBOOST:
            print("Optimizing XGBoost hyperparameters...")
            xgb_optimized = self.optimize_hyperparameters(X_train, y_winner_train, 'xgb', n_trials=20)
            if not xgb_optimized:
                xgb_optimized = {}  # Use defaults if optimization fails
        else:
            xgb_optimized = {}
            
        # Optimize LightGBM parameters
        if HAS_LIGHTGBM:
            print("Optimizing LightGBM hyperparameters...")
            lgbm_optimized = self.optimize_hyperparameters(X_train, y_winner_train, 'lgbm', n_trials=20)
            if not lgbm_optimized:
                lgbm_optimized = {}
        else:
            lgbm_optimized = {}
            
        # Optimize CatBoost parameters
        if HAS_CATBOOST:
            print("Optimizing CatBoost hyperparameters...")
            catboost_optimized = self.optimize_hyperparameters(X_train, y_winner_train, 'catboost', n_trials=20)
            if not catboost_optimized:
                catboost_optimized = {}
        else:
            catboost_optimized = {}

        # Build base models for stacking with optimized parameters
        # GPU Usage Strategy:
        # Sequential execution: XGBoost + LightGBM (GPU) + CatBoost + RandomForest + MLP (CPU)
        base_models = []

        if HAS_XGBOOST:
            print("\n✓ XGBoost available (GPU - parallel with LightGBM)")
            # Use balanced weights since data augmentation eliminates bias
            scale_pos = 1.0  # Balanced since we have equal red/blue representation after augmentation

            # Create XGBoost classifier with optimized parameters
            xgb_params = {
                "n_estimators": xgb_optimized.get("n_estimators", 1000),
                "max_depth": xgb_optimized.get("max_depth", 10),
                "learning_rate": xgb_optimized.get("learning_rate", 0.01),
                "subsample": xgb_optimized.get("subsample", 0.9),
                "colsample_bytree": xgb_optimized.get("colsample_bytree", 0.9),
                "colsample_bylevel": xgb_optimized.get("colsample_bylevel", 0.9),
                "n_jobs": -1,
                "reg_alpha": xgb_optimized.get("reg_alpha", 0.05),
                "reg_lambda": xgb_optimized.get("reg_lambda", 0.5),
                "min_child_weight": xgb_optimized.get("min_child_weight", 2),
                "gamma": xgb_optimized.get("gamma", 0.05),
                "scale_pos_weight": scale_pos,
                "random_state": 42,
                "eval_metric": "logloss",
                "tree_method": "hist",
                "device": "cpu",  # CPU for deterministic results
                "seed": 42,
                "enable_categorical": True,
                "max_delta_step": 1,
            }

            xgb_classifier = XGBClassifier(**xgb_params)

            xgb_model = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectPercentile(
                            f_classif, percentile=75
                        ),  # 60% -> 75% (better accuracy)
                    ),  # Match Class Weighting
                    ("classifier", xgb_classifier),
                ]
            )
            base_models.append(("xgb", xgb_model))

        if HAS_LIGHTGBM:
            print("✓ LightGBM available (GPU - parallel with XGBoost)")
            lgbm_model = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectPercentile(
                            f_classif, percentile=75
                        ),  # 60% -> 75% (better accuracy)
                    ),  # Match Class Weighting
                    (
                        "classifier",
                        LGBMClassifier(
                            n_estimators=lgbm_optimized.get("n_estimators", 400),
                            max_depth=lgbm_optimized.get("max_depth", 7),
                            learning_rate=lgbm_optimized.get("learning_rate", 0.03),
                            num_leaves=lgbm_optimized.get("num_leaves", 40),
                            subsample=lgbm_optimized.get("subsample", 0.8),
                            colsample_bytree=lgbm_optimized.get("colsample_bytree", 0.8),
                            reg_alpha=lgbm_optimized.get("reg_alpha", 0.1),
                            reg_lambda=lgbm_optimized.get("reg_lambda", 0.8),
                            min_child_weight=lgbm_optimized.get("min_child_weight", 3),
                            device="cpu",  # CPU for deterministic results
                            random_state=42,
                            verbose=-1,
                        ),
                    ),
                ]
            )
            base_models.append(("lgbm", lgbm_model))

        # CPU Models: CatBoost, RandomForest, MLP (parallel with each other and GPU models)
        if HAS_CATBOOST:
            print("✓ CatBoost available (CPU - parallel with other models)")
            catboost_model = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectPercentile(
                            f_classif, percentile=75
                        ),  # 60% -> 75% (better accuracy)
                    ),  # Match Class Weighting
                    (
                        "classifier",
                        CatBoostClassifier(
                            iterations=catboost_optimized.get("iterations", 400),
                            depth=catboost_optimized.get("depth", 7),
                            learning_rate=catboost_optimized.get("learning_rate", 0.03),
                            l2_leaf_reg=catboost_optimized.get("l2_leaf_reg", 0.8),
                            task_type="CPU",  # Use CPU to avoid persistent GPU conflicts
                            random_state=42,
                            verbose=0,
                        ),
                    ),
                ]
            )
            base_models.append(("catboost", catboost_model))

        # Random Forest (always available)
        print("✓ Random Forest available")
        rf_model = Pipeline(
            [
                ("preprocessor", preprocessor),
                (
                    "feature_selector",
                    SelectPercentile(f_classif, percentile=75),
                ),  # Match Class Weighting
                (
                    "classifier",
                    RandomForestClassifier(
                        n_estimators=600,  # More trees for better accuracy
                        max_depth=15,  # Optimal depth for winner prediction
                        min_samples_split=6,  # Balanced for accuracy
                        min_samples_leaf=2,  # Balanced for accuracy
                        random_state=42,
                        n_jobs=-1,
                        class_weight="balanced",
                    ),
                ),
            ]
        )
        base_models.append(("rf", rf_model))

        # Neural Network
        if self.use_neural_net:
            print("✓ Neural Network enabled")
            nn_model = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectPercentile(
                            f_classif, percentile=75
                        ),  # 60% -> 75% (better accuracy)
                    ),  # Match Class Weighting
                    (
                        "classifier",
                        MLPClassifier(
                            hidden_layer_sizes=(
                                256,
                                128,
                                64,
                            ),  # Deeper network for better accuracy
                            activation="relu",
                            solver="adam",
                            alpha=0.0005,  # Lower regularization for better accuracy
                            batch_size=32,  # Smaller batch size for better learning
                            learning_rate="adaptive",
                            max_iter=400,  # Balanced iterations
                            early_stopping=True,
                            random_state=42,
                        ),
                    ),
                ]
            )
            base_models.append(("nn", nn_model))

        print(f"\nBuilding stacked ensemble with {len(base_models)} base models...")

        # Enhanced meta-learner with multiple algorithms
        meta_learners = []

        # XGBoost meta-learner
        if HAS_XGBOOST:
            meta_learners.append(
                (
                    "xgb_meta",
                    XGBClassifier(
                        n_estimators=200,
                        max_depth=4,
                        learning_rate=0.05,
                        subsample=0.9,
                        colsample_bytree=0.9,
                        n_jobs=-1,  # Use all CPU cores for faster training (causes multiple windows in .exe)
                        reg_alpha=0.2,
                        reg_lambda=1,
                        random_state=42,
                        eval_metric="logloss",
                        tree_method="hist",  # Use histogram method for consistency
                        seed=42,  # Additional XGBoost seed
                    ),
                )
            )

        # LightGBM meta-learner
        if HAS_LIGHTGBM:
            meta_learners.append(
                (
                    "lgbm_meta",
                    LGBMClassifier(
                        n_estimators=200,
                        max_depth=4,
                        learning_rate=0.05,
                        num_leaves=20,
                        subsample=0.9,
                        colsample_bytree=0.9,
                        reg_alpha=0.2,
                        reg_lambda=1,
                        random_state=42,
                        verbose=-1,
                    ),
                )
            )

        # Neural Network meta-learner
        meta_learners.append(
            (
                "nn_meta",
                MLPClassifier(
                    hidden_layer_sizes=(64, 32),
                    activation="relu",
                    solver="adam",
                    alpha=0.001,
                    batch_size=64,
                    learning_rate="adaptive",
                    max_iter=300,
                    early_stopping=True,
                    random_state=42,
                ),
            )
        )

        # Logistic Regression meta-learner
        meta_learners.append(
            ("lr_meta", LogisticRegression(C=0.5, max_iter=1000, random_state=42))
        )

        # Create optimized voting ensemble of meta-learners with weights
        meta_weights = [1.2, 1.0, 0.9] if len(meta_learners) == 3 else None
        voting_meta = VotingClassifier(
            estimators=meta_learners, voting="soft", weights=meta_weights
        )

        # MULTI-LEVEL STACKING: Level 1 -> Level 2 -> Final Blender
        if self.use_ensemble and len(base_models) > 1:
            print("\n🔧 Building multi-level stacking ensemble...")

            # Level 1: Base models
            level1_models = base_models

            # Level 2: Meta-learners (current approach)
            level2_meta = voting_meta

            # Level 3: Final blender with more sophisticated approach
            final_blender = VotingClassifier(
                estimators=[
                    (
                        "xgb_final",
                        XGBClassifier(
                            n_estimators=200,
                            max_depth=6,
                            learning_rate=0.1,
                            subsample=0.8,
                            colsample_bytree=0.8,
                            reg_alpha=0.1,
                            reg_lambda=1.0,
                            random_state=42,
                            eval_metric="logloss",
                        ),
                    ),
                    (
                        "lgbm_final",
                        LGBMClassifier(
                            n_estimators=200,
                            max_depth=6,
                            learning_rate=0.1,
                            num_leaves=31,
                            subsample=0.8,
                            colsample_bytree=0.8,
                            reg_alpha=0.1,
                            reg_lambda=1.0,
                            random_state=42,
                            verbose=-1,
                        ),
                    ),
                    (
                        "rf_final",
                        RandomForestClassifier(
                            n_estimators=300,
                            max_depth=10,
                            min_samples_split=4,
                            min_samples_leaf=2,
                            random_state=42,
                            n_jobs=-1,
                        ),
                    ),
                    (
                        "lr_final",
                        LogisticRegression(C=1.0, max_iter=1000, random_state=42),
                    ),
                ],
                voting="soft",
                weights=[
                    0.4,
                    0.3,
                    0.2,
                    0.1,
                ],  # XGBoost, LightGBM, RandomForest, LogisticRegression
            )

            # Create the multi-level stacking
            self.winner_model = StackingClassifier(
                estimators=level1_models,
                final_estimator=StackingClassifier(
                    estimators=[("level2", level2_meta)],
                    final_estimator=final_blender,
                    cv=3,  # Fewer folds for final level
                    n_jobs=-1,
                    stack_method="predict_proba",
                    passthrough=False,
                ),
                cv=5,  # 10 -> 5 folds (faster training)
                n_jobs=-1,
                stack_method="predict_proba",
                passthrough=True,  # Include original features in first level
            )

            # Enhanced calibration for the entire stack
            self.winner_model = CalibratedClassifierCV(
                self.winner_model,
                method="isotonic",
                cv=5,  # 10 -> 5 folds (faster training)
            )
        else:
            # Fallback to single best model
            self.winner_model = base_models[0][1]
            self.winner_model = CalibratedClassifierCV(
                self.winner_model,
                method="isotonic",
                cv=3,  # 5 -> 3 folds (1.7x faster)
            )

        # Ensure we have proper DataFrame format for scikit-learn
        if not isinstance(X_train, pd.DataFrame):
            X_train = pd.DataFrame(X_train, columns=feature_columns)
        if not isinstance(y_winner_train, (pd.Series, np.ndarray)):
            y_winner_train = np.array(y_winner_train)

        print("\nTraining winner prediction model...")
        self.winner_model.fit(X_train, y_winner_train)

        # ============================================================================
        # TRAIN METHOD PREDICTION MODEL
        # ============================================================================
        print("\n" + "=" * 80)
        print("TRAINING METHOD PREDICTION MODEL")
        print("=" * 80)

        print("\nTraining enhanced method prediction ensemble...")

        # Create multiple method prediction models
        method_models = []

        # XGBoost method model
        if HAS_XGBOOST:
            xgb_method = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectPercentile(
                            f_classif, percentile=75
                        ),  # 60% -> 75% (better accuracy)
                    ),  # Match Class Weighting
                    (
                        "classifier",
                        XGBClassifier(
                            n_estimators=600,  # 500 -> 600 (better accuracy)
                            max_depth=9,  # 8 -> 9 (better accuracy)
                            learning_rate=0.02,  # 0.025 -> 0.02 (better convergence)
                            subsample=0.85,  # 0.8 -> 0.85 (better accuracy)
                            colsample_bytree=0.85,  # 0.8 -> 0.85 (better accuracy)
                            n_jobs=-1,
                            reg_alpha=0.15,
                            reg_lambda=0.8,
                            random_state=42,
                            objective="multi:softprob",
                            tree_method="hist",
                            seed=42,
                            # early_stopping_rounds=50,  # Removed - requires validation set
                        ),
                    ),
                ]
            )
            method_models.append(("xgb_method", xgb_method))

        # LightGBM method model
        if HAS_LIGHTGBM:
            lgbm_method = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectPercentile(
                            f_classif, percentile=75
                        ),  # 60% -> 75% (better accuracy)
                    ),  # Match Class Weighting
                    (
                        "classifier",
                        LGBMClassifier(
                            n_estimators=400,  # 600 -> 400 (faster training)
                            max_depth=7,  # 9 -> 7 (faster training)
                            learning_rate=0.03,  # 0.02 -> 0.03 (faster convergence)
                            num_leaves=40,  # 60 -> 40 (faster training)
                            subsample=0.8,  # 0.85 -> 0.8 (faster training)
                            colsample_bytree=0.8,  # 0.85 -> 0.8 (faster training)
                            reg_alpha=0.15,
                            reg_lambda=0.8,
                            device="cpu",  # CPU for deterministic results
                            random_state=42,
                            verbose=-1,
                        ),
                    ),
                ]
            )
            method_models.append(("lgbm_method", lgbm_method))

        # Random Forest method model
        rf_method = Pipeline(
            [
                ("preprocessor", preprocessor),
                (
                    "feature_selector",
                    SelectPercentile(f_classif, percentile=75),
                ),  # Match Class Weighting
                (
                    "classifier",
                    RandomForestClassifier(
                        n_estimators=600,  # 500 -> 600 (better accuracy)
                        max_depth=18,  # 15 -> 18 (better accuracy)
                        min_samples_split=5,  # 6 -> 5 (better accuracy)
                        min_samples_leaf=2,
                        random_state=42,
                        n_jobs=-1,
                        class_weight="balanced",
                    ),
                ),
            ]
        )
        method_models.append(("rf_method", rf_method))

        # Neural Network method model
        nn_method = Pipeline(
            [
                ("preprocessor", preprocessor),
                (
                    "feature_selector",
                    SelectPercentile(f_classif, percentile=75),
                ),  # Match Class Weighting
                (
                    "classifier",
                    MLPClassifier(
                        hidden_layer_sizes=(
                            128,
                            64,
                            32,
                        ),  # (128, 64) -> (128, 64, 32) (better accuracy)
                        activation="relu",
                        solver="adam",
                        alpha=0.0005,
                        batch_size=32,
                        learning_rate="adaptive",
                        max_iter=400,  # 300 -> 400 (better accuracy)
                        early_stopping=True,
                        random_state=42,
                    ),
                ),
            ]
        )
        method_models.append(("nn_method", nn_method))

        # Create enhanced voting ensemble for method prediction with data augmentation
        # Use optimized weighted voting based on model performance characteristics
        if len(method_models) == 4:
            # Optimized weights based on data augmentation performance
            method_weights = [
                1.3,
                1.2,
                1.0,
                0.8,
            ]  # XGBoost, LightGBM, RandomForest, Neural Network
        elif len(method_models) == 3:
            method_weights = [1.3, 1.2, 1.0]  # XGBoost, LightGBM, RandomForest
        else:
            method_weights = None

        self.method_model = VotingClassifier(
            estimators=method_models, voting="soft", weights=method_weights
        )

        # Calibrate the method model
        self.method_model = CalibratedClassifierCV(
            self.method_model,
            method="isotonic",
            cv=3,  # 5 -> 3 folds (faster training)
        )

        self.method_model.fit(X_train, y_method_train)
        print("Enhanced method prediction ensemble training complete")

        # ============================================================================
        # TRAIN DEEP LEARNING MODEL (if enabled)
        # ============================================================================
        if self.use_deep_learning and HAS_TENSORFLOW:
            print("\n" + "=" * 80)
            print("TRAINING DEEP LEARNING MODEL WITH FIGHTER EMBEDDINGS")
            print("=" * 80)

            # Create fighter encoder
            all_fighters = pd.concat([df["r_fighter"], df["b_fighter"]]).unique()

            self.fighter_encoder = LabelEncoder()
            self.fighter_encoder.fit(all_fighters)
            self.num_fighters = len(all_fighters)

            # Encode fighters
            df["r_fighter_encoded"] = self.fighter_encoder.transform(df["r_fighter"])
            df["b_fighter_encoded"] = self.fighter_encoder.transform(df["b_fighter"])

            # Prepare data for deep learning with safe access
            def safe_get_values(data, index, column):
                try:
                    # Check if it's a pandas DataFrame
                    if hasattr(data, "loc") and hasattr(data, "columns"):
                        if column in data.columns:
                            return data.loc[index, column].values
                        else:
                            return np.zeros(len(index))
                    # Check if it's a pandas Series
                    elif hasattr(data, "values") and hasattr(data, "index"):
                        if column in data.index:
                            return data[column].values
                        else:
                            return np.zeros(len(index))
                    # Fallback for other data types
                    else:
                        return np.zeros(len(index))
                except (IndexError, KeyError, AttributeError):
                    return np.zeros(len(index))

            X_dl_train = [
                safe_get_values(df, X_train.index, "r_fighter_encoded"),
                safe_get_values(df, X_train.index, "b_fighter_encoded"),
                preprocessor.fit_transform(X_train),
            ]

            X_dl_test = [
                safe_get_values(df, X_test.index, "r_fighter_encoded"),
                safe_get_values(df, X_test.index, "b_fighter_encoded"),
                preprocessor.transform(X_test),
            ]

            # Map method labels to 6 classes: Red_KO/TKO, Red_Submission, Red_Decision, Blue_KO/TKO, Blue_Submission, Blue_Decision
            method_map = {}
            for i, label in enumerate(
                self.label_encoders["winner_method_encoder"].classes_
            ):
                winner = label.split("_")[0]
                method_type = label.split("_")[1]

                if winner == "Red":
                    if method_type == "KO/TKO":
                        method_map[i] = 0
                    elif method_type == "Submission":
                        method_map[i] = 1
                    else:  # Decision
                        method_map[i] = 2
                else:  # Blue
                    if method_type == "KO/TKO":
                        method_map[i] = 3
                    elif method_type == "Submission":
                        method_map[i] = 4
                    else:  # Decision
                        method_map[i] = 5

            # Convert to categorical for deep learning
            try:
                from tensorflow.keras.utils import to_categorical
            except ImportError:
                from keras.utils import to_categorical

            y_method_train_dl = to_categorical(
                [method_map[y] for y in y_method_train], num_classes=6
            )
            y_method_test_dl = to_categorical(
                [method_map[y] for y in y_method_test], num_classes=6
            )

            # Build and train deep learning model
            self.deep_learning_model = self.build_deep_learning_model(
                num_features=preprocessor.transform(X_train).shape[1],
                num_fighters=self.num_fighters,
            )

            print("\nTraining enhanced deep learning model...")

            # Enhanced callbacks for better training
            callbacks = [
                keras.callbacks.EarlyStopping(
                    monitor="val_loss",
                    patience=15,
                    restore_best_weights=True,
                    min_delta=0.001,
                ),
                keras.callbacks.ModelCheckpoint(
                    "best_dl_model.h5",
                    monitor="val_loss",
                    save_best_only=True,
                    verbose=0,
                ),
            ]

            history = self.deep_learning_model.fit(
                X_dl_train,
                {
                    "winner": y_winner_train.values
                    if hasattr(y_winner_train, "values")
                    else y_winner_train,
                    "method": y_method_train_dl,
                },
                validation_split=0.2,
                epochs=100,  # Increased epochs
                batch_size=64,  # Increased batch size
                verbose=0,
                callbacks=callbacks,
            )

            dl_results = self.deep_learning_model.evaluate(
                X_dl_test,
                {
                    "winner": y_winner_test.values
                    if hasattr(y_winner_test, "values")
                    else y_winner_test,
                    "method": y_method_test_dl,
                },
                verbose=0,
            )

            print(f"\nDeep Learning Results:")
            print(f"  Winner Accuracy: {dl_results[3]:.4f}")
            print(f"  Method Accuracy: {dl_results[4]:.4f}")

        # Enhanced Time-based cross-validation with parallel processing
        tscv = TimeSeriesSplit(n_splits=5)  # Match Class Weighting

        def train_fold(fold_data):
            """Train a single fold - designed for parallel execution"""
            fold, train_idx, val_idx, X, y_winner, preprocessor, feature_columns = (
                fold_data
            )

            # Safe indexing for both pandas and numpy arrays
            if hasattr(X, "iloc"):
                X_fold_train, X_fold_val = X.iloc[train_idx], X.iloc[val_idx]
            else:
                X_fold_train, X_fold_val = X[train_idx], X[val_idx]

            if hasattr(y_winner, "iloc"):
                y_fold_train, y_fold_val = (
                    y_winner.iloc[train_idx],
                    y_winner.iloc[val_idx],
                )
            else:
                y_fold_train, y_fold_val = y_winner[train_idx], y_winner[val_idx]

            # Ensure proper DataFrame format for cross-validation
            if not isinstance(X_fold_train, pd.DataFrame):
                X_fold_train = pd.DataFrame(X_fold_train, columns=feature_columns)
            if not isinstance(y_fold_train, (pd.Series, np.ndarray)):
                y_fold_train = np.array(y_fold_train)

            # Use the same enhanced model as in training
            if HAS_XGBOOST:
                # Create XGBoost parameters for cross-validation
                fold_xgb_params = {
                    "n_estimators": 800,
                    "max_depth": 10,
                    "learning_rate": 0.015,
                    "subsample": 0.85,
                    "colsample_bytree": 0.85,
                    "n_jobs": -1,
                    "reg_alpha": 0.1,
                    "reg_lambda": 0.8,
                    "random_state": 42,
                    "eval_metric": "logloss",
                    "tree_method": "hist",
                    "device": "cpu",  # CPU for deterministic results
                    "seed": 42,
                }

                fold_classifier = XGBClassifier(**fold_xgb_params)

                fold_model = Pipeline(
                    [
                        ("preprocessor", preprocessor),
                        (
                            "feature_selector",
                            SelectPercentile(
                                f_classif, percentile=75
                            ),  # 60% -> 75% (better accuracy)
                        ),  # Match Class Weighting
                        ("classifier", fold_classifier),
                    ]
                )
            else:
                fold_model = Pipeline(
                    [
                        ("preprocessor", preprocessor),
                        (
                            "feature_selector",
                            SelectPercentile(
                                f_classif, percentile=75
                            ),  # 60% -> 75% (better accuracy)
                        ),  # Match Class Weighting
                        (
                            "classifier",
                            RandomForestClassifier(
                                n_estimators=600,
                                max_depth=20,
                                min_samples_split=8,  # Enhanced parameters
                                min_samples_leaf=2,
                                random_state=42,
                                n_jobs=-1,
                                class_weight="balanced",  # Balanced since data augmentation eliminates bias
                            ),
                        ),
                    ]
                )

            fold_model.fit(X_fold_train, y_fold_train)
            score = fold_model.score(X_fold_val, y_fold_val)
            return fold, score

        print(
            "\nRunning enhanced time-based cross-validation with parallel processing..."
        )

        # Prepare fold data for parallel processing
        fold_data = []
        for fold, (train_idx, val_idx) in enumerate(tscv.split(X)):
            fold_data.append(
                (fold, train_idx, val_idx, X, y_winner, preprocessor, feature_columns)
            )

        # Run folds in parallel with optimized backend (5-8x faster)
        winner_cv_scores = []
        results = Parallel(n_jobs=min(12, mp.cpu_count()), backend="loky")(
            delayed(train_fold)(data) for data in fold_data
        )

        # Sort results by fold number and extract scores
        results.sort(key=lambda x: x[0])
        for fold, score in results:
            winner_cv_scores.append(score)
            print(f"  Fold {fold + 1}/7: {score:.4f}")

        # Calculate additional metrics
        winner_cv_mean = np.mean(winner_cv_scores)
        winner_cv_std = np.std(winner_cv_scores)

        print(f"\nWinner Model Cross-Validation Results:")
        print(f"  Mean Score: {winner_cv_mean:.4f} (+/- {winner_cv_std:.4f})")
        print(f"  Individual Scores: {[f'{score:.4f}' for score in winner_cv_scores]}")

        print(f"\n{'=' * 80}")
        print(f"Time-Based CV Accuracy: {winner_cv_mean:.4f} ± {winner_cv_std:.4f}")
        print(
            f"Test Set Accuracy: {self.winner_model.score(X_test, y_winner_test):.4f}"
        )
        print(f"{'=' * 80}\n")

        return feature_columns

    def get_fighter_latest_stats(self, fighter_name):
        """Get latest stats for fighter"""
        red_fights = self.df_train[
            self.df_train["r_fighter"] == fighter_name
        ].sort_values("event_date", ascending=False)
        blue_fights = self.df_train[
            self.df_train["b_fighter"] == fighter_name
        ].sort_values("event_date", ascending=False)

        if len(red_fights) > 0 and len(blue_fights) > 0:
            # Safe access for both pandas and numpy arrays
            if hasattr(red_fights, "iloc"):
                red_latest = red_fights.iloc[0]
                blue_latest = blue_fights.iloc[0]
            else:
                red_latest = red_fights[0]
                blue_latest = blue_fights[0]

            latest = (
                red_latest
                if red_latest["event_date"] > blue_latest["event_date"]
                else blue_latest
            )
            prefix = (
                "r" if red_latest["event_date"] > blue_latest["event_date"] else "b"
            )
        elif len(red_fights) > 0:
            if hasattr(red_fights, "iloc"):
                latest, prefix = red_fights.iloc[0], "r"
            else:
                latest, prefix = red_fights[0], "r"
        elif len(blue_fights) > 0:
            if hasattr(blue_fights, "iloc"):
                latest, prefix = blue_fights.iloc[0], "b"
            else:
                latest, prefix = blue_fights[0], "b"
        else:
            return None

        stats = {
            k: latest[f"{prefix}_{k}_corrected"]
            for k in [
                "wins",
                "losses",
                "draws",
                "win_loss_ratio",
                "pro_SLpM",
                "pro_sig_str_acc",
                "pro_SApM",
                "pro_str_def",
                "pro_total_str_pM",
                "pro_total_str_acc",
                "pro_total_str_absorbed_pM",
                "pro_td_avg",
                "pro_td_acc",
                "pro_td_def",
                "pro_sub_avg",
                "pro_kd_pM",
                "ko_rate",
                "sub_rate",
                "dec_rate",
                "recent_form",
                "head_pct",
                "body_pct",
                "leg_pct",
                "distance_pct",
                "clinch_pct",
                "ground_pct",
                "win_streak",
                "loss_streak",
                "last_5_wins",
                "days_since_last_fight",
                "recent_finish_rate",
                "durability",
                "fight_time_minutes",
                "slpm_trend",
                "td_avg_trend",
                "age_adjusted_performance",
                "peak_indicator",
            ]
        }

        stats.update(
            {
                k: latest[f"{prefix}_{k}"]
                for k in [
                    "height",
                    "reach",
                    "weight",
                    "age_at_event",
                    "stance",
                    "ape_index",
                ]
            }
        )

        # Update with last fight result
        if pd.notna(latest["winner"]):
            fighter_won = (latest["winner"] == "Red" and prefix == "r") or (
                latest["winner"] == "Blue" and prefix == "b"
            )
            fighter_lost = (latest["winner"] == "Red" and prefix == "b") or (
                latest["winner"] == "Blue" and prefix == "r"
            )

            if fighter_won:
                stats["wins"] += 1
            elif fighter_lost:
                stats["losses"] += 1
            else:
                stats["draws"] += 1

            stats["win_loss_ratio"] = stats["wins"] / max(stats["losses"], 1)

            total_fights = stats["wins"] + stats["losses"]
            if total_fights > 0:
                method = str(latest["method"]).lower()
                if "ko" in method or "tko" in method:
                    method_type = "ko"
                elif "sub" in method:
                    method_type = "sub"
                else:
                    method_type = "dec"

                prev_total = total_fights - 1
                if prev_total > 0:
                    prev_ko = int(stats["ko_rate"] * prev_total)
                    prev_sub = int(stats["sub_rate"] * prev_total)
                    prev_dec = int(stats["dec_rate"] * prev_total)
                else:
                    prev_ko = prev_sub = prev_dec = 0

                if fighter_won:
                    if method_type == "ko":
                        prev_ko += 1
                    elif method_type == "sub":
                        prev_sub += 1
                    else:
                        prev_dec += 1

                stats["ko_rate"] = prev_ko / total_fights
                stats["sub_rate"] = prev_sub / total_fights
                stats["dec_rate"] = prev_dec / total_fights

        return stats

    def prepare_upcoming_fight(self, fight, feature_columns):
        """Prepare upcoming fight with all advanced features"""
        r_stats = self.get_fighter_latest_stats(fight["red_fighter"])
        b_stats = self.get_fighter_latest_stats(fight["blue_fighter"])

        if not r_stats or not b_stats:
            return None, None, None

        r_total_fights = r_stats["wins"] + r_stats["losses"]
        b_total_fights = b_stats["wins"] + b_stats["losses"]

        fight_features = {
            "r_fighter": fight["red_fighter"],  # For deep learning
            "b_fighter": fight["blue_fighter"],  # For deep learning
            "height_diff": (r_stats["height"] - b_stats["height"])
            if r_stats["height"] and b_stats["height"]
            else 0,
            "reach_diff": (r_stats["reach"] - b_stats["reach"])
            if r_stats["reach"] and b_stats["reach"]
            else 0,
            "weight_diff": (r_stats["weight"] - b_stats["weight"])
            if r_stats["weight"] and b_stats["weight"]
            else 0,
            "age_at_event_diff": (r_stats["age_at_event"] - b_stats["age_at_event"])
            if r_stats["age_at_event"] and b_stats["age_at_event"]
            else 0,
            "ape_index_diff": (r_stats["ape_index"] - b_stats["ape_index"])
            if r_stats["ape_index"] and b_stats["ape_index"]
            else 0,
            **{
                f"{k}_diff_corrected": r_stats[k] - b_stats[k]
                for k in [
                    "wins",
                    "losses",
                    "win_loss_ratio",
                    "pro_SLpM",
                    "pro_sig_str_acc",
                    "pro_SApM",
                    "pro_str_def",
                    "pro_total_str_pM",
                    "pro_total_str_acc",
                    "pro_total_str_absorbed_pM",
                    "pro_td_avg",
                    "pro_td_acc",
                    "pro_td_def",
                    "pro_sub_avg",
                    "pro_kd_pM",
                    "ko_rate",
                    "sub_rate",
                    "dec_rate",
                    "recent_form",
                    "head_pct",
                    "body_pct",
                    "leg_pct",
                    "distance_pct",
                    "clinch_pct",
                    "ground_pct",
                    "win_streak",
                    "loss_streak",
                    "last_5_wins",
                    "days_since_last_fight",
                    "recent_finish_rate",
                    "durability",
                    "fight_time_minutes",
                    "slpm_trend",
                    "td_avg_trend",
                    "age_adjusted_performance",
                    "peak_indicator",
                ]
            },
            "h2h_advantage": 0,
            "total_rounds": fight["total_rounds"],
            "is_title_bout": 1 if fight["total_rounds"] == 5 else 0,
            "r_total_fights": r_total_fights,
            "b_total_fights": b_total_fights,
        }

        # All individual stats
        for prefix, stats in [("r", r_stats), ("b", b_stats)]:
            for stat in [
                "wins",
                "losses",
                "ko_rate",
                "sub_rate",
                "dec_rate",
                "pro_str_def",
                "pro_td_def",
                "durability",
                "pro_SLpM",
                "pro_sig_str_acc",
                "pro_SApM",
                "pro_td_avg",
                "pro_td_acc",
                "pro_sub_avg",
                "pro_kd_pM",
                "head_pct",
                "distance_pct",
                "clinch_pct",
                "ground_pct",
                "fight_time_minutes",
                "recent_form",
                "recent_finish_rate",
                "win_streak",
                "loss_streak",
            ]:
                fight_features[f"{prefix}_{stat}_corrected"] = stats[stat]

        # Derived features
        r_avg_fight_time = r_stats["fight_time_minutes"] / (r_total_fights + 1)
        b_avg_fight_time = b_stats["fight_time_minutes"] / (b_total_fights + 1)

        fight_features.update(
            {
                "net_striking_advantage": (r_stats["pro_SLpM"] - r_stats["pro_SApM"])
                - (b_stats["pro_SLpM"] - b_stats["pro_SApM"]),
                "striking_efficiency": (
                    r_stats["pro_SLpM"] * r_stats["pro_sig_str_acc"]
                )
                - (b_stats["pro_SLpM"] * b_stats["pro_sig_str_acc"]),
                "defensive_striking": (r_stats["pro_str_def"] - r_stats["pro_SApM"])
                - (b_stats["pro_str_def"] - b_stats["pro_SApM"]),
                "grappling_control": (r_stats["pro_td_avg"] * r_stats["pro_td_acc"])
                - (b_stats["pro_td_avg"] * b_stats["pro_td_acc"]),
                "grappling_defense": (
                    r_stats["pro_td_def"] - r_stats["pro_sub_avg"] / 5
                )
                - (b_stats["pro_td_def"] - b_stats["pro_sub_avg"] / 5),
                "offensive_output": (
                    r_stats["pro_SLpM"] + r_stats["pro_td_avg"] + r_stats["pro_sub_avg"]
                )
                - (
                    b_stats["pro_SLpM"] + b_stats["pro_td_avg"] + b_stats["pro_sub_avg"]
                ),
                "defensive_composite": (
                    (r_stats["pro_str_def"] + r_stats["pro_td_def"]) / 2
                )
                - ((b_stats["pro_str_def"] + b_stats["pro_td_def"]) / 2),
                "ko_specialist_gap": (r_stats["ko_rate"] * r_stats["pro_SLpM"])
                - (b_stats["ko_rate"] * b_stats["pro_SLpM"]),
                "submission_specialist_gap": (
                    r_stats["sub_rate"] * r_stats["pro_sub_avg"]
                )
                - (b_stats["sub_rate"] * b_stats["pro_sub_avg"]),
                "experience_gap": r_total_fights - b_total_fights,
                "skill_momentum": (
                    (r_stats["pro_SLpM"] - b_stats["pro_SLpM"])
                    * (r_stats["recent_form"] - b_stats["recent_form"])
                ),
                "finish_threat": (r_stats["ko_rate"] + r_stats["sub_rate"])
                - (b_stats["ko_rate"] + b_stats["sub_rate"]),
                "momentum_advantage": (
                    (r_stats["win_streak"] - b_stats["win_streak"])
                    - (r_stats["loss_streak"] - b_stats["loss_streak"])
                ),
                "inactivity_penalty": -1
                if r_stats["days_since_last_fight"] - b_stats["days_since_last_fight"]
                > 365
                else 1
                if b_stats["days_since_last_fight"] - r_stats["days_since_last_fight"]
                > 365
                else 0,
                "pace_differential": (r_stats["pro_SLpM"] + r_stats["pro_td_avg"])
                - (b_stats["pro_SLpM"] + b_stats["pro_td_avg"]),
                "experience_ratio": r_total_fights / (b_total_fights + 1),
                "avg_fight_time_diff": r_avg_fight_time - b_avg_fight_time,
            }
        )

        # Enhanced experience
        r_elite_fight_ratio = (
            r_stats["wins"] / r_total_fights if r_total_fights > 0 else 0
        )
        b_elite_fight_ratio = (
            b_stats["wins"] / b_total_fights if b_total_fights > 0 else 0
        )

        fight_features["quality_experience_gap"] = (
            r_total_fights * r_elite_fight_ratio - b_total_fights * b_elite_fight_ratio
        )

        r_championship_exp = (
            1
            if r_total_fights > 0
            and (r_stats["fight_time_minutes"] / r_total_fights) > 12
            else 0
        )
        b_championship_exp = (
            1
            if b_total_fights > 0
            and (b_stats["fight_time_minutes"] / b_total_fights) > 12
            else 0
        )
        fight_features["championship_exp_diff"] = (
            r_championship_exp - b_championship_exp
        )

        r_adversity = r_stats["losses"] * (r_stats["win_loss_ratio"] - 1)
        b_adversity = b_stats["losses"] * (b_stats["win_loss_ratio"] - 1)
        fight_features["adversity_exp_diff"] = r_adversity - b_adversity

        fight_features["experience_skill_interaction"] = (
            fight_features["experience_gap"]
            * (r_stats["pro_SLpM"] - b_stats["pro_SLpM"])
            / 50
        )

        fight_features["veteran_edge"] = (
            fight_features["experience_gap"]
            * (r_stats["recent_form"] - b_stats["recent_form"])
            if abs(fight_features["experience_gap"]) > 5
            else 0
        )

        fight_features["novice_vulnerability"] = (
            -abs(fight_features["experience_gap"]) * 0.5
            if r_total_fights < 10 or b_total_fights < 10
            else 0
        )

        fight_features["exp_gap_historical_win_rate"] = 0.5

        # Style matchups
        r_striker_score = (
            r_stats["distance_pct"] * 1.2
            + r_stats["head_pct"] * 0.8
            - r_stats["ground_pct"] * 1.0
        )
        b_striker_score = (
            b_stats["distance_pct"] * 1.2
            + b_stats["head_pct"] * 0.8
            - b_stats["ground_pct"] * 1.0
        )
        fight_features["striker_advantage"] = r_striker_score - b_striker_score
        fight_features["r_striker_score"] = r_striker_score
        fight_features["b_striker_score"] = b_striker_score

        r_grappler_score = (
            r_stats["pro_td_avg"] * 0.4
            + r_stats["pro_sub_avg"] * 0.3
            + r_stats["ground_pct"] * 0.3
        )
        b_grappler_score = (
            b_stats["pro_td_avg"] * 0.4
            + b_stats["pro_sub_avg"] * 0.3
            + b_stats["ground_pct"] * 0.3
        )
        fight_features["grappler_advantage"] = r_grappler_score - b_grappler_score
        fight_features["r_grappler_score"] = r_grappler_score
        fight_features["b_grappler_score"] = b_grappler_score

        fight_features["effective_reach_advantage"] = (
            fight_features["reach_diff"] * 1.5
            if fight_features["striker_advantage"] > 0.2
            else fight_features["reach_diff"]
        )

        if (
            "stance_encoder" in self.label_encoders
            and r_stats["stance"]
            and b_stats["stance"]
        ):
            r_enc = self.label_encoders["stance_encoder"].transform(
                [r_stats["stance"]]
            )[0]
            b_enc = self.label_encoders["stance_encoder"].transform(
                [b_stats["stance"]]
            )[0]
            fight_features["stance_diff"] = r_enc - b_enc

            fight_features["orthodox_southpaw_matchup"] = (
                0.05
                if (r_stats["stance"] == "Orthodox" and b_stats["stance"] == "Southpaw")
                else -0.05
                if (r_stats["stance"] == "Southpaw" and b_stats["stance"] == "Orthodox")
                else 0
            )
        else:
            fight_features["stance_diff"] = 0
            fight_features["orthodox_southpaw_matchup"] = 0

        return fight_features, r_stats, b_stats

    def get_dynamic_method_weights(self, fight_data, ml_probs, rule_probs, method_type):
        """Calculate optimal weights dynamically based on context and confidence"""

        # Method-specific optimal weights (tuned based on method predictability)
        method_weights = {
            "KO/TKO": {"ml": 0.7, "rule": 0.3},  # ML better at KO patterns
            "Submission": {"ml": 0.3, "rule": 0.7},  # Rules better for grappling
            "Decision": {"ml": 0.5, "rule": 0.5},  # Balanced approach
        }

        # Start with method-specific weights
        weights = method_weights.get(method_type, {"ml": 0.4, "rule": 0.6})

        # Adjust based on confidence difference
        ml_confidence = max(ml_probs) if ml_probs else 0
        rule_confidence = max(rule_probs) if rule_probs else 0
        conf_diff = abs(ml_confidence - rule_confidence)

        if conf_diff > 0.2:  # One approach is much more confident
            if ml_confidence > rule_confidence:
                weights["ml"] = min(0.8, weights["ml"] + 0.2)
                weights["rule"] = max(0.2, weights["rule"] - 0.2)
            else:
                weights["ml"] = max(0.2, weights["ml"] - 0.2)
                weights["rule"] = min(0.8, weights["rule"] + 0.2)

        # Adjust based on fight context - handle both DataFrame and dict inputs
        if hasattr(fight_data, "iloc") and hasattr(fight_data, "columns"):  # DataFrame
            is_title_bout = (
                fight_data["is_title_bout"].iloc[0]
                if "is_title_bout" in fight_data.columns
                else 0
            )
            exp_gap = (
                abs(fight_data["experience_gap"].iloc[0])
                if "experience_gap" in fight_data.columns
                else 0
            )
            striker_adv = (
                abs(fight_data["striker_advantage"].iloc[0])
                if "striker_advantage" in fight_data.columns
                else 0
            )
        else:  # Dictionary or NumPy array
            is_title_bout = (
                fight_data.get("is_title_bout", 0) if hasattr(fight_data, "get") else 0
            )
            exp_gap = abs(fight_data.get("experience_gap", 0))
            striker_adv = abs(fight_data.get("striker_advantage", 0))

        # Adjust based on fight context
        if is_title_bout == 1:  # Title fight - trust ML more
            weights["ml"] = min(0.8, weights["ml"] + 0.1)
            weights["rule"] = max(0.2, weights["rule"] - 0.1)

        # Adjust based on experience gap
        if exp_gap > 10:  # Large experience gap - trust ML more
            weights["ml"] = min(0.8, weights["ml"] + 0.1)
            weights["rule"] = max(0.2, weights["rule"] - 0.1)

        # Adjust based on style matchup
        if striker_adv > 0.5:  # Clear style matchup - trust rules more
            weights["ml"] = max(0.2, weights["ml"] - 0.1)
            weights["rule"] = min(0.8, weights["rule"] + 0.1)

        return weights

    def predict_method_type(self, fight_data):
        """Predict the most likely method type based on fighter characteristics"""

        # Extract key indicators - handle both DataFrame and dict inputs
        if hasattr(fight_data, "iloc") and hasattr(fight_data, "columns"):  # DataFrame
            r_ko_rate = (
                fight_data["r_ko_rate_corrected"].iloc[0]
                if "r_ko_rate_corrected" in fight_data.columns
                else 0
            )
            b_ko_rate = (
                fight_data["b_ko_rate_corrected"].iloc[0]
                if "b_ko_rate_corrected" in fight_data.columns
                else 0
            )
            r_sub_rate = (
                fight_data["r_sub_rate_corrected"].iloc[0]
                if "r_sub_rate_corrected" in fight_data.columns
                else 0
            )
            b_sub_rate = (
                fight_data["b_sub_rate_corrected"].iloc[0]
                if "b_sub_rate_corrected" in fight_data.columns
                else 0
            )
            r_dec_rate = (
                fight_data["r_dec_rate_corrected"].iloc[0]
                if "r_dec_rate_corrected" in fight_data.columns
                else 0
            )
            b_dec_rate = (
                fight_data["b_dec_rate_corrected"].iloc[0]
                if "b_dec_rate_corrected" in fight_data.columns
                else 0
            )
        else:  # Dictionary or NumPy array
            r_ko_rate = (
                fight_data.get("r_ko_rate_corrected", 0)
                if hasattr(fight_data, "get")
                else 0
            )
            b_ko_rate = fight_data.get("b_ko_rate_corrected", 0)
            r_sub_rate = fight_data.get("r_sub_rate_corrected", 0)
            b_sub_rate = fight_data.get("b_sub_rate_corrected", 0)
            r_dec_rate = fight_data.get("r_dec_rate_corrected", 0)
            b_dec_rate = fight_data.get("b_dec_rate_corrected", 0)

        # Calculate average rates
        avg_ko_rate = (r_ko_rate + b_ko_rate) / 2
        avg_sub_rate = (r_sub_rate + b_sub_rate) / 2
        avg_dec_rate = (r_dec_rate + b_dec_rate) / 2

        # Predict based on highest rate
        if avg_ko_rate > avg_sub_rate and avg_ko_rate > avg_dec_rate:
            return "KO/TKO"
        elif avg_sub_rate > avg_ko_rate and avg_sub_rate > avg_dec_rate:
            return "Submission"
        else:
            return "Decision"

    def calculate_enhanced_method_adjustments(
        self, fight_data, winner_prefix, loser_prefix
    ):
        """Comprehensive method prediction using all available stats"""

        # Convert winner_prefix to column prefix (Red -> r, Blue -> b)
        w_prefix = "r" if winner_prefix == "Red" else "b"
        l_prefix = "r" if loser_prefix == "Red" else "b"

        # Extract all relevant stats with proper type checking
        def safe_get_value(data, key, default=0.0):
            try:
                # Check if it's a pandas DataFrame/Series
                if (
                    hasattr(data, "iloc")
                    and hasattr(data, "columns")
                    and hasattr(data, "index")
                ):
                    if key in data.columns:
                        return data[key].iloc[0] if len(data) > 0 else default
                    else:
                        return default
                # Check if it's a pandas Series with values
                elif hasattr(data, "values") and hasattr(data, "index"):
                    if key in data.index:
                        return (
                            data[key].values[0]
                            if len(data[key].values) > 0
                            else default
                        )
                    else:
                        return default
                # Check if it's a dictionary-like object
                elif hasattr(data, "get"):
                    return data.get(key, default)
                # Check if it's a NumPy array with structured access
                elif hasattr(data, "dtype") and hasattr(data, "shape"):
                    if hasattr(data, "columns") and key in data.columns:
                        return data[key][0] if len(data[key]) > 0 else default
                    else:
                        return default
                else:
                    return default
            except (IndexError, KeyError, AttributeError):
                return default

        w_slpm = safe_get_value(fight_data, f"{w_prefix}_pro_SLpM_corrected")
        w_sig_acc = safe_get_value(fight_data, f"{w_prefix}_pro_sig_str_acc_corrected")
        w_td_avg = safe_get_value(fight_data, f"{w_prefix}_pro_td_avg_corrected")
        w_td_acc = safe_get_value(fight_data, f"{w_prefix}_pro_td_acc_corrected")
        w_sub_avg = safe_get_value(fight_data, f"{w_prefix}_pro_sub_avg_corrected")
        w_ko_rate = safe_get_value(fight_data, f"{w_prefix}_ko_rate_corrected")
        w_sub_rate = safe_get_value(fight_data, f"{w_prefix}_sub_rate_corrected")
        w_dec_rate = safe_get_value(fight_data, f"{w_prefix}_dec_rate_corrected")
        w_kd_rate = safe_get_value(fight_data, f"{w_prefix}_pro_kd_pM_corrected")

        l_str_def = safe_get_value(fight_data, f"{l_prefix}_pro_str_def_corrected")
        l_sapm = safe_get_value(fight_data, f"{l_prefix}_pro_SApM_corrected")
        l_td_def = safe_get_value(fight_data, f"{l_prefix}_pro_td_def_corrected")
        l_durability = safe_get_value(fight_data, f"{l_prefix}_durability_corrected")

        w_head_pct = safe_get_value(fight_data, f"{w_prefix}_head_pct_corrected")
        w_distance_pct = safe_get_value(
            fight_data, f"{w_prefix}_distance_pct_corrected"
        )
        w_clinch_pct = safe_get_value(fight_data, f"{w_prefix}_clinch_pct_corrected")
        w_ground_pct = safe_get_value(fight_data, f"{w_prefix}_ground_pct_corrected")

        l_distance_pct = safe_get_value(
            fight_data, f"{l_prefix}_distance_pct_corrected"
        )
        l_ground_pct = safe_get_value(fight_data, f"{l_prefix}_ground_pct_corrected")

        total_rounds = safe_get_value(fight_data, "total_rounds")

        # ENHANCED KO/TKO PROBABILITY
        ko_base = w_ko_rate

        # Advanced striking analysis
        striking_volume_factor = min(w_slpm / 6.0, 1.5)
        accuracy_factor = 1 + (w_sig_acc - 0.45) * 2.5  # Increased sensitivity
        accuracy_factor = max(0.4, min(accuracy_factor, 2.0))

        # Enhanced head hunting analysis
        head_hunting_factor = 1 + (w_head_pct - 0.5) * 0.8  # Increased sensitivity
        head_hunting_factor = max(0.6, min(head_hunting_factor, 1.6))

        # Distance fighting preference
        distance_factor = 1 + (w_distance_pct - 0.6) * 0.6  # Increased sensitivity
        distance_factor = max(0.7, min(distance_factor, 1.5))

        # Enhanced knockdown threat analysis
        kd_threat_factor = 1 + min(w_kd_rate / 0.4, 1.5)  # Increased sensitivity
        kd_threat_factor = max(0.8, kd_threat_factor)

        # Advanced opponent vulnerability analysis
        opp_vulnerability = (
            (1 - l_str_def) * 0.45 + (l_sapm / 6.0) * 0.35 + (1 - l_durability) * 0.2
        )
        opp_vulnerability = min(opp_vulnerability, 1.2)  # Allow higher vulnerability

        # Enhanced power differential analysis
        power_differential = (w_slpm * w_sig_acc) - (l_sapm * (1 - l_str_def))
        power_differential_factor = 1 + max(
            0, power_differential / 4.0
        )  # Increased sensitivity
        power_differential_factor = min(power_differential_factor, 2.0)

        # Recent form impact on finishing ability
        # Extract recent form data (with fallback if not available)
        try:
            w_recent_form = safe_get_value(
                fight_data, f"{w_prefix}_recent_form_corrected"
            )
        except (KeyError, IndexError):
            w_recent_form = 0.5  # Default neutral value

        recent_form_factor = 1 + (w_recent_form - 0.5) * 0.4
        recent_form_factor = max(0.7, min(recent_form_factor, 1.3))

        # ENHANCED METHOD PREDICTION FEATURES
        # Momentum and confidence factors
        w_win_streak = safe_get_value(fight_data, f"{w_prefix}_win_streak_corrected")
        w_loss_streak = safe_get_value(fight_data, f"{w_prefix}_loss_streak_corrected")
        momentum_factor = 1 + (w_win_streak - w_loss_streak) * 0.1
        momentum_factor = max(0.8, min(momentum_factor, 1.4))

        # Pressure and championship experience
        is_title_bout = safe_get_value(fight_data, "is_title_bout", 0)
        championship_factor = 1 + is_title_bout * 0.15

        # Style matchup impact on method
        striker_vs_grappler = safe_get_value(fight_data, "striker_vs_grappler", 0)
        style_factor = 1 + abs(striker_vs_grappler) * 0.1

        # Physical advantage impact
        height_advantage = safe_get_value(fight_data, "height_diff", 0)
        reach_advantage = safe_get_value(fight_data, "reach_diff", 0)
        physical_factor = 1 + (height_advantage + reach_advantage) * 0.05
        physical_factor = max(0.9, min(physical_factor, 1.2))

        ko_prob = ko_base * (
            striking_volume_factor * 0.15
            + accuracy_factor * 0.18
            + head_hunting_factor * 0.12
            + distance_factor * 0.08
            + opp_vulnerability * 0.18
            + power_differential_factor * 0.10
            + kd_threat_factor * 0.10
            + recent_form_factor * 0.12
            + momentum_factor * 0.08
            + championship_factor * 0.05
            + style_factor * 0.06
            + physical_factor * 0.08
        )

        if total_rounds == 5:
            ko_prob *= 1.15

        ko_prob = min(ko_prob, 0.85)

        # SUBMISSION PROBABILITY
        sub_base = w_sub_rate

        grappling_control = w_td_avg * w_td_acc
        control_factor = 1 + min(grappling_control / 2.0, 1.0)

        sub_attempt_factor = 1 + (w_sub_avg / 2.0)
        sub_attempt_factor = min(sub_attempt_factor, 2.0)

        ground_preference_factor = 1 + (w_ground_pct - 0.2) * 0.8
        ground_preference_factor = max(0.6, min(ground_preference_factor, 1.6))

        opp_grappling_weakness = (
            (1 - l_td_def) * 0.5 + (1 - l_durability) * 0.3 + l_ground_pct * 0.2
        )

        td_differential = (w_td_avg * w_td_acc) - (l_td_def * 2.0)
        td_factor = 1 + max(0, td_differential / 3.0)
        td_factor = min(td_factor, 1.7)

        sub_prob = sub_base * (
            control_factor * 0.20
            + sub_attempt_factor * 0.18
            + ground_preference_factor * 0.12
            + opp_grappling_weakness * 0.22
            + td_factor * 0.12
            + momentum_factor * 0.08
            + championship_factor * 0.04
            + style_factor * 0.06
            + physical_factor * 0.05
        )

        if total_rounds == 5:
            sub_prob *= 1.10

        sub_prob = min(sub_prob, 0.75)

        # DECISION PROBABILITY
        dec_base = w_dec_rate

        # Factors that favor decisions
        decision_factors = []

        # High output but low finishing rate
        if w_slpm > 4.0 and w_ko_rate < 0.15:
            decision_factors.append(1.3)

        # High takedown rate but low submission rate
        if w_td_avg > 2.0 and w_sub_rate < 0.10:
            decision_factors.append(1.2)

        # Opponent durability
        if l_durability > 0.7:
            decision_factors.append(1.4)

        # Distance fighting preference
        if w_distance_pct > 0.7:
            decision_factors.append(1.1)

        # Clinch fighting (often leads to decisions)
        if w_clinch_pct > 0.3:
            decision_factors.append(1.15)

        decision_multiplier = np.mean(decision_factors) if decision_factors else 1.0

        # Enhanced decision probability with new factors
        dec_prob = (
            dec_base
            * decision_multiplier
            * (
                0.85  # Base multiplier
                + momentum_factor * 0.05
                + championship_factor * 0.03
                + style_factor * 0.04
                + physical_factor * 0.03
            )
        )

        # Additional context-based adjustments
        # Weight class adjustments - handle both DataFrame and dict inputs
        if hasattr(fight_data, "iloc") and hasattr(fight_data, "columns"):  # DataFrame
            weight_class = (
                fight_data["weight_class"].iloc[0]
                if "weight_class" in fight_data.columns
                else "Unknown"
            )
            w_recent_form = (
                fight_data[f"{w_prefix}_recent_form_corrected"].iloc[0]
                if f"{w_prefix}_recent_form_corrected" in fight_data.columns
                else 0.5
            )
            l_durability = (
                fight_data[f"{l_prefix}_durability_corrected"].iloc[0]
                if f"{l_prefix}_durability_corrected" in fight_data.columns
                else 0.5
            )
        else:  # Dictionary or NumPy array
            weight_class = (
                fight_data.get("weight_class", "Unknown")
                if hasattr(fight_data, "get")
                else "Unknown"
            )
            w_recent_form = fight_data.get(f"{w_prefix}_recent_form_corrected", 0.5)
            l_durability = fight_data.get(f"{l_prefix}_durability_corrected", 0.5)

        if weight_class in ["Heavyweight", "Light Heavyweight"]:
            ko_prob *= 1.2  # Higher KO rate in heavier divisions
            sub_prob *= 0.8  # Lower sub rate in heavier divisions
        elif weight_class in ["Flyweight", "Bantamweight"]:
            ko_prob *= 0.9  # Lower KO rate in lighter divisions
            sub_prob *= 1.1  # Higher sub rate in lighter divisions

        # Round length adjustments
        if total_rounds == 5:  # Championship rounds
            ko_prob *= 1.1  # Slightly higher KO chance with more time
            sub_prob *= 1.05  # Slightly higher sub chance
            dec_prob *= 1.15  # Much higher decision chance

        # Recent form adjustments
        if w_recent_form > 0.7:  # Hot streak
            ko_prob *= 1.15
            sub_prob *= 1.1
        elif w_recent_form < 0.3:  # Cold streak
            ko_prob *= 0.9
            sub_prob *= 0.85
            dec_prob *= 1.1

        # Durability adjustments
        if l_durability < 0.3:  # Fragile opponent
            ko_prob *= 1.25
        elif l_durability > 0.8:  # Durable opponent
            ko_prob *= 0.8
            dec_prob *= 1.2

        # Normalize probabilities
        total_prob = ko_prob + sub_prob + dec_prob
        if total_prob > 0:
            ko_prob = ko_prob / total_prob
            sub_prob = sub_prob / total_prob
            dec_prob = dec_prob / total_prob

        return {
            f"{winner_prefix}_KO/TKO": ko_prob,
            f"{winner_prefix}_Submission": sub_prob,
            f"{winner_prefix}_Decision": dec_prob,
        }

    def predict_fight(self, fight_data, feature_columns):
        """Enhanced fight prediction with comprehensive method adjustments"""
        # Ensure consistent random state before prediction
        self.set_random_seeds()

        X = fight_data[feature_columns]

        # Get winner prediction
        winner_proba = self.winner_model.predict_proba(X)[0]
        winner_pred = self.winner_model.predict(X)[0]
        winner_name = "Red" if winner_pred == 1 else "Blue"

        # If deep learning is available, ensemble it with traditional model
        if self.use_deep_learning and self.deep_learning_model and self.fighter_encoder:
            try:
                # Get fighter encodings with safe access
                def safe_get_fighter(data, key):
                    try:
                        # Check if it's a pandas DataFrame/Series
                        if (
                            hasattr(data, "iloc")
                            and hasattr(data, "columns")
                            and hasattr(data, "index")
                        ):
                            if key in data.columns:
                                return data[key].iloc[0] if len(data) > 0 else None
                            else:
                                return None
                        # Check if it's a pandas Series with values
                        elif hasattr(data, "values") and hasattr(data, "index"):
                            if key in data.index:
                                return (
                                    data[key].values[0]
                                    if len(data[key].values) > 0
                                    else None
                                )
                            else:
                                return None
                        # Check if it's a dictionary-like object
                        elif hasattr(data, "get"):
                            return data.get(key, None)
                        # Check if it's a NumPy array with structured access
                        elif hasattr(data, "dtype") and hasattr(data, "shape"):
                            if hasattr(data, "columns") and key in data.columns:
                                return data[key][0] if len(data[key]) > 0 else None
                            else:
                                return None
                        else:
                            return None
                    except (IndexError, KeyError, AttributeError):
                        return None

                r_fighter = safe_get_fighter(fight_data, "r_fighter")
                b_fighter = safe_get_fighter(fight_data, "b_fighter")

                if (
                    r_fighter
                    and b_fighter
                    and r_fighter in self.fighter_encoder.classes_
                    and b_fighter in self.fighter_encoder.classes_
                ):
                    r_fighter_encoded = self.fighter_encoder.transform([r_fighter])[0]
                    b_fighter_encoded = self.fighter_encoder.transform([b_fighter])[0]

                    # Prepare DL input
                    from sklearn.compose import ColumnTransformer
                    from sklearn.impute import SimpleImputer
                    from sklearn.preprocessing import StandardScaler
                    from sklearn.pipeline import Pipeline

                    numeric_transformer = Pipeline(
                        [
                            ("imputer", SimpleImputer(strategy="median")),
                            ("scaler", StandardScaler()),
                        ]
                    )
                    preprocessor = ColumnTransformer(
                        [("num", numeric_transformer, feature_columns)]
                    )
                    X_scaled = preprocessor.fit_transform(X)

                    X_dl = [
                        np.array([r_fighter_encoded]),
                        np.array([b_fighter_encoded]),
                        X_scaled,
                    ]

                    # Get DL predictions
                    dl_winner_proba, dl_method_proba = self.deep_learning_model.predict(
                        X_dl, verbose=0
                    )

                    # Convert winner prediction to 2-class format
                    dl_winner_2class = np.array(
                        [[1 - dl_winner_proba[0][0], dl_winner_proba[0][0]]]
                    )

                    # Ensemble: 60% traditional, 40% deep learning
                    winner_proba = winner_proba * 0.6 + dl_winner_2class[0] * 0.4
                    winner_pred = 1 if winner_proba[1] > 0.5 else 0
                    winner_name = "Red" if winner_pred == 1 else "Blue"
            except Exception:
                # Fallback to traditional model if DL fails
                pass

        # COMPREHENSIVE METHOD PREDICTION WITH DYNAMIC WEIGHTING
        loser_prefix = "Blue" if winner_name == "Red" else "Red"

        # Get base method probabilities from model
        method_proba = self.method_model.predict_proba(X)[0]
        method_labels = self.label_encoders["winner_method_encoder"].classes_

        # Get comprehensive method adjustments
        method_adjustments = self.calculate_enhanced_method_adjustments(
            fight_data, winner_name, loser_prefix
        )

        # Predict method type for dynamic weighting
        predicted_method_type = self.predict_method_type(fight_data)

        # Prepare ML and rule probabilities for weighting
        ml_probs = {}
        rule_probs = {}
        for i, label in enumerate(method_labels):
            if label.startswith(winner_name):
                method_name = label.split("_")[1]
                ml_probs[method_name] = method_proba[i]
                rule_probs[method_name] = method_adjustments.get(label, 0.0)

        # Get dynamic weights based on context and confidence
        weights = self.get_dynamic_method_weights(
            fight_data,
            list(ml_probs.values()),
            list(rule_probs.values()),
            predicted_method_type,
        )

        # Debug information (can be removed in production)
        if hasattr(self, "debug_mode") and self.debug_mode:
            print(
                f"Dynamic Weights - Method: {predicted_method_type}, ML: {weights['ml']:.2f}, Rule: {weights['rule']:.2f}"
            )
            print(f"ML Probs: {ml_probs}")
            print(f"Rule Probs: {rule_probs}")

        # Combine model predictions with comprehensive adjustments using dynamic weights
        final_method_probs = {}
        for i, label in enumerate(method_labels):
            if label.startswith(winner_name):
                method_name = label.split("_")[1]
                base_prob = method_proba[i]
                adjustment_prob = method_adjustments.get(label, 0.0)

                # Dynamic weighted combination based on context and confidence
                final_prob = (
                    base_prob * weights["ml"] + adjustment_prob * weights["rule"]
                )
                final_method_probs[label] = final_prob

        # Normalize probabilities
        total = sum(final_method_probs.values())
        if total > 0:
            final_method_probs = {k: v / total for k, v in final_method_probs.items()}

        final_method = max(final_method_probs, key=final_method_probs.get)

        return {
            "winner": winner_name,
            "winner_confidence": winner_proba[winner_pred],
            "red_prob": winner_proba[1],
            "blue_prob": winner_proba[0],
            "method": final_method.split("_")[1],
            "method_probabilities": final_method_probs,
            "combined_method_prob": final_method_probs[final_method],
        }

    def predict_upcoming_fights(self, upcoming_fights, feature_columns):
        """Predict upcoming fights"""
        predictions = []
        skipped_fights = []

        for fight in upcoming_fights:
            result = self.prepare_upcoming_fight(fight, feature_columns)
            if not result[0]:
                skipped_fights.append(
                    f"{fight['red_fighter']} vs {fight['blue_fighter']}"
                )
                continue

            fight_features, r_stats, b_stats = result
            fight_df = pd.DataFrame([fight_features])

            for col in feature_columns:
                if col not in fight_df.columns:
                    fight_df[col] = 0

            pred = self.predict_fight(fight_df, feature_columns)

            winner_prefix = pred["winner"]
            ko_prob = pred["method_probabilities"].get(f"{winner_prefix}_KO/TKO", 0)
            sub_prob = pred["method_probabilities"].get(
                f"{winner_prefix}_Submission", 0
            )
            dec_prob = pred["method_probabilities"].get(f"{winner_prefix}_Decision", 0)

            predictions.append(
                {
                    "Red Fighter": fight["red_fighter"],
                    "Blue Fighter": fight["blue_fighter"],
                    "Weight Class": fight["weight_class"],
                    "Winner": fight["red_fighter"]
                    if pred["winner"] == "Red"
                    else fight["blue_fighter"],
                    "Win%": pred["winner_confidence"],
                    "Method": pred["method"],
                    "Method%": pred["combined_method_prob"],
                    "KO/TKO%": ko_prob,
                    "Submission%": sub_prob,
                    "Decision%": dec_prob,
                }
            )

        return predictions, skipped_fights

    def export_predictions_to_excel(self, predictions, filename="predictions.xlsx"):
        """Export predictions to formatted Excel file"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        df = pd.DataFrame(predictions)

        # Ensure percentage columns are in decimal format (0-1 range)
        percentage_columns = ["Win%", "Method%", "KO/TKO%", "Submission%", "Decision%"]
        for col in percentage_columns:
            if col in df.columns:
                # If values are > 1, they're likely already percentages, convert to decimal
                if df[col].max() > 1:
                    df[col] = df[col] / 100
                df[col] = df[col].round(4)

        df.to_excel(filename, index=False, engine="openpyxl")

        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        header_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        header_font = Font(bold=True, size=11, color="000000")
        left_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Apply formatting to ALL cells
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.alignment = left_alignment
                cell.border = thin_border

        # Apply header-specific formatting
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Format percentage columns: Win%(5), Method%(7), KO/TKO%(8), Submission%(9), Decision%(10)
        percentage_col_indices = [5, 7, 8, 9, 10]
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in percentage_col_indices:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = "0.00%"

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value is None:
                        cell_length = 0
                    elif isinstance(cell.value, (int, float)):
                        cell_length = (
                            10
                            if cell.number_format == "0.00%"
                            else len(str(cell.value))
                        )
                    else:
                        cell_length = len(str(cell.value))

                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass

            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(filename)
        print(f"\nPredictions exported to: {filename}")
        return filename


class UFCPredictorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("UFC Fight Predictor")
        self.root.geometry("1000x800")
        self.root.minsize(700, 550)

        self.data_file_path = tk.StringVar(value=fight_data_path)
        self.output_file_path = tk.StringVar(value="UFC_predictions_3.xlsx")
        self.use_deep_learning = tk.BooleanVar(value=True)
        self.predictor = None
        self.create_widgets()

    def create_widgets(self):
        title_frame = tk.Frame(self.root, bg="#D20A0A")
        title_frame.pack(fill=tk.X)
        tk.Label(
            title_frame,
            text="UFC FIGHT PREDICTOR",
            font=("Arial", 16, "bold"),
            fg="white",
            bg="#D20A0A",
        ).pack(pady=(10, 8))

        file_frame = ttk.LabelFrame(self.root, text="Data File", padding="5")
        file_frame.pack(fill=tk.X, padx=10, pady=3)
        ttk.Entry(file_frame, textvariable=self.data_file_path, width=70).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(file_frame, text="Browse", command=self.browse_data_file).pack(
            side=tk.LEFT
        )

        input_frame = ttk.LabelFrame(
            self.root, text="Enter Fights to Predict", padding="5"
        )
        input_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=3)
        instructions = """Enter fights in CSV format (one per line):
Format: red_fighter,blue_fighter,weight_class,gender,total_rounds

Example:
Max Holloway,Dustin Poirier,Lightweight,Men,5
Ilia Topuria,Charles Oliveira,Lightweight,Men,5
Tatiana Suarez,Amanda Lemos,Strawweight,Women,3"""
        ttk.Label(
            input_frame,
            text=instructions,
            font=("Arial", 8),
            foreground="gray",
            justify=tk.LEFT,
        ).pack(anchor=tk.W, pady=(0, 3))

        self.fights_text = scrolledtext.ScrolledText(
            input_frame, height=8, width=95, wrap=tk.WORD
        )
        self.fights_text.pack(fill=tk.BOTH, expand=True, pady=3)

        output_frame = ttk.LabelFrame(self.root, text="Output File", padding="5")
        output_frame.pack(fill=tk.X, padx=10, pady=3)
        ttk.Entry(output_frame, textvariable=self.output_file_path, width=70).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(output_frame, text="Browse", command=self.browse_output_file).pack(
            side=tk.LEFT
        )

        # Options frame
        options_frame = ttk.LabelFrame(self.root, text="Model Options", padding="5")
        options_frame.pack(fill=tk.X, padx=10, pady=3)
        ttk.Checkbutton(
            options_frame,
            text="Enable Deep Learning (TensorFlow)",
            variable=self.use_deep_learning,
        ).pack(anchor=tk.W, padx=5)

        button_frame = ttk.Frame(self.root, padding="5")
        button_frame.pack(fill=tk.X, padx=10)
        ttk.Button(button_frame, text="Load Sample", command=self.load_sample).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(button_frame, text="Clear", command=self.clear_input).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(
            button_frame, text="Generate Predictions", command=self.run_predictions
        ).pack(side=tk.RIGHT, padx=3)

        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(
            self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W
        )
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)

    def browse_data_file(self):
        filename = filedialog.askopenfilename(
            title="Select Fight Data CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if filename:
            self.data_file_path.set(filename)

    def browse_output_file(self):
        filename = filedialog.asksaveasfilename(
            title="Save Predictions As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if filename:
            self.output_file_path.set(filename)

    def load_sample(self):
        sample = """Max Holloway,Dustin Poirier,Lightweight,Men,5
Ilia Topuria,Charles Oliveira,Lightweight,Men,5
Tatiana Suarez,Amanda Lemos,Strawweight,Women,3"""
        self.fights_text.delete("1.0", tk.END)
        self.fights_text.insert("1.0", sample)

    def clear_input(self):
        self.fights_text.delete("1.0", tk.END)

    def parse_fights_input(self, text):
        text = text.strip()
        fights = []

        for line in text.split("\n"):
            line = line.strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split(",")]

            if len(parts) != 5:
                raise ValueError(
                    f"Invalid CSV format. Expected 5 fields, got {len(parts)} in line: {line}"
                )

            try:
                fights.append(
                    {
                        "red_fighter": parts[0],
                        "blue_fighter": parts[1],
                        "weight_class": parts[2],
                        "gender": parts[3],
                        "total_rounds": int(parts[4]),
                    }
                )
            except ValueError as e:
                raise ValueError(
                    f"Error parsing line: {line}\nTotal rounds must be a number. {str(e)}"
                )

        if not fights:
            raise ValueError("No valid fight data found.")
        return fights

    def run_predictions(self):
        try:
            use_dl = self.use_deep_learning.get()
            dl_status = " with Deep Learning" if use_dl else ""
            self.status_var.set(
                f"Loading data and training advanced ensemble{dl_status}..."
            )
            self.root.update()

            fights_text = self.fights_text.get("1.0", tk.END)
            upcoming_fights = self.parse_fights_input(fights_text)

            data_file = self.data_file_path.get()
            if not os.path.exists(data_file):
                raise FileNotFoundError(f"Data file not found: {data_file}")

            df = pd.read_csv(data_file)
            self.status_var.set(
                f"Loaded {len(df)} fights. Training ensemble{dl_status}..."
            )
            self.root.update()

            self.predictor = AdvancedUFCPredictor(
                use_ensemble=True,
                use_neural_net=False,
                use_deep_learning=use_dl,
                debug_mode=False,  # Set to True for debugging
            )

            f = io.StringIO()
            with redirect_stdout(f):
                df = self.predictor.fix_data_leakage(df)
                self.predictor.df_train = df
                feature_columns = self.predictor.train_models(df)

            self.status_var.set("Generating predictions...")
            self.root.update()

            predictions, skipped_fights = self.predictor.predict_upcoming_fights(
                upcoming_fights, feature_columns
            )

            if not predictions:
                self.status_var.set("No predictions generated")
                messagebox.showerror(
                    "No Predictions",
                    "All fights were skipped due to insufficient fighter data.",
                )
                return

            output_file = self.output_file_path.get()

            with redirect_stdout(io.StringIO()):
                self.predictor.export_predictions_to_excel(predictions, output_file)

            # Clean up temporary files immediately after training
            cleanup_temp_files()

            success_msg = f"Predictions generated!\n\nSaved to: {output_file}\n\n{len(predictions)} fight(s) predicted"
            if use_dl:
                success_msg += "\n✓ Deep Learning enabled"

            if skipped_fights:
                success_msg += f"\n{len(skipped_fights)} fight(s) skipped"
                self.status_var.set(
                    f"Success! {len(predictions)} predictions, {len(skipped_fights)} skipped"
                )
                skipped_msg = (
                    success_msg
                    + "\n\nSkipped:\n"
                    + "\n".join(f"• {fight}" for fight in skipped_fights)
                )
                messagebox.showwarning("Predictions Complete", skipped_msg)
            else:
                self.status_var.set(
                    f"Success! All {len(predictions)} predictions saved"
                )
                messagebox.showinfo("Success", success_msg)

        except Exception as e:
            import traceback

            error_details = traceback.format_exc()
            self.status_var.set("Error occurred")
            messagebox.showerror(
                "Error", f"Error:\n\n{str(e)}\n\nDetails:\n{error_details}"
            )


def main():
    """Main function to run the UFC Predictor GUI"""
    try:
        # Check if already running to prevent multiple instances
        if hasattr(main, "_running"):
            return
        main._running = True

        root = tk.Tk()
        app = UFCPredictorGUI(root)
        root.mainloop()
    except Exception as e:
        print(f"Error starting UFC Predictor: {e}")
        import traceback

        traceback.print_exc()
    finally:
        # Clean up temporary files when GUI is closed
        cleanup_temp_files()
        # Reset the flag when done
        if hasattr(main, "_running"):
            delattr(main, "_running")


# Safe execution for IDE "Run Code" feature
if __name__ == "__main__":
    try:
        # Handle multiprocessing for PyInstaller - must be called before any multiprocessing operations
        if hasattr(sys, "frozen") and sys.frozen:
            import multiprocessing

            multiprocessing.freeze_support()
        main()
    except Exception as e:
        print(f"Error in UFC Predictor execution: {e}")
        # Don't exit, just print the error for debugging