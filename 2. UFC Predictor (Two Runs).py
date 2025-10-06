import sys
import os
import random

# Disable multiprocessing in ML libraries to prevent subprocess creation
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["VECLIB_MAXIMUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"
os.environ["NUMEXPR_MAX_THREADS"] = "1"
os.environ["XGBOOST_DISABLE_MULTIPROCESSING"] = "1"

# Enable multiprocessing for our custom parallel operations
os.environ["JOBLIB_MULTIPROCESSING"] = "1"
os.environ["LOKY_MAX_WORKERS"] = "8"

# Note: multiprocessing.freeze_support() will be called later in the main execution block

# Get the directory where this script is located
# Handle both regular Python execution and PyInstaller executable
if getattr(sys, "frozen", False):
    # Running as PyInstaller executable
    script_dir = os.path.dirname(sys.executable)
else:
    # Running as regular Python script
    script_dir = os.path.dirname(os.path.abspath(__file__))

# Check for fight_data.csv in the same directory as the script/executable
fight_data_path = os.path.join(script_dir, "fight_data.csv")
if not os.path.exists(fight_data_path):
    print(f"Warning: fight_data.csv not found in script directory: {script_dir}")

import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog, messagebox

# Import multiprocessing only when needed
import json
import pandas as pd
import numpy as np

# Set random seeds for reproducibility
random.seed(42)
np.random.seed(42)

# Additional deterministic settings
os.environ["PYTHONHASHSEED"] = "42"
os.environ["TF_DETERMINISTIC_OPS"] = "1"
os.environ["TF_CUDNN_DETERMINISTIC"] = "1"
import io
from contextlib import redirect_stdout, redirect_stderr
from sklearn.model_selection import train_test_split, TimeSeriesSplit, cross_val_score
from sklearn.ensemble import (
    RandomForestClassifier,
    VotingClassifier,
    StackingClassifier,
)
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
from sklearn.feature_selection import SelectPercentile, f_classif
from sklearn.metrics import classification_report, accuracy_score
from sklearn.calibration import CalibratedClassifierCV
from sklearn.linear_model import LogisticRegression
from sklearn.neural_network import MLPClassifier
from scipy.stats import linregress
import warnings
import multiprocessing as mp
from joblib import Parallel, delayed
import shutil
import atexit

warnings.filterwarnings("ignore")

def cleanup_temp_files():
    """Clean up temporary files and folders created during training"""
    try:
        # Remove catboost_info folder if it exists
        if os.path.exists("catboost_info"):
            shutil.rmtree("catboost_info")
            print("Cleaned up catboost_info folder")
        
        # Remove best_dl_model.h5 file if it exists
        if os.path.exists("best_dl_model.h5"):
            os.remove("best_dl_model.h5")
            print("Cleaned up best_dl_model.h5 file")
            
    except Exception as e:
        print(f"Warning: Could not clean up temporary files: {e}")

# Register cleanup function to run on exit
atexit.register(cleanup_temp_files)

try:
    from xgboost import XGBClassifier

    HAS_XGBOOST = True
except ImportError:
    HAS_XGBOOST = False
    print("XGBoost not available.")

try:
    from lightgbm import LGBMClassifier

    HAS_LIGHTGBM = True
except ImportError:
    HAS_LIGHTGBM = False
    print("LightGBM not available. Install with: pip install lightgbm")

try:
    from catboost import CatBoostClassifier

    HAS_CATBOOST = True
except ImportError:
    HAS_CATBOOST = False
    print("CatBoost not available. Install with: pip install catboost")

try:
    import tensorflow as tf
    from tensorflow import keras

    HAS_TENSORFLOW = True
except ImportError:
    HAS_TENSORFLOW = False
    print("TensorFlow not available. Install with: pip install tensorflow")


class AdvancedUFCPredictor:
    def __init__(
        self,
        use_ensemble=True,
        use_neural_net=False,
        use_deep_learning=False,
        debug_mode=False,
    ):
        self.winner_model = None
        self.method_model = None  # Single method prediction model
        self.deep_learning_model = None  # TensorFlow model
        self.label_encoders = {}
        self.use_ensemble = use_ensemble
        self.use_neural_net = use_neural_net
        self.use_deep_learning = use_deep_learning and HAS_TENSORFLOW
        self.debug_mode = debug_mode
        self.df_train = None
        self.fighter_style_cache = {}
        self.fighter_encoder = None  # For fighter embeddings
        self.num_fighters = 0
        
        # Performance optimization attributes
        self.feature_cache = {}
        self.preprocessor_cache = None
        self.feature_columns_cache = None
        self.early_stopping_patience = 50
        self.best_score = 0
        self.patience_counter = 0

        # Set all random seeds for maximum reproducibility
        self.set_random_seeds()

    def set_random_seeds(self):
        """Set all random seeds for maximum reproducibility while maintaining accuracy"""
        # Python random
        random.seed(42)

        # NumPy random
        np.random.seed(42)

        # Set environment variables for deterministic operations
        os.environ["PYTHONHASHSEED"] = "42"
        os.environ["TF_DETERMINISTIC_OPS"] = "1"
        os.environ["TF_CUDNN_DETERMINISTIC"] = "1"

        # XGBoost specific random seed settings
        os.environ["XGBOOST_DISABLE_MULTIPROCESSING"] = "1"

        # Set TensorFlow random seed if available
        if HAS_TENSORFLOW:
            try:
                import tensorflow as tf

                tf.random.set_seed(42)
                # Enable deterministic operations for consistency
                tf.config.experimental.enable_op_determinism()
            except:
                pass

    def calculate_streak(self, recent_wins, count_wins=True):
        """Calculate current win or loss streak"""
        if not recent_wins:
            return 0

        streak = 0
        target = 1 if count_wins else 0

        for result in reversed(recent_wins):
            if result == target:
                streak += 1
            else:
                break

        return streak

    def calculate_trajectory(self, values):
        """Calculate linear trend of recent performance"""
        if len(values) < 2:
            return 0
        try:
            x = np.arange(len(values))
            slope, _, _, _, _ = linregress(x, values)
            return slope
        except:
            return 0

    def classify_fighter_style(self, stats):
        """Classify fighter as striker, grappler, or balanced"""
        striker_score = (
            stats.get("pro_SLpM_corrected", 0) * 0.4
            + stats.get("distance_pct_corrected", 0) * 0.3
            + stats.get("head_pct_corrected", 0) * 0.3
            - stats.get("ground_pct_corrected", 0) * 0.4
        )

        grappler_score = (
            stats.get("pro_td_avg_corrected", 0) * 0.4
            + stats.get("pro_sub_avg_corrected", 0) * 0.3
            + stats.get("ground_pct_corrected", 0) * 0.3
            - stats.get("distance_pct_corrected", 0) * 0.2
        )

        if striker_score > 0.3 and striker_score > grappler_score:
            return "striker"
        elif grappler_score > 0.3 and grappler_score > striker_score:
            return "grappler"
        else:
            return "balanced"

    def calculate_age_curve_factor(self, age):
        """Calculate performance multiplier based on age"""
        if age < 25:
            return 0.92  # Still developing
        elif 25 <= age <= 32:
            return 1.0  # Prime years
        elif 33 <= age <= 36:
            return 0.94  # Slight decline
        else:
            return 0.85  # Significant decline

    def is_fighter_at_peak(self, wins, losses, age, recent_form):
        """Determine if fighter is at career peak"""
        total_fights = wins + losses
        experience_factor = min(total_fights / 20, 1.0)  # Peaks around 20 fights
        age_factor = self.calculate_age_curve_factor(age)
        form_factor = recent_form

        peak_score = experience_factor * 0.4 + age_factor * 0.3 + form_factor * 0.3
        return peak_score

    def build_deep_learning_model(self, num_features, num_fighters):
        """Build enhanced TensorFlow deep learning model with fighter embeddings and method prediction"""
        if not HAS_TENSORFLOW:
            return None

        # Ensure deterministic operations for consistency
        self.set_random_seeds()

        # Fighter embedding input
        fighter_r_input = keras.Input(shape=(1,), name="fighter_r")
        fighter_b_input = keras.Input(shape=(1,), name="fighter_b")

        # Enhanced embedding layer with larger dimensions
        fighter_embedding = keras.layers.Embedding(
            input_dim=num_fighters,
            output_dim=min(64, num_fighters // 4),
            name="fighter_embedding",
        )

        fighter_r_embed = keras.layers.Flatten()(fighter_embedding(fighter_r_input))
        fighter_b_embed = keras.layers.Flatten()(fighter_embedding(fighter_b_input))

        # Statistical features input with enhanced processing
        stats_input = keras.Input(shape=(num_features,), name="stats")
        stats_dense = keras.layers.Dense(256, activation="relu")(stats_input)
        stats_dense = keras.layers.BatchNormalization()(stats_dense)
        stats_dense = keras.layers.Dropout(0.3)(stats_dense)

        stats_dense = keras.layers.Dense(128, activation="relu")(stats_dense)
        stats_dense = keras.layers.Dropout(0.2)(stats_dense)

        # Combine all inputs
        combined = keras.layers.Concatenate()(
            [fighter_r_embed, fighter_b_embed, stats_dense]
        )

        # Enhanced shared layers
        x = keras.layers.Dense(512, activation="relu")(combined)
        x = keras.layers.BatchNormalization()(x)
        x = keras.layers.Dropout(0.3)(x)

        x = keras.layers.Dense(256, activation="relu")(x)
        x = keras.layers.BatchNormalization()(x)
        x = keras.layers.Dropout(0.3)(x)

        x = keras.layers.Dense(128, activation="relu")(x)
        x = keras.layers.Dropout(0.2)(x)

        # Winner prediction branch
        winner_branch = keras.layers.Dense(64, activation="relu")(x)
        winner_branch = keras.layers.Dropout(0.1)(winner_branch)
        winner_output = keras.layers.Dense(1, activation="sigmoid", name="winner")(
            winner_branch
        )

        # Enhanced method prediction branch
        method_branch = keras.layers.Dense(128, activation="relu")(x)
        method_branch = keras.layers.BatchNormalization()(method_branch)
        method_branch = keras.layers.Dropout(0.2)(method_branch)

        # Specialized method prediction layers
        method_branch = keras.layers.Dense(64, activation="relu")(method_branch)
        method_branch = keras.layers.Dropout(0.1)(method_branch)

        # Enhanced method prediction with residual connections
        method_residual = keras.layers.Dense(64, activation="relu")(method_branch)
        method_residual = keras.layers.Dropout(0.1)(method_residual)
        method_branch = keras.layers.Add()([method_branch, method_residual])

        # Additional method-specific processing
        method_branch = keras.layers.Dense(128, activation="relu")(method_branch)
        method_branch = keras.layers.BatchNormalization()(method_branch)
        method_branch = keras.layers.Dropout(0.2)(method_branch)

        # Method output (6 classes: Red_KO/TKO, Red_Submission, Red_Decision, Blue_KO/TKO, Blue_Submission, Blue_Decision)
        method_output = keras.layers.Dense(6, activation="softmax", name="method")(
            method_branch
        )

        model = keras.Model(
            inputs=[fighter_r_input, fighter_b_input, stats_input],
            outputs=[winner_output, method_output],
        )

        # Enhanced optimizer with fixed learning rate
        model.compile(
            optimizer=keras.optimizers.Adam(learning_rate=0.001),
            loss={
                "winner": "binary_crossentropy",
                "method": "categorical_crossentropy",
            },
            loss_weights={
                "winner": 1.0,
                "method": 1.5,  # Increased weight for method prediction
            },
            metrics={"winner": "accuracy", "method": "accuracy"},
        )

        return model

    def fix_data_leakage(self, df):
        """Recalculate comprehensive fighter statistics chronologically"""
        print("Fixing data leakage with advanced feature tracking...\n")

        import copy

        df["event_date"] = pd.to_datetime(df["event_date"])
        df = df.sort_values("event_date").reset_index(drop=True)

        fighter_stats = {}

        stats_to_track = {
            "wins": 0,
            "losses": 0,
            "draws": 0,
            "sig_str_total": 0,
            "sig_str_att_total": 0,
            "sig_str_absorbed_total": 0,
            "total_str_landed": 0,
            "total_str_att": 0,
            "total_str_absorbed": 0,
            "str_def_hits": 0,
            "str_def_att": 0,
            "td_total": 0,
            "td_att_total": 0,
            "td_def_success": 0,
            "td_def_att": 0,
            "sub_att_total": 0,
            "kd_total": 0,
            "kd_absorbed_total": 0,
            "fight_time_minutes": 0,
            "fight_count": 0,
            "ko_wins": 0,
            "sub_wins": 0,
            "dec_wins": 0,
            "ko_losses": 0,
            "sub_losses": 0,
            "recent_wins": [],
            "recent_finishes": [],
            "last_fight_date": None,
            "head_pct_sum": 0,
            "body_pct_sum": 0,
            "leg_pct_sum": 0,
            "location_fight_count": 0,
            "distance_pct_sum": 0,
            "clinch_pct_sum": 0,
            "ground_pct_sum": 0,
            "position_fight_count": 0,
            # Trajectory tracking
            "slpm_history": [],
            "td_avg_history": [],
            "win_history": [],
        }

        for prefix in ["r", "b"]:
            for stat in [
                "wins",
                "losses",
                "draws",
                "win_loss_ratio",
                "pro_SLpM",
                "pro_sig_str_acc",
                "pro_SApM",
                "pro_str_def",
                "pro_total_str_pM",
                "pro_total_str_acc",
                "pro_total_str_absorbed_pM",
                "pro_td_avg",
                "pro_td_acc",
                "pro_td_def",
                "pro_sub_avg",
                "pro_kd_pM",
                "ko_rate",
                "sub_rate",
                "dec_rate",
                "recent_form",
                "head_pct",
                "body_pct",
                "leg_pct",
                "distance_pct",
                "clinch_pct",
                "ground_pct",
                "win_streak",
                "loss_streak",
                "last_5_wins",
                "days_since_last_fight",
                "recent_finish_rate",
                "durability",
                "fight_time_minutes",
                "slpm_trend",
                "td_avg_trend",
                "age_adjusted_performance",
                "peak_indicator",
            ]:
                df[f"{prefix}_{stat}_corrected"] = 0.0

        df["h2h_advantage"] = 0.0
        fighter_h2h = {}

        for idx, row in df.iterrows():
            if idx % 500 == 0:
                print(f"   Processing fight {idx}/{len(df)}...")

            r_fighter, b_fighter = row["r_fighter"], row["b_fighter"]

            # Head-to-head tracking
            h2h_key = (r_fighter, b_fighter)
            h2h_key_reverse = (b_fighter, r_fighter)

            if h2h_key in fighter_h2h:
                df.at[idx, "h2h_advantage"] = fighter_h2h[h2h_key]
            elif h2h_key_reverse in fighter_h2h:
                df.at[idx, "h2h_advantage"] = -fighter_h2h[h2h_key_reverse]

            if r_fighter not in fighter_stats:
                fighter_stats[r_fighter] = copy.deepcopy(stats_to_track)
            if b_fighter not in fighter_stats:
                fighter_stats[b_fighter] = copy.deepcopy(stats_to_track)

            for fighter, prefix in [(r_fighter, "r"), (b_fighter, "b")]:
                stats = fighter_stats[fighter]
                df.at[idx, f"{prefix}_wins_corrected"] = stats["wins"]
                df.at[idx, f"{prefix}_losses_corrected"] = stats["losses"]
                df.at[idx, f"{prefix}_draws_corrected"] = stats["draws"]
                df.at[idx, f"{prefix}_win_loss_ratio_corrected"] = stats["wins"] / max(
                    stats["losses"], 1
                )

                # Durability
                total_losses = stats["losses"]
                finish_losses = stats["ko_losses"] + stats["sub_losses"]
                df.at[idx, f"{prefix}_durability_corrected"] = (
                    1.0 / (1 + finish_losses) if total_losses > 0 else 1.0
                )

                total_fights = stats["wins"] + stats["losses"]
                if total_fights > 0:
                    df.at[idx, f"{prefix}_ko_rate_corrected"] = (
                        stats["ko_wins"] / total_fights
                    )
                    df.at[idx, f"{prefix}_sub_rate_corrected"] = (
                        stats["sub_wins"] / total_fights
                    )
                    df.at[idx, f"{prefix}_dec_rate_corrected"] = (
                        stats["dec_wins"] / total_fights
                    )

                # Recent finish rate
                if len(stats["recent_finishes"]) >= 10:
                    df.at[idx, f"{prefix}_recent_finish_rate_corrected"] = (
                        sum(stats["recent_finishes"][-10:]) / 10
                    )
                elif len(stats["recent_finishes"]) >= 8:
                    df.at[idx, f"{prefix}_recent_finish_rate_corrected"] = (
                        sum(stats["recent_finishes"][-8:]) / 8
                    )
                elif len(stats["recent_finishes"]) >= 6:
                    df.at[idx, f"{prefix}_recent_finish_rate_corrected"] = (
                        sum(stats["recent_finishes"][-6:]) / 6
                    )
                elif len(stats["recent_finishes"]) >= 4:
                    df.at[idx, f"{prefix}_recent_finish_rate_corrected"] = (
                        sum(stats["recent_finishes"][-4:]) / 4
                    )
                elif len(stats["recent_finishes"]) >= 2:
                    df.at[idx, f"{prefix}_recent_finish_rate_corrected"] = (
                        sum(stats["recent_finishes"][-2:]) / 2
                    )

                # Recent form
                if len(stats["recent_wins"]) >= 10:
                    df.at[idx, f"{prefix}_recent_form_corrected"] = (
                        sum(stats["recent_wins"][-10:]) / 10
                    )
                elif len(stats["recent_wins"]) >= 8:
                    df.at[idx, f"{prefix}_recent_form_corrected"] = (
                        sum(stats["recent_wins"][-8:]) / 8
                    )
                elif len(stats["recent_wins"]) >= 6:
                    df.at[idx, f"{prefix}_recent_form_corrected"] = (
                        sum(stats["recent_wins"][-6:]) / 6
                    )
                elif len(stats["recent_wins"]) >= 4:
                    df.at[idx, f"{prefix}_recent_form_corrected"] = (
                        sum(stats["recent_wins"][-4:]) / 4
                    )
                elif len(stats["recent_wins"]) >= 2:
                    df.at[idx, f"{prefix}_recent_form_corrected"] = (
                        sum(stats["recent_wins"][-2:]) / 2
                    )

                # Momentum features
                df.at[idx, f"{prefix}_win_streak_corrected"] = self.calculate_streak(
                    stats["recent_wins"], True
                )
                df.at[idx, f"{prefix}_loss_streak_corrected"] = self.calculate_streak(
                    stats["recent_wins"], False
                )

                if len(stats["recent_wins"]) >= 10:
                    df.at[idx, f"{prefix}_last_5_wins_corrected"] = sum(
                        stats["recent_wins"][-10:]
                    )
                elif len(stats["recent_wins"]) >= 8:
                    df.at[idx, f"{prefix}_last_5_wins_corrected"] = sum(
                        stats["recent_wins"][-8:]
                    )
                elif len(stats["recent_wins"]) >= 6:
                    df.at[idx, f"{prefix}_last_5_wins_corrected"] = sum(
                        stats["recent_wins"][-6:]
                    )
                elif len(stats["recent_wins"]) >= 4:
                    df.at[idx, f"{prefix}_last_5_wins_corrected"] = sum(
                        stats["recent_wins"][-4:]
                    )
                elif len(stats["recent_wins"]) >= 2:
                    df.at[idx, f"{prefix}_last_5_wins_corrected"] = sum(
                        stats["recent_wins"][-2:]
                    )

                # Days since last fight
                if stats["last_fight_date"]:
                    days_off = (row["event_date"] - stats["last_fight_date"]).days
                    df.at[idx, f"{prefix}_days_since_last_fight_corrected"] = days_off

                # Fight time
                df.at[idx, f"{prefix}_fight_time_minutes_corrected"] = stats[
                    "fight_time_minutes"
                ]

                if stats["fight_time_minutes"] > 0:
                    df.at[idx, f"{prefix}_pro_SLpM_corrected"] = (
                        stats["sig_str_total"] / stats["fight_time_minutes"]
                    )
                    df.at[idx, f"{prefix}_pro_SApM_corrected"] = (
                        stats["sig_str_absorbed_total"] / stats["fight_time_minutes"]
                    )
                    df.at[idx, f"{prefix}_pro_total_str_pM_corrected"] = (
                        stats["total_str_landed"] / stats["fight_time_minutes"]
                    )
                    df.at[idx, f"{prefix}_pro_total_str_absorbed_pM_corrected"] = (
                        stats["total_str_absorbed"] / stats["fight_time_minutes"]
                    )
                    df.at[idx, f"{prefix}_pro_td_avg_corrected"] = (
                        stats["td_total"] / stats["fight_time_minutes"]
                    ) * 15
                    df.at[idx, f"{prefix}_pro_sub_avg_corrected"] = (
                        stats["sub_att_total"] / stats["fight_time_minutes"]
                    ) * 15
                    df.at[idx, f"{prefix}_pro_kd_pM_corrected"] = (
                        stats["kd_total"] / stats["fight_time_minutes"]
                    )

                if stats["sig_str_att_total"] > 0:
                    df.at[idx, f"{prefix}_pro_sig_str_acc_corrected"] = (
                        stats["sig_str_total"] / stats["sig_str_att_total"]
                    )
                if stats["total_str_att"] > 0:
                    df.at[idx, f"{prefix}_pro_total_str_acc_corrected"] = (
                        stats["total_str_landed"] / stats["total_str_att"]
                    )
                if stats["str_def_att"] > 0:
                    df.at[idx, f"{prefix}_pro_str_def_corrected"] = (
                        stats["str_def_hits"] / stats["str_def_att"]
                    )
                if stats["td_att_total"] > 0:
                    df.at[idx, f"{prefix}_pro_td_acc_corrected"] = (
                        stats["td_total"] / stats["td_att_total"]
                    )
                if stats["td_def_att"] > 0:
                    df.at[idx, f"{prefix}_pro_td_def_corrected"] = (
                        stats["td_def_success"] / stats["td_def_att"]
                    )

                # Strike location percentages
                if stats["location_fight_count"] > 0:
                    df.at[idx, f"{prefix}_head_pct_corrected"] = (
                        stats["head_pct_sum"] / stats["location_fight_count"]
                    )
                    df.at[idx, f"{prefix}_body_pct_corrected"] = (
                        stats["body_pct_sum"] / stats["location_fight_count"]
                    )
                    df.at[idx, f"{prefix}_leg_pct_corrected"] = (
                        stats["leg_pct_sum"] / stats["location_fight_count"]
                    )

                if stats["position_fight_count"] > 0:
                    df.at[idx, f"{prefix}_distance_pct_corrected"] = (
                        stats["distance_pct_sum"] / stats["position_fight_count"]
                    )
                    df.at[idx, f"{prefix}_clinch_pct_corrected"] = (
                        stats["clinch_pct_sum"] / stats["position_fight_count"]
                    )
                    df.at[idx, f"{prefix}_ground_pct_corrected"] = (
                        stats["ground_pct_sum"] / stats["position_fight_count"]
                    )

                # TRAJECTORY FEATURES
                df.at[idx, f"{prefix}_slpm_trend_corrected"] = (
                    self.calculate_trajectory(
                        stats["slpm_history"][-5:]
                        if len(stats["slpm_history"]) >= 5
                        else stats["slpm_history"]
                    )
                )
                df.at[idx, f"{prefix}_td_avg_trend_corrected"] = (
                    self.calculate_trajectory(
                        stats["td_avg_history"][-5:]
                        if len(stats["td_avg_history"]) >= 5
                        else stats["td_avg_history"]
                    )
                )

                # Age-adjusted performance
                age = (
                    row[f"{prefix}_age_at_event"]
                    if pd.notna(row.get(f"{prefix}_age_at_event"))
                    else 30
                )
                age_curve = self.calculate_age_curve_factor(age)
                recent_form = df.at[idx, f"{prefix}_recent_form_corrected"]
                df.at[idx, f"{prefix}_age_adjusted_performance_corrected"] = (
                    recent_form * age_curve
                )

                # Peak indicator
                df.at[idx, f"{prefix}_peak_indicator_corrected"] = (
                    self.is_fighter_at_peak(
                        stats["wins"], stats["losses"], age, recent_form
                    )
                )

            # Update stats after fight
            if pd.notna(row["winner"]):
                fight_time_min = (
                    row["total_fight_time_sec"] / 60
                    if pd.notna(row["total_fight_time_sec"])
                    else 0
                )
                method = str(row["method"]).lower()
                method_cat = (
                    "ko"
                    if "ko" in method or "tko" in method
                    else "sub"
                    if "sub" in method
                    else "dec"
                )
                is_finish = method_cat in ["ko", "sub"]

                if row["winner"] == "Red":
                    fighter_stats[r_fighter]["wins"] += 1
                    fighter_stats[b_fighter]["losses"] += 1
                    fighter_stats[r_fighter][f"{method_cat}_wins"] += 1
                    fighter_stats[r_fighter]["recent_wins"].append(1)
                    fighter_stats[b_fighter]["recent_wins"].append(0)
                    fighter_stats[r_fighter]["recent_finishes"].append(
                        1 if is_finish else 0
                    )
                    fighter_stats[b_fighter]["recent_finishes"].append(0)
                    if is_finish:
                        fighter_stats[b_fighter][f"{method_cat}_losses"] += 1
                    fighter_h2h[(r_fighter, b_fighter)] = (
                        fighter_h2h.get((r_fighter, b_fighter), 0) + 1
                    )
                elif row["winner"] == "Blue":
                    fighter_stats[b_fighter]["wins"] += 1
                    fighter_stats[r_fighter]["losses"] += 1
                    fighter_stats[b_fighter][f"{method_cat}_wins"] += 1
                    fighter_stats[b_fighter]["recent_wins"].append(1)
                    fighter_stats[r_fighter]["recent_wins"].append(0)
                    fighter_stats[b_fighter]["recent_finishes"].append(
                        1 if is_finish else 0
                    )
                    fighter_stats[r_fighter]["recent_finishes"].append(0)
                    if is_finish:
                        fighter_stats[r_fighter][f"{method_cat}_losses"] += 1
                    fighter_h2h[(r_fighter, b_fighter)] = (
                        fighter_h2h.get((r_fighter, b_fighter), 0) - 1
                    )

                for fighter in [r_fighter, b_fighter]:
                    if len(fighter_stats[fighter]["recent_wins"]) > 10:
                        fighter_stats[fighter]["recent_wins"] = fighter_stats[fighter][
                            "recent_wins"
                        ][-10:]
                    if len(fighter_stats[fighter]["recent_finishes"]) > 10:
                        fighter_stats[fighter]["recent_finishes"] = fighter_stats[
                            fighter
                        ]["recent_finishes"][-10:]
                    fighter_stats[fighter]["last_fight_date"] = row["event_date"]

                for fighter, f_prefix, opp_prefix in [
                    (r_fighter, "r", "b"),
                    (b_fighter, "b", "r"),
                ]:
                    for stat_pair in [
                        ("sig_str", "sig_str_total"),
                        ("str", "total_str_landed"),
                    ]:
                        if pd.notna(row[f"{f_prefix}_{stat_pair[0]}"]):
                            fighter_stats[fighter][stat_pair[1]] += row[
                                f"{f_prefix}_{stat_pair[0]}"
                            ]

                    if pd.notna(row[f"{opp_prefix}_sig_str"]):
                        fighter_stats[fighter]["sig_str_absorbed_total"] += row[
                            f"{opp_prefix}_sig_str"
                        ]
                    if pd.notna(row[f"{opp_prefix}_str"]):
                        fighter_stats[fighter]["total_str_absorbed"] += row[
                            f"{opp_prefix}_str"
                        ]

                    for att in ["sig_str_att", "str_att", "td_att", "sub_att"]:
                        col = f"{f_prefix}_{att}"
                        if pd.notna(row.get(col)):
                            key = (
                                att + "_total" if att != "str_att" else "total_str_att"
                            )
                            fighter_stats[fighter][key] += row[col]

                    if pd.notna(row[f"{f_prefix}_td"]):
                        fighter_stats[fighter]["td_total"] += row[f"{f_prefix}_td"]

                    if pd.notna(row.get(f"{f_prefix}_kd")):
                        fighter_stats[fighter]["kd_total"] += row[f"{f_prefix}_kd"]
                    if pd.notna(row.get(f"{opp_prefix}_kd")):
                        fighter_stats[fighter]["kd_absorbed_total"] += row[
                            f"{opp_prefix}_kd"
                        ]

                    if pd.notna(row[f"{opp_prefix}_sig_str_att"]):
                        fighter_stats[fighter]["str_def_att"] += row[
                            f"{opp_prefix}_sig_str_att"
                        ]
                        if pd.notna(row[f"{opp_prefix}_sig_str"]):
                            fighter_stats[fighter]["str_def_hits"] += (
                                row[f"{opp_prefix}_sig_str_att"]
                                - row[f"{opp_prefix}_sig_str"]
                            )

                    if pd.notna(row[f"{opp_prefix}_td_att"]):
                        fighter_stats[fighter]["td_def_att"] += row[
                            f"{opp_prefix}_td_att"
                        ]
                        if pd.notna(row[f"{opp_prefix}_td"]):
                            fighter_stats[fighter]["td_def_success"] += (
                                row[f"{opp_prefix}_td_att"] - row[f"{opp_prefix}_td"]
                            )

                    # Strike location
                    if (
                        pd.notna(row.get(f"{f_prefix}_head"))
                        or pd.notna(row.get(f"{f_prefix}_body"))
                        or pd.notna(row.get(f"{f_prefix}_leg"))
                    ):
                        if pd.notna(row.get(f"{f_prefix}_head")):
                            fighter_stats[fighter]["head_pct_sum"] += row[
                                f"{f_prefix}_head"
                            ]
                        if pd.notna(row.get(f"{f_prefix}_body")):
                            fighter_stats[fighter]["body_pct_sum"] += row[
                                f"{f_prefix}_body"
                            ]
                        if pd.notna(row.get(f"{f_prefix}_leg")):
                            fighter_stats[fighter]["leg_pct_sum"] += row[
                                f"{f_prefix}_leg"
                            ]
                        fighter_stats[fighter]["location_fight_count"] += 1

                    if (
                        pd.notna(row.get(f"{f_prefix}_distance"))
                        or pd.notna(row.get(f"{f_prefix}_clinch"))
                        or pd.notna(row.get(f"{f_prefix}_ground"))
                    ):
                        if pd.notna(row.get(f"{f_prefix}_distance")):
                            fighter_stats[fighter]["distance_pct_sum"] += row[
                                f"{f_prefix}_distance"
                            ]
                        if pd.notna(row.get(f"{f_prefix}_clinch")):
                            fighter_stats[fighter]["clinch_pct_sum"] += row[
                                f"{f_prefix}_clinch"
                            ]
                        if pd.notna(row.get(f"{f_prefix}_ground")):
                            fighter_stats[fighter]["ground_pct_sum"] += row[
                                f"{f_prefix}_ground"
                            ]
                        fighter_stats[fighter]["position_fight_count"] += 1

                    fighter_stats[fighter]["fight_time_minutes"] += fight_time_min
                    fighter_stats[fighter]["fight_count"] += 1

                    # Track trajectory
                    current_slpm = df.at[idx, f"{f_prefix}_pro_SLpM_corrected"]
                    current_td_avg = df.at[idx, f"{f_prefix}_pro_td_avg_corrected"]
                    fighter_stats[fighter]["slpm_history"].append(current_slpm)
                    fighter_stats[fighter]["td_avg_history"].append(current_td_avg)

                    if len(fighter_stats[fighter]["slpm_history"]) > 10:
                        fighter_stats[fighter]["slpm_history"] = fighter_stats[fighter][
                            "slpm_history"
                        ][-10:]
                    if len(fighter_stats[fighter]["td_avg_history"]) > 10:
                        fighter_stats[fighter]["td_avg_history"] = fighter_stats[
                            fighter
                        ]["td_avg_history"][-10:]

        diff_stats = [
            "wins",
            "losses",
            "draws",
            "win_loss_ratio",
            "pro_SLpM",
            "pro_sig_str_acc",
            "pro_SApM",
            "pro_str_def",
            "pro_total_str_pM",
            "pro_total_str_acc",
            "pro_total_str_absorbed_pM",
            "pro_td_avg",
            "pro_td_acc",
            "pro_td_def",
            "pro_sub_avg",
            "pro_kd_pM",
            "ko_rate",
            "sub_rate",
            "dec_rate",
            "recent_form",
            "head_pct",
            "body_pct",
            "leg_pct",
            "distance_pct",
            "clinch_pct",
            "ground_pct",
            "win_streak",
            "loss_streak",
            "last_5_wins",
            "days_since_last_fight",
            "recent_finish_rate",
            "durability",
            "fight_time_minutes",
            "slpm_trend",
            "td_avg_trend",
            "age_adjusted_performance",
            "peak_indicator",
        ]

        for stat in diff_stats:
            df[f"{stat}_diff_corrected"] = (
                df[f"r_{stat}_corrected"] - df[f"b_{stat}_corrected"]
            )

        return df

    def prepare_features(self, df):
        """Prepare enhanced features with all advanced metrics"""
        # Check feature cache first
        cache_key = f"{len(df)}_{hash(str(df.columns.tolist()))}"
        if cache_key in self.feature_cache:
            print("Using cached features...")
            return self.feature_cache[cache_key]
        
        # Ensure consistent random state for feature preparation
        self.set_random_seeds()

        print("\nPreparing advanced features...")

        df = df[df["winner"].isin(["Red", "Blue"])].copy()

        method_mapping = {
            "KO/TKO": "KO/TKO",
            "Submission": "Submission",
            "Decision - Unanimous": "Decision",
            "Decision - Split": "Decision",
            "Decision - Majority": "Decision",
            "TKO - Doctor's Stoppage": "KO/TKO",
            "Could Not Continue": "KO/TKO",
            "DQ": "Decision",
            "Overturned": "Decision",
        }

        df["method_simple"] = df["method"].map(method_mapping).fillna("Decision")
        df["winner_method_simple"] = df["winner"] + "_" + df["method_simple"]

        feature_columns = [
            "height_diff",
            "reach_diff",
            "weight_diff",
            "age_at_event_diff",
            "ape_index_diff",
            "wins_diff_corrected",
            "losses_diff_corrected",
            "win_loss_ratio_diff_corrected",
            "pro_SLpM_diff_corrected",
            "pro_sig_str_acc_diff_corrected",
            "pro_SApM_diff_corrected",
            "pro_str_def_diff_corrected",
            "pro_total_str_pM_diff_corrected",
            "pro_total_str_acc_diff_corrected",
            "pro_total_str_absorbed_pM_diff_corrected",
            "pro_td_avg_diff_corrected",
            "pro_td_acc_diff_corrected",
            "pro_td_def_diff_corrected",
            "pro_sub_avg_diff_corrected",
            "pro_kd_pM_diff_corrected",
            "ko_rate_diff_corrected",
            "sub_rate_diff_corrected",
            "dec_rate_diff_corrected",
            "recent_form_diff_corrected",
            "head_pct_diff_corrected",
            "body_pct_diff_corrected",
            "leg_pct_diff_corrected",
            "distance_pct_diff_corrected",
            "clinch_pct_diff_corrected",
            "ground_pct_diff_corrected",
            "win_streak_diff_corrected",
            "loss_streak_diff_corrected",
            "last_5_wins_diff_corrected",
            "days_since_last_fight_diff_corrected",
            "recent_finish_rate_diff_corrected",
            "durability_diff_corrected",
            "slpm_trend_diff_corrected",
            "td_avg_trend_diff_corrected",
            "age_adjusted_performance_diff_corrected",
            "peak_indicator_diff_corrected",
            "h2h_advantage",
            "total_rounds",
            "is_title_bout",
        ]

        # Core derived features
        df["net_striking_advantage"] = (
            df["r_pro_SLpM_corrected"] - df["r_pro_SApM_corrected"]
        ) - (df["b_pro_SLpM_corrected"] - df["b_pro_SApM_corrected"])

        df["striking_efficiency"] = (
            df["r_pro_SLpM_corrected"] * df["r_pro_sig_str_acc_corrected"]
        ) - (df["b_pro_SLpM_corrected"] * df["b_pro_sig_str_acc_corrected"])

        df["defensive_striking"] = (
            df["r_pro_str_def_corrected"] - df["r_pro_SApM_corrected"]
        ) - (df["b_pro_str_def_corrected"] - df["b_pro_SApM_corrected"])

        df["grappling_control"] = (
            df["r_pro_td_avg_corrected"] * df["r_pro_td_acc_corrected"]
        ) - (df["b_pro_td_avg_corrected"] * df["b_pro_td_acc_corrected"])

        df["grappling_defense"] = (
            df["r_pro_td_def_corrected"] - df["r_pro_sub_avg_corrected"] / 5
        ) - (df["b_pro_td_def_corrected"] - df["b_pro_sub_avg_corrected"] / 5)

        df["offensive_output"] = (
            df["r_pro_SLpM_corrected"]
            + df["r_pro_td_avg_corrected"]
            + df["r_pro_sub_avg_corrected"]
        ) - (
            df["b_pro_SLpM_corrected"]
            + df["b_pro_td_avg_corrected"]
            + df["b_pro_sub_avg_corrected"]
        )

        df["defensive_composite"] = (
            (df["r_pro_str_def_corrected"] + df["r_pro_td_def_corrected"]) / 2
        ) - ((df["b_pro_str_def_corrected"] + df["b_pro_td_def_corrected"]) / 2)

        df["ko_specialist_gap"] = (
            df["r_ko_rate_corrected"] * df["r_pro_SLpM_corrected"]
        ) - (df["b_ko_rate_corrected"] * df["b_pro_SLpM_corrected"])

        df["submission_specialist_gap"] = (
            df["r_sub_rate_corrected"] * df["r_pro_sub_avg_corrected"]
        ) - (df["b_sub_rate_corrected"] * df["b_pro_sub_avg_corrected"])

        # Experience features
        df["r_total_fights"] = df["r_wins_corrected"] + df["r_losses_corrected"]
        df["b_total_fights"] = df["b_wins_corrected"] + df["b_losses_corrected"]
        df["experience_gap"] = df["r_total_fights"] - df["b_total_fights"]
        df["experience_ratio"] = df["r_total_fights"] / (df["b_total_fights"] + 1)

        df["r_avg_fight_time"] = df["r_fight_time_minutes_corrected"] / (
            df["r_total_fights"] + 1
        )
        df["b_avg_fight_time"] = df["b_fight_time_minutes_corrected"] / (
            df["b_total_fights"] + 1
        )
        df["avg_fight_time_diff"] = df["r_avg_fight_time"] - df["b_avg_fight_time"]

        df["skill_momentum"] = (
            df["pro_SLpM_diff_corrected"] * df["recent_form_diff_corrected"]
        )
        df["finish_threat"] = (
            df["r_ko_rate_corrected"] + df["r_sub_rate_corrected"]
        ) - (df["b_ko_rate_corrected"] + df["b_sub_rate_corrected"])

        df["momentum_advantage"] = (
            df["win_streak_diff_corrected"] - df["loss_streak_diff_corrected"]
        )

        df["inactivity_penalty"] = np.where(
            df["days_since_last_fight_diff_corrected"] > 365,
            -1,
            np.where(df["days_since_last_fight_diff_corrected"] < -365, 1, 0),
        )

        df["pace_differential"] = (
            df["r_pro_SLpM_corrected"] + df["r_pro_td_avg_corrected"]
        ) - (df["b_pro_SLpM_corrected"] + df["b_pro_td_avg_corrected"])

        # Enhanced experience features
        df["r_elite_fight_ratio"] = np.where(
            df["r_total_fights"] > 0, df["r_wins_corrected"] / df["r_total_fights"], 0
        )
        df["b_elite_fight_ratio"] = np.where(
            df["b_total_fights"] > 0, df["b_wins_corrected"] / df["b_total_fights"], 0
        )

        df["quality_experience_gap"] = (
            df["r_total_fights"] * df["r_elite_fight_ratio"]
            - df["b_total_fights"] * df["b_elite_fight_ratio"]
        )

        df["r_championship_experience"] = np.where(
            df["r_total_fights"] > 0,
            (df["r_fight_time_minutes_corrected"] / df["r_total_fights"]) > 12,
            0,
        ).astype(int)
        df["b_championship_experience"] = np.where(
            df["b_total_fights"] > 0,
            (df["b_fight_time_minutes_corrected"] / df["b_total_fights"]) > 12,
            0,
        ).astype(int)
        df["championship_exp_diff"] = (
            df["r_championship_experience"] - df["b_championship_experience"]
        )

        df["r_adversity_experience"] = df["r_losses_corrected"] * (
            df["r_win_loss_ratio_corrected"] - 1
        )
        df["b_adversity_experience"] = df["b_losses_corrected"] * (
            df["b_win_loss_ratio_corrected"] - 1
        )
        df["adversity_exp_diff"] = (
            df["r_adversity_experience"] - df["b_adversity_experience"]
        )

        df["experience_skill_interaction"] = (
            df["experience_gap"] * df["pro_SLpM_diff_corrected"] / 50
        )

        df["veteran_edge"] = np.where(
            np.abs(df["experience_gap"]) > 5,
            df["experience_gap"] * df["recent_form_diff_corrected"],
            0,
        )

        df["novice_vulnerability"] = np.where(
            (df["r_total_fights"] < 10) | (df["b_total_fights"] < 10),
            -np.abs(df["experience_gap"]) * 0.5,
            0,
        )

        # Experience gap historical win rate
        bins = [-np.inf, -10, -5, 0, 5, 10, np.inf]
        bucket_indices = np.digitize(df["experience_gap"], bins[1:-1])
        bucket_labels = [
            "huge_blue",
            "mod_blue",
            "slight_blue",
            "slight_red",
            "mod_red",
            "huge_red",
        ]
        df["exp_gap_bucket_temp"] = [bucket_labels[i] for i in bucket_indices]

        exp_gap_win_rate = (
            df.groupby("exp_gap_bucket_temp")["winner"]
            .apply(lambda x: (x == "Red").mean())
            .to_dict()
        )

        df["exp_gap_historical_win_rate"] = (
            df["exp_gap_bucket_temp"].map(exp_gap_win_rate).fillna(0.5)
        )
        df.drop("exp_gap_bucket_temp", axis=1, inplace=True)

        # STYLE MATCHUP FEATURES
        df["r_striker_score"] = (
            df["r_distance_pct_corrected"] * 1.2
            + df["r_head_pct_corrected"] * 0.8
            - df["r_ground_pct_corrected"] * 1.0
        )
        df["b_striker_score"] = (
            df["b_distance_pct_corrected"] * 1.2
            + df["b_head_pct_corrected"] * 0.8
            - df["b_ground_pct_corrected"] * 1.0
        )
        df["striker_advantage"] = df["r_striker_score"] - df["b_striker_score"]

        df["r_grappler_score"] = (
            df["r_pro_td_avg_corrected"] * 0.4
            + df["r_pro_sub_avg_corrected"] * 0.3
            + df["r_ground_pct_corrected"] * 0.3
        )
        df["b_grappler_score"] = (
            df["b_pro_td_avg_corrected"] * 0.4
            + df["b_pro_sub_avg_corrected"] * 0.3
            + df["b_ground_pct_corrected"] * 0.3
        )
        df["grappler_advantage"] = df["r_grappler_score"] - df["b_grappler_score"]

        # Reach advantage in striking context
        df["effective_reach_advantage"] = np.where(
            df["striker_advantage"] > 0.2, df["reach_diff"] * 1.5, df["reach_diff"]
        )

        # Stance matchup
        if "r_stance" in df.columns and "b_stance" in df.columns:
            if "stance_encoder" not in self.label_encoders:
                self.label_encoders["stance_encoder"] = LabelEncoder()
                all_stances = pd.concat([df["r_stance"], df["b_stance"]]).unique()
                self.label_encoders["stance_encoder"].fit(all_stances)

            df["r_stance_encoded"] = self.label_encoders["stance_encoder"].transform(
                df["r_stance"].fillna("Orthodox")
            )
            df["b_stance_encoded"] = self.label_encoders["stance_encoder"].transform(
                df["b_stance"].fillna("Orthodox")
            )
            df["stance_diff"] = df["r_stance_encoded"] - df["b_stance_encoded"]

            # Orthodox vs Southpaw advantage
            df["orthodox_southpaw_matchup"] = np.where(
                (df["r_stance"] == "Orthodox") & (df["b_stance"] == "Southpaw"),
                0.05,
                np.where(
                    (df["r_stance"] == "Southpaw") & (df["b_stance"] == "Orthodox"),
                    -0.05,
                    0,
                ),
            )

            feature_columns.extend(["stance_diff", "orthodox_southpaw_matchup"])

        # ENHANCED FEATURE ENGINEERING FOR METHOD PREDICTION

        # Fighter-specific finishing pattern analysis
        df["r_ko_specialist"] = np.where(
            (df["r_ko_rate_corrected"] > 0.4) & (df["r_pro_SLpM_corrected"] > 4.0), 1, 0
        )
        df["b_ko_specialist"] = np.where(
            (df["b_ko_rate_corrected"] > 0.4) & (df["b_pro_SLpM_corrected"] > 4.0), 1, 0
        )
        df["ko_specialist_matchup"] = df["r_ko_specialist"] - df["b_ko_specialist"]

        df["r_sub_specialist"] = np.where(
            (df["r_sub_rate_corrected"] > 0.3) & (df["r_pro_td_avg_corrected"] > 2.0),
            1,
            0,
        )
        df["b_sub_specialist"] = np.where(
            (df["b_sub_rate_corrected"] > 0.3) & (df["b_pro_td_avg_corrected"] > 2.0),
            1,
            0,
        )
        df["sub_specialist_matchup"] = df["r_sub_specialist"] - df["b_sub_specialist"]

        # Durability and chin analysis
        df["r_durability_score"] = (
            df["r_durability_corrected"] * 0.4
            + (1 - df["r_pro_SApM_corrected"] / 6.0) * 0.3
            + (df["r_pro_str_def_corrected"]) * 0.3
        )
        df["b_durability_score"] = (
            df["b_durability_corrected"] * 0.4
            + (1 - df["b_pro_SApM_corrected"] / 6.0) * 0.3
            + (df["b_pro_str_def_corrected"]) * 0.3
        )
        df["durability_advantage"] = df["r_durability_score"] - df["b_durability_score"]

        # Knockdown resistance analysis
        df["r_kd_resistance"] = 1 - (df["r_pro_SApM_corrected"] / 6.0) * (
            1 - df["r_pro_str_def_corrected"]
        )
        df["b_kd_resistance"] = 1 - (df["b_pro_SApM_corrected"] / 6.0) * (
            1 - df["b_pro_str_def_corrected"]
        )
        df["kd_resistance_advantage"] = df["r_kd_resistance"] - df["b_kd_resistance"]

        # Style matchup analysis for method prediction
        df["striker_vs_grappler"] = np.where(
            (df["striker_advantage"] > 0.3) & (df["grappler_advantage"] < -0.3),
            1,
            np.where(
                (df["striker_advantage"] < -0.3) & (df["grappler_advantage"] > 0.3),
                -1,
                0,
            ),
        )

        # Distance vs ground fighting preference
        df["distance_ground_preference"] = (
            df["r_distance_pct_corrected"] - df["r_ground_pct_corrected"]
        ) - (df["b_distance_pct_corrected"] - df["b_ground_pct_corrected"])

        # Clinch fighting advantage
        df["clinch_advantage"] = (
            df["r_clinch_pct_corrected"] * df["r_pro_SLpM_corrected"]
        ) - (df["b_clinch_pct_corrected"] * df["b_pro_SLpM_corrected"])

        # Head hunting vs body work preference
        df["head_hunting_advantage"] = (
            df["r_head_pct_corrected"] * df["r_pro_SLpM_corrected"]
        ) - (df["b_head_pct_corrected"] * df["b_pro_SLpM_corrected"])

        # Opponent vulnerability scoring
        df["r_opponent_vulnerability"] = (
            (1 - df["b_pro_str_def_corrected"]) * 0.3
            + (df["b_pro_SApM_corrected"] / 6.0) * 0.2
            + (1 - df["b_durability_corrected"]) * 0.3
            + (df["b_pro_kd_pM_corrected"] / 0.5) * 0.2
        )
        df["b_opponent_vulnerability"] = (
            (1 - df["r_pro_str_def_corrected"]) * 0.3
            + (df["r_pro_SApM_corrected"] / 6.0) * 0.2
            + (1 - df["r_durability_corrected"]) * 0.3
            + (df["r_pro_kd_pM_corrected"] / 0.5) * 0.2
        )
        df["vulnerability_advantage"] = (
            df["r_opponent_vulnerability"] - df["b_opponent_vulnerability"]
        )

        # Contextual fight factors
        df["title_fight_ko_factor"] = np.where(df["is_title_bout"] == 1, 1.2, 1.0)
        df["five_round_advantage"] = np.where(df["total_rounds"] == 5, 1.15, 1.0)

        # Recent form and momentum analysis
        df["r_momentum_score"] = (
            df["r_recent_form_corrected"] * 0.4
            + df["r_win_streak_corrected"] / 10.0 * 0.3
            + df["r_recent_finish_rate_corrected"] * 0.3
        )
        df["b_momentum_score"] = (
            df["b_recent_form_corrected"] * 0.4
            + df["b_win_streak_corrected"] / 10.0 * 0.3
            + df["b_recent_finish_rate_corrected"] * 0.3
        )
        # momentum_advantage already defined above

        # Age and experience interaction
        df["age_experience_interaction"] = (
            df["age_at_event_diff"] * df["experience_gap"] / 100.0
        )

        # Power vs technique analysis
        df["r_power_technique_ratio"] = (
            df["r_pro_SLpM_corrected"] * df["r_pro_sig_str_acc_corrected"]
        ) / (df["r_pro_SApM_corrected"] + 0.1)
        df["b_power_technique_ratio"] = (
            df["b_pro_SLpM_corrected"] * df["b_pro_sig_str_acc_corrected"]
        ) / (df["b_pro_SApM_corrected"] + 0.1)
        df["power_technique_advantage"] = (
            df["r_power_technique_ratio"] - df["b_power_technique_ratio"]
        )

        # Grappling control vs submission threat
        df["r_grappling_threat"] = (
            df["r_pro_td_avg_corrected"] * df["r_pro_td_acc_corrected"] * 0.6
            + df["r_pro_sub_avg_corrected"] * 0.4
        )
        df["b_grappling_threat"] = (
            df["b_pro_td_avg_corrected"] * df["b_pro_td_acc_corrected"] * 0.6
            + df["b_pro_sub_avg_corrected"] * 0.4
        )
        df["grappling_threat_advantage"] = (
            df["r_grappling_threat"] - df["b_grappling_threat"]
        )

        # ============================================================================
        # ADVANCED FEATURE ENGINEERING FOR HIGHER ACCURACY
        # ============================================================================

        # Fighter-specific finishing patterns over time
        df["r_recent_ko_trend"] = np.where(
            df["r_recent_finish_rate_corrected"] > df["r_ko_rate_corrected"], 1, 0
        )
        df["b_recent_ko_trend"] = np.where(
            df["b_recent_finish_rate_corrected"] > df["b_ko_rate_corrected"], 1, 0
        )
        df["recent_ko_trend_advantage"] = (
            df["r_recent_ko_trend"] - df["b_recent_ko_trend"]
        )

        # Opponent-specific finishing success rates
        df["r_opponent_ko_susceptibility"] = (
            df["b_pro_SApM_corrected"]
            * (1 - df["b_pro_str_def_corrected"])
            * (1 - df["b_durability_corrected"])
        )
        df["b_opponent_ko_susceptibility"] = (
            df["r_pro_SApM_corrected"]
            * (1 - df["r_pro_str_def_corrected"])
            * (1 - df["r_durability_corrected"])
        )
        df["ko_susceptibility_advantage"] = (
            df["r_opponent_ko_susceptibility"] - df["b_opponent_ko_susceptibility"]
        )

        # Submission opportunity analysis
        df["r_sub_opportunity_score"] = (
            df["r_pro_td_avg_corrected"]
            * df["r_pro_td_acc_corrected"]
            * df["r_sub_rate_corrected"]
            * (1 - df["b_pro_td_def_corrected"])
        )
        df["b_sub_opportunity_score"] = (
            df["b_pro_td_avg_corrected"]
            * df["b_pro_td_acc_corrected"]
            * df["b_sub_rate_corrected"]
            * (1 - df["r_pro_td_def_corrected"])
        )
        df["sub_opportunity_advantage"] = (
            df["r_sub_opportunity_score"] - df["b_sub_opportunity_score"]
        )

        # Fight pace and cardio analysis
        df["r_cardio_factor"] = (
            df["r_avg_fight_time"]
            * df["r_pro_SLpM_corrected"]
            / (df["r_pro_SApM_corrected"] + 0.1)
        )
        df["b_cardio_factor"] = (
            df["b_avg_fight_time"]
            * df["b_pro_SLpM_corrected"]
            / (df["b_pro_SApM_corrected"] + 0.1)
        )
        df["cardio_advantage"] = df["r_cardio_factor"] - df["b_cardio_factor"]

        # Weight class specific adjustments
        df["weight_class_ko_factor"] = np.where(
            df["weight_class"].isin(["Heavyweight", "Light Heavyweight"]),
            1.3,
            np.where(
                df["weight_class"].isin(["Middleweight", "Welterweight"]), 1.1, 1.0
            ),
        )

        # Championship experience and pressure
        df["r_championship_pressure"] = (
            df["r_championship_experience"]
            * df["is_title_bout"]
            * df["r_recent_form_corrected"]
        )
        df["b_championship_pressure"] = (
            df["b_championship_experience"]
            * df["is_title_bout"]
            * df["b_recent_form_corrected"]
        )
        df["championship_pressure_advantage"] = (
            df["r_championship_pressure"] - df["b_championship_pressure"]
        )

        # Recent opponent quality analysis (using available features)
        df["r_opponent_quality"] = (
            df["r_recent_form_corrected"] * 0.6
            + df["r_recent_finish_rate_corrected"] * 0.4
        )
        df["b_opponent_quality"] = (
            df["b_recent_form_corrected"] * 0.6
            + df["b_recent_finish_rate_corrected"] * 0.4
        )
        df["opponent_quality_advantage"] = (
            df["r_opponent_quality"] - df["b_opponent_quality"]
        )

        # Fight ending probability by round
        df["early_finish_probability"] = (
            (df["r_ko_rate_corrected"] + df["r_sub_rate_corrected"])
            * (df["b_ko_rate_corrected"] + df["b_sub_rate_corrected"])
            * 0.5
        )

        # Technical striking vs brawling analysis
        df["r_technical_striker"] = np.where(
            (df["r_pro_sig_str_acc_corrected"] > 0.5)
            & (df["r_pro_SLpM_corrected"] > 3.0),
            1,
            0,
        )
        df["b_technical_striker"] = np.where(
            (df["b_pro_sig_str_acc_corrected"] > 0.5)
            & (df["b_pro_SLpM_corrected"] > 3.0),
            1,
            0,
        )
        df["technical_striker_advantage"] = (
            df["r_technical_striker"] - df["b_technical_striker"]
        )

        # Clinch and dirty boxing effectiveness
        df["r_clinch_effectiveness"] = (
            df["r_clinch_pct_corrected"]
            * df["r_pro_SLpM_corrected"]
            * df["r_pro_sig_str_acc_corrected"]
        )
        df["b_clinch_effectiveness"] = (
            df["b_clinch_pct_corrected"]
            * df["b_pro_SLpM_corrected"]
            * df["b_pro_sig_str_acc_corrected"]
        )
        df["clinch_effectiveness_advantage"] = (
            df["r_clinch_effectiveness"] - df["b_clinch_effectiveness"]
        )

        # ============================================================================
        # NEW HIGH-IMPACT FEATURES FOR BETTER ACCURACY
        # ============================================================================

        # APE INDEX ADVANTAGE (Reach - Height) - Critical for striking range
        df["ape_index_advantage"] = df["ape_index_diff"]

        # ENHANCED MOMENTUM FEATURES
        df["momentum_velocity"] = (
            df["r_recent_form_corrected"] - df["b_recent_form_corrected"]
        ) * (df["r_win_streak_corrected"] - df["b_win_streak_corrected"])

        # FIGHTER STYLE MATCHUP DEPTH
        df["style_matchup_depth"] = (
            abs(df["striker_advantage"])
            + abs(df["grappler_advantage"])
            + abs(df["r_pro_SLpM_corrected"] - df["b_pro_SLpM_corrected"]) * 0.3
        )

        # CHAMPIONSHIP EXPERIENCE IMPACT
        df["championship_impact"] = (
            df["championship_exp_diff"]
            * df["is_title_bout"]
            * (df["r_recent_form_corrected"] - df["b_recent_form_corrected"])
        )

        # OPPONENT QUALITY MOMENTUM
        df["opponent_quality_momentum"] = (
            df["r_opponent_quality"] - df["b_opponent_quality"]
        ) * (df["r_recent_form_corrected"] - df["b_recent_form_corrected"])

        # FINISHING PRESSURE UNDER STRESS
        df["finishing_pressure_stress"] = (
            df["r_ko_rate_corrected"] + df["r_sub_rate_corrected"]
        ) * (
            df["r_recent_form_corrected"] * 0.6 + df["r_win_streak_corrected"] * 0.4
        ) - (df["b_ko_rate_corrected"] + df["b_sub_rate_corrected"]) * (
            df["b_recent_form_corrected"] * 0.6 + df["b_win_streak_corrected"] * 0.4
        )

        # RING RUST FACTOR - Time since last fight
        df["ring_rust_factor"] = np.where(
            df["days_since_last_fight_diff_corrected"] > 0,
            -0.1
            * (df["days_since_last_fight_diff_corrected"] / 365),  # Red corner rust
            np.where(
                df["days_since_last_fight_diff_corrected"] < 0,
                0.1
                * (
                    abs(df["days_since_last_fight_diff_corrected"]) / 365
                ),  # Blue corner rust
                0,
            ),
        )

        # RING RUST VS MOMENTUM
        df["ring_rust_vs_momentum"] = df["ring_rust_factor"] * (
            df["r_recent_form_corrected"] - df["b_recent_form_corrected"]
        )

        # WEIGHT CLASS FACTOR - Different dynamics in each division
        weight_class_factors = {
            "Heavyweight": 1.2,
            "Light Heavyweight": 1.1,
            "Middleweight": 1.0,
            "Welterweight": 0.95,
            "Lightweight": 0.9,
            "Featherweight": 0.85,
            "Bantamweight": 0.8,
            "Flyweight": 0.75,
            "Women's Bantamweight": 0.8,
            "Women's Flyweight": 0.75,
            "Women's Strawweight": 0.7,
        }
        df["weight_class_factor"] = (
            df["weight_class"].map(weight_class_factors).fillna(1.0)
        )

        # WEIGHT CLASS ADAPTATION
        df["weight_class_adaptation"] = df["weight_class_factor"] * (
            df["r_pro_SLpM_corrected"] - df["b_pro_SLpM_corrected"]
        )

        # STYLE CLASH SEVERITY - How different the fighting styles are
        df["style_clash_severity"] = abs(
            df["striker_vs_grappler"] * 0.5
            + (df["r_pro_SLpM_corrected"] - df["b_pro_SLpM_corrected"]) * 0.3
            + (df["r_pro_td_avg_corrected"] - df["b_pro_td_avg_corrected"]) * 0.2
        )

        # ENHANCED STANCE MATCHUP ADVANTAGE (including Switch stance)
        if "r_stance" in df.columns and "b_stance" in df.columns:
            # Create comprehensive stance matchup matrix
            def calculate_stance_advantage(r_stance, b_stance):
                # Based on real-world fight data analysis
                if r_stance == "Orthodox" and b_stance == "Southpaw":
                    return -0.06  # Southpaw has 6% advantage (52% vs 46%)
                elif r_stance == "Southpaw" and b_stance == "Orthodox":
                    return 0.06  # Southpaw has 6% advantage (52% vs 46%)
                elif r_stance == "Switch" and b_stance == "Orthodox":
                    return 0.05  # Switch has slight advantage
                elif r_stance == "Switch" and b_stance == "Southpaw":
                    return 0.05  # Switch has slight advantage
                elif r_stance == "Orthodox" and b_stance == "Switch":
                    return -0.05  # Switch has slight advantage
                elif r_stance == "Southpaw" and b_stance == "Switch":
                    return -0.05  # Switch has slight advantage
                else:
                    return 0  # Same stance or Switch vs Switch

            df["stance_matchup_advantage"] = df.apply(
                lambda row: calculate_stance_advantage(
                    row["r_stance"], row["b_stance"]
                ),
                axis=1,
            )

            # STANCE VERSATILITY IMPACT
            df["stance_versatility_advantage"] = np.where(
                df["r_stance"] == "Switch",
                0.05,
                np.where(df["b_stance"] == "Switch", -0.05, 0),
            )

            df["stance_versatility_impact"] = (
                df["stance_versatility_advantage"] * df["style_matchup_depth"]
            )
        else:
            df["stance_matchup_advantage"] = 0
            df["stance_versatility_advantage"] = 0
            df["stance_versatility_impact"] = 0

        # CHAMPIONSHIP PRESSURE - Title fight experience and pressure
        df["championship_pressure"] = (
            df["championship_exp_diff"]
            * df["is_title_bout"]
            * (df["r_recent_form_corrected"] - df["b_recent_form_corrected"])
        )

        # MOMENTUM SWING - Recent performance trends
        df["momentum_swing"] = (
            df["r_recent_form_corrected"]
            - df["b_recent_form_corrected"]
            + (df["r_win_streak_corrected"] - df["b_win_streak_corrected"]) * 0.1
        )

        # POWER VS TECHNIQUE - Striking power vs technical striking
        df["power_vs_technique"] = (
            df["r_pro_SLpM_corrected"] - df["b_pro_SLpM_corrected"]
        ) * 0.6 + (
            df["r_pro_sig_str_acc_corrected"] - df["b_pro_sig_str_acc_corrected"]
        ) * 0.4

        # CARDIO ADVANTAGE - Endurance and pace (already defined above)

        # FINISH PRESSURE - Finishing ability under pressure
        df["finish_pressure"] = (
            df["r_ko_rate_corrected"] + df["r_sub_rate_corrected"]
        ) * (df["r_recent_form_corrected"] + df["r_win_streak_corrected"] * 0.1) - (
            df["b_ko_rate_corrected"] + df["b_sub_rate_corrected"]
        ) * (df["b_recent_form_corrected"] + df["b_win_streak_corrected"] * 0.1)

        # OPPONENT QUALITY GAP - Quality of recent opponents
        df["opponent_quality_gap"] = df["r_opponent_quality"] - df["b_opponent_quality"]

        # RECENT OPPONENT STRENGTH - Strength of recent competition
        df["recent_opponent_strength"] = (
            df["r_recent_form_corrected"] * df["r_recent_finish_rate_corrected"]
        ) - (df["b_recent_form_corrected"] * df["b_recent_finish_rate_corrected"])

        # UPSET POTENTIAL - Likelihood of upset based on experience and form
        df["upset_potential"] = np.where(
            (df["experience_gap"] < -5) & (df["recent_form_diff_corrected"] > 0.3),
            0.3,
            np.where(
                (df["experience_gap"] > 5) & (df["recent_form_diff_corrected"] < -0.3),
                -0.3,
                0,
            ),
        )

        # Add all new features to column list
        feature_columns.extend(
            [
                "net_striking_advantage",
                "striking_efficiency",
                "defensive_striking",
                "grappling_control",
                "grappling_defense",
                "offensive_output",
                "defensive_composite",
                "ko_specialist_gap",
                "submission_specialist_gap",
                "experience_gap",
                "skill_momentum",
                "finish_threat",
                "momentum_advantage",
                "inactivity_penalty",
                "pace_differential",
                "r_total_fights",
                "b_total_fights",
                "experience_ratio",
                "avg_fight_time_diff",
                "quality_experience_gap",
                "championship_exp_diff",
                "adversity_exp_diff",
                "experience_skill_interaction",
                "veteran_edge",
                "novice_vulnerability",
                "exp_gap_historical_win_rate",
                "striker_advantage",
                "grappler_advantage",
                "effective_reach_advantage",
                "ko_specialist_matchup",
                "sub_specialist_matchup",
                "durability_advantage",
                "kd_resistance_advantage",
                "striker_vs_grappler",
                "distance_ground_preference",
                "clinch_advantage",
                "head_hunting_advantage",
                "vulnerability_advantage",
                "title_fight_ko_factor",
                "five_round_advantage",
                "age_experience_interaction",
                "power_technique_advantage",
                "grappling_threat_advantage",
                "recent_ko_trend_advantage",
                "ko_susceptibility_advantage",
                "sub_opportunity_advantage",
                "cardio_advantage",
                "weight_class_ko_factor",
                "championship_pressure_advantage",
                "opponent_quality_advantage",
                "early_finish_probability",
                "technical_striker_advantage",
                "clinch_effectiveness_advantage",
                # NEW HIGH-IMPACT FEATURES
                "ape_index_advantage",
                "stance_matchup_advantage",
                "stance_versatility_advantage",
                "weight_class_factor",
                "ring_rust_factor",
                "championship_pressure",
                "momentum_swing",
                "style_clash_severity",
                "power_vs_technique",
                "finish_pressure",
                "opponent_quality_gap",
                "recent_opponent_strength",
                "upset_potential",
                # ENHANCED ACCURACY FEATURES
                "momentum_velocity",
                "style_matchup_depth",
                "championship_impact",
                "opponent_quality_momentum",
                "finishing_pressure_stress",
                "ring_rust_vs_momentum",
                "weight_class_adaptation",
                "stance_versatility_impact",
            ]
        )

        for col in feature_columns:
            if col in df.columns:
                df[col] = df[col].fillna(0)

        df = df.replace([np.inf, -np.inf], [1e6, -1e6])

        print(f"Total features: {len(feature_columns)}")
        
        # Cache the results
        self.feature_cache[cache_key] = (df, feature_columns)
        return df, feature_columns

    def train_models(self, df):
        """Train stacked ensemble with specialized method models and deep learning"""
        # Ensure consistent random state before training
        self.set_random_seeds()

        df, feature_columns = self.prepare_features(df)

        X = df[feature_columns]
        y_winner = (df["winner"] == "Red").astype(int)

        if "winner_method_encoder" not in self.label_encoders:
            self.label_encoders["winner_method_encoder"] = LabelEncoder()
        y_method = self.label_encoders["winner_method_encoder"].fit_transform(
            df["winner_method_simple"]
        )

        (
            X_train,
            X_test,
            y_winner_train,
            y_winner_test,
            y_method_train,
            y_method_test,
        ) = train_test_split(
            X, y_winner, y_method, test_size=0.2, random_state=42, stratify=y_winner
        )

        print("\n" + "=" * 80)
        print("TRAINING ADVANCED STACKED ENSEMBLE MODEL")
        print("=" * 80)

        numeric_transformer = Pipeline(
            [
                ("imputer", SimpleImputer(strategy="median")),
                ("scaler", StandardScaler()),
            ]
        )

        preprocessor = ColumnTransformer(
            [("num", numeric_transformer, feature_columns)]
        )

        # Build base models for stacking
        base_models = []

        if HAS_XGBOOST:
            print("\n✓ XGBoost available")
            scale_pos = len(y_winner[y_winner == 0]) / max(
                len(y_winner[y_winner == 1]), 1
            )
            xgb_model = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectPercentile(f_classif, percentile=75),
                    ),  # Increased from 75 to 85
                    (
                        "classifier",
                        XGBClassifier(
                            n_estimators=600,
                            max_depth=9,
                            learning_rate=0.02,
                            subsample=0.85,
                            colsample_bytree=0.85,
                            colsample_bylevel=0.85,
                            n_jobs=-1,
                            reg_alpha=0.15,
                            reg_lambda=0.8,
                            min_child_weight=3,
                            gamma=0.1,
                            scale_pos_weight=scale_pos,
                            random_state=42,
                            eval_metric="logloss",
                            tree_method="hist",
                            seed=42,
                        ),
                    ),
                ]
            )
            base_models.append(("xgb", xgb_model))

        if HAS_LIGHTGBM:
            print("✓ LightGBM available")
            lgbm_model = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectPercentile(f_classif, percentile=75),
                    ),  # Increased from 75 to 85
                    (
                        "classifier",
                        LGBMClassifier(
                            n_estimators=600,
                            max_depth=9,
                            learning_rate=0.02,
                            num_leaves=60,
                            subsample=0.85,
                            colsample_bytree=0.85,
                            reg_alpha=0.15,
                            reg_lambda=0.8,
                            min_child_weight=3,
                            random_state=42,
                            verbose=-1,
                        ),
                    ),
                ]
            )
            base_models.append(("lgbm", lgbm_model))

        if HAS_CATBOOST:
            print("✓ CatBoost available")
            catboost_model = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectPercentile(f_classif, percentile=75),
                    ),  # Increased from 75 to 85
                    (
                        "classifier",
                        CatBoostClassifier(
                            iterations=600,
                            depth=9,
                            learning_rate=0.02,
                            l2_leaf_reg=0.8,
                            random_state=42,
                            verbose=0,
                        ),
                    ),
                ]
            )
            base_models.append(("catboost", catboost_model))

        # Random Forest (always available)
        print("✓ Random Forest available")
        rf_model = Pipeline(
            [
                ("preprocessor", preprocessor),
                (
                    "feature_selector",
                    SelectPercentile(f_classif, percentile=75),
                ),  # Increased from 75 to 85
                (
                    "classifier",
                    RandomForestClassifier(
                        n_estimators=600,
                        max_depth=18,
                        min_samples_split=5,
                        min_samples_leaf=2,
                        random_state=42,
                        n_jobs=-1,
                    ),
                ),
            ]
        )
        base_models.append(("rf", rf_model))

        # Neural Network
        if self.use_neural_net:
            print("✓ Neural Network enabled")
            nn_model = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectPercentile(f_classif, percentile=75),
                    ),  # Increased from 75 to 85
                    (
                        "classifier",
                        MLPClassifier(
                            hidden_layer_sizes=(256, 128, 64),
                            activation="relu",
                            solver="adam",
                            alpha=0.0005,
                            batch_size=32,
                            learning_rate="adaptive",
                            max_iter=400,
                            early_stopping=True,
                            random_state=42,
                        ),
                    ),
                ]
            )
            base_models.append(("nn", nn_model))

        print(f"\nBuilding stacked ensemble with {len(base_models)} base models...")

        # Enhanced meta-learner with multiple algorithms
        meta_learners = []

        # XGBoost meta-learner
        if HAS_XGBOOST:
            meta_learners.append(
                (
                    "xgb_meta",
                    XGBClassifier(
                        n_estimators=200,
                        max_depth=4,
                        learning_rate=0.05,
                        subsample=0.9,
                        colsample_bytree=0.9,
                        n_jobs=-1,  # Use all CPU cores for faster training (causes multiple windows in .exe)
                        reg_alpha=0.2,
                        reg_lambda=1,
                        random_state=42,
                        eval_metric="logloss",
                        tree_method="hist",  # Use histogram method for consistency
                        seed=42,  # Additional XGBoost seed
                    ),
                )
            )

        # LightGBM meta-learner
        if HAS_LIGHTGBM:
            meta_learners.append(
                (
                    "lgbm_meta",
                    LGBMClassifier(
                        n_estimators=200,
                        max_depth=4,
                        learning_rate=0.05,
                        num_leaves=20,
                        subsample=0.9,
                        colsample_bytree=0.9,
                        reg_alpha=0.2,
                        reg_lambda=1,
                        random_state=42,
                        verbose=-1,
                    ),
                )
            )

        # Neural Network meta-learner
        meta_learners.append(
            (
                "nn_meta",
                MLPClassifier(
                    hidden_layer_sizes=(64, 32),
                    activation="relu",
                    solver="adam",
                    alpha=0.001,
                    batch_size=64,
                    learning_rate="adaptive",
                    max_iter=300,
                    early_stopping=True,
                    random_state=42,
                ),
            )
        )

        # Logistic Regression meta-learner
        meta_learners.append(
            ("lr_meta", LogisticRegression(C=0.5, max_iter=1000, random_state=42))
        )

        # Create voting ensemble of meta-learners
        voting_meta = VotingClassifier(estimators=meta_learners, voting="soft")

        # Stack models with enhanced meta-learner
        if self.use_ensemble and len(base_models) > 1:
            self.winner_model = StackingClassifier(
                estimators=base_models,
                final_estimator=voting_meta,
                cv=5,  # Reduced for speed while maintaining quality
                n_jobs=-1,
                stack_method="predict_proba",
            )
            self.winner_model = CalibratedClassifierCV(
                self.winner_model, method="isotonic", cv=3
            )
        else:
            # Fallback to single best model
            self.winner_model = base_models[0][1]
            self.winner_model = CalibratedClassifierCV(
                self.winner_model, method="isotonic", cv=5
            )

        print("\nTraining winner prediction model...")
        self.winner_model.fit(X_train, y_winner_train)

        # ============================================================================
        # TRAIN METHOD PREDICTION MODEL
        # ============================================================================
        print("\n" + "=" * 80)
        print("TRAINING METHOD PREDICTION MODEL")
        print("=" * 80)

        print("\nTraining enhanced method prediction ensemble...")

        # Create multiple method prediction models
        method_models = []

        # XGBoost method model
        if HAS_XGBOOST:
            xgb_method = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectPercentile(f_classif, percentile=75),
                    ),  # Increased from 75 to 85
                    (
                        "classifier",
                        XGBClassifier(
                            n_estimators=600,
                            max_depth=9,
                            learning_rate=0.02,
                            subsample=0.85,
                            colsample_bytree=0.85,
                            n_jobs=-1,
                            reg_alpha=0.15,
                            reg_lambda=0.8,
                            random_state=42,
                            objective="multi:softprob",
                            tree_method="hist",
                            seed=42,
                        ),
                    ),
                ]
            )
            method_models.append(("xgb_method", xgb_method))

        # LightGBM method model
        if HAS_LIGHTGBM:
            lgbm_method = Pipeline(
                [
                    ("preprocessor", preprocessor),
                    (
                        "feature_selector",
                        SelectPercentile(f_classif, percentile=75),
                    ),  # Increased from 75 to 85
                    (
                        "classifier",
                        LGBMClassifier(
                            n_estimators=600,
                            max_depth=9,
                            learning_rate=0.02,
                            num_leaves=60,
                            subsample=0.85,
                            colsample_bytree=0.85,
                            reg_alpha=0.15,
                            reg_lambda=0.8,
                            random_state=42,
                            verbose=-1,
                        ),
                    ),
                ]
            )
            method_models.append(("lgbm_method", lgbm_method))

        # Random Forest method model
        rf_method = Pipeline(
            [
                ("preprocessor", preprocessor),
                (
                    "feature_selector",
                    SelectPercentile(f_classif, percentile=75),
                ),  # Increased from 75 to 85
                (
                    "classifier",
                    RandomForestClassifier(
                        n_estimators=600,
                        max_depth=18,
                        min_samples_split=5,
                        min_samples_leaf=2,
                        random_state=42,
                        n_jobs=-1,
                    ),
                ),
            ]
        )
        method_models.append(("rf_method", rf_method))

        # Neural Network method model
        nn_method = Pipeline(
            [
                ("preprocessor", preprocessor),
                (
                    "feature_selector",
                    SelectPercentile(f_classif, percentile=75),
                ),  # Increased from 75 to 85
                (
                    "classifier",
                    MLPClassifier(
                        hidden_layer_sizes=(128, 64, 32),
                        activation="relu",
                        solver="adam",
                        alpha=0.0005,
                        batch_size=32,
                        learning_rate="adaptive",
                        max_iter=400,
                        early_stopping=True,
                        random_state=42,
                    ),
                ),
            ]
        )
        method_models.append(("nn_method", nn_method))

        # Create voting ensemble for method prediction
        self.method_model = VotingClassifier(estimators=method_models, voting="soft")

        # Calibrate the method model
        self.method_model = CalibratedClassifierCV(
            self.method_model, method="isotonic", cv=3
        )

        self.method_model.fit(X_train, y_method_train)
        print("Enhanced method prediction ensemble training complete")

        # ============================================================================
        # TRAIN DEEP LEARNING MODEL (if enabled)
        # ============================================================================
        if self.use_deep_learning and HAS_TENSORFLOW:
            print("\n" + "=" * 80)
            print("TRAINING DEEP LEARNING MODEL WITH FIGHTER EMBEDDINGS")
            print("=" * 80)

            # Create fighter encoder
            all_fighters = pd.concat([df["r_fighter"], df["b_fighter"]]).unique()

            self.fighter_encoder = LabelEncoder()
            self.fighter_encoder.fit(all_fighters)
            self.num_fighters = len(all_fighters)

            # Encode fighters
            df["r_fighter_encoded"] = self.fighter_encoder.transform(df["r_fighter"])
            df["b_fighter_encoded"] = self.fighter_encoder.transform(df["b_fighter"])

            # Prepare data for deep learning
            X_dl_train = [
                df.loc[X_train.index, "r_fighter_encoded"].values,
                df.loc[X_train.index, "b_fighter_encoded"].values,
                preprocessor.fit_transform(X_train),
            ]

            X_dl_test = [
                df.loc[X_test.index, "r_fighter_encoded"].values,
                df.loc[X_test.index, "b_fighter_encoded"].values,
                preprocessor.transform(X_test),
            ]

            # Map method labels to 6 classes: Red_KO/TKO, Red_Submission, Red_Decision, Blue_KO/TKO, Blue_Submission, Blue_Decision
            method_map = {}
            for i, label in enumerate(
                self.label_encoders["winner_method_encoder"].classes_
            ):
                winner = label.split("_")[0]
                method_type = label.split("_")[1]

                if winner == "Red":
                    if method_type == "KO/TKO":
                        method_map[i] = 0
                    elif method_type == "Submission":
                        method_map[i] = 1
                    else:  # Decision
                        method_map[i] = 2
                else:  # Blue
                    if method_type == "KO/TKO":
                        method_map[i] = 3
                    elif method_type == "Submission":
                        method_map[i] = 4
                    else:  # Decision
                        method_map[i] = 5

            # Convert to categorical for deep learning
            from tensorflow.keras.utils import to_categorical

            y_method_train_dl = to_categorical(
                [method_map[y] for y in y_method_train], num_classes=6
            )
            y_method_test_dl = to_categorical(
                [method_map[y] for y in y_method_test], num_classes=6
            )

            # Build and train deep learning model
            self.deep_learning_model = self.build_deep_learning_model(
                num_features=preprocessor.transform(X_train).shape[1],
                num_fighters=self.num_fighters,
            )

            print("\nTraining enhanced deep learning model...")

            # Enhanced callbacks for better training
            callbacks = [
                keras.callbacks.EarlyStopping(
                    monitor="val_loss",
                    patience=15,
                    restore_best_weights=True,
                    min_delta=0.001,
                ),
                keras.callbacks.ModelCheckpoint(
                    "best_dl_model.h5",
                    monitor="val_loss",
                    save_best_only=True,
                    verbose=0,
                ),
            ]

            history = self.deep_learning_model.fit(
                X_dl_train,
                {"winner": y_winner_train.values, "method": y_method_train_dl},
                validation_split=0.2,
                epochs=80,  # Reduced for speed
                batch_size=64,
                verbose=0,
                callbacks=callbacks,
            )

            dl_results = self.deep_learning_model.evaluate(
                X_dl_test,
                {"winner": y_winner_test.values, "method": y_method_test_dl},
                verbose=0,
            )

            print(f"\nDeep Learning Results:")
            print(f"  Winner Accuracy: {dl_results[3]:.4f}")
            print(f"  Method Accuracy: {dl_results[4]:.4f}")

        # Time-based cross-validation with parallel processing
        tscv = TimeSeriesSplit(n_splits=5)
        
        def train_fold(fold_data):
            train_idx, val_idx, X, y_winner, preprocessor = fold_data
            X_fold_train, X_fold_val = X.iloc[train_idx], X.iloc[val_idx]
            y_fold_train, y_fold_val = y_winner.iloc[train_idx], y_winner.iloc[val_idx]

            if HAS_XGBOOST:
                fold_model = Pipeline(
                    [
                        ("preprocessor", preprocessor),
                        ("feature_selector", SelectPercentile(f_classif, percentile=75)),
                        (
                            "classifier",
                            XGBClassifier(
                                n_estimators=500,
                                max_depth=8,
                                learning_rate=0.025,
                                subsample=0.85,
                                colsample_bytree=0.85,
                                n_jobs=1,  # Single job for parallel processing
                                reg_alpha=0.2,
                                reg_lambda=1,
                                random_state=42,
                                eval_metric="logloss",
                                tree_method="hist",
                                seed=42,
                            ),
                        ),
                    ]
                )
            else:
                fold_model = Pipeline(
                    [
                        ("preprocessor", preprocessor),
                        ("feature_selector", SelectPercentile(f_classif, percentile=75)),
                        (
                            "classifier",
                            RandomForestClassifier(
                                n_estimators=500,
                                max_depth=15,
                                min_samples_split=6,
                                min_samples_leaf=2,
                                random_state=42,
                                n_jobs=1,  # Single job for parallel processing
                            ),
                        ),
                    ]
                )

            fold_model.fit(X_fold_train, y_fold_train)
            return fold_model.score(X_fold_val, y_fold_val)

        # Prepare fold data for parallel processing
        fold_data = [(train_idx, val_idx, X, y_winner, preprocessor) 
                     for train_idx, val_idx in tscv.split(X)]
        
        # Use parallel processing for cross-validation
        n_jobs = min(12, mp.cpu_count())  # Use up to 12 cores
        winner_cv_scores = Parallel(n_jobs=n_jobs, backend="loky")(
            delayed(train_fold)(fold) for fold in fold_data
        )

        print(f"\n{'=' * 80}")
        print(
            f"Time-Based CV Accuracy: {np.mean(winner_cv_scores):.4f} ± {np.std(winner_cv_scores):.4f}"
        )
        print(
            f"Test Set Accuracy: {self.winner_model.score(X_test, y_winner_test):.4f}"
        )
        print(f"{'=' * 80}\n")

        return feature_columns

    def get_fighter_latest_stats(self, fighter_name):
        """Get latest stats for fighter"""
        red_fights = self.df_train[
            self.df_train["r_fighter"] == fighter_name
        ].sort_values("event_date", ascending=False)
        blue_fights = self.df_train[
            self.df_train["b_fighter"] == fighter_name
        ].sort_values("event_date", ascending=False)

        if len(red_fights) > 0 and len(blue_fights) > 0:
            latest = (
                red_fights.iloc[0]
                if red_fights.iloc[0]["event_date"] > blue_fights.iloc[0]["event_date"]
                else blue_fights.iloc[0]
            )
            prefix = (
                "r"
                if red_fights.iloc[0]["event_date"] > blue_fights.iloc[0]["event_date"]
                else "b"
            )
        elif len(red_fights) > 0:
            latest, prefix = red_fights.iloc[0], "r"
        elif len(blue_fights) > 0:
            latest, prefix = blue_fights.iloc[0], "b"
        else:
            return None

        stats = {
            k: latest[f"{prefix}_{k}_corrected"]
            for k in [
                "wins",
                "losses",
                "draws",
                "win_loss_ratio",
                "pro_SLpM",
                "pro_sig_str_acc",
                "pro_SApM",
                "pro_str_def",
                "pro_total_str_pM",
                "pro_total_str_acc",
                "pro_total_str_absorbed_pM",
                "pro_td_avg",
                "pro_td_acc",
                "pro_td_def",
                "pro_sub_avg",
                "pro_kd_pM",
                "ko_rate",
                "sub_rate",
                "dec_rate",
                "recent_form",
                "head_pct",
                "body_pct",
                "leg_pct",
                "distance_pct",
                "clinch_pct",
                "ground_pct",
                "win_streak",
                "loss_streak",
                "last_5_wins",
                "days_since_last_fight",
                "recent_finish_rate",
                "durability",
                "fight_time_minutes",
                "slpm_trend",
                "td_avg_trend",
                "age_adjusted_performance",
                "peak_indicator",
            ]
        }

        stats.update(
            {
                k: latest[f"{prefix}_{k}"]
                for k in [
                    "height",
                    "reach",
                    "weight",
                    "age_at_event",
                    "stance",
                    "ape_index",
                ]
            }
        )

        # Update with last fight result
        if pd.notna(latest["winner"]):
            fighter_won = (latest["winner"] == "Red" and prefix == "r") or (
                latest["winner"] == "Blue" and prefix == "b"
            )
            fighter_lost = (latest["winner"] == "Red" and prefix == "b") or (
                latest["winner"] == "Blue" and prefix == "r"
            )

            if fighter_won:
                stats["wins"] += 1
            elif fighter_lost:
                stats["losses"] += 1
            else:
                stats["draws"] += 1

            stats["win_loss_ratio"] = stats["wins"] / max(stats["losses"], 1)

            total_fights = stats["wins"] + stats["losses"]
            if total_fights > 0:
                method = str(latest["method"]).lower()
                if "ko" in method or "tko" in method:
                    method_type = "ko"
                elif "sub" in method:
                    method_type = "sub"
                else:
                    method_type = "dec"

                prev_total = total_fights - 1
                if prev_total > 0:
                    prev_ko = int(stats["ko_rate"] * prev_total)
                    prev_sub = int(stats["sub_rate"] * prev_total)
                    prev_dec = int(stats["dec_rate"] * prev_total)
                else:
                    prev_ko = prev_sub = prev_dec = 0

                if fighter_won:
                    if method_type == "ko":
                        prev_ko += 1
                    elif method_type == "sub":
                        prev_sub += 1
                    else:
                        prev_dec += 1

                stats["ko_rate"] = prev_ko / total_fights
                stats["sub_rate"] = prev_sub / total_fights
                stats["dec_rate"] = prev_dec / total_fights

        return stats

    def prepare_upcoming_fight(self, fight, feature_columns):
        """Prepare upcoming fight with all advanced features"""
        r_stats = self.get_fighter_latest_stats(fight["red_fighter"])
        b_stats = self.get_fighter_latest_stats(fight["blue_fighter"])

        if not r_stats or not b_stats:
            return None, None, None

        r_total_fights = r_stats["wins"] + r_stats["losses"]
        b_total_fights = b_stats["wins"] + b_stats["losses"]

        fight_features = {
            "r_fighter": fight["red_fighter"],  # For deep learning
            "b_fighter": fight["blue_fighter"],  # For deep learning
            "height_diff": (r_stats["height"] - b_stats["height"])
            if r_stats["height"] and b_stats["height"]
            else 0,
            "reach_diff": (r_stats["reach"] - b_stats["reach"])
            if r_stats["reach"] and b_stats["reach"]
            else 0,
            "weight_diff": (r_stats["weight"] - b_stats["weight"])
            if r_stats["weight"] and b_stats["weight"]
            else 0,
            "age_at_event_diff": (r_stats["age_at_event"] - b_stats["age_at_event"])
            if r_stats["age_at_event"] and b_stats["age_at_event"]
            else 0,
            "ape_index_diff": (r_stats["ape_index"] - b_stats["ape_index"])
            if r_stats["ape_index"] and b_stats["ape_index"]
            else 0,
            **{
                f"{k}_diff_corrected": r_stats[k] - b_stats[k]
                for k in [
                    "wins",
                    "losses",
                    "win_loss_ratio",
                    "pro_SLpM",
                    "pro_sig_str_acc",
                    "pro_SApM",
                    "pro_str_def",
                    "pro_total_str_pM",
                    "pro_total_str_acc",
                    "pro_total_str_absorbed_pM",
                    "pro_td_avg",
                    "pro_td_acc",
                    "pro_td_def",
                    "pro_sub_avg",
                    "pro_kd_pM",
                    "ko_rate",
                    "sub_rate",
                    "dec_rate",
                    "recent_form",
                    "head_pct",
                    "body_pct",
                    "leg_pct",
                    "distance_pct",
                    "clinch_pct",
                    "ground_pct",
                    "win_streak",
                    "loss_streak",
                    "last_5_wins",
                    "days_since_last_fight",
                    "recent_finish_rate",
                    "durability",
                    "fight_time_minutes",
                    "slpm_trend",
                    "td_avg_trend",
                    "age_adjusted_performance",
                    "peak_indicator",
                ]
            },
            "h2h_advantage": 0,
            "total_rounds": fight["total_rounds"],
            "is_title_bout": 1 if fight["total_rounds"] == 5 else 0,
            "r_total_fights": r_total_fights,
            "b_total_fights": b_total_fights,
        }

        # All individual stats
        for prefix, stats in [("r", r_stats), ("b", b_stats)]:
            for stat in [
                "wins",
                "losses",
                "ko_rate",
                "sub_rate",
                "dec_rate",
                "pro_str_def",
                "pro_td_def",
                "durability",
                "pro_SLpM",
                "pro_sig_str_acc",
                "pro_SApM",
                "pro_td_avg",
                "pro_td_acc",
                "pro_sub_avg",
                "pro_kd_pM",
                "head_pct",
                "distance_pct",
                "clinch_pct",
                "ground_pct",
                "fight_time_minutes",
                "recent_form",
                "recent_finish_rate",
                "win_streak",
                "loss_streak",
            ]:
                fight_features[f"{prefix}_{stat}_corrected"] = stats[stat]

        # Derived features
        r_avg_fight_time = r_stats["fight_time_minutes"] / (r_total_fights + 1)
        b_avg_fight_time = b_stats["fight_time_minutes"] / (b_total_fights + 1)

        fight_features.update(
            {
                "net_striking_advantage": (r_stats["pro_SLpM"] - r_stats["pro_SApM"])
                - (b_stats["pro_SLpM"] - b_stats["pro_SApM"]),
                "striking_efficiency": (
                    r_stats["pro_SLpM"] * r_stats["pro_sig_str_acc"]
                )
                - (b_stats["pro_SLpM"] * b_stats["pro_sig_str_acc"]),
                "defensive_striking": (r_stats["pro_str_def"] - r_stats["pro_SApM"])
                - (b_stats["pro_str_def"] - b_stats["pro_SApM"]),
                "grappling_control": (r_stats["pro_td_avg"] * r_stats["pro_td_acc"])
                - (b_stats["pro_td_avg"] * b_stats["pro_td_acc"]),
                "grappling_defense": (
                    r_stats["pro_td_def"] - r_stats["pro_sub_avg"] / 5
                )
                - (b_stats["pro_td_def"] - b_stats["pro_sub_avg"] / 5),
                "offensive_output": (
                    r_stats["pro_SLpM"] + r_stats["pro_td_avg"] + r_stats["pro_sub_avg"]
                )
                - (
                    b_stats["pro_SLpM"] + b_stats["pro_td_avg"] + b_stats["pro_sub_avg"]
                ),
                "defensive_composite": (
                    (r_stats["pro_str_def"] + r_stats["pro_td_def"]) / 2
                )
                - ((b_stats["pro_str_def"] + b_stats["pro_td_def"]) / 2),
                "ko_specialist_gap": (r_stats["ko_rate"] * r_stats["pro_SLpM"])
                - (b_stats["ko_rate"] * b_stats["pro_SLpM"]),
                "submission_specialist_gap": (
                    r_stats["sub_rate"] * r_stats["pro_sub_avg"]
                )
                - (b_stats["sub_rate"] * b_stats["pro_sub_avg"]),
                "experience_gap": r_total_fights - b_total_fights,
                "skill_momentum": (
                    (r_stats["pro_SLpM"] - b_stats["pro_SLpM"])
                    * (r_stats["recent_form"] - b_stats["recent_form"])
                ),
                "finish_threat": (r_stats["ko_rate"] + r_stats["sub_rate"])
                - (b_stats["ko_rate"] + b_stats["sub_rate"]),
                "momentum_advantage": (
                    (r_stats["win_streak"] - b_stats["win_streak"])
                    - (r_stats["loss_streak"] - b_stats["loss_streak"])
                ),
                "inactivity_penalty": -1
                if r_stats["days_since_last_fight"] - b_stats["days_since_last_fight"]
                > 365
                else 1
                if b_stats["days_since_last_fight"] - r_stats["days_since_last_fight"]
                > 365
                else 0,
                "pace_differential": (r_stats["pro_SLpM"] + r_stats["pro_td_avg"])
                - (b_stats["pro_SLpM"] + b_stats["pro_td_avg"]),
                "experience_ratio": r_total_fights / (b_total_fights + 1),
                "avg_fight_time_diff": r_avg_fight_time - b_avg_fight_time,
            }
        )

        # Enhanced experience
        r_elite_fight_ratio = (
            r_stats["wins"] / r_total_fights if r_total_fights > 0 else 0
        )
        b_elite_fight_ratio = (
            b_stats["wins"] / b_total_fights if b_total_fights > 0 else 0
        )

        fight_features["quality_experience_gap"] = (
            r_total_fights * r_elite_fight_ratio - b_total_fights * b_elite_fight_ratio
        )

        r_championship_exp = (
            1
            if r_total_fights > 0
            and (r_stats["fight_time_minutes"] / r_total_fights) > 12
            else 0
        )
        b_championship_exp = (
            1
            if b_total_fights > 0
            and (b_stats["fight_time_minutes"] / b_total_fights) > 12
            else 0
        )
        fight_features["championship_exp_diff"] = (
            r_championship_exp - b_championship_exp
        )

        r_adversity = r_stats["losses"] * (r_stats["win_loss_ratio"] - 1)
        b_adversity = b_stats["losses"] * (b_stats["win_loss_ratio"] - 1)
        fight_features["adversity_exp_diff"] = r_adversity - b_adversity

        fight_features["experience_skill_interaction"] = (
            fight_features["experience_gap"]
            * (r_stats["pro_SLpM"] - b_stats["pro_SLpM"])
            / 50
        )

        fight_features["veteran_edge"] = (
            fight_features["experience_gap"]
            * (r_stats["recent_form"] - b_stats["recent_form"])
            if abs(fight_features["experience_gap"]) > 5
            else 0
        )

        fight_features["novice_vulnerability"] = (
            -abs(fight_features["experience_gap"]) * 0.5
            if r_total_fights < 10 or b_total_fights < 10
            else 0
        )

        fight_features["exp_gap_historical_win_rate"] = 0.5

        # Style matchups
        r_striker_score = (
            r_stats["distance_pct"] * 1.2
            + r_stats["head_pct"] * 0.8
            - r_stats["ground_pct"] * 1.0
        )
        b_striker_score = (
            b_stats["distance_pct"] * 1.2
            + b_stats["head_pct"] * 0.8
            - b_stats["ground_pct"] * 1.0
        )
        fight_features["striker_advantage"] = r_striker_score - b_striker_score
        fight_features["r_striker_score"] = r_striker_score
        fight_features["b_striker_score"] = b_striker_score

        r_grappler_score = (
            r_stats["pro_td_avg"] * 0.4
            + r_stats["pro_sub_avg"] * 0.3
            + r_stats["ground_pct"] * 0.3
        )
        b_grappler_score = (
            b_stats["pro_td_avg"] * 0.4
            + b_stats["pro_sub_avg"] * 0.3
            + b_stats["ground_pct"] * 0.3
        )
        fight_features["grappler_advantage"] = r_grappler_score - b_grappler_score
        fight_features["r_grappler_score"] = r_grappler_score
        fight_features["b_grappler_score"] = b_grappler_score

        fight_features["effective_reach_advantage"] = (
            fight_features["reach_diff"] * 1.5
            if fight_features["striker_advantage"] > 0.2
            else fight_features["reach_diff"]
        )

        if (
            "stance_encoder" in self.label_encoders
            and r_stats["stance"]
            and b_stats["stance"]
        ):
            r_enc = self.label_encoders["stance_encoder"].transform(
                [r_stats["stance"]]
            )[0]
            b_enc = self.label_encoders["stance_encoder"].transform(
                [b_stats["stance"]]
            )[0]
            fight_features["stance_diff"] = r_enc - b_enc

            fight_features["orthodox_southpaw_matchup"] = (
                0.05
                if (r_stats["stance"] == "Orthodox" and b_stats["stance"] == "Southpaw")
                else -0.05
                if (r_stats["stance"] == "Southpaw" and b_stats["stance"] == "Orthodox")
                else 0
            )
        else:
            fight_features["stance_diff"] = 0
            fight_features["orthodox_southpaw_matchup"] = 0

        return fight_features, r_stats, b_stats

    def get_dynamic_method_weights(self, fight_data, ml_probs, rule_probs, method_type):
        """Calculate optimal weights dynamically based on context and confidence"""

        # Method-specific optimal weights (tuned based on method predictability)
        method_weights = {
            "KO/TKO": {"ml": 0.7, "rule": 0.3},  # ML better at KO patterns
            "Submission": {"ml": 0.3, "rule": 0.7},  # Rules better for grappling
            "Decision": {"ml": 0.5, "rule": 0.5},  # Balanced approach
        }

        # Start with method-specific weights
        weights = method_weights.get(method_type, {"ml": 0.4, "rule": 0.6})

        # Adjust based on confidence difference
        ml_confidence = max(ml_probs) if ml_probs else 0
        rule_confidence = max(rule_probs) if rule_probs else 0
        conf_diff = abs(ml_confidence - rule_confidence)

        if conf_diff > 0.2:  # One approach is much more confident
            if ml_confidence > rule_confidence:
                weights["ml"] = min(0.8, weights["ml"] + 0.2)
                weights["rule"] = max(0.2, weights["rule"] - 0.2)
            else:
                weights["ml"] = max(0.2, weights["ml"] - 0.2)
                weights["rule"] = min(0.8, weights["rule"] + 0.2)

        # Adjust based on fight context - handle both DataFrame and dict inputs
        if hasattr(fight_data, "iloc"):  # DataFrame
            is_title_bout = (
                fight_data["is_title_bout"].iloc[0]
                if "is_title_bout" in fight_data.columns
                else 0
            )
            exp_gap = (
                abs(fight_data["experience_gap"].iloc[0])
                if "experience_gap" in fight_data.columns
                else 0
            )
            striker_adv = (
                abs(fight_data["striker_advantage"].iloc[0])
                if "striker_advantage" in fight_data.columns
                else 0
            )
        else:  # Dictionary
            is_title_bout = fight_data.get("is_title_bout", 0)
            exp_gap = abs(fight_data.get("experience_gap", 0))
            striker_adv = abs(fight_data.get("striker_advantage", 0))

        # Adjust based on fight context
        if is_title_bout == 1:  # Title fight - trust ML more
            weights["ml"] = min(0.8, weights["ml"] + 0.1)
            weights["rule"] = max(0.2, weights["rule"] - 0.1)

        # Adjust based on experience gap
        if exp_gap > 10:  # Large experience gap - trust ML more
            weights["ml"] = min(0.8, weights["ml"] + 0.1)
            weights["rule"] = max(0.2, weights["rule"] - 0.1)

        # Adjust based on style matchup
        if striker_adv > 0.5:  # Clear style matchup - trust rules more
            weights["ml"] = max(0.2, weights["ml"] - 0.1)
            weights["rule"] = min(0.8, weights["rule"] + 0.1)

        return weights

    def predict_method_type(self, fight_data):
        """Predict the most likely method type based on fighter characteristics"""

        # Extract key indicators - handle both DataFrame and dict inputs
        if hasattr(fight_data, "iloc"):  # DataFrame
            r_ko_rate = (
                fight_data["r_ko_rate_corrected"].iloc[0]
                if "r_ko_rate_corrected" in fight_data.columns
                else 0
            )
            b_ko_rate = (
                fight_data["b_ko_rate_corrected"].iloc[0]
                if "b_ko_rate_corrected" in fight_data.columns
                else 0
            )
            r_sub_rate = (
                fight_data["r_sub_rate_corrected"].iloc[0]
                if "r_sub_rate_corrected" in fight_data.columns
                else 0
            )
            b_sub_rate = (
                fight_data["b_sub_rate_corrected"].iloc[0]
                if "b_sub_rate_corrected" in fight_data.columns
                else 0
            )
            r_dec_rate = (
                fight_data["r_dec_rate_corrected"].iloc[0]
                if "r_dec_rate_corrected" in fight_data.columns
                else 0
            )
            b_dec_rate = (
                fight_data["b_dec_rate_corrected"].iloc[0]
                if "b_dec_rate_corrected" in fight_data.columns
                else 0
            )
        else:  # Dictionary
            r_ko_rate = fight_data.get("r_ko_rate_corrected", 0)
            b_ko_rate = fight_data.get("b_ko_rate_corrected", 0)
            r_sub_rate = fight_data.get("r_sub_rate_corrected", 0)
            b_sub_rate = fight_data.get("b_sub_rate_corrected", 0)
            r_dec_rate = fight_data.get("r_dec_rate_corrected", 0)
            b_dec_rate = fight_data.get("b_dec_rate_corrected", 0)

        # Calculate average rates
        avg_ko_rate = (r_ko_rate + b_ko_rate) / 2
        avg_sub_rate = (r_sub_rate + b_sub_rate) / 2
        avg_dec_rate = (r_dec_rate + b_dec_rate) / 2

        # Predict based on highest rate
        if avg_ko_rate > avg_sub_rate and avg_ko_rate > avg_dec_rate:
            return "KO/TKO"
        elif avg_sub_rate > avg_ko_rate and avg_sub_rate > avg_dec_rate:
            return "Submission"
        else:
            return "Decision"

    def calculate_enhanced_method_adjustments(
        self, fight_data, winner_prefix, loser_prefix
    ):
        """Comprehensive method prediction using all available stats"""

        # Convert winner_prefix to column prefix (Red -> r, Blue -> b)
        w_prefix = "r" if winner_prefix == "Red" else "b"
        l_prefix = "r" if loser_prefix == "Red" else "b"

        # Safe data extraction function
        def safe_get_value(data, key, default=0.0):
            try:
                return data[key].values[0] if hasattr(data[key], 'values') else data[key]
            except (KeyError, IndexError, AttributeError):
                return default

        # Extract all relevant stats using safe function
        w_slpm = safe_get_value(fight_data, f"{w_prefix}_pro_SLpM_corrected")
        w_sig_acc = safe_get_value(fight_data, f"{w_prefix}_pro_sig_str_acc_corrected")
        w_td_avg = safe_get_value(fight_data, f"{w_prefix}_pro_td_avg_corrected")
        w_td_acc = safe_get_value(fight_data, f"{w_prefix}_pro_td_acc_corrected")
        w_sub_avg = safe_get_value(fight_data, f"{w_prefix}_pro_sub_avg_corrected")
        w_ko_rate = safe_get_value(fight_data, f"{w_prefix}_ko_rate_corrected")
        w_sub_rate = safe_get_value(fight_data, f"{w_prefix}_sub_rate_corrected")
        w_dec_rate = safe_get_value(fight_data, f"{w_prefix}_dec_rate_corrected")
        w_kd_rate = safe_get_value(fight_data, f"{w_prefix}_pro_kd_pM_corrected")

        l_str_def = safe_get_value(fight_data, f"{l_prefix}_pro_str_def_corrected")
        l_sapm = safe_get_value(fight_data, f"{l_prefix}_pro_SApM_corrected")
        l_td_def = safe_get_value(fight_data, f"{l_prefix}_pro_td_def_corrected")
        l_durability = safe_get_value(fight_data, f"{l_prefix}_durability_corrected")

        w_head_pct = safe_get_value(fight_data, f"{w_prefix}_head_pct_corrected")
        w_distance_pct = safe_get_value(fight_data, f"{w_prefix}_distance_pct_corrected")
        w_clinch_pct = safe_get_value(fight_data, f"{w_prefix}_clinch_pct_corrected")
        w_ground_pct = safe_get_value(fight_data, f"{w_prefix}_ground_pct_corrected")

        l_distance_pct = safe_get_value(fight_data, f"{l_prefix}_distance_pct_corrected")
        l_ground_pct = safe_get_value(fight_data, f"{l_prefix}_ground_pct_corrected")

        # Extract recent form data (with fallback if not available)
        w_recent_form = safe_get_value(fight_data, f"{w_prefix}_recent_form_corrected", 0.5)

        total_rounds = safe_get_value(fight_data, "total_rounds", 3)

        # ENHANCED KO/TKO PROBABILITY
        ko_base = w_ko_rate

        # Advanced striking analysis
        striking_volume_factor = min(w_slpm / 6.0, 1.5)
        accuracy_factor = 1 + (w_sig_acc - 0.45) * 2.5  # Increased sensitivity
        accuracy_factor = max(0.4, min(accuracy_factor, 2.0))

        # Enhanced head hunting analysis
        head_hunting_factor = 1 + (w_head_pct - 0.5) * 0.8  # Increased sensitivity
        head_hunting_factor = max(0.6, min(head_hunting_factor, 1.6))

        # Distance fighting preference
        distance_factor = 1 + (w_distance_pct - 0.6) * 0.6  # Increased sensitivity
        distance_factor = max(0.7, min(distance_factor, 1.5))

        # Enhanced knockdown threat analysis
        kd_threat_factor = 1 + min(w_kd_rate / 0.4, 1.5)  # Increased sensitivity
        kd_threat_factor = max(0.8, kd_threat_factor)

        # Advanced opponent vulnerability analysis
        opp_vulnerability = (
            (1 - l_str_def) * 0.45 + (l_sapm / 6.0) * 0.35 + (1 - l_durability) * 0.2
        )
        opp_vulnerability = min(opp_vulnerability, 1.2)  # Allow higher vulnerability

        # Enhanced power differential analysis
        power_differential = (w_slpm * w_sig_acc) - (l_sapm * (1 - l_str_def))
        power_differential_factor = 1 + max(0, power_differential / 4.0)  # Increased sensitivity
        power_differential_factor = min(power_differential_factor, 2.0)

        # Recent form impact on finishing ability
        recent_form_factor = 1 + (w_recent_form - 0.5) * 0.4
        recent_form_factor = max(0.7, min(recent_form_factor, 1.3))

        ko_prob = ko_base * (
            striking_volume_factor * 0.18
            + accuracy_factor * 0.20
            + head_hunting_factor * 0.15
            + distance_factor * 0.10
            + opp_vulnerability * 0.20
            + power_differential_factor * 0.10
            + kd_threat_factor * 0.12
            + recent_form_factor * 0.15
        )

        if total_rounds == 5:
            ko_prob *= 1.15

        ko_prob = min(ko_prob, 0.85)

        # SUBMISSION PROBABILITY
        sub_base = w_sub_rate

        grappling_control = w_td_avg * w_td_acc
        control_factor = 1 + min(grappling_control / 2.0, 1.0)

        sub_attempt_factor = 1 + (w_sub_avg / 2.0)
        sub_attempt_factor = min(sub_attempt_factor, 2.0)

        ground_preference_factor = 1 + (w_ground_pct - 0.2) * 0.8
        ground_preference_factor = max(0.6, min(ground_preference_factor, 1.6))

        opp_grappling_weakness = (
            (1 - l_td_def) * 0.5 + (1 - l_durability) * 0.3 + l_ground_pct * 0.2
        )

        td_differential = (w_td_avg * w_td_acc) - (l_td_def * 2.0)
        td_factor = 1 + max(0, td_differential / 3.0)
        td_factor = min(td_factor, 1.7)

        sub_prob = sub_base * (
            control_factor * 0.25
            + sub_attempt_factor * 0.20
            + ground_preference_factor * 0.15
            + opp_grappling_weakness * 0.25
            + td_factor * 0.15
        )

        if total_rounds == 5:
            sub_prob *= 1.10

        sub_prob = min(sub_prob, 0.75)

        # DECISION PROBABILITY
        dec_base = w_dec_rate

        # Factors that favor decisions
        decision_factors = []

        # High output but low finishing rate
        if w_slpm > 4.0 and w_ko_rate < 0.15:
            decision_factors.append(1.3)

        # High takedown rate but low submission rate
        if w_td_avg > 2.0 and w_sub_rate < 0.10:
            decision_factors.append(1.2)

        # Opponent durability
        if l_durability > 0.7:
            decision_factors.append(1.4)

        # Distance fighting preference
        if w_distance_pct > 0.7:
            decision_factors.append(1.1)

        # Clinch fighting (often leads to decisions)
        if w_clinch_pct > 0.3:
            decision_factors.append(1.15)

        decision_multiplier = np.mean(decision_factors) if decision_factors else 1.0
        dec_prob = dec_base * decision_multiplier

        # Additional context-based adjustments
        # Weight class adjustments - handle both DataFrame and dict inputs
        if hasattr(fight_data, "iloc"):  # DataFrame
            weight_class = (
                fight_data["weight_class"].iloc[0]
                if "weight_class" in fight_data.columns
                else "Unknown"
            )
            w_recent_form = (
                fight_data[f"{w_prefix}_recent_form_corrected"].iloc[0]
                if f"{w_prefix}_recent_form_corrected" in fight_data.columns
                else 0.5
            )
            l_durability = (
                fight_data[f"{l_prefix}_durability_corrected"].iloc[0]
                if f"{l_prefix}_durability_corrected" in fight_data.columns
                else 0.5
            )
        else:  # Dictionary
            weight_class = fight_data.get("weight_class", "Unknown")
            w_recent_form = fight_data.get(f"{w_prefix}_recent_form_corrected", 0.5)
            l_durability = fight_data.get(f"{l_prefix}_durability_corrected", 0.5)

        if weight_class in ["Heavyweight", "Light Heavyweight"]:
            ko_prob *= 1.2  # Higher KO rate in heavier divisions
            sub_prob *= 0.8  # Lower sub rate in heavier divisions
        elif weight_class in ["Flyweight", "Bantamweight"]:
            ko_prob *= 0.9  # Lower KO rate in lighter divisions
            sub_prob *= 1.1  # Higher sub rate in lighter divisions

        # Round length adjustments
        if total_rounds == 5:  # Championship rounds
            ko_prob *= 1.1  # Slightly higher KO chance with more time
            sub_prob *= 1.05  # Slightly higher sub chance
            dec_prob *= 1.15  # Much higher decision chance

        # Recent form adjustments
        if w_recent_form > 0.7:  # Hot streak
            ko_prob *= 1.15
            sub_prob *= 1.1
        elif w_recent_form < 0.3:  # Cold streak
            ko_prob *= 0.9
            sub_prob *= 0.85
            dec_prob *= 1.1

        # Durability adjustments
        if l_durability < 0.3:  # Fragile opponent
            ko_prob *= 1.25
        elif l_durability > 0.8:  # Durable opponent
            ko_prob *= 0.8
            dec_prob *= 1.2

        # Normalize probabilities
        total_prob = ko_prob + sub_prob + dec_prob
        if total_prob > 0:
            ko_prob = ko_prob / total_prob
            sub_prob = sub_prob / total_prob
            dec_prob = dec_prob / total_prob

        return {
            f"{winner_prefix}_KO/TKO": ko_prob,
            f"{winner_prefix}_Submission": sub_prob,
            f"{winner_prefix}_Decision": dec_prob,
        }

    def swap_fight_data(self, fight_data):
        """Swap red and blue corner data to eliminate bias"""
        if hasattr(fight_data, "iloc"):  # DataFrame
            swapped_data = fight_data.copy()

            # Swap fighter names
            if (
                "r_fighter" in swapped_data.columns
                and "b_fighter" in swapped_data.columns
            ):
                swapped_data["r_fighter"], swapped_data["b_fighter"] = (
                    swapped_data["b_fighter"],
                    swapped_data["r_fighter"],
                )

            # Swap all red/blue prefixed columns
            for col in swapped_data.columns:
                if col.startswith("r_") and not col.endswith("_diff"):
                    b_col = col.replace("r_", "b_", 1)
                    if b_col in swapped_data.columns:
                        swapped_data[col], swapped_data[b_col] = (
                            swapped_data[b_col],
                            swapped_data[col],
                        )

            # Negate all difference columns (since we swapped the order)
            for col in swapped_data.columns:
                if col.endswith("_diff") or col.endswith("_diff_corrected"):
                    swapped_data[col] = -swapped_data[col]

            # Swap total fights
            if (
                "r_total_fights" in swapped_data.columns
                and "b_total_fights" in swapped_data.columns
            ):
                swapped_data["r_total_fights"], swapped_data["b_total_fights"] = (
                    swapped_data["b_total_fights"],
                    swapped_data["r_total_fights"],
                )

            return swapped_data
        else:  # Dictionary
            swapped_data = fight_data.copy()

            # Swap fighter names
            if "r_fighter" in swapped_data and "b_fighter" in swapped_data:
                swapped_data["r_fighter"], swapped_data["b_fighter"] = (
                    swapped_data["b_fighter"],
                    swapped_data["r_fighter"],
                )

            # Swap all red/blue prefixed keys
            keys_to_swap = []
            for key in swapped_data.keys():
                if key.startswith("r_") and not key.endswith("_diff"):
                    b_key = key.replace("r_", "b_", 1)
                    if b_key in swapped_data:
                        keys_to_swap.append((key, b_key))

            for r_key, b_key in keys_to_swap:
                swapped_data[r_key], swapped_data[b_key] = (
                    swapped_data[b_key],
                    swapped_data[r_key],
                )

            # Negate all difference columns
            for key in swapped_data.keys():
                if key.endswith("_diff") or key.endswith("_diff_corrected"):
                    swapped_data[key] = -swapped_data[key]

            # Swap total fights
            if "r_total_fights" in swapped_data and "b_total_fights" in swapped_data:
                swapped_data["r_total_fights"], swapped_data["b_total_fights"] = (
                    swapped_data["b_total_fights"],
                    swapped_data["r_total_fights"],
                )

            return swapped_data

    def predict_fight(self, fight_data, feature_columns):
        """Enhanced fight prediction with comprehensive method adjustments and bias correction"""
        # Ensure consistent random state before prediction
        self.set_random_seeds()

        # Run prediction with original data
        X = fight_data[feature_columns]
        winner_proba_original = self.winner_model.predict_proba(X)[0]
        winner_pred_original = self.winner_model.predict(X)[0]

        # Run prediction with swapped data to eliminate red corner bias
        swapped_data = self.swap_fight_data(fight_data)
        X_swapped = swapped_data[feature_columns]
        winner_proba_swapped = self.winner_model.predict_proba(X_swapped)[0]

        # Select the prediction with higher confidence - match manual process
        red_prob_original = winner_proba_original[1]
        blue_prob_original = winner_proba_original[0]
        red_prob_swapped = winner_proba_swapped[1]
        blue_prob_swapped = winner_proba_swapped[0]

        # Compare Win% predictions for each fighter across both runs
        # Red fighter's best Win% is max(red_prob_original, blue_prob_swapped)
        # Blue fighter's best Win% is max(blue_prob_original, red_prob_swapped)
        red_fighter_best_prob = max(red_prob_original, blue_prob_swapped)
        blue_fighter_best_prob = max(blue_prob_original, red_prob_swapped)

        if red_fighter_best_prob > blue_fighter_best_prob:
            winner_name = "Red"
            winner_proba = [blue_fighter_best_prob, red_fighter_best_prob]
            winner_pred = 1
            # Use original data if red wins in original, swapped data if red wins in swapped
            if red_prob_original > blue_prob_original:
                use_original_data = True
            else:
                use_original_data = False
        else:
            winner_name = "Blue"
            winner_proba = [blue_fighter_best_prob, red_fighter_best_prob]
            winner_pred = 0
            # Use original data if blue wins in original, swapped data if blue wins in swapped
            if blue_prob_original > red_prob_original:
                use_original_data = True
            else:
                use_original_data = False

        # Use the appropriate dataset for method prediction
        if use_original_data:
            X_for_method = X
            fight_data_for_method = fight_data
        else:
            X_for_method = X_swapped
            fight_data_for_method = swapped_data

        # If deep learning is available, ensemble it with traditional model
        if self.use_deep_learning and self.deep_learning_model and self.fighter_encoder:
            try:
                # Get fighter encodings from the data used for method prediction
                r_fighter = (
                    fight_data_for_method["r_fighter"].values[0]
                    if "r_fighter" in fight_data_for_method
                    else None
                )
                b_fighter = (
                    fight_data_for_method["b_fighter"].values[0]
                    if "b_fighter" in fight_data_for_method
                    else None
                )

                if (
                    r_fighter
                    and b_fighter
                    and r_fighter in self.fighter_encoder.classes_
                    and b_fighter in self.fighter_encoder.classes_
                ):
                    r_fighter_encoded = self.fighter_encoder.transform([r_fighter])[0]
                    b_fighter_encoded = self.fighter_encoder.transform([b_fighter])[0]

                    # Prepare DL input
                    from sklearn.compose import ColumnTransformer
                    from sklearn.impute import SimpleImputer
                    from sklearn.preprocessing import StandardScaler
                    from sklearn.pipeline import Pipeline

                    numeric_transformer = Pipeline(
                        [
                            ("imputer", SimpleImputer(strategy="median")),
                            ("scaler", StandardScaler()),
                        ]
                    )
                    preprocessor = ColumnTransformer(
                        [("num", numeric_transformer, feature_columns)]
                    )
                    X_scaled = preprocessor.fit_transform(X_for_method)

                    X_dl = [
                        np.array([r_fighter_encoded]),
                        np.array([b_fighter_encoded]),
                        X_scaled,
                    ]

                    # Get DL predictions
                    dl_winner_proba, dl_method_proba = self.deep_learning_model.predict(
                        X_dl, verbose=0
                    )

                    # Convert winner prediction to 2-class format
                    dl_winner_2class = np.array(
                        [[1 - dl_winner_proba[0][0], dl_winner_proba[0][0]]]
                    )

                    # Ensemble: 60% traditional, 40% deep learning
                    winner_proba = winner_proba * 0.6 + dl_winner_2class[0] * 0.4
                    winner_pred = 1 if winner_proba[1] > 0.5 else 0
                    winner_name = "Red" if winner_pred == 1 else "Blue"
            except Exception:
                # Fallback to traditional model if DL fails
                pass

        # COMPREHENSIVE METHOD PREDICTION WITH DYNAMIC WEIGHTING
        loser_prefix = "Blue" if winner_name == "Red" else "Red"

        # Get base method probabilities from model using the correct data
        method_proba = self.method_model.predict_proba(X_for_method)[0]
        method_labels = self.label_encoders["winner_method_encoder"].classes_

        # Get comprehensive method adjustments using the correct data
        method_adjustments = self.calculate_enhanced_method_adjustments(
            fight_data_for_method, winner_name, loser_prefix
        )

        # Predict method type for dynamic weighting using the correct data
        predicted_method_type = self.predict_method_type(fight_data_for_method)

        # Prepare ML and rule probabilities for weighting
        ml_probs = {}
        rule_probs = {}
        for i, label in enumerate(method_labels):
            if label.startswith(winner_name):
                method_name = label.split("_")[1]
                ml_probs[method_name] = method_proba[i]
                rule_probs[method_name] = method_adjustments.get(label, 0.0)

        # Get dynamic weights based on context and confidence using the correct data
        weights = self.get_dynamic_method_weights(
            fight_data_for_method,
            list(ml_probs.values()),
            list(rule_probs.values()),
            predicted_method_type,
        )

        # Debug information (can be removed in production)
        if hasattr(self, "debug_mode") and self.debug_mode:
            print(
                f"Dynamic Weights - Method: {predicted_method_type}, ML: {weights['ml']:.2f}, Rule: {weights['rule']:.2f}"
            )
            print(f"ML Probs: {ml_probs}")
            print(f"Rule Probs: {rule_probs}")

        # Combine model predictions with comprehensive adjustments using dynamic weights
        final_method_probs = {}
        for i, label in enumerate(method_labels):
            if label.startswith(winner_name):
                method_name = label.split("_")[1]
                base_prob = method_proba[i]
                adjustment_prob = method_adjustments.get(label, 0.0)

                # Dynamic weighted combination based on context and confidence
                final_prob = (
                    base_prob * weights["ml"] + adjustment_prob * weights["rule"]
                )
                final_method_probs[label] = final_prob

        # Normalize probabilities
        total = sum(final_method_probs.values())
        if total > 0:
            final_method_probs = {k: v / total for k, v in final_method_probs.items()}

        final_method = max(final_method_probs, key=final_method_probs.get)

        return {
            "winner": winner_name,
            "winner_confidence": winner_proba[winner_pred],
            "red_prob": winner_proba[1],
            "blue_prob": winner_proba[0],
            "method": final_method.split("_")[1],
            "method_probabilities": final_method_probs,
            "combined_method_prob": final_method_probs[final_method],
        }

    def predict_upcoming_fights(self, upcoming_fights, feature_columns):
        """Predict upcoming fights"""
        predictions = []
        skipped_fights = []

        for fight in upcoming_fights:
            result = self.prepare_upcoming_fight(fight, feature_columns)
            if not result[0]:
                skipped_fights.append(
                    f"{fight['red_fighter']} vs {fight['blue_fighter']}"
                )
                continue

            fight_features, r_stats, b_stats = result
            fight_df = pd.DataFrame([fight_features])

            for col in feature_columns:
                if col not in fight_df.columns:
                    fight_df[col] = 0

            pred = self.predict_fight(fight_df, feature_columns)

            winner_prefix = pred["winner"]
            ko_prob = pred["method_probabilities"].get(f"{winner_prefix}_KO/TKO", 0)
            sub_prob = pred["method_probabilities"].get(
                f"{winner_prefix}_Submission", 0
            )
            dec_prob = pred["method_probabilities"].get(f"{winner_prefix}_Decision", 0)

            predictions.append(
                {
                    "Red Fighter": fight["red_fighter"],
                    "Blue Fighter": fight["blue_fighter"],
                    "Weight Class": fight["weight_class"],
                    "Winner": fight["red_fighter"]
                    if pred["winner"] == "Red"
                    else fight["blue_fighter"],
                    "Win%": pred["winner_confidence"],
                    "Method": pred["method"],
                    "Method%": pred["combined_method_prob"],
                    "KO/TKO%": ko_prob,
                    "Submission%": sub_prob,
                    "Decision%": dec_prob,
                }
            )

        return predictions, skipped_fights

    def export_predictions_to_excel(self, predictions, filename="predictions.xlsx"):
        """Export predictions to formatted Excel file"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        df = pd.DataFrame(predictions)

        # Ensure percentage columns are in decimal format (0-1 range)
        percentage_columns = ["Win%", "Method%", "KO/TKO%", "Submission%", "Decision%"]
        for col in percentage_columns:
            if col in df.columns:
                # If values are > 1, they're likely already percentages, convert to decimal
                if df[col].max() > 1:
                    df[col] = df[col] / 100
                df[col] = df[col].round(4)

        df.to_excel(filename, index=False, engine="openpyxl")

        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        header_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        header_font = Font(bold=True, size=11, color="000000")
        left_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Apply formatting to ALL cells
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.alignment = left_alignment
                cell.border = thin_border

        # Apply header-specific formatting
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Format percentage columns: Win%(5), Method%(7), KO/TKO%(8), Submission%(9), Decision%(10)
        percentage_col_indices = [5, 7, 8, 9, 10]
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in percentage_col_indices:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = "0.00%"

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value is None:
                        cell_length = 0
                    elif isinstance(cell.value, (int, float)):
                        cell_length = (
                            10
                            if cell.number_format == "0.00%"
                            else len(str(cell.value))
                        )
                    else:
                        cell_length = len(str(cell.value))

                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass

            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(filename)
        print(f"\nPredictions exported to: {filename}")
        return filename


class UFCPredictorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("UFC Fight Predictor")
        self.root.geometry("1000x800")
        self.root.minsize(700, 550)

        self.data_file_path = tk.StringVar(value=fight_data_path)
        self.output_file_path = tk.StringVar(value="UFC_predictions_2.xlsx")
        self.use_deep_learning = tk.BooleanVar(value=True)
        self.predictor = None
        self.create_widgets()

    def create_widgets(self):
        title_frame = tk.Frame(self.root, bg="#D20A0A")
        title_frame.pack(fill=tk.X)
        tk.Label(
            title_frame,
            text="UFC FIGHT PREDICTOR",
            font=("Arial", 16, "bold"),
            fg="white",
            bg="#D20A0A",
        ).pack(pady=(10, 8))

        file_frame = ttk.LabelFrame(self.root, text="Data File", padding="5")
        file_frame.pack(fill=tk.X, padx=10, pady=3)
        ttk.Entry(file_frame, textvariable=self.data_file_path, width=70).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(file_frame, text="Browse", command=self.browse_data_file).pack(
            side=tk.LEFT
        )

        input_frame = ttk.LabelFrame(
            self.root, text="Enter Fights to Predict", padding="5"
        )
        input_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=3)
        instructions = """Enter fights in CSV format (one per line):
Format: red_fighter,blue_fighter,weight_class,gender,total_rounds

Example:
Max Holloway,Dustin Poirier,Lightweight,Men,5
Ilia Topuria,Charles Oliveira,Lightweight,Men,5
Tatiana Suarez,Amanda Lemos,Strawweight,Women,3"""
        ttk.Label(
            input_frame,
            text=instructions,
            font=("Arial", 8),
            foreground="gray",
            justify=tk.LEFT,
        ).pack(anchor=tk.W, pady=(0, 3))

        self.fights_text = scrolledtext.ScrolledText(
            input_frame, height=8, width=95, wrap=tk.WORD
        )
        self.fights_text.pack(fill=tk.BOTH, expand=True, pady=3)

        output_frame = ttk.LabelFrame(self.root, text="Output File", padding="5")
        output_frame.pack(fill=tk.X, padx=10, pady=3)
        ttk.Entry(output_frame, textvariable=self.output_file_path, width=70).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(output_frame, text="Browse", command=self.browse_output_file).pack(
            side=tk.LEFT
        )

        # Options frame
        options_frame = ttk.LabelFrame(self.root, text="Model Options", padding="5")
        options_frame.pack(fill=tk.X, padx=10, pady=3)
        ttk.Checkbutton(
            options_frame,
            text="Enable Deep Learning (TensorFlow)",
            variable=self.use_deep_learning,
        ).pack(anchor=tk.W, padx=5)

        button_frame = ttk.Frame(self.root, padding="5")
        button_frame.pack(fill=tk.X, padx=10)
        ttk.Button(button_frame, text="Load Sample", command=self.load_sample).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(button_frame, text="Clear", command=self.clear_input).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(
            button_frame, text="Generate Predictions", command=self.run_predictions
        ).pack(side=tk.RIGHT, padx=3)

        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(
            self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W
        )
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)

    def browse_data_file(self):
        filename = filedialog.askopenfilename(
            title="Select Fight Data CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if filename:
            self.data_file_path.set(filename)

    def browse_output_file(self):
        filename = filedialog.asksaveasfilename(
            title="Save Predictions As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if filename:
            self.output_file_path.set(filename)

    def load_sample(self):
        sample = """Max Holloway,Dustin Poirier,Lightweight,Men,5
Ilia Topuria,Charles Oliveira,Lightweight,Men,5
Tatiana Suarez,Amanda Lemos,Strawweight,Women,3"""
        self.fights_text.delete("1.0", tk.END)
        self.fights_text.insert("1.0", sample)

    def clear_input(self):
        self.fights_text.delete("1.0", tk.END)

    def parse_fights_input(self, text):
        text = text.strip()
        fights = []

        for line in text.split("\n"):
            line = line.strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split(",")]

            if len(parts) != 5:
                raise ValueError(
                    f"Invalid CSV format. Expected 5 fields, got {len(parts)} in line: {line}"
                )

            try:
                fights.append(
                    {
                        "red_fighter": parts[0],
                        "blue_fighter": parts[1],
                        "weight_class": parts[2],
                        "gender": parts[3],
                        "total_rounds": int(parts[4]),
                    }
                )
            except ValueError as e:
                raise ValueError(
                    f"Error parsing line: {line}\nTotal rounds must be a number. {str(e)}"
                )

        if not fights:
            raise ValueError("No valid fight data found.")
        return fights

    def run_predictions(self):
        try:
            use_dl = self.use_deep_learning.get()
            dl_status = " with Deep Learning" if use_dl else ""
            self.status_var.set(
                f"Loading data and training advanced ensemble{dl_status}..."
            )
            self.root.update()

            fights_text = self.fights_text.get("1.0", tk.END)
            upcoming_fights = self.parse_fights_input(fights_text)

            data_file = self.data_file_path.get()
            if not os.path.exists(data_file):
                raise FileNotFoundError(f"Data file not found: {data_file}")

            df = pd.read_csv(data_file)
            self.status_var.set(
                f"Loaded {len(df)} fights. Training ensemble{dl_status}..."
            )
            self.root.update()

            self.predictor = AdvancedUFCPredictor(
                use_ensemble=True,
                use_neural_net=False,
                use_deep_learning=use_dl,
                debug_mode=False,  # Set to True for debugging
            )

            f = io.StringIO()
            with redirect_stdout(f):
                df = self.predictor.fix_data_leakage(df)
                self.predictor.df_train = df
                feature_columns = self.predictor.train_models(df)

            self.status_var.set("Generating predictions...")
            self.root.update()

            predictions, skipped_fights = self.predictor.predict_upcoming_fights(
                upcoming_fights, feature_columns
            )

            if not predictions:
                self.status_var.set("No predictions generated")
                messagebox.showerror(
                    "No Predictions",
                    "All fights were skipped due to insufficient fighter data.",
                )
                return

            output_file = self.output_file_path.get()

            with redirect_stdout(io.StringIO()):
                self.predictor.export_predictions_to_excel(predictions, output_file)

            # Clean up temporary files immediately after training
            cleanup_temp_files()

            success_msg = f"Predictions generated!\n\nSaved to: {output_file}\n\n{len(predictions)} fight(s) predicted"
            if use_dl:
                success_msg += "\n✓ Deep Learning enabled"

            if skipped_fights:
                success_msg += f"\n{len(skipped_fights)} fight(s) skipped"
                self.status_var.set(
                    f"Success! {len(predictions)} predictions, {len(skipped_fights)} skipped"
                )
                skipped_msg = (
                    success_msg
                    + "\n\nSkipped:\n"
                    + "\n".join(f"• {fight}" for fight in skipped_fights)
                )
                messagebox.showwarning("Predictions Complete", skipped_msg)
            else:
                self.status_var.set(
                    f"Success! All {len(predictions)} predictions saved"
                )
                messagebox.showinfo("Success", success_msg)

        except Exception as e:
            import traceback

            error_details = traceback.format_exc()
            self.status_var.set("Error occurred")
            messagebox.showerror(
                "Error", f"Error:\n\n{str(e)}\n\nDetails:\n{error_details}"
            )


def main():
    """Main function to run the UFC Predictor GUI"""
    try:
        # Check if already running to prevent multiple instances
        if hasattr(main, "_running"):
            return
        main._running = True

        root = tk.Tk()
        app = UFCPredictorGUI(root)
        root.mainloop()
    except Exception as e:
        print(f"Error starting UFC Predictor: {e}")
        import traceback

        traceback.print_exc()
    finally:
        # Clean up temporary files when GUI is closed
        cleanup_temp_files()
        # Reset the flag when done
        if hasattr(main, "_running"):
            delattr(main, "_running")


# Safe execution for IDE "Run Code" feature
if __name__ == "__main__":
    try:
        # Handle multiprocessing for PyInstaller - must be called before any multiprocessing operations
        if hasattr(sys, "frozen") and sys.frozen:
            import multiprocessing

            multiprocessing.freeze_support()
        main()
    except Exception as e:
        print(f"Error in UFC Predictor execution: {e}")
        # Don't exit, just print the error for debugging


# ~ 7 Minutes
# ✅ Winners correct: 141 / 218 → 64.7%
# ✅ Winner + method correct: 80 / 218 → 36.7%