import sys
import os
import random
import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog, messagebox
import pandas as pd
import numpy as np
import io
from contextlib import redirect_stdout
from sklearn.model_selection import train_test_split, TimeSeriesSplit, StratifiedKFold
from sklearn.ensemble import RandomForestClassifier, VotingClassifier
from sklearn.feature_selection import SelectPercentile, f_classif, RFECV
from sklearn.metrics import log_loss, accuracy_score, roc_auc_score, classification_report
from sklearn.calibration import CalibratedClassifierCV
from sklearn.linear_model import LogisticRegression
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
# permutation_importance removed - using model built-in importance instead
from scipy.stats import linregress
import warnings
import shutil
import atexit
from joblib import Parallel, delayed
import multiprocessing as mp
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Enable multiprocessing for faster training
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"

# Get the directory where this script is located
if getattr(sys, "frozen", False):
    script_dir = os.path.dirname(sys.executable)
else:
    script_dir = os.path.dirname(os.path.abspath(__file__))

fight_data_path = os.path.join(script_dir, "fight_data.csv")
if not os.path.exists(fight_data_path):
    print(f"Warning: fight_data.csv not found in script directory: {script_dir}")

# Set random seeds for reproducibility
random.seed(42)
np.random.seed(42)
os.environ["PYTHONHASHSEED"] = "42"

warnings.filterwarnings("ignore")
warnings.filterwarnings("ignore", category=UserWarning, module="joblib")

def cleanup_temp_files():
    """Clean up temporary files and folders created during training"""
    try:
        if os.path.exists("catboost_info"):
            shutil.rmtree("catboost_info")
        if os.path.exists("best_dl_model.h5"):
            os.remove("best_dl_model.h5")
    except Exception as e:
        print(f"Warning: Could not clean up temporary files: {e}")

atexit.register(cleanup_temp_files)

# Try to import optional libraries
try:
    from xgboost import XGBClassifier
    HAS_XGBOOST = True
except ImportError:
    HAS_XGBOOST = False
    print("XGBoost not available.")

try:
    from lightgbm import LGBMClassifier
    HAS_LIGHTGBM = True
except ImportError:
    HAS_LIGHTGBM = False
    print("LightGBM not available.")

try:
    from catboost import CatBoostClassifier
    HAS_CATBOOST = True
except ImportError:
    HAS_CATBOOST = False
    print("CatBoost not available. Install with: pip install catboost")

try:
    from sklearn.ensemble import StackingClassifier
    HAS_STACKING = True
except ImportError:
    HAS_STACKING = False
    print("StackingClassifier not available.")

try:
    import optuna
    HAS_OPTUNA = True
    optuna.logging.set_verbosity(optuna.logging.WARNING)
except ImportError:
    HAS_OPTUNA = False
    print("Optuna not available. Install with: pip install optuna")

# GPU Detection and Configuration
def detect_gpu():
    """Detect if GPU is available for XGBoost, LightGBM, and CatBoost"""
    gpu_available = {
        'xgboost': False,
        'lightgbm': False,
        'catboost': False
    }

    # Test XGBoost GPU
    if HAS_XGBOOST:
        try:
            test_model = XGBClassifier(n_estimators=1, device='cuda')
            test_model.fit([[1, 2], [3, 4]], [0, 1])
            gpu_available['xgboost'] = True
        except Exception:
            pass

    # Test LightGBM GPU
    if HAS_LIGHTGBM:
        try:
            test_model = LGBMClassifier(n_estimators=1, device='gpu', verbose=-1)
            test_model.fit([[1, 2], [3, 4]], [0, 1])
            gpu_available['lightgbm'] = True
        except Exception:
            pass

    # Test CatBoost GPU
    if HAS_CATBOOST:
        try:
            test_model = CatBoostClassifier(iterations=1, task_type='GPU', devices='0', verbose=0)
            test_model.fit([[1, 2], [3, 4]], [0, 1])
            gpu_available['catboost'] = True
        except Exception:
            pass

    return gpu_available

# Detect GPU availability at startup
GPU_AVAILABLE = detect_gpu()

# Print GPU status
print("\n" + "="*80)
print("GPU ACCELERATION STATUS")
print("="*80)
if HAS_XGBOOST:
    print(f"  XGBoost GPU:  {'✓ ENABLED' if GPU_AVAILABLE['xgboost'] else '✗ DISABLED (using CPU)'}")
if HAS_LIGHTGBM:
    print(f"  LightGBM GPU: {'✓ ENABLED' if GPU_AVAILABLE['lightgbm'] else '✗ DISABLED (using CPU)'}")
if HAS_CATBOOST:
    print(f"  CatBoost GPU: {'✓ ENABLED' if GPU_AVAILABLE['catboost'] else '✗ DISABLED (using CPU)'}")
if not any(GPU_AVAILABLE.values()):
    print("  No GPU acceleration available. Install CUDA and GPU-enabled libraries for speedup.")
    print("  XGBoost: pip install xgboost (requires CUDA)")
    print("  LightGBM: pip install lightgbm --config-settings=cmake.define.USE_GPU=ON")
    print("  CatBoost: pip install catboost (GPU auto-detected)")
print("="*80 + "\n")


class ImprovedUFCPredictor:
    def __init__(self, use_ensemble=True, debug_mode=False):
        self.winner_model = None
        self.label_encoders = {}
        self.use_ensemble = use_ensemble
        self.debug_mode = debug_mode
        self.df_train = None
        self.feature_importance = None
        self.best_params = None
        self.calibrated_model = None
        self.feature_columns = None  # Store feature columns for predictions
        self.fighter_elos = {}  # Store fighter ELO ratings
        self.elo_history = {}  # Store ELO history for predictions

        # Set random seeds
        self.set_random_seeds()

    def set_random_seeds(self):
        """Set all random seeds for reproducibility"""
        random.seed(42)
        np.random.seed(42)
        os.environ["PYTHONHASHSEED"] = "42"

    # ===== CORE FEATURES: ~111 RESEARCH-BACKED FEATURES =====
    # Original 94 + 17 new features from RF/GBDT/SVM top-10 analysis

    def get_core_feature_names(self):
        """Define ~111 features including research-backed top-performers (Date, CLINCH, GROUND, CTRL, REV)"""
        return [
            # TIER 0: ELO Features (3 features) - MOST IMPORTANT
            "elo_diff",
            "r_elo_pre_fight",
            "b_elo_pre_fight",

            # TIER 1: Core Performance (15 features)
            "recent_form_diff_corrected",
            "win_streak_diff_corrected",
            "loss_streak_diff_corrected",
            "net_striking_advantage",
            "striking_efficiency",
            "defensive_striking",
            "grappling_control",
            "finish_rate_diff",
            "win_loss_ratio_diff_corrected",
            "last_5_wins_diff_corrected",
            "pro_SLpM_diff_corrected",
            "pro_SApM_diff_corrected",
            "pro_td_avg_diff_corrected",
            "ko_rate_diff_corrected",
            "sub_rate_diff_corrected",

            # TIER 2: Physical/Style (13 features - expanded stance features)
            "reach_diff",
            "age_at_event_diff",
            "orthodox_vs_southpaw_advantage",  # Directional: +1 if red ortho & blue south, -1 if opposite
            "orthodox_vs_switch_advantage",    # Directional: +1 if red ortho & blue switch, -1 if opposite
            "southpaw_vs_switch_advantage",    # Directional: +1 if red south & blue switch, -1 if opposite
            "mirror_matchup",                  # Symmetric: 1 if both same stance, 0 otherwise
            "striker_vs_grappler",
            "experience_gap",
            "height_diff",
            "ape_index_diff",
            "weight_diff",
            "striker_advantage",
            "grappler_advantage",

            # TIER 3: Context (8 features)
            "opponent_quality_diff",
            "days_since_last_fight_diff_corrected",
            "is_title_bout",
            "total_rounds",
            "r_trajectory_3",
            "b_trajectory_3",
            "ring_rust_factor",
            "weight_class_factor",

            # TIER 4: Advanced (15-20 features)
            "momentum_swing",
            "style_clash_severity",
            "power_vs_technique",
            "finish_pressure",
            "upset_potential",
            "pro_str_def_diff_corrected",
            "pro_td_def_diff_corrected",
            "pro_sig_str_acc_diff_corrected",
            "recent_finish_rate_diff_corrected",
            "durability_diff_corrected",
            "h2h_advantage",
            "clutch_factor_diff",
            "momentum_quality_diff",
            "pressure_performance_diff",
            "form_consistency_diff",

            # Additional high-value features
            "offensive_output",
            "defensive_composite",
            "ko_specialist_gap",
            "submission_specialist_gap",
            "skill_momentum",

            # Interaction features (combining key metrics)
            "elo_x_form",
            "reach_x_striking",
            "experience_x_age",

            # Advanced interaction features (data-driven from importance analysis)
            "age_x_striking",
            "age_x_grappling",
            "age_x_durability",
            "td_x_defense",
            "grappling_x_experience",
            "striking_x_accuracy",
            "striking_x_defense",
            "form_x_experience",
            "finish_x_momentum",
            "height_x_reach",
            "physical_x_striking",

            # Elite interaction features (based on top 20 analysis)
            "elo_x_win_ratio",
            "win_ratio_x_form",
            "durability_x_striking",
            "elo_x_durability",
            "submission_x_grappling",
            "ko_power_x_striking",
            "momentum_x_win_streak",
            "streak_differential",

            # NEW: Strategic interactions based on 60.79% model analysis
            "age_x_win_streak",
            "elo_x_sub_threat",
            "form_x_durability",
            "striking_x_grappling_matchup",
            "momentum_combo",

            # NEW HIGH-IMPACT: Elite compound features
            "elite_finisher",
            "veteran_advantage",
            "complete_fighter",
            "total_finish_threat",
            "unstoppable_streak",
            "age_prime_advantage",
            "freshness_advantage",
            "desperation_diff",
            "momentum_sustainability_diff",
            "elo_x_finish",
            "adversity_experience_diff",

            # RESEARCH-BACKED: Top features from RF/GBDT/SVM analysis
            # NEW: Opponent quality (was placeholder returning 0!)
            "avg_opponent_elo_diff_corrected",

            # NEW: Recent momentum vs career average (improving/declining fighters)
            "distance_pct_momentum_diff_corrected",
            "slpm_momentum_diff_corrected",
            "ctrl_sec_momentum_diff_corrected",

            # NEW: Rounds-based strategy adjustments
            "rounds_x_cardio",
            "rounds_x_finish_rate",
            "rounds_x_durability",

            # NEW: Enhanced top feature interactions
            "h2h_x_elo",
            "h2h_x_form",
            "age_prime_score_diff",
            "age_x_experience",
            "win_ratio_x_finish",
            "win_ratio_x_durability",

            # Positional striking CAREER AVERAGES (CLINCH #3, GROUND #5, DISTANCE #6)
            "distance_pct_diff_corrected",
            "clinch_pct_diff_corrected",
            "ground_pct_diff_corrected",
            "positional_striking_advantage",

            # Target area CAREER AVERAGES distribution
            "head_pct_diff_corrected",
            "body_pct_diff_corrected",
            "leg_pct_diff_corrected",
            "target_distribution_advantage",

            # Control and reversals CAREER AVERAGES (CTRL #5 GBDT, REV #8/#9)
            "avg_ctrl_sec_diff_corrected",
            "avg_rev_diff_corrected",
            "control_dominance",

            # Compound features with career-based positionals
            "clinch_x_grappling",
            "distance_x_striking",
            "ground_x_control",
            "positional_mastery",

            # NEW FEATURE: Main event experience
            "main_event_experience_diff",  # Count of 5-round fights (championship experience)

            # ========== PHASE 1A: POLYNOMIAL FEATURES (42 features) ==========
            # Squared terms for non-linear effects
            "elo_diff_squared",
            "age_diff_squared",
            "age_diff_cubic",
            "win_loss_ratio_diff_squared",
            "reach_diff_squared",
            "recent_form_diff_squared",
            "win_streak_diff_squared",
            "momentum_swing_squared",
            "pro_SLpM_diff_squared",
            "pro_sig_str_acc_diff_squared",
            "pro_td_avg_diff_squared",
            "net_striking_advantage_squared",
            "height_diff_squared",
            "weight_diff_squared",
            "ape_index_diff_squared",
            "experience_gap_squared",
            "days_since_last_fight_squared",
            "ko_rate_diff_squared",
            "sub_rate_diff_squared",
            "finish_rate_diff_squared",
            "pro_str_def_diff_squared",
            "pro_td_def_diff_squared",
            "durability_diff_squared",
            "clinch_pct_diff_squared",
            "ground_pct_diff_squared",
            "distance_pct_diff_squared",
            "avg_ctrl_sec_diff_squared",
            "elo_x_form_squared",
            "striker_advantage_squared",
            "grappler_advantage_squared",
            "slpm_momentum_diff_squared",
            "ctrl_sec_momentum_diff_squared",
            "elite_finisher_squared",
            "veteran_advantage_squared",
            "unstoppable_streak_squared",
            "h2h_advantage_squared",
            "avg_opponent_elo_diff_squared",
            "main_event_experience_diff_squared",
            # Cubic terms for complex curves
            "elo_diff_cubic",
            "win_loss_ratio_diff_cubic",
            "reach_diff_cubic",
            "experience_gap_cubic",

            # ========== PHASE 1B: ROLLING STATISTICS FEATURES (27 features) ==========
            # Rolling averages (3/5/10 fight windows)
            "rolling_slpm_3_diff", "rolling_slpm_5_diff", "rolling_slpm_10_diff",
            "rolling_sapm_3_diff", "rolling_sapm_5_diff", "rolling_sapm_10_diff",
            "rolling_td_acc_3_diff", "rolling_td_acc_5_diff", "rolling_td_acc_10_diff",
            "rolling_td_def_3_diff", "rolling_td_def_5_diff", "rolling_td_def_10_diff",
            "rolling_ctrl_3_diff", "rolling_ctrl_5_diff", "rolling_ctrl_10_diff",
            "rolling_finish_rate_3_diff", "rolling_finish_rate_5_diff", "rolling_finish_rate_10_diff",
            "rolling_strike_acc_3_diff", "rolling_strike_acc_5_diff", "rolling_strike_acc_10_diff",
            "rolling_damage_3_diff", "rolling_damage_5_diff", "rolling_damage_10_diff",
            # Variance/consistency metrics
            "slpm_variance_5_diff",
            "sapm_variance_5_diff",
            "performance_consistency_5_diff",
            # Trend features (improving/declining)
            "slpm_trend_5_diff",
            "td_acc_trend_5_diff",
            "finish_rate_trend_5_diff",

            # ========== PHASE 1C: OPPONENT-ADJUSTED FEATURES (17 features) ==========
            # Base opponent-adjusted metrics (9 features)
            "win_rate_vs_elite_diff",
            "win_rate_vs_strikers_diff",
            "win_rate_vs_grapplers_diff",
            "win_rate_vs_durable_diff",
            "win_rate_vs_finishers_diff",
            "finish_rate_vs_elite_diff",
            "recent_opponent_quality_5_diff",
            "style_versatility_diff",
            "step_up_performance_diff",
            # Interaction features (6 features)
            "step_up_x_elo",
            "versatility_x_form",
            "elite_wins_x_opp_quality",
            "elite_finish_x_power",
            "striker_killer_metric",
            "grappler_killer_metric",
            # Composite metrics (2 features)
            "championship_readiness",
            "competition_level",

            # ========== PHASE 1D: STATISTICAL RATIO FEATURES (29 features) ==========
            # Efficiency ratios (6 features)
            "striking_output_quality_diff",
            "grappling_output_quality_diff",
            "damage_ratio_diff",
            "defense_offense_balance_diff",
            "td_defense_offense_balance_diff",
            "finish_efficiency_diff",
            # Quality over quantity (3 features)
            "precision_striking_diff",
            "quality_grappling_diff",
            "submission_threat_ratio_diff",
            # Defensive efficiency (3 features)
            "damage_absorption_efficiency_diff",
            "total_defense_index_diff",
            "defense_versatility_diff",
            # Offensive versatility (3 features)
            "total_offense_index_diff",
            "offensive_versatility_diff",
            "striker_index_diff",
            # Win quality (3 features)
            "win_loss_ratio_squared_diff",
            "experience_quality_diff",
            "win_efficiency_diff",
            # Recent form ratios (1 feature)
            "recent_form_ratio_diff",
            # Physical efficiency (3 features)
            "reach_efficiency_diff",
            "size_adjusted_striking_diff",
            "size_adjusted_grappling_diff",
            # Advanced composites (4 features)
            "complete_fighter_index_diff",
            "pressure_fighter_index_diff",
            "counter_fighter_index_diff",
            "finishing_threat_composite_diff",
            # Rolling performance ratios (3 features)
            "recent_vs_career_striking_diff",
            "striking_consistency_ratio_diff",
            "improvement_trajectory_ratio_diff",

            # ========== PHASE 2: ADVANCED CAREER & MOMENTUM FEATURES (53 features) ==========
            # Base differentials for new tracked stats (12 features)
            "total_rounds_fought_diff_corrected",
            "total_fights_fought_diff_corrected",
            "title_fights_diff_corrected",
            "five_round_fights_diff_corrected",
            "last_fight_was_finish_diff_corrected",
            "last_fight_was_win_diff_corrected",
            "last_fight_dominance_diff_corrected",
            "early_finish_rate_diff_corrected",
            "late_finish_rate_diff_corrected",
            "first_round_ko_rate_diff_corrected",
            "fights_last_24_months_diff_corrected",
            "chin_deterioration_diff_corrected",

            # Career stage features (4 features)
            "prime_years_advantage",
            "declining_phase_diff",
            "rising_prospect_advantage",
            "age_experience_ratio_diff",

            # Last fight momentum features (3 features)
            "last_fight_finish_momentum",
            "last_fight_dominance",
            "last_fight_complete_momentum",

            # Damage trends (1 feature)
            "chin_deterioration",

            # Fight context features (3 features)
            "title_fight_experience",
            "five_round_experience",
            "five_round_cardio_advantage",

            # Activity & ring rust features (3 features)
            "activity_level_diff",
            "layoff_severity_diff",
            "optimal_activity_advantage",

            # Finish timing specialization (4 features)
            "early_finisher_advantage",
            "late_finisher_advantage",
            "first_round_killer_advantage",
            "finish_timing_rounds_matchup",

            # Advanced momentum features (1 feature)
            "finish_momentum_acceleration",

            # Grappling efficiency (1 feature)
            "ctrl_time_per_td_diff",

            # Pace & pressure features (2 features)
            "total_output_pace_diff",
            "pressure_differential",

            # Physical interactions (1 feature)
            "reach_height_interaction_diff",

            # Polynomial features for PHASE 2 stats (18 features)
            "total_rounds_fought_diff_squared",
            "total_fights_fought_diff_squared",
            "title_fights_diff_squared",
            "five_round_fights_diff_squared",
            "last_fight_dominance_squared",
            "last_fight_complete_momentum_squared",
            "early_finish_rate_diff_squared",
            "late_finish_rate_diff_squared",
            "first_round_ko_rate_diff_squared",
            "fights_last_24_months_diff_squared",
            "prime_years_advantage_squared",
            "age_experience_ratio_diff_squared",
            "chin_deterioration_squared",
            "layoff_severity_diff_squared",
            "finish_momentum_acceleration_squared",
            "ctrl_time_per_td_diff_squared",
            "total_output_pace_diff_squared",
            "pressure_differential_squared",

            # ========== PHASE 3: ADVANCED OPPONENT QUALITY, VOLATILITY & INTERACTION FEATURES (27 features) ==========
            # Base differentials (20 features)
            "avg_opponent_elo_l5_diff",
            "elo_momentum_vs_competition_diff",
            "performance_vs_ranked_opponents_diff",
            "performance_volatility_l10_diff",
            "finish_rate_acceleration_diff",
            "slpm_coefficient_of_variation_diff",
            "performance_decline_velocity_diff",
            "mileage_adjusted_age_diff",
            "prime_exit_risk_diff",
            "career_inflection_point_diff",
            "title_shot_proximity_score_diff",
            "tactical_evolution_score_diff",
            "finish_method_diversity_diff",
            "aging_power_striker_penalty_diff",
            "elo_volatility_interaction_diff",
            "layoff_veteran_interaction_diff",
            "bayesian_finish_rate_diff",
            "confidence_weighted_damage_ratio_diff",
            "distance_from_career_peak_diff",
            "elite_performance_frequency_l10_diff",
            # Polynomial features (7 features)
            "avg_opponent_elo_l5_diff_squared",
            "elo_momentum_vs_competition_diff_squared",
            "performance_decline_velocity_diff_squared",
            "mileage_adjusted_age_diff_squared",
            "layoff_veteran_interaction_diff_squared",
            "performance_volatility_l10_diff_squared",
            "finish_rate_acceleration_diff_squared",
        ]

    def calculate_streak(self, recent_wins, count_wins=True):
        """Calculate current win or loss streak"""
        if not recent_wins:
            return 0
        streak = 0
        target = 1 if count_wins else 0
        for result in reversed(recent_wins):
            if result == target:
                streak += 1
            else:
                break
        return streak

    def calculate_trajectory(self, values):
        """Calculate linear trend of recent performance"""
        if len(values) < 2:
            return 0
        try:
            x = np.arange(len(values))
            slope, _, _, _, _ = linregress(x, values)
            return slope
        except:
            return 0

    def calculate_age_curve_factor(self, age):
        """Calculate performance multiplier based on age"""
        if age < 25:
            return 0.92
        elif 25 <= age <= 32:
            return 1.0
        elif 33 <= age <= 36:
            return 0.94
        else:
            return 0.85

    # ===== ELO RATING SYSTEM (COPIED FROM UFC ELO SYSTEM) =====

    def elo_expected_score(self, fighter_elo, opponent_elo):
        """
        Calculate expected score using ELO formula
        Returns probability of winning (0-1)
        """
        return 1 / (1 + 10 ** ((opponent_elo - fighter_elo) / 400))

    def elo_calculate_k_factor(self, row, winner_fighter, fighter_result, fighter_elos, elo_history):
        """
        Calculate dynamic K-factor based on fight circumstances

        K-factor determines how much ELO changes after a fight.
        Higher K-factor means more volatile ratings.
        """
        k_base = 40  # Increased K-factor for wider ELO spread and better differentiation
        k = k_base

        # 1. CHAMPIONSHIP BOUT WEIGHT (title fights matter more)
        is_title = row.get('is_title_bout', 0) == 1
        if is_title:
            k *= 2.0
            if fighter_result == 'Win':
                k *= 1.5  # Winning title fights is even more valuable

        # 2. RECENCY WEIGHT (more recent fights are weighted higher)
        # Calculate days since fight
        if 'event_date' in row.index and pd.notna(row['event_date']):
            from datetime import datetime
            event_date = pd.to_datetime(row['event_date'])
            days_ago = (datetime.now() - event_date).days
            # Apply exponential decay: fights in last year = 100%, 2 years ago = 75%, etc.
            recency_multiplier = np.exp(-days_ago / 365 / 2)  # Half-life of 2 years
            if recency_multiplier < 0.5:  # Cap at 50% for very old fights
                recency_multiplier = 0.5
            k *= recency_multiplier

        # 3. DOMINANCE BONUS (finishing early matters more)
        if 'finish_round' in row.index and pd.notna(row['finish_round']):
            round_finished = row['finish_round']
            if round_finished <= 1:
                k *= 1.5  # First round finish
            elif round_finished == 2:
                k *= 1.3  # Second round finish
            elif round_finished == 3:
                k *= 1.2  # Third round finish

        # 4. OPPONENT STRENGTH ADJUSTMENT
        opponent = row['b_fighter'] if row['r_fighter'] == winner_fighter else row['r_fighter']
        fighter_elo = fighter_elos.get(winner_fighter, 1500)

        if opponent in fighter_elos:
            opponent_elo = fighter_elos[opponent]

            # Bonus for beating above-average opponents
            if opponent_elo > 1500:
                if opponent_elo > 1600:
                    k *= 1.2
                if opponent_elo > 1700:
                    k *= 1.2

            # Upset bonuses/penalties
            elo_diff = fighter_elo - opponent_elo
            if fighter_result == 'Win' and elo_diff < -100:
                # Upset bonus: beating much higher rated opponent
                upset_factor = abs(elo_diff) / 100 * 0.1
                k *= (1 + upset_factor)
            elif fighter_result == 'Loss' and elo_diff > 100:
                # Upset penalty: losing to much lower rated opponent
                upset_penalty = elo_diff / 100 * 0.15
                k *= (1 + upset_penalty)

        # 5. WIN/LOSS STREAK MULTIPLIER
        fighter_streak = self.elo_get_recent_streak(winner_fighter, elo_history)
        if fighter_streak >= 3:  # Win streak
            k *= 1.15
        elif fighter_streak <= -3:  # Loss streak
            k *= 1.25

        # 6. METHOD OF VICTORY BONUS
        if fighter_result == 'Win' and 'method' in row.index and pd.notna(row['method']):
            method = str(row['method']).upper()
            if 'KO' in method or 'TKO' in method:
                k *= 1.1  # Finishing via KO is more impressive
            elif 'SUBMISSION' in method:
                k *= 1.05  # Submissions are also dominant

        return k

    def elo_get_recent_streak(self, fighter, elo_history):
        """Get fighter's recent win/loss streak from ELO history"""
        if fighter not in elo_history or len(elo_history[fighter]) == 0:
            return 0

        # Look at last 5 results
        history = elo_history[fighter][-5:]

        if len(history) == 0:
            return 0

        streak = 0
        for entry in reversed(history):
            result = entry.get('result', None)
            if result == 'Win':
                streak = (streak if streak > 0 else 0) + 1
            elif result == 'Loss':
                streak = (streak if streak < 0 else 0) - 1
            else:
                streak = 0

        return streak

    def elo_update_ratings(self, r_fighter, b_fighter, winner, k_factor, fighter_elos, elo_history, fight_date):
        """
        Update ELO ratings for both fighters after a fight

        Args:
            r_fighter: Red corner fighter name
            b_fighter: Blue corner fighter name
            winner: 'Red', 'Blue', or 'Draw'
            k_factor: K-factor for this fight
            fighter_elos: Dict of current ELO ratings
            elo_history: Dict of fighter history
            fight_date: Date of the fight
        """
        # Get current ELOs
        r_elo = fighter_elos.get(r_fighter, 1500)
        b_elo = fighter_elos.get(b_fighter, 1500)

        # Calculate expected scores
        r_expected = self.elo_expected_score(r_elo, b_elo)
        b_expected = self.elo_expected_score(b_elo, r_elo)

        # Determine actual scores based on winner
        if winner == 'Red':
            r_actual = 1.0
            b_actual = 0.0
            r_result = 'Win'
            b_result = 'Loss'
        elif winner == 'Blue':
            r_actual = 0.0
            b_actual = 1.0
            r_result = 'Loss'
            b_result = 'Win'
        elif winner == 'Draw':
            r_actual = 0.5
            b_actual = 0.5
            r_result = 'Draw'
            b_result = 'Draw'
            k_factor *= 0.5  # Draws have less impact
        else:
            # No Contest or invalid
            return

        # Calculate ELO changes
        r_change = k_factor * (r_actual - r_expected)
        b_change = k_factor * (b_actual - b_expected)

        # Update ELO ratings
        fighter_elos[r_fighter] = r_elo + r_change
        fighter_elos[b_fighter] = b_elo + b_change

        # Record in history
        if r_fighter not in elo_history:
            elo_history[r_fighter] = []
        if b_fighter not in elo_history:
            elo_history[b_fighter] = []

        elo_history[r_fighter].append({
            'date': fight_date,
            'elo': fighter_elos[r_fighter],
            'result': r_result
        })
        elo_history[b_fighter].append({
            'date': fight_date,
            'elo': fighter_elos[b_fighter],
            'result': b_result
        })

    def calculate_elo_ratings(self, df):
        """
        Calculate ELO ratings for all fighters chronologically
        This ensures no data leakage - each fight uses pre-fight ELO ratings

        Returns:
            df: DataFrame with ELO columns added (r_elo_pre_fight, b_elo_pre_fight, elo_diff)
        """
        print("\n" + "="*80)
        print("CALCULATING ELO RATINGS (CHRONOLOGICAL - NO DATA LEAKAGE)")
        print("="*80 + "\n")

        # Initialize ELO system
        initial_elo = 1500
        fighter_elos = {}
        elo_history = {}

        # Sort by date to process chronologically
        df = df.sort_values('event_date').reset_index(drop=True)

        # Initialize ELO columns
        df['r_elo_pre_fight'] = 0.0
        df['b_elo_pre_fight'] = 0.0
        df['elo_diff'] = 0.0
        df['r_prob_pre_fight'] = 0.0
        df['b_prob_pre_fight'] = 0.0

        for idx, row in df.iterrows():
            if idx % 500 == 0:
                print(f"   Processing fight {idx}/{len(df)} for ELO...")

            r_fighter = row['r_fighter']
            b_fighter = row['b_fighter']

            # Initialize fighters if not seen before
            if r_fighter not in fighter_elos:
                fighter_elos[r_fighter] = initial_elo
                elo_history[r_fighter] = []
            if b_fighter not in fighter_elos:
                fighter_elos[b_fighter] = initial_elo
                elo_history[b_fighter] = []

            # CRITICAL: Capture PRE-FIGHT ELO ratings (before this fight is processed)
            r_elo_pre = fighter_elos[r_fighter]
            b_elo_pre = fighter_elos[b_fighter]

            # Calculate pre-fight probabilities
            r_prob = self.elo_expected_score(r_elo_pre, b_elo_pre)
            b_prob = self.elo_expected_score(b_elo_pre, r_elo_pre)

            # Store pre-fight ELO ratings and probabilities
            df.at[idx, 'r_elo_pre_fight'] = r_elo_pre
            df.at[idx, 'b_elo_pre_fight'] = b_elo_pre
            df.at[idx, 'elo_diff'] = r_elo_pre - b_elo_pre
            df.at[idx, 'r_prob_pre_fight'] = r_prob
            df.at[idx, 'b_prob_pre_fight'] = b_prob

            # Now update ELO ratings based on fight result
            winner = row.get('winner')

            if pd.notna(winner) and winner in ['Red', 'Blue', 'Draw']:
                # Determine who won for K-factor calculation
                if winner == 'Red':
                    winner_fighter = r_fighter
                    fighter_result = 'Win'
                elif winner == 'Blue':
                    winner_fighter = b_fighter
                    fighter_result = 'Win'
                else:
                    winner_fighter = r_fighter
                    fighter_result = 'Draw'

                # Calculate K-factor
                k_factor = self.elo_calculate_k_factor(
                    row, winner_fighter, fighter_result, fighter_elos, elo_history
                )

                # Update ELO ratings (this modifies fighter_elos for future fights)
                self.elo_update_ratings(
                    r_fighter, b_fighter, winner, k_factor,
                    fighter_elos, elo_history, row['event_date']
                )

        print("\nELO calculation complete!")
        print(f"   Processed {len(df)} fights")
        print(f"   Tracked {len(fighter_elos)} unique fighters")
        print(f"   Average ELO: {np.mean(list(fighter_elos.values())):.1f}")
        print(f"   ELO Range: {min(fighter_elos.values()):.1f} - {max(fighter_elos.values()):.1f}")

        # Store fighter ELOs for prediction use
        self.fighter_elos = fighter_elos
        self.elo_history = elo_history

        return df

    def build_swapped(self, df):
        """
        Create swap-mirror version for antisymmetrization.

        This is a cleaner alternative to swap_corners that:
        1. Swaps r_* ↔ b_* columns
        2. Flips winner label
        3. Recomputes features WITHOUT recalculating ELO (preserves chronological integrity)
        4. Manually handles ELO parity (flip sign of elo_diff)
        5. Manually handles identity-based features (h2h_advantage)

        Used by directional_and_invariant() for automatic feature directionalization.
        """
        df_s = df.copy()

        # Swap r_* ↔ b_* columns
        for c in df.columns:
            if c.startswith("r_") and ("b_" + c[2:]) in df.columns:
                # Use .copy() to avoid SettingWithCopyWarning
                df_s[c], df_s["b_" + c[2:]] = df["b_" + c[2:]].copy(), df[c].copy()

        # Flip winner label
        if "winner" in df_s:
            df_s["winner"] = df["winner"].map({"Red": "Blue", "Blue": "Red"})

        # Recompute features from swapped bases, but DON'T recompute ELO
        # ELO is chronologically dependent - we handle it separately below
        df_s = self.build_features(df_s, recompute_elo=False)

        # Manual ELO parity (since ELO is chronologically ordered)
        # When corners swap: elo_diff_red = -elo_diff_blue
        if "elo_diff" in df:
            df_s["elo_diff"] = -df["elo_diff"].values
            if "r_elo_pre_fight" in df and "b_elo_pre_fight" in df:
                df_s["r_elo_pre_fight"] = df["b_elo_pre_fight"].values
                df_s["b_elo_pre_fight"] = df["r_elo_pre_fight"].values

        # Manual handling of identity-based features
        # h2h_advantage is fighter-pair specific, so it flips sign when corners swap
        if "h2h_advantage" in df_s:
            df_s["h2h_advantage"] = -df_s["h2h_advantage"]

        return df_s

    def directional_and_invariant(self, X, Xs):
        """
        Antisymmetrize features into directional and invariant components.

        Directional features (D): flip sign when corners swap → encode red-relative advantage
        Invariant features (I): stay same when corners swap → encode fight characteristics

        Math:
            D = 0.5 * (X_orig - X_swap)   # Antisymmetric: D(swap) = -D(orig)
            I = 0.5 * (X_orig + X_swap)   # Symmetric: I(swap) = I(orig)

        Args:
            X: Original feature matrix
            Xs: Swapped feature matrix (from build_swapped)

        Returns:
            D: Directional features (DataFrame)
            I: Invariant features (DataFrame)
        """
        # Find common numeric columns
        common = sorted(set(X.columns) & set(Xs.columns))
        Xo = X[common].select_dtypes(include='number')
        Xsw = Xs[common].select_dtypes(include='number')

        # Antisymmetrize: directional features flip on swap
        D = 0.5 * (Xo - Xsw)
        D.columns = [f"{c}" for c in D.columns]  # Keep original names

        # Symmetrize: invariant features stay same on swap
        I = 0.5 * (Xo + Xsw)
        I.columns = [f"{c}_inv" for c in I.columns]  # Mark as invariant

        return D, I

    # ===== FIGHTER-SPECIFIC ADVANCED FEATURES =====

    def calculate_fighter_vs_similar_opponents(self, df, fighter_col):
        """How has fighter performed against similar style opponents?"""
        performance_scores = []

        for idx, row in df.iterrows():
            fighter = row[fighter_col]
            if pd.isna(fighter):
                performance_scores.append(0.5)
                continue

            try:
                # Get all fights for this fighter
                fighter_fights = df[(df['r_fighter'] == fighter) | (df['b_fighter'] == fighter)]

                if len(fighter_fights) < 3:
                    performance_scores.append(0.5)
                    continue

                # Calculate recent performance vs striker/grappler types
                recent_fights = fighter_fights.tail(5)
                wins_vs_strikers = 0
                total_vs_strikers = 0

                for _, fight in recent_fights.iterrows():
                    opponent = fight['b_fighter'] if fight['r_fighter'] == fighter else fight['r_fighter']
                    opponent_prefix = 'b' if fight['r_fighter'] == fighter else 'r'

                    # Check if opponent is striker or grappler
                    slpm = fight.get(f'{opponent_prefix}_pro_SLpM_corrected', 0)
                    td_avg = fight.get(f'{opponent_prefix}_pro_td_avg_corrected', 0)

                    if slpm > td_avg * 2:  # Striker
                        total_vs_strikers += 1
                        if fight.get('winner') == ('Red' if fight['r_fighter'] == fighter else 'Blue'):
                            wins_vs_strikers += 1

                if total_vs_strikers > 0:
                    performance_scores.append(wins_vs_strikers / total_vs_strikers)
                else:
                    performance_scores.append(0.5)

            except:
                performance_scores.append(0.5)

        return np.array(performance_scores)

    def calculate_elite_competition_performance(self, df, fighter_col):
        """Recent performance vs elite competition (top-ranked opponents)"""
        elite_performance = []

        for idx, row in df.iterrows():
            fighter = row[fighter_col]
            if pd.isna(fighter):
                elite_performance.append(0.5)
                continue

            try:
                # Get fighter's recent fights
                fighter_fights = df[(df['r_fighter'] == fighter) | (df['b_fighter'] == fighter)]
                recent_fights = fighter_fights.tail(5)

                elite_wins = 0
                elite_fights_count = 0

                for _, fight in recent_fights.iterrows():
                    opponent = fight['b_fighter'] if fight['r_fighter'] == fighter else fight['r_fighter']
                    opponent_prefix = 'b' if fight['r_fighter'] == fighter else 'r'

                    # Consider elite if win rate > 70%
                    opponent_wins = fight.get(f'{opponent_prefix}_wins_corrected', 0)
                    opponent_losses = fight.get(f'{opponent_prefix}_losses_corrected', 1)
                    opponent_winrate = opponent_wins / max(opponent_wins + opponent_losses, 1)

                    if opponent_winrate > 0.7:  # Elite opponent
                        elite_fights_count += 1
                        if fight.get('winner') == ('Red' if fight['r_fighter'] == fighter else 'Blue'):
                            elite_wins += 1

                if elite_fights_count >= 2:
                    elite_performance.append(elite_wins / elite_fights_count)
                else:
                    elite_performance.append(0.5)

            except:
                elite_performance.append(0.5)

        return np.array(elite_performance)

    def calculate_fighter_improvement_rate(self, df, fighter_col):
        """Calculate if fighter is improving or declining"""
        improvement_rates = []

        for idx, row in df.iterrows():
            fighter = row[fighter_col]
            if pd.isna(fighter):
                improvement_rates.append(0.0)
                continue

            try:
                # Get all fights for this fighter
                fighter_fights = df[(df['r_fighter'] == fighter) | (df['b_fighter'] == fighter)]

                if len(fighter_fights) < 6:
                    improvement_rates.append(0.0)
                    continue

                # Compare first 1/3 vs last 1/3 of career
                third = len(fighter_fights) // 3
                early_fights = fighter_fights.head(third)
                recent_fights = fighter_fights.tail(third)

                # Calculate win rates
                early_wins = 0
                recent_wins = 0

                for _, fight in early_fights.iterrows():
                    if fight.get('winner') == ('Red' if fight['r_fighter'] == fighter else 'Blue'):
                        early_wins += 1

                for _, fight in recent_fights.iterrows():
                    if fight.get('winner') == ('Red' if fight['r_fighter'] == fighter else 'Blue'):
                        recent_wins += 1

                early_winrate = early_wins / len(early_fights)
                recent_winrate = recent_wins / len(recent_fights)

                improvement_rates.append(recent_winrate - early_winrate)

            except:
                improvement_rates.append(0.0)

        return np.array(improvement_rates)

    # ===== DATA LEAKAGE PREVENTION =====

    def fix_data_leakage(self, df):
        """Recalculate comprehensive fighter statistics chronologically"""
        print("\n" + "="*80)
        print("FIXING DATA LEAKAGE - CHRONOLOGICAL RECALCULATION")
        print("="*80 + "\n")

        import copy

        df["event_date"] = pd.to_datetime(df["event_date"])
        df = df.sort_values("event_date").reset_index(drop=True)

        # Initialize ELO tracking for PHASE 3 (must match calculate_elo_ratings logic)
        self.fighter_elos = {}
        self.elo_history = {}

        fighter_stats = {}

        stats_to_track = {
            "wins": 0, "losses": 0, "draws": 0,
            "sig_str_total": 0, "sig_str_att_total": 0, "sig_str_absorbed_total": 0,
            "str_def_success": 0, "str_def_att": 0,  # Striking defense tracking
            "td_total": 0, "td_att_total": 0, "td_def_success": 0, "td_def_att": 0,
            "sub_att_total": 0, "kd_total": 0, "kd_absorbed": 0,  # kd_absorbed = knockdowns received
            "fight_time_minutes": 0, "fight_count": 0,
            "ko_wins": 0, "sub_wins": 0, "dec_wins": 0,
            "ko_losses": 0, "sub_losses": 0,
            "recent_wins": [], "recent_finishes": [],
            "last_fight_date": None,
            # NEW: Positional striking tracking (distance/clinch/ground)
            "distance_strikes": 0, "clinch_strikes": 0, "ground_strikes": 0, "total_sig_strikes": 0,
            # NEW: Target area tracking (head/body/leg)
            "head_strikes": 0, "body_strikes": 0, "leg_strikes": 0,
            # NEW: Control and reversals
            "ctrl_sec_total": 0, "rev_total": 0,
            # NEW: Opponent quality tracking
            "total_opponent_elo": 0, "opponents_faced": 0,
            # NEW: Recent stats for momentum (last 3 fights)
            "recent_distance_pct": [], "recent_clinch_pct": [], "recent_ground_pct": [],
            "recent_slpm": [], "recent_ctrl_sec": [],
            # NEW: Main event experience (5-round fights)
            "main_event_fights": 0,
            # PHASE 1B: Rolling statistics tracking (lists of recent values)
            "rolling_slpm": [],  # Last 10 fights SLpM
            "rolling_sapm": [],  # Last 10 fights SApM
            "rolling_td_acc": [],  # Last 10 fights TD accuracy
            "rolling_td_def": [],  # Last 10 fights TD defense
            "rolling_ctrl_time": [],  # Last 10 fights control time
            "rolling_finishes": [],  # Last 10 fights finish indicator (1=finish, 0=decision)
            "rolling_strike_acc": [],  # Last 10 fights striking accuracy
            "rolling_damage": [],  # Last 10 fights damage (sig_str - sig_str_absorbed)
            # PHASE 1C: Opponent-adjusted performance tracking
            "vs_elite_record": {"wins": 0, "fights": 0},  # Record vs top 25% ELO opponents
            "vs_striker_record": {"wins": 0, "fights": 0},  # Record vs strikers (high SLpM)
            "vs_grappler_record": {"wins": 0, "fights": 0},  # Record vs grapplers (high TD avg)
            "vs_durable_record": {"wins": 0, "fights": 0},  # Record vs durable fighters
            "vs_finisher_record": {"wins": 0, "fights": 0},  # Record vs fighters with high finish rate
            "finish_vs_elite": {"finishes": 0, "fights": 0},  # Finish rate vs elite
            "recent_opponent_elos": [],  # Last 5 opponent ELOs
            "recent_opponent_styles": [],  # Last 5 opponent styles (striker=1, grappler=-1, balanced=0)
            # PHASE 2: Advanced career tracking
            "total_rounds_fought": 0,  # Cumulative rounds across all fights
            "total_fights_fought": 0,  # Total MMA fights (wins + losses + draws)
            "title_fights": 0,  # Number of title fights
            "five_round_fights": 0,  # Total 5-round fights (title + main events)
            # Last fight tracking for momentum
            "last_fight_method": None,  # Last fight finish method (KO/TKO, Submission, Decision)
            "last_fight_finish_round": None,  # Round fight ended (1-5)
            "last_fight_was_finish": False,  # Boolean: was it a finish?
            "last_fight_was_win": False,  # Boolean: did they win?
            "last_fight_time_seconds": 0,  # Time to finish (if finish)
            # Finish timing patterns
            "early_finishes": 0,  # Finishes in rounds 1-2
            "late_finishes": 0,  # Finishes in rounds 3-5
            "first_round_kos": 0,  # First round KO/TKOs specifically
            # Activity tracking
            "fight_dates": [],  # List of fight dates for activity calculation
            # Recent finish history (last 5 fights)
            "recent_finish_methods": [],  # List of last 5 finish methods
            "recent_finish_rounds": [],  # List of last 5 rounds fights ended
            # ELO rating (tracked chronologically)
            "elo": 1500,  # Current ELO rating (starts at 1500)
            # PHASE 3: Advanced opponent quality, volatility, and interaction features
            # CLUSTER 1: Opponent Quality Tracking
            "opponent_elo_history": [],  # Last 10 opponent ELOs
            "opponent_ranking_history": [],  # Last 10 opponent rankings (percentile)
            "damage_ratio_history": [],  # Last 10 damage ratios for volatility
            "slpm_history": [],  # Last 10 SLpM values for CV calculation
            # CLUSTER 2: Volatility Tracking
            "finish_rate_history_5": [],  # Track last 5 finish outcomes (1 or 0)
            "finish_rate_history_10": [],  # Track last 10 finish outcomes
            # Weight class tracking
            "weight_class_history": [],  # Track all weight classes fought in
            # CLUSTER 3: Career Peak Tracking
            "career_best_damage_ratio": 0.0,  # Best damage ratio ever achieved
            "elite_performance_count_l10": 0,  # Count of elite performances in last 10
            "damage_ratio_threshold_80th": None,  # 80th percentile of own damage ratios
            # CLUSTER 4: Fight Urgency Tracking
            "current_losing_streak": 0,  # Current consecutive losses
            # CLUSTER 6: Interaction Terms (uses existing data, no new tracking needed)
            # CLUSTER 7: Bayesian Tracking (uses existing data, no new tracking needed)
        }

        # Initialize corrected columns
        for prefix in ["r", "b"]:
            for stat in ["wins", "losses", "win_loss_ratio", "pro_SLpM", "pro_SApM",
                        "pro_sig_str_acc", "pro_str_def", "pro_td_avg", "pro_td_acc",
                        "pro_td_def", "pro_sub_avg", "ko_rate", "sub_rate", "dec_rate",
                        "recent_form", "win_streak", "loss_streak", "last_5_wins",
                        "days_since_last_fight", "recent_finish_rate", "durability",
                        # NEW: Career-based positional/target/control metrics
                        "distance_pct", "clinch_pct", "ground_pct",
                        "head_pct", "body_pct", "leg_pct",
                        "avg_ctrl_sec", "avg_rev",
                        # NEW: Opponent quality
                        "avg_opponent_elo",
                        # NEW: Recent momentum vs career
                        "distance_pct_momentum", "slpm_momentum", "ctrl_sec_momentum",
                        # NEW: Main event experience
                        "main_event_fights",
                        # PHASE 1B: Rolling statistics (averages over last 3/5/10 fights)
                        "rolling_slpm_3", "rolling_slpm_5", "rolling_slpm_10",
                        "rolling_sapm_3", "rolling_sapm_5", "rolling_sapm_10",
                        "rolling_td_acc_3", "rolling_td_acc_5", "rolling_td_acc_10",
                        "rolling_td_def_3", "rolling_td_def_5", "rolling_td_def_10",
                        "rolling_ctrl_3", "rolling_ctrl_5", "rolling_ctrl_10",
                        "rolling_finish_rate_3", "rolling_finish_rate_5", "rolling_finish_rate_10",
                        "rolling_strike_acc_3", "rolling_strike_acc_5", "rolling_strike_acc_10",
                        "rolling_damage_3", "rolling_damage_5", "rolling_damage_10",
                        # Variance/consistency metrics
                        "slpm_variance_5", "sapm_variance_5",
                        "performance_consistency_5",
                        # Trend features (improving/declining)
                        "slpm_trend_5", "td_acc_trend_5", "finish_rate_trend_5",
                        # PHASE 1C: Opponent-adjusted metrics
                        "win_rate_vs_elite", "win_rate_vs_strikers", "win_rate_vs_grapplers",
                        "win_rate_vs_durable", "win_rate_vs_finishers", "finish_rate_vs_elite",
                        "recent_opponent_quality_5", "style_versatility", "step_up_performance",
                        # PHASE 2: Advanced career metrics
                        "total_rounds_fought", "total_fights_fought", "title_fights", "five_round_fights",
                        "last_fight_was_finish", "last_fight_was_win", "last_fight_dominance",
                        "early_finish_rate", "late_finish_rate", "first_round_ko_rate",
                        "fights_last_24_months", "avg_finish_time_last_3", "chin_deterioration",
                        # PHASE 3: Advanced features
                        "avg_opponent_elo_l5", "elo_momentum_vs_competition",
                        "performance_vs_ranked_opponents", "performance_volatility_l10",
                        "finish_rate_acceleration", "slpm_coefficient_of_variation",
                        "performance_decline_velocity", "mileage_adjusted_age",
                        "prime_exit_risk", "career_inflection_point",
                        "title_shot_proximity_score", "tactical_evolution_score",
                        "finish_method_diversity", "aging_power_striker_penalty",
                        "elo_volatility_interaction", "layoff_veteran_interaction",
                        "bayesian_finish_rate", "confidence_weighted_damage_ratio",
                        "distance_from_career_peak", "elite_performance_frequency_l10"]:
                df[f"{prefix}_{stat}_corrected"] = 0.0

        df["h2h_advantage"] = 0.0
        df["opponent_quality_diff"] = 0.0
        fighter_h2h = {}

        for idx, row in df.iterrows():
            if idx % 500 == 0:
                print(f"   Processing fight {idx}/{len(df)}...")

            r_fighter, b_fighter = row["r_fighter"], row["b_fighter"]

            # H2H tracking
            h2h_key = (r_fighter, b_fighter)
            h2h_key_reverse = (b_fighter, r_fighter)

            if h2h_key in fighter_h2h:
                df.at[idx, "h2h_advantage"] = fighter_h2h[h2h_key]
            elif h2h_key_reverse in fighter_h2h:
                df.at[idx, "h2h_advantage"] = -fighter_h2h[h2h_key_reverse]

            if r_fighter not in fighter_stats:
                fighter_stats[r_fighter] = copy.deepcopy(stats_to_track)
            if b_fighter not in fighter_stats:
                fighter_stats[b_fighter] = copy.deepcopy(stats_to_track)

            # Record stats before fight
            for fighter, prefix in [(r_fighter, "r"), (b_fighter, "b")]:
                stats = fighter_stats[fighter]
                df.at[idx, f"{prefix}_wins_corrected"] = stats["wins"]
                df.at[idx, f"{prefix}_losses_corrected"] = stats["losses"]
                df.at[idx, f"{prefix}_win_loss_ratio_corrected"] = stats["wins"] / max(stats["losses"], 1)

                total_fights = stats["wins"] + stats["losses"]
                if total_fights > 0:
                    df.at[idx, f"{prefix}_ko_rate_corrected"] = stats["ko_wins"] / total_fights
                    df.at[idx, f"{prefix}_sub_rate_corrected"] = stats["sub_wins"] / total_fights
                    df.at[idx, f"{prefix}_dec_rate_corrected"] = stats["dec_wins"] / total_fights

                # Recent form
                if len(stats["recent_wins"]) >= 5:
                    df.at[idx, f"{prefix}_recent_form_corrected"] = sum(stats["recent_wins"][-5:]) / 5
                elif len(stats["recent_wins"]) >= 3:
                    df.at[idx, f"{prefix}_recent_form_corrected"] = sum(stats["recent_wins"][-3:]) / 3

                # Streaks
                df.at[idx, f"{prefix}_win_streak_corrected"] = self.calculate_streak(stats["recent_wins"], True)
                df.at[idx, f"{prefix}_loss_streak_corrected"] = self.calculate_streak(stats["recent_wins"], False)

                # Last 5 wins count
                if len(stats["recent_wins"]) >= 5:
                    df.at[idx, f"{prefix}_last_5_wins_corrected"] = sum(stats["recent_wins"][-5:])
                elif len(stats["recent_wins"]) > 0:
                    df.at[idx, f"{prefix}_last_5_wins_corrected"] = sum(stats["recent_wins"])

                # Trajectory (momentum indicator based on last 3 fights, weighted by recency)
                # Converts wins/losses to 1/-1, weights recent fights more heavily
                if len(stats["recent_wins"]) >= 3:
                    last_3 = stats["recent_wins"][-3:]
                    # Convert 1/0 to 1/-1 for wins/losses
                    weighted = [(1 if w else -1) for w in last_3]
                    # Weight by position: oldest=1x, middle=2x, newest=3x
                    df.at[idx, f"{prefix}_trajectory_3"] = (weighted[0] * 1) + (weighted[1] * 2) + (weighted[2] * 3)
                elif len(stats["recent_wins"]) > 0:
                    # For fighters with < 3 fights, use what we have
                    trajectory = sum([(1 if w else -1) * (i + 1) for i, w in enumerate(stats["recent_wins"])])
                    df.at[idx, f"{prefix}_trajectory_3"] = trajectory

                # Days since last fight
                if stats["last_fight_date"]:
                    days_off = (row["event_date"] - stats["last_fight_date"]).days
                    df.at[idx, f"{prefix}_days_since_last_fight_corrected"] = days_off

                # Per-minute stats
                if stats["fight_time_minutes"] > 0:
                    df.at[idx, f"{prefix}_pro_SLpM_corrected"] = stats["sig_str_total"] / stats["fight_time_minutes"]
                    df.at[idx, f"{prefix}_pro_SApM_corrected"] = stats["sig_str_absorbed_total"] / stats["fight_time_minutes"]
                    df.at[idx, f"{prefix}_pro_td_avg_corrected"] = (stats["td_total"] / stats["fight_time_minutes"]) * 15
                    df.at[idx, f"{prefix}_pro_sub_avg_corrected"] = (stats["sub_att_total"] / stats["fight_time_minutes"]) * 15

                if stats["sig_str_att_total"] > 0:
                    df.at[idx, f"{prefix}_pro_sig_str_acc_corrected"] = stats["sig_str_total"] / stats["sig_str_att_total"]

                if stats["td_att_total"] > 0:
                    df.at[idx, f"{prefix}_pro_td_acc_corrected"] = stats["td_total"] / stats["td_att_total"]

                if stats["td_def_att"] > 0:
                    df.at[idx, f"{prefix}_pro_td_def_corrected"] = stats["td_def_success"] / stats["td_def_att"]

                # Striking defense percentage
                if stats["str_def_att"] > 0:
                    df.at[idx, f"{prefix}_pro_str_def_corrected"] = stats["str_def_success"] / stats["str_def_att"]

                # Durability
                finish_losses = stats["ko_losses"] + stats["sub_losses"]
                df.at[idx, f"{prefix}_durability_corrected"] = 1.0 / (1 + finish_losses)

                # Recent finish rate (last 5 fights)
                if len(stats["recent_finishes"]) >= 5:
                    df.at[idx, f"{prefix}_recent_finish_rate_corrected"] = sum(stats["recent_finishes"][-5:]) / 5
                elif len(stats["recent_finishes"]) >= 3:
                    df.at[idx, f"{prefix}_recent_finish_rate_corrected"] = sum(stats["recent_finishes"][-3:]) / 3
                elif len(stats["recent_finishes"]) > 0:
                    df.at[idx, f"{prefix}_recent_finish_rate_corrected"] = sum(stats["recent_finishes"]) / len(stats["recent_finishes"])

                # Weight class comfort: 1.0 if at natural weight, penalty if moving up/down
                # Natural weight = most common weight class in history
                current_weight = row.get("weight_class", "Unknown")
                if len(stats["weight_class_history"]) > 0 and pd.notna(current_weight):
                    from collections import Counter
                    weight_counts = Counter(stats["weight_class_history"])
                    natural_weight = weight_counts.most_common(1)[0][0]

                    if current_weight == natural_weight:
                        weight_comfort = 1.0  # At natural weight
                    else:
                        # Penalty for fighting outside natural weight class
                        # Larger penalty if they have little experience at this weight
                        fights_at_current = weight_counts.get(current_weight, 0)
                        total_fights = len(stats["weight_class_history"])
                        weight_comfort = 1.0 - (0.15 * (1 - fights_at_current / total_fights))  # 0.85 to 1.0
                    df.at[idx, f"{prefix}_weight_comfort"] = weight_comfort

                # NEW: Career-based positional striking percentages
                if stats["total_sig_strikes"] > 0:
                    df.at[idx, f"{prefix}_distance_pct_corrected"] = stats["distance_strikes"] / stats["total_sig_strikes"]
                    df.at[idx, f"{prefix}_clinch_pct_corrected"] = stats["clinch_strikes"] / stats["total_sig_strikes"]
                    df.at[idx, f"{prefix}_ground_pct_corrected"] = stats["ground_strikes"] / stats["total_sig_strikes"]
                    df.at[idx, f"{prefix}_head_pct_corrected"] = stats["head_strikes"] / stats["total_sig_strikes"]
                    df.at[idx, f"{prefix}_body_pct_corrected"] = stats["body_strikes"] / stats["total_sig_strikes"]
                    df.at[idx, f"{prefix}_leg_pct_corrected"] = stats["leg_strikes"] / stats["total_sig_strikes"]
                else:
                    # Defaults based on typical UFC averages
                    df.at[idx, f"{prefix}_distance_pct_corrected"] = 0.60  # Most strikes from distance
                    df.at[idx, f"{prefix}_clinch_pct_corrected"] = 0.25
                    df.at[idx, f"{prefix}_ground_pct_corrected"] = 0.15
                    df.at[idx, f"{prefix}_head_pct_corrected"] = 0.55
                    df.at[idx, f"{prefix}_body_pct_corrected"] = 0.25
                    df.at[idx, f"{prefix}_leg_pct_corrected"] = 0.20

                # NEW: Career-based control time and reversals per fight
                if stats["fight_count"] > 0:
                    df.at[idx, f"{prefix}_avg_ctrl_sec_corrected"] = stats["ctrl_sec_total"] / stats["fight_count"]
                    df.at[idx, f"{prefix}_avg_rev_corrected"] = stats["rev_total"] / stats["fight_count"]
                else:
                    df.at[idx, f"{prefix}_avg_ctrl_sec_corrected"] = 0
                    df.at[idx, f"{prefix}_avg_rev_corrected"] = 0

                # NEW: Opponent quality (average ELO of opponents faced)
                if stats["opponents_faced"] > 0:
                    df.at[idx, f"{prefix}_avg_opponent_elo_corrected"] = stats["total_opponent_elo"] / stats["opponents_faced"]
                else:
                    df.at[idx, f"{prefix}_avg_opponent_elo_corrected"] = 1500  # Default ELO

                # NEW: Recent momentum vs career average
                # Distance percentage momentum
                if len(stats["recent_distance_pct"]) >= 2:
                    recent_avg = sum(stats["recent_distance_pct"][-3:]) / len(stats["recent_distance_pct"][-3:])
                    career_avg = stats["distance_strikes"] / stats["total_sig_strikes"] if stats["total_sig_strikes"] > 0 else 0.6
                    df.at[idx, f"{prefix}_distance_pct_momentum_corrected"] = recent_avg - career_avg

                # Striking output momentum
                if len(stats["recent_slpm"]) >= 2:
                    recent_avg = sum(stats["recent_slpm"][-3:]) / len(stats["recent_slpm"][-3:])
                    career_avg = stats["sig_str_total"] / stats["fight_time_minutes"] if stats["fight_time_minutes"] > 0 else 0
                    df.at[idx, f"{prefix}_slpm_momentum_corrected"] = recent_avg - career_avg

                # Control time momentum
                if len(stats["recent_ctrl_sec"]) >= 2:
                    recent_avg = sum(stats["recent_ctrl_sec"][-3:]) / len(stats["recent_ctrl_sec"][-3:])
                    career_avg = stats["ctrl_sec_total"] / stats["fight_count"] if stats["fight_count"] > 0 else 0
                    df.at[idx, f"{prefix}_ctrl_sec_momentum_corrected"] = recent_avg - career_avg

                # NEW: Main event experience (5-round fights)
                df.at[idx, f"{prefix}_main_event_fights_corrected"] = stats["main_event_fights"]

                # PHASE 1B: Helper function to calculate rolling average (defined early for use below)
                def rolling_avg(values, window):
                    if len(values) >= window:
                        return sum(values[-window:]) / window
                    elif len(values) > 0:
                        return sum(values) / len(values)
                    return 0

                # PHASE 2: Advanced career tracking
                df.at[idx, f"{prefix}_total_rounds_fought_corrected"] = stats["total_rounds_fought"]
                df.at[idx, f"{prefix}_total_fights_fought_corrected"] = stats["total_fights_fought"]
                df.at[idx, f"{prefix}_title_fights_corrected"] = stats["title_fights"]
                df.at[idx, f"{prefix}_five_round_fights_corrected"] = stats["five_round_fights"]

                # Last fight momentum indicators
                df.at[idx, f"{prefix}_last_fight_was_finish_corrected"] = 1.0 if stats["last_fight_was_finish"] else 0.0
                df.at[idx, f"{prefix}_last_fight_was_win_corrected"] = 1.0 if stats["last_fight_was_win"] else 0.0

                # Last fight dominance score (5 = round 1 finish, 1 = round 5 finish/decision)
                if stats["last_fight_was_finish"] and stats["last_fight_finish_round"]:
                    dominance = 6 - stats["last_fight_finish_round"]  # R1=5, R2=4, R3=3, R4=2, R5=1
                else:
                    dominance = 0  # Decision or no previous fight
                df.at[idx, f"{prefix}_last_fight_dominance_corrected"] = dominance

                # Finish timing rates
                total_fights = stats["wins"] + stats["losses"]
                if total_fights > 0:
                    df.at[idx, f"{prefix}_early_finish_rate_corrected"] = stats["early_finishes"] / total_fights
                    df.at[idx, f"{prefix}_late_finish_rate_corrected"] = stats["late_finishes"] / total_fights
                    df.at[idx, f"{prefix}_first_round_ko_rate_corrected"] = stats["first_round_kos"] / total_fights

                # Activity level (fights in last 24 months)
                fight_dates_24mo = [d for d in stats["fight_dates"] if (row["event_date"] - d).days <= 730]
                df.at[idx, f"{prefix}_fights_last_24_months_corrected"] = len(fight_dates_24mo)

                # Average finish time for last 3 finishes
                finish_times = []
                for i, method in enumerate(stats["recent_finish_methods"][-3:]):
                    if method in ["KO/TKO", "Submission"] and i < len(stats["recent_finish_rounds"]):
                        round_num = stats["recent_finish_rounds"][i]
                        # Estimate time: (round - 1) * 300 + 150 (assume mid-round)
                        if round_num:
                            finish_times.append((round_num - 1) * 300 + 150)
                if finish_times:
                    df.at[idx, f"{prefix}_avg_finish_time_last_3_corrected"] = sum(finish_times) / len(finish_times)

                # Chin deterioration (recent damage vs career damage)
                career_sapm = stats["sig_str_absorbed_total"] / stats["fight_time_minutes"] if stats["fight_time_minutes"] > 0 else 0
                recent_sapm = rolling_avg(stats["rolling_sapm"], 3)
                if career_sapm > 0:
                    df.at[idx, f"{prefix}_chin_deterioration_corrected"] = (recent_sapm - career_sapm) / career_sapm
                else:
                    df.at[idx, f"{prefix}_chin_deterioration_corrected"] = 0

                # PHASE 3: Record advanced stats before fight
                # CLUSTER 1: Opponent Quality Features
                if len(stats["opponent_elo_history"]) >= 5:
                    df.at[idx, f"{prefix}_avg_opponent_elo_l5_corrected"] = np.mean(stats["opponent_elo_history"][-5:])
                else:
                    df.at[idx, f"{prefix}_avg_opponent_elo_l5_corrected"] = np.mean(stats["opponent_elo_history"]) if stats["opponent_elo_history"] else 1500

                # elo_momentum_vs_competition
                current_elo = stats["elo"]  # Use tracked ELO
                if len(stats["opponent_elo_history"]) >= 3:
                    avg_opp_elo_l3 = np.mean(stats["opponent_elo_history"][-3:])
                    recent_win_rate_3 = sum(stats["recent_wins"][-3:]) / 3 if len(stats["recent_wins"]) >= 3 else 0.5
                    df.at[idx, f"{prefix}_elo_momentum_vs_competition_corrected"] = (current_elo - avg_opp_elo_l3) * recent_win_rate_3
                else:
                    df.at[idx, f"{prefix}_elo_momentum_vs_competition_corrected"] = 0

                # performance_vs_ranked_opponents
                if len(stats["opponent_ranking_history"]) > 0:
                    top_50_fights = sum(1 for rank in stats["opponent_ranking_history"] if rank >= 0.5)
                    top_50_wins = 0
                    for i, rank in enumerate(stats["opponent_ranking_history"]):
                        if rank >= 0.5 and i < len(stats["recent_wins"]) and stats["recent_wins"][-(len(stats["opponent_ranking_history"]) - i)]:
                            top_50_wins += 1
                    baseline_win_rate = stats["wins"] / max(stats["wins"] + stats["losses"], 1)
                    win_rate_vs_ranked = top_50_wins / max(top_50_fights, 1) if top_50_fights > 0 else baseline_win_rate
                    df.at[idx, f"{prefix}_performance_vs_ranked_opponents_corrected"] = win_rate_vs_ranked - baseline_win_rate
                else:
                    df.at[idx, f"{prefix}_performance_vs_ranked_opponents_corrected"] = 0

                # CLUSTER 2: Volatility Features
                if len(stats["damage_ratio_history"]) >= 3:
                    df.at[idx, f"{prefix}_performance_volatility_l10_corrected"] = np.std(stats["damage_ratio_history"][-10:])
                else:
                    df.at[idx, f"{prefix}_performance_volatility_l10_corrected"] = 0

                # finish_rate_acceleration
                if len(stats["finish_rate_history_10"]) >= 10:
                    finish_rate_l5 = np.mean(stats["finish_rate_history_10"][-5:])
                    finish_rate_l6_to_10 = np.mean(stats["finish_rate_history_10"][-10:-5])
                    df.at[idx, f"{prefix}_finish_rate_acceleration_corrected"] = finish_rate_l5 - finish_rate_l6_to_10
                elif len(stats["finish_rate_history_10"]) >= 5:
                    df.at[idx, f"{prefix}_finish_rate_acceleration_corrected"] = np.mean(stats["finish_rate_history_10"][-5:]) - 0.3
                else:
                    df.at[idx, f"{prefix}_finish_rate_acceleration_corrected"] = 0

                # slpm_coefficient_of_variation
                if len(stats["slpm_history"]) >= 3:
                    mean_slpm = np.mean(stats["slpm_history"][-10:])
                    if mean_slpm > 0:
                        df.at[idx, f"{prefix}_slpm_coefficient_of_variation_corrected"] = np.std(stats["slpm_history"][-10:]) / mean_slpm
                    else:
                        df.at[idx, f"{prefix}_slpm_coefficient_of_variation_corrected"] = 0
                else:
                    df.at[idx, f"{prefix}_slpm_coefficient_of_variation_corrected"] = 0

                # CLUSTER 3: Career Phase & Decline
                if len(stats["damage_ratio_history"]) >= 10:
                    dr_l3 = np.mean(stats["damage_ratio_history"][-3:])
                    dr_l10 = np.mean(stats["damage_ratio_history"][-10:])
                    age = row[f"{prefix}_age_at_event"]
                    age_factor = max(0, age - 32) * 0.15
                    df.at[idx, f"{prefix}_performance_decline_velocity_corrected"] = ((dr_l3 - dr_l10) / 7) * (1 + age_factor)
                elif len(stats["damage_ratio_history"]) >= 3:
                    dr_l3 = np.mean(stats["damage_ratio_history"][-3:])
                    age = row[f"{prefix}_age_at_event"]
                    age_factor = max(0, age - 32) * 0.15
                    df.at[idx, f"{prefix}_performance_decline_velocity_corrected"] = (dr_l3 - 1.0) * (1 + age_factor)
                else:
                    df.at[idx, f"{prefix}_performance_decline_velocity_corrected"] = 0

                # mileage_adjusted_age
                age = row[f"{prefix}_age_at_event"]
                total_rounds = stats["total_rounds_fought"]
                kd_absorbed = stats["kd_absorbed"]
                df.at[idx, f"{prefix}_mileage_adjusted_age_corrected"] = age + (total_rounds / 30) + (kd_absorbed * 0.5)

                # prime_exit_risk
                age = row[f"{prefix}_age_at_event"]
                if age > 33:
                    recent_win_rate_5 = sum(stats["recent_wins"][-5:]) / 5 if len(stats["recent_wins"]) >= 5 else 0.5
                    df.at[idx, f"{prefix}_prime_exit_risk_corrected"] = ((age - 33) ** 2) * (1 - recent_win_rate_5)
                else:
                    df.at[idx, f"{prefix}_prime_exit_risk_corrected"] = 0

                # CLUSTER 4: Fight Urgency
                losing_streak = stats["current_losing_streak"]
                wins = stats["wins"]
                age = row[f"{prefix}_age_at_event"]
                age_penalty = 1 + max(0, age - 35) * 0.3
                df.at[idx, f"{prefix}_career_inflection_point_corrected"] = losing_streak * (1 / (wins + 1)) * age_penalty

                # title_shot_proximity_score
                current_elo = stats["elo"]  # Use tracked ELO
                division_median_elo = 1500
                win_streak = self.calculate_streak(stats["recent_wins"], True)
                title_fights = stats["title_fights"]
                title_multiplier = 0.5 if title_fights > 0 else 1.0
                df.at[idx, f"{prefix}_title_shot_proximity_score_corrected"] = (current_elo / division_median_elo) * (win_streak / 3) * title_multiplier

                # CLUSTER 5: Style Evolution
                if len(stats["rolling_slpm"]) >= 10:
                    striking_rate_l3 = np.mean(stats["rolling_slpm"][-3:])
                    striking_rate_l10 = np.mean(stats["rolling_slpm"][-10:])
                    td_rate_l3 = np.mean(stats["rolling_td_acc"][-3:]) if len(stats["rolling_td_acc"]) >= 3 else 0
                    td_rate_l10 = np.mean(stats["rolling_td_acc"][-10:]) if len(stats["rolling_td_acc"]) >= 10 else 0
                    df.at[idx, f"{prefix}_tactical_evolution_score_corrected"] = abs(striking_rate_l3 - striking_rate_l10) + abs(td_rate_l3 - td_rate_l10)
                else:
                    df.at[idx, f"{prefix}_tactical_evolution_score_corrected"] = 0

                # finish_method_diversity
                total_fights = stats["wins"] + stats["losses"]
                if total_fights > 0:
                    ko_rate = stats["ko_wins"] / total_fights
                    sub_rate = stats["sub_wins"] / total_fights
                    dec_rate = stats["dec_wins"] / total_fights
                    rates = [ko_rate, sub_rate, dec_rate]
                    rates = [r for r in rates if r > 0]
                    entropy = -sum(r * np.log(r + 1e-10) for r in rates) if rates else 0
                    df.at[idx, f"{prefix}_finish_method_diversity_corrected"] = entropy
                else:
                    df.at[idx, f"{prefix}_finish_method_diversity_corrected"] = 0

                # CLUSTER 6: Interaction Terms
                age = row[f"{prefix}_age_at_event"]
                total_fights = stats["wins"] + stats["losses"]
                ko_rate = stats["ko_wins"] / total_fights if total_fights > 0 else 0
                if age > 30:
                    df.at[idx, f"{prefix}_aging_power_striker_penalty_corrected"] = (age - 30) * ko_rate * (-1)
                else:
                    df.at[idx, f"{prefix}_aging_power_striker_penalty_corrected"] = 0

                # elo_volatility_interaction
                current_elo = stats["elo"]  # Use tracked ELO
                performance_volatility = df.at[idx, f"{prefix}_performance_volatility_l10_corrected"]
                df.at[idx, f"{prefix}_elo_volatility_interaction_corrected"] = current_elo * (1 / (1 + performance_volatility))

                # layoff_veteran_interaction
                days_since_last = df.at[idx, f"{prefix}_days_since_last_fight_corrected"]
                layoff_severity = max(0, (days_since_last - 365) / 30)
                total_fights = stats["total_fights_fought"]
                age = row[f"{prefix}_age_at_event"]
                age_factor = 1 + max(0, age - 32) * 0.1
                df.at[idx, f"{prefix}_layoff_veteran_interaction_corrected"] = layoff_severity * (total_fights / 20) * age_factor

                # CLUSTER 7: Bayesian Adjustments
                total_fights = stats["wins"] + stats["losses"]
                finish_count = stats["ko_wins"] + stats["sub_wins"]
                df.at[idx, f"{prefix}_bayesian_finish_rate_corrected"] = (finish_count + 3) / (total_fights + 10)

                # confidence_weighted_damage_ratio
                if len(stats["damage_ratio_history"]) > 0:
                    recent_dr = np.mean(stats["damage_ratio_history"][-3:])
                else:
                    recent_dr = 1.0
                total_fights = stats["total_fights_fought"]
                confidence = min(1.0, total_fights / 15)
                df.at[idx, f"{prefix}_confidence_weighted_damage_ratio_corrected"] = recent_dr * confidence

                # CLUSTER 9: Peak Performance
                career_best = stats["career_best_damage_ratio"]
                if len(stats["damage_ratio_history"]) >= 3 and career_best > 0:
                    current_l3 = np.mean(stats["damage_ratio_history"][-3:])
                    df.at[idx, f"{prefix}_distance_from_career_peak_corrected"] = (career_best - current_l3) / career_best
                else:
                    df.at[idx, f"{prefix}_distance_from_career_peak_corrected"] = 0

                # elite_performance_frequency_l10
                if len(stats["damage_ratio_history"]) >= 10:
                    elite_count = stats["elite_performance_count_l10"]
                    df.at[idx, f"{prefix}_elite_performance_frequency_l10_corrected"] = elite_count / 10
                else:
                    df.at[idx, f"{prefix}_elite_performance_frequency_l10_corrected"] = 0

                # PHASE 1B: Rolling statistics (last 3/5/10 fights averages)

                # SLpM rolling averages
                df.at[idx, f"{prefix}_rolling_slpm_3_corrected"] = rolling_avg(stats["rolling_slpm"], 3)
                df.at[idx, f"{prefix}_rolling_slpm_5_corrected"] = rolling_avg(stats["rolling_slpm"], 5)
                df.at[idx, f"{prefix}_rolling_slpm_10_corrected"] = rolling_avg(stats["rolling_slpm"], 10)

                # SApM rolling averages
                df.at[idx, f"{prefix}_rolling_sapm_3_corrected"] = rolling_avg(stats["rolling_sapm"], 3)
                df.at[idx, f"{prefix}_rolling_sapm_5_corrected"] = rolling_avg(stats["rolling_sapm"], 5)
                df.at[idx, f"{prefix}_rolling_sapm_10_corrected"] = rolling_avg(stats["rolling_sapm"], 10)

                # TD accuracy rolling averages
                df.at[idx, f"{prefix}_rolling_td_acc_3_corrected"] = rolling_avg(stats["rolling_td_acc"], 3)
                df.at[idx, f"{prefix}_rolling_td_acc_5_corrected"] = rolling_avg(stats["rolling_td_acc"], 5)
                df.at[idx, f"{prefix}_rolling_td_acc_10_corrected"] = rolling_avg(stats["rolling_td_acc"], 10)

                # TD defense rolling averages
                df.at[idx, f"{prefix}_rolling_td_def_3_corrected"] = rolling_avg(stats["rolling_td_def"], 3)
                df.at[idx, f"{prefix}_rolling_td_def_5_corrected"] = rolling_avg(stats["rolling_td_def"], 5)
                df.at[idx, f"{prefix}_rolling_td_def_10_corrected"] = rolling_avg(stats["rolling_td_def"], 10)

                # Control time rolling averages
                df.at[idx, f"{prefix}_rolling_ctrl_3_corrected"] = rolling_avg(stats["rolling_ctrl_time"], 3)
                df.at[idx, f"{prefix}_rolling_ctrl_5_corrected"] = rolling_avg(stats["rolling_ctrl_time"], 5)
                df.at[idx, f"{prefix}_rolling_ctrl_10_corrected"] = rolling_avg(stats["rolling_ctrl_time"], 10)

                # Finish rate rolling averages
                df.at[idx, f"{prefix}_rolling_finish_rate_3_corrected"] = rolling_avg(stats["rolling_finishes"], 3)
                df.at[idx, f"{prefix}_rolling_finish_rate_5_corrected"] = rolling_avg(stats["rolling_finishes"], 5)
                df.at[idx, f"{prefix}_rolling_finish_rate_10_corrected"] = rolling_avg(stats["rolling_finishes"], 10)

                # Strike accuracy rolling averages
                df.at[idx, f"{prefix}_rolling_strike_acc_3_corrected"] = rolling_avg(stats["rolling_strike_acc"], 3)
                df.at[idx, f"{prefix}_rolling_strike_acc_5_corrected"] = rolling_avg(stats["rolling_strike_acc"], 5)
                df.at[idx, f"{prefix}_rolling_strike_acc_10_corrected"] = rolling_avg(stats["rolling_strike_acc"], 10)

                # Damage rolling averages (net striking)
                df.at[idx, f"{prefix}_rolling_damage_3_corrected"] = rolling_avg(stats["rolling_damage"], 3)
                df.at[idx, f"{prefix}_rolling_damage_5_corrected"] = rolling_avg(stats["rolling_damage"], 5)
                df.at[idx, f"{prefix}_rolling_damage_10_corrected"] = rolling_avg(stats["rolling_damage"], 10)

                # Variance/Consistency metrics (last 5 fights)
                if len(stats["rolling_slpm"]) >= 5:
                    values = stats["rolling_slpm"][-5:]
                    mean_val = sum(values) / 5
                    variance = sum((x - mean_val) ** 2 for x in values) / 5
                    df.at[idx, f"{prefix}_slpm_variance_5_corrected"] = variance

                if len(stats["rolling_sapm"]) >= 5:
                    values = stats["rolling_sapm"][-5:]
                    mean_val = sum(values) / 5
                    variance = sum((x - mean_val) ** 2 for x in values) / 5
                    df.at[idx, f"{prefix}_sapm_variance_5_corrected"] = variance

                # Performance consistency (inverse of variance in damage output)
                if len(stats["rolling_damage"]) >= 5:
                    values = stats["rolling_damage"][-5:]
                    mean_val = sum(values) / 5
                    variance = sum((x - mean_val) ** 2 for x in values) / 5
                    # Consistency = 1 / (1 + variance) - higher = more consistent
                    df.at[idx, f"{prefix}_performance_consistency_5_corrected"] = 1.0 / (1.0 + variance)

                # Trend features (improving/declining) - linear regression slope
                def calculate_trend(values):
                    """Calculate linear trend slope"""
                    if len(values) < 3:
                        return 0
                    n = len(values)
                    x = list(range(n))
                    x_mean = sum(x) / n
                    y_mean = sum(values) / n
                    numerator = sum((x[i] - x_mean) * (values[i] - y_mean) for i in range(n))
                    denominator = sum((x[i] - x_mean) ** 2 for i in range(n))
                    return numerator / denominator if denominator != 0 else 0

                if len(stats["rolling_slpm"]) >= 3:
                    df.at[idx, f"{prefix}_slpm_trend_5_corrected"] = calculate_trend(stats["rolling_slpm"][-5:] if len(stats["rolling_slpm"]) >= 5 else stats["rolling_slpm"])

                if len(stats["rolling_td_acc"]) >= 3:
                    df.at[idx, f"{prefix}_td_acc_trend_5_corrected"] = calculate_trend(stats["rolling_td_acc"][-5:] if len(stats["rolling_td_acc"]) >= 5 else stats["rolling_td_acc"])

                if len(stats["rolling_finishes"]) >= 3:
                    df.at[idx, f"{prefix}_finish_rate_trend_5_corrected"] = calculate_trend(stats["rolling_finishes"][-5:] if len(stats["rolling_finishes"]) >= 5 else stats["rolling_finishes"])

                # PHASE 1C: Opponent-adjusted performance metrics
                # Win rate vs elite opponents (top 25% ELO > ~1600)
                if stats["vs_elite_record"]["fights"] > 0:
                    df.at[idx, f"{prefix}_win_rate_vs_elite_corrected"] = stats["vs_elite_record"]["wins"] / stats["vs_elite_record"]["fights"]

                # Win rate vs strikers (high SLpM opponents)
                if stats["vs_striker_record"]["fights"] > 0:
                    df.at[idx, f"{prefix}_win_rate_vs_strikers_corrected"] = stats["vs_striker_record"]["wins"] / stats["vs_striker_record"]["fights"]

                # Win rate vs grapplers (high TD avg opponents)
                if stats["vs_grappler_record"]["fights"] > 0:
                    df.at[idx, f"{prefix}_win_rate_vs_grapplers_corrected"] = stats["vs_grappler_record"]["wins"] / stats["vs_grappler_record"]["fights"]

                # Win rate vs durable opponents (high durability)
                if stats["vs_durable_record"]["fights"] > 0:
                    df.at[idx, f"{prefix}_win_rate_vs_durable_corrected"] = stats["vs_durable_record"]["wins"] / stats["vs_durable_record"]["fights"]

                # Win rate vs finishers (high finish rate opponents)
                if stats["vs_finisher_record"]["fights"] > 0:
                    df.at[idx, f"{prefix}_win_rate_vs_finishers_corrected"] = stats["vs_finisher_record"]["wins"] / stats["vs_finisher_record"]["fights"]

                # Finish rate vs elite opponents
                if stats["finish_vs_elite"]["fights"] > 0:
                    df.at[idx, f"{prefix}_finish_rate_vs_elite_corrected"] = stats["finish_vs_elite"]["finishes"] / stats["finish_vs_elite"]["fights"]

                # Recent opponent quality (avg ELO of last 5 opponents)
                if len(stats["recent_opponent_elos"]) > 0:
                    df.at[idx, f"{prefix}_recent_opponent_quality_5_corrected"] = sum(stats["recent_opponent_elos"]) / len(stats["recent_opponent_elos"])
                else:
                    df.at[idx, f"{prefix}_recent_opponent_quality_5_corrected"] = 1500  # Default

                # Style versatility (can beat both strikers and grapplers)
                striker_success = stats["vs_striker_record"]["wins"] / stats["vs_striker_record"]["fights"] if stats["vs_striker_record"]["fights"] > 0 else 0.5
                grappler_success = stats["vs_grappler_record"]["wins"] / stats["vs_grappler_record"]["fights"] if stats["vs_grappler_record"]["fights"] > 0 else 0.5
                df.at[idx, f"{prefix}_style_versatility_corrected"] = min(striker_success, grappler_success)  # Worst matchup determines versatility

                # Step-up performance (performance vs elite - performance vs non-elite)
                elite_rate = stats["vs_elite_record"]["wins"] / stats["vs_elite_record"]["fights"] if stats["vs_elite_record"]["fights"] > 0 else 0.5
                overall_rate = stats["wins"] / max(stats["wins"] + stats["losses"], 1)
                df.at[idx, f"{prefix}_step_up_performance_corrected"] = elite_rate - overall_rate

            # Calculate opponent quality differential
            r_opp_elo = df.at[idx, "r_avg_opponent_elo_corrected"]
            b_opp_elo = df.at[idx, "b_avg_opponent_elo_corrected"]
            df.at[idx, "opponent_quality_diff"] = r_opp_elo - b_opp_elo

            # Update stats after fight
            if pd.notna(row["winner"]):
                method = str(row.get("method", "")).lower()
                method_cat = "ko" if "ko" in method or "tko" in method else "sub" if "sub" in method else "dec"
                is_finish = method_cat in ["ko", "sub"]

                if row["winner"] == "Red":
                    fighter_stats[r_fighter]["wins"] += 1
                    fighter_stats[b_fighter]["losses"] += 1
                    fighter_stats[r_fighter][f"{method_cat}_wins"] += 1
                    fighter_stats[r_fighter]["recent_wins"].append(1)
                    fighter_stats[b_fighter]["recent_wins"].append(0)
                    # Track recent finishes for both fighters
                    fighter_stats[r_fighter]["recent_finishes"].append(1 if is_finish else 0)
                    fighter_stats[b_fighter]["recent_finishes"].append(0)  # Loser didn't finish
                    if is_finish:
                        fighter_stats[b_fighter][f"{method_cat}_losses"] += 1
                    fighter_h2h[(r_fighter, b_fighter)] = fighter_h2h.get((r_fighter, b_fighter), 0) + 1
                elif row["winner"] == "Blue":
                    fighter_stats[b_fighter]["wins"] += 1
                    fighter_stats[r_fighter]["losses"] += 1
                    fighter_stats[b_fighter][f"{method_cat}_wins"] += 1
                    fighter_stats[b_fighter]["recent_wins"].append(1)
                    fighter_stats[r_fighter]["recent_wins"].append(0)
                    # Track recent finishes for both fighters
                    fighter_stats[b_fighter]["recent_finishes"].append(1 if is_finish else 0)
                    fighter_stats[r_fighter]["recent_finishes"].append(0)  # Loser didn't finish
                    if is_finish:
                        fighter_stats[r_fighter][f"{method_cat}_losses"] += 1
                    fighter_h2h[(r_fighter, b_fighter)] = fighter_h2h.get((r_fighter, b_fighter), 0) - 1

                # Trim recent history
                for fighter in [r_fighter, b_fighter]:
                    if len(fighter_stats[fighter]["recent_wins"]) > 10:
                        fighter_stats[fighter]["recent_wins"] = fighter_stats[fighter]["recent_wins"][-10:]
                    if len(fighter_stats[fighter]["recent_finishes"]) > 10:
                        fighter_stats[fighter]["recent_finishes"] = fighter_stats[fighter]["recent_finishes"][-10:]
                    fighter_stats[fighter]["last_fight_date"] = row["event_date"]

                # Track weight classes for both fighters
                weight_class = row.get("weight_class", "Unknown")
                if pd.notna(weight_class):
                    fighter_stats[r_fighter]["weight_class_history"].append(weight_class)
                    fighter_stats[b_fighter]["weight_class_history"].append(weight_class)

                # Update ELO ratings (using proper UFC ELO system for PHASE 3 features)
                # Initialize fighters in ELO system if not seen before
                if r_fighter not in self.fighter_elos:
                    self.fighter_elos[r_fighter] = 1500
                    self.elo_history[r_fighter] = []
                if b_fighter not in self.fighter_elos:
                    self.fighter_elos[b_fighter] = 1500
                    self.elo_history[b_fighter] = []

                # Sync tracked ELO with main ELO system
                fighter_stats[r_fighter]["elo"] = self.fighter_elos[r_fighter]
                fighter_stats[b_fighter]["elo"] = self.fighter_elos[b_fighter]

                # Determine winner for K-factor calculation
                if row["winner"] == "Red":
                    winner_fighter = r_fighter
                    fighter_result = "Win"
                elif row["winner"] == "Blue":
                    winner_fighter = b_fighter
                    fighter_result = "Win"
                else:  # Draw
                    winner_fighter = r_fighter
                    fighter_result = "Draw"

                # Calculate K-factor using the proper UFC ELO system
                k_factor = self.elo_calculate_k_factor(
                    row, winner_fighter, fighter_result, self.fighter_elos, self.elo_history
                )

                # Update ELO ratings using the proper UFC ELO system
                self.elo_update_ratings(
                    r_fighter, b_fighter, row["winner"], k_factor,
                    self.fighter_elos, self.elo_history, row["event_date"]
                )

                # NOTE: Do NOT update tracked ELO yet - opponent classification below needs PRE-FIGHT values
                # Post-fight ELO will be synced after all opponent-based stats are calculated

                # Update fight stats
                for fighter, f_prefix, opp_prefix in [(r_fighter, "r", "b"), (b_fighter, "b", "r")]:
                    if pd.notna(row.get(f"{f_prefix}_sig_str")):
                        fighter_stats[fighter]["sig_str_total"] += row[f"{f_prefix}_sig_str"]
                    if pd.notna(row.get(f"{f_prefix}_sig_str_att")):
                        fighter_stats[fighter]["sig_str_att_total"] += row[f"{f_prefix}_sig_str_att"]
                    if pd.notna(row.get(f"{opp_prefix}_sig_str")):
                        fighter_stats[fighter]["sig_str_absorbed_total"] += row[f"{opp_prefix}_sig_str"]

                    # Track striking defense: opponent's strikes attempted against you
                    if pd.notna(row.get(f"{opp_prefix}_sig_str_att")):
                        fighter_stats[fighter]["str_def_att"] += row[f"{opp_prefix}_sig_str_att"]
                        # Strikes successfully defended = opponent attempts - opponent landed
                        opp_sig_str = row.get(f"{opp_prefix}_sig_str", 0) if pd.notna(row.get(f"{opp_prefix}_sig_str")) else 0
                        fighter_stats[fighter]["str_def_success"] += (row[f"{opp_prefix}_sig_str_att"] - opp_sig_str)

                    if pd.notna(row.get(f"{f_prefix}_td")):
                        fighter_stats[fighter]["td_total"] += row[f"{f_prefix}_td"]
                    if pd.notna(row.get(f"{f_prefix}_td_att")):
                        fighter_stats[fighter]["td_att_total"] += row[f"{f_prefix}_td_att"]

                    # Track TD defense: opponent's TDs attempted against you
                    if pd.notna(row.get(f"{opp_prefix}_td_att")):
                        fighter_stats[fighter]["td_def_att"] += row[f"{opp_prefix}_td_att"]
                        # TDs successfully defended = opponent attempts - opponent successes
                        opp_td = row.get(f"{opp_prefix}_td", 0) if pd.notna(row.get(f"{opp_prefix}_td")) else 0
                        fighter_stats[fighter]["td_def_success"] += (row[f"{opp_prefix}_td_att"] - opp_td)

                    if pd.notna(row.get(f"{f_prefix}_sub_att")):
                        fighter_stats[fighter]["sub_att_total"] += row[f"{f_prefix}_sub_att"]

                    # Track knockdowns scored and absorbed
                    if pd.notna(row.get(f"{f_prefix}_kd")):
                        fighter_stats[fighter]["kd_total"] += row[f"{f_prefix}_kd"]
                    if pd.notna(row.get(f"{opp_prefix}_kd")):
                        fighter_stats[fighter]["kd_absorbed"] += row[f"{opp_prefix}_kd"]

                    fight_time = row.get("total_fight_time_sec", 0) / 60 if pd.notna(row.get("total_fight_time_sec")) else 0
                    fighter_stats[fighter]["fight_time_minutes"] += fight_time
                    fighter_stats[fighter]["fight_count"] += 1

                    # NEW: Track positional striking (distance/clinch/ground)
                    sig_str = row.get(f"{f_prefix}_sig_str", 0)
                    if pd.notna(sig_str) and sig_str > 0:
                        fighter_stats[fighter]["total_sig_strikes"] += sig_str

                        # Positional percentages from this fight
                        distance_pct = row.get(f"{f_prefix}_distance", 0) if pd.notna(row.get(f"{f_prefix}_distance")) else 0
                        clinch_pct = row.get(f"{f_prefix}_clinch", 0) if pd.notna(row.get(f"{f_prefix}_clinch")) else 0
                        ground_pct = row.get(f"{f_prefix}_ground", 0) if pd.notna(row.get(f"{f_prefix}_ground")) else 0

                        # Add strikes from each position (percentage x total strikes)
                        fighter_stats[fighter]["distance_strikes"] += sig_str * distance_pct
                        fighter_stats[fighter]["clinch_strikes"] += sig_str * clinch_pct
                        fighter_stats[fighter]["ground_strikes"] += sig_str * ground_pct

                        # NEW: Track target areas (head/body/leg)
                        head_pct = row.get(f"{f_prefix}_head", 0) if pd.notna(row.get(f"{f_prefix}_head")) else 0
                        body_pct = row.get(f"{f_prefix}_body", 0) if pd.notna(row.get(f"{f_prefix}_body")) else 0
                        leg_pct = row.get(f"{f_prefix}_leg", 0) if pd.notna(row.get(f"{f_prefix}_leg")) else 0

                        fighter_stats[fighter]["head_strikes"] += sig_str * head_pct
                        fighter_stats[fighter]["body_strikes"] += sig_str * body_pct
                        fighter_stats[fighter]["leg_strikes"] += sig_str * leg_pct

                    # NEW: Track control time and reversals
                    if pd.notna(row.get(f"{f_prefix}_ctrl_sec")):
                        fighter_stats[fighter]["ctrl_sec_total"] += row[f"{f_prefix}_ctrl_sec"]
                    if pd.notna(row.get(f"{f_prefix}_rev")):
                        fighter_stats[fighter]["rev_total"] += row[f"{f_prefix}_rev"]

                # NEW: Track opponent ELO for quality metric (after both fighters updated)
                # Red corner tracks blue's ELO, blue tracks red's ELO
                if hasattr(self, 'fighter_elos'):
                    r_elo = self.fighter_elos.get(r_fighter, 1500)
                    b_elo = self.fighter_elos.get(b_fighter, 1500)

                    fighter_stats[r_fighter]["total_opponent_elo"] += b_elo
                    fighter_stats[r_fighter]["opponents_faced"] += 1
                    fighter_stats[b_fighter]["total_opponent_elo"] += r_elo
                    fighter_stats[b_fighter]["opponents_faced"] += 1

                # NEW: Track main event fights (5-round fights)
                if row.get("total_rounds", 3) == 5:
                    fighter_stats[r_fighter]["main_event_fights"] += 1
                    fighter_stats[b_fighter]["main_event_fights"] += 1

                # PHASE 2: Track new advanced stats
                finish_round = row.get("finish_round", row.get("total_rounds", 3))
                fight_method = str(row.get("method", "Decision"))

                # Track rounds fought for both fighters
                for fighter in [r_fighter, b_fighter]:
                    fighter_stats[fighter]["total_rounds_fought"] += finish_round if pd.notna(finish_round) else 3
                    fighter_stats[fighter]["total_fights_fought"] += 1

                    # Title fight tracking
                    if row.get("is_title_bout", 0) == 1:
                        fighter_stats[fighter]["title_fights"] += 1

                    # Five round fights
                    if row.get("total_rounds", 3) == 5:
                        fighter_stats[fighter]["five_round_fights"] += 1

                    # Fight dates for activity tracking
                    fighter_stats[fighter]["fight_dates"].append(row["event_date"])
                    if len(fighter_stats[fighter]["fight_dates"]) > 24:  # Keep last 24 fights
                        fighter_stats[fighter]["fight_dates"] = fighter_stats[fighter]["fight_dates"][-24:]

                # Track last fight details for winner
                winner_fighter = r_fighter if row["winner"] == "Red" else b_fighter if row["winner"] == "Blue" else None
                loser_fighter = b_fighter if row["winner"] == "Red" else r_fighter if row["winner"] == "Blue" else None

                if winner_fighter:
                    fighter_stats[winner_fighter]["last_fight_method"] = fight_method
                    fighter_stats[winner_fighter]["last_fight_finish_round"] = finish_round
                    fighter_stats[winner_fighter]["last_fight_was_finish"] = is_finish
                    fighter_stats[winner_fighter]["last_fight_was_win"] = True

                    # Track finish timing (early vs late)
                    if is_finish and pd.notna(finish_round):
                        if finish_round <= 2:
                            fighter_stats[winner_fighter]["early_finishes"] += 1
                        elif finish_round >= 3:
                            fighter_stats[winner_fighter]["late_finishes"] += 1

                        # First round KO tracking
                        if finish_round == 1 and method_cat == "ko":
                            fighter_stats[winner_fighter]["first_round_kos"] += 1

                    # Track recent finish methods
                    fighter_stats[winner_fighter]["recent_finish_methods"].append(fight_method)
                    fighter_stats[winner_fighter]["recent_finish_rounds"].append(finish_round)
                    if len(fighter_stats[winner_fighter]["recent_finish_methods"]) > 5:
                        fighter_stats[winner_fighter]["recent_finish_methods"] = fighter_stats[winner_fighter]["recent_finish_methods"][-5:]
                        fighter_stats[winner_fighter]["recent_finish_rounds"] = fighter_stats[winner_fighter]["recent_finish_rounds"][-5:]

                if loser_fighter:
                    fighter_stats[loser_fighter]["last_fight_method"] = fight_method
                    fighter_stats[loser_fighter]["last_fight_finish_round"] = finish_round
                    fighter_stats[loser_fighter]["last_fight_was_finish"] = is_finish
                    fighter_stats[loser_fighter]["last_fight_was_win"] = False

                    # Track recent finish methods (even for losses)
                    fighter_stats[loser_fighter]["recent_finish_methods"].append(fight_method)
                    fighter_stats[loser_fighter]["recent_finish_rounds"].append(finish_round)
                    if len(fighter_stats[loser_fighter]["recent_finish_methods"]) > 5:
                        fighter_stats[loser_fighter]["recent_finish_methods"] = fighter_stats[loser_fighter]["recent_finish_methods"][-5:]
                        fighter_stats[loser_fighter]["recent_finish_rounds"] = fighter_stats[loser_fighter]["recent_finish_rounds"][-5:]

                # NEW: Track recent stats for momentum (last 3 fights)
                for fighter, f_prefix in [(r_fighter, "r"), (b_fighter, "b")]:
                    # Track recent positional percentages (use per-fight data, not cumulative)
                    this_fight_distance = row.get(f"{f_prefix}_distance", 0) if pd.notna(row.get(f"{f_prefix}_distance")) else 0
                    fighter_stats[fighter]["recent_distance_pct"].append(this_fight_distance)
                    if len(fighter_stats[fighter]["recent_distance_pct"]) > 3:
                        fighter_stats[fighter]["recent_distance_pct"] = fighter_stats[fighter]["recent_distance_pct"][-3:]

                    # Track recent striking output
                    if fight_time > 0:
                        this_fight_slpm = row.get(f"{f_prefix}_sig_str", 0) / (fight_time / 60) if pd.notna(row.get(f"{f_prefix}_sig_str")) else 0
                        fighter_stats[fighter]["recent_slpm"].append(this_fight_slpm)
                        if len(fighter_stats[fighter]["recent_slpm"]) > 3:
                            fighter_stats[fighter]["recent_slpm"] = fighter_stats[fighter]["recent_slpm"][-3:]

                    # Track recent control time
                    this_fight_ctrl = row.get(f"{f_prefix}_ctrl_sec", 0) if pd.notna(row.get(f"{f_prefix}_ctrl_sec")) else 0
                    fighter_stats[fighter]["recent_ctrl_sec"].append(this_fight_ctrl)
                    if len(fighter_stats[fighter]["recent_ctrl_sec"]) > 3:
                        fighter_stats[fighter]["recent_ctrl_sec"] = fighter_stats[fighter]["recent_ctrl_sec"][-3:]

                    # PHASE 1B: Update rolling statistics (keep last 10 fights)
                    # SLpM for this fight
                    if fight_time > 0:
                        this_fight_slpm = row.get(f"{f_prefix}_sig_str", 0) / fight_time if pd.notna(row.get(f"{f_prefix}_sig_str")) else 0
                        fighter_stats[fighter]["rolling_slpm"].append(this_fight_slpm)
                        if len(fighter_stats[fighter]["rolling_slpm"]) > 10:
                            fighter_stats[fighter]["rolling_slpm"] = fighter_stats[fighter]["rolling_slpm"][-10:]

                    # SApM for this fight
                    if fight_time > 0:
                        opp_prefix = "b" if f_prefix == "r" else "r"
                        this_fight_sapm = row.get(f"{opp_prefix}_sig_str", 0) / fight_time if pd.notna(row.get(f"{opp_prefix}_sig_str")) else 0
                        fighter_stats[fighter]["rolling_sapm"].append(this_fight_sapm)
                        if len(fighter_stats[fighter]["rolling_sapm"]) > 10:
                            fighter_stats[fighter]["rolling_sapm"] = fighter_stats[fighter]["rolling_sapm"][-10:]

                    # TD accuracy for this fight
                    td_att = row.get(f"{f_prefix}_td_att", 0)
                    if pd.notna(td_att) and td_att > 0:
                        this_fight_td_acc = row.get(f"{f_prefix}_td", 0) / td_att
                    else:
                        this_fight_td_acc = 0
                    fighter_stats[fighter]["rolling_td_acc"].append(this_fight_td_acc)
                    if len(fighter_stats[fighter]["rolling_td_acc"]) > 10:
                        fighter_stats[fighter]["rolling_td_acc"] = fighter_stats[fighter]["rolling_td_acc"][-10:]

                    # TD defense for this fight
                    opp_prefix = "b" if f_prefix == "r" else "r"
                    opp_td_att = row.get(f"{opp_prefix}_td_att", 0)
                    if pd.notna(opp_td_att) and opp_td_att > 0:
                        opp_td = row.get(f"{opp_prefix}_td", 0)
                        this_fight_td_def = 1.0 - (opp_td / opp_td_att)
                    else:
                        this_fight_td_def = 1.0  # Perfect defense if opponent had no attempts
                    fighter_stats[fighter]["rolling_td_def"].append(this_fight_td_def)
                    if len(fighter_stats[fighter]["rolling_td_def"]) > 10:
                        fighter_stats[fighter]["rolling_td_def"] = fighter_stats[fighter]["rolling_td_def"][-10:]

                    # Control time for this fight
                    this_fight_ctrl_sec = row.get(f"{f_prefix}_ctrl_sec", 0) if pd.notna(row.get(f"{f_prefix}_ctrl_sec")) else 0
                    fighter_stats[fighter]["rolling_ctrl_time"].append(this_fight_ctrl_sec)
                    if len(fighter_stats[fighter]["rolling_ctrl_time"]) > 10:
                        fighter_stats[fighter]["rolling_ctrl_time"] = fighter_stats[fighter]["rolling_ctrl_time"][-10:]

                    # Finish indicator (1 = finish, 0 = decision)
                    method = str(row.get("method", "")).lower()
                    is_finish = 1 if ("ko" in method or "tko" in method or "sub" in method) else 0
                    fighter_stats[fighter]["rolling_finishes"].append(is_finish)
                    if len(fighter_stats[fighter]["rolling_finishes"]) > 10:
                        fighter_stats[fighter]["rolling_finishes"] = fighter_stats[fighter]["rolling_finishes"][-10:]

                    # Strike accuracy for this fight
                    sig_str_att = row.get(f"{f_prefix}_sig_str_att", 0)
                    if pd.notna(sig_str_att) and sig_str_att > 0:
                        this_fight_strike_acc = row.get(f"{f_prefix}_sig_str", 0) / sig_str_att
                    else:
                        this_fight_strike_acc = 0
                    fighter_stats[fighter]["rolling_strike_acc"].append(this_fight_strike_acc)
                    if len(fighter_stats[fighter]["rolling_strike_acc"]) > 10:
                        fighter_stats[fighter]["rolling_strike_acc"] = fighter_stats[fighter]["rolling_strike_acc"][-10:]

                    # Damage (net striking) for this fight
                    sig_str = row.get(f"{f_prefix}_sig_str", 0) if pd.notna(row.get(f"{f_prefix}_sig_str")) else 0
                    opp_sig_str = row.get(f"{opp_prefix}_sig_str", 0) if pd.notna(row.get(f"{opp_prefix}_sig_str")) else 0
                    this_fight_damage = sig_str - opp_sig_str
                    fighter_stats[fighter]["rolling_damage"].append(this_fight_damage)
                    if len(fighter_stats[fighter]["rolling_damage"]) > 10:
                        fighter_stats[fighter]["rolling_damage"] = fighter_stats[fighter]["rolling_damage"][-10:]

                    # PHASE 1C: Update opponent-adjusted performance records
                    # Classify opponent based on THEIR pre-fight stats
                    # CRITICAL FIX: Use pre-computed _corrected values from the row to avoid data leakage
                    # These values were calculated BEFORE this fight and stored in the row
                    opp_prefix = "b" if f_prefix == "r" else "r"

                    # Get opponent's pre-fight ELO (from pre-computed value in row, NOT tracked stats)
                    # Use avg_opponent_elo as proxy since we track opponent ELOs, not fighter's own ELO in row
                    # The tracked stats["elo"] at this point is still PRE-FIGHT (we haven't synced post-fight yet)
                    opponent = r_fighter if f_prefix == "b" else b_fighter
                    opp_stats = fighter_stats.get(opponent, {})
                    opp_elo = opp_stats.get("elo", 1500)  # Still PRE-FIGHT because not yet synced

                    if pd.notna(opp_elo):
                        fighter_stats[fighter]["recent_opponent_elos"].append(opp_elo)
                        if len(fighter_stats[fighter]["recent_opponent_elos"]) > 5:
                            fighter_stats[fighter]["recent_opponent_elos"] = fighter_stats[fighter]["recent_opponent_elos"][-5:]

                    # Classify opponent as elite (strong opponent threshold)
                    # Elite threshold: 1600 (top ~25% of active fighters)
                    is_elite_opponent = opp_elo >= 1600 if pd.notna(opp_elo) else False

                    # Classify opponent style based on PRE-COMPUTED career averages (chronologically safe)
                    opp_slpm = row.get(f"{opp_prefix}_pro_SLpM_corrected", 0)
                    opp_td_avg = row.get(f"{opp_prefix}_pro_td_avg_corrected", 0)

                    is_striker = False
                    is_grappler = False
                    opponent_style = 0  # -1=grappler, 0=balanced, 1=striker

                    if pd.notna(opp_slpm) and pd.notna(opp_td_avg):
                        # High SLpM (>4.0) and low TD avg (<2.0) = Striker
                        if opp_slpm > 4.0 and opp_td_avg < 2.0:
                            is_striker = True
                            opponent_style = 1
                        # High TD avg (>3.0) and lower SLpM = Grappler
                        elif opp_td_avg > 3.0 and opp_slpm < 4.5:
                            is_grappler = True
                            opponent_style = -1

                    fighter_stats[fighter]["recent_opponent_styles"].append(opponent_style)
                    if len(fighter_stats[fighter]["recent_opponent_styles"]) > 5:
                        fighter_stats[fighter]["recent_opponent_styles"] = fighter_stats[fighter]["recent_opponent_styles"][-5:]

                    # Classify opponent as durable using PRE-COMPUTED values
                    opp_wins = row.get(f"{opp_prefix}_wins_corrected", 0)
                    opp_losses = row.get(f"{opp_prefix}_losses_corrected", 0)
                    opp_fights = opp_wins + opp_losses if pd.notna(opp_wins) and pd.notna(opp_losses) else 0
                    opp_str_def = row.get(f"{opp_prefix}_pro_str_def_corrected", 0)
                    is_durable = (opp_fights >= 15 and opp_str_def > 0.55) if pd.notna(opp_str_def) else False

                    # Classify opponent as finisher using PRE-COMPUTED rates
                    opp_ko_rate = row.get(f"{opp_prefix}_ko_rate_corrected", 0)
                    opp_sub_rate = row.get(f"{opp_prefix}_sub_rate_corrected", 0)
                    opp_finish_rate = opp_ko_rate + opp_sub_rate if pd.notna(opp_ko_rate) and pd.notna(opp_sub_rate) else 0
                    is_finisher = opp_finish_rate > 0.65

                    # Determine if fighter won this fight
                    winner = row.get("winner")
                    did_win = (winner == "Red" and f_prefix == "r") or (winner == "Blue" and f_prefix == "b")

                    # Update vs_elite record
                    if is_elite_opponent:
                        fighter_stats[fighter]["vs_elite_record"]["fights"] += 1
                        if did_win:
                            fighter_stats[fighter]["vs_elite_record"]["wins"] += 1

                        # Track finishes vs elite
                        fighter_stats[fighter]["finish_vs_elite"]["fights"] += 1
                        method = str(row.get("method", "")).lower()
                        if did_win and ("ko" in method or "tko" in method or "sub" in method):
                            fighter_stats[fighter]["finish_vs_elite"]["finishes"] += 1

                    # Update vs_striker record
                    if is_striker:
                        fighter_stats[fighter]["vs_striker_record"]["fights"] += 1
                        if did_win:
                            fighter_stats[fighter]["vs_striker_record"]["wins"] += 1

                    # Update vs_grappler record
                    if is_grappler:
                        fighter_stats[fighter]["vs_grappler_record"]["fights"] += 1
                        if did_win:
                            fighter_stats[fighter]["vs_grappler_record"]["wins"] += 1

                    # Update vs_durable record
                    if is_durable:
                        fighter_stats[fighter]["vs_durable_record"]["fights"] += 1
                        if did_win:
                            fighter_stats[fighter]["vs_durable_record"]["wins"] += 1

                    # Update vs_finisher record
                    if is_finisher:
                        fighter_stats[fighter]["vs_finisher_record"]["fights"] += 1
                        if did_win:
                            fighter_stats[fighter]["vs_finisher_record"]["wins"] += 1

                # PHASE 3: Update advanced tracking after fight
                for fighter, f_prefix, opp_fighter, opp_prefix in [(r_fighter, "r", b_fighter, "b"), (b_fighter, "b", r_fighter, "r")]:
                    # CLUSTER 1: Track opponent quality
                    # Get opponent's ELO before this fight
                    opponent_elo = fighter_stats[opp_fighter]["elo"]  # Use tracked ELO
                    fighter_stats[fighter]["opponent_elo_history"].append(opponent_elo)
                    if len(fighter_stats[fighter]["opponent_elo_history"]) > 10:
                        fighter_stats[fighter]["opponent_elo_history"] = fighter_stats[fighter]["opponent_elo_history"][-10:]

                    # Track opponent ranking (percentile in division)
                    # Estimate percentile based on ELO (1500 = median)
                    opponent_percentile = min(1.0, max(0.0, (opponent_elo - 1300) / 400))
                    fighter_stats[fighter]["opponent_ranking_history"].append(opponent_percentile)
                    if len(fighter_stats[fighter]["opponent_ranking_history"]) > 10:
                        fighter_stats[fighter]["opponent_ranking_history"] = fighter_stats[fighter]["opponent_ranking_history"][-10:]

                    # CLUSTER 2: Track performance volatility
                    # Calculate this fight's damage ratio
                    my_sig_str = row.get(f"{f_prefix}_sig_str", 0)
                    opp_sig_str = row.get(f"{opp_prefix}_sig_str", 0)
                    damage_ratio = my_sig_str / max(opp_sig_str, 1) if opp_sig_str > 0 else my_sig_str / max(my_sig_str, 1)

                    fighter_stats[fighter]["damage_ratio_history"].append(damage_ratio)
                    if len(fighter_stats[fighter]["damage_ratio_history"]) > 10:
                        fighter_stats[fighter]["damage_ratio_history"] = fighter_stats[fighter]["damage_ratio_history"][-10:]

                    # Update career best damage ratio
                    if damage_ratio > fighter_stats[fighter]["career_best_damage_ratio"]:
                        fighter_stats[fighter]["career_best_damage_ratio"] = damage_ratio

                    # Track if this was an elite performance (top 20% of career)
                    if len(fighter_stats[fighter]["damage_ratio_history"]) >= 5:
                        # Calculate 80th percentile threshold
                        threshold_80 = np.percentile(fighter_stats[fighter]["damage_ratio_history"], 80)
                        fighter_stats[fighter]["damage_ratio_threshold_80th"] = threshold_80

                        # Update elite performance count for last 10 fights
                        elite_count = sum(1 for dr in fighter_stats[fighter]["damage_ratio_history"][-10:] if dr >= threshold_80)
                        fighter_stats[fighter]["elite_performance_count_l10"] = elite_count

                    # Track SLpM for coefficient of variation
                    if fight_time > 0:
                        this_fight_slpm = row.get(f"{f_prefix}_sig_str", 0) / (fight_time / 60) if pd.notna(row.get(f"{f_prefix}_sig_str")) else 0
                        fighter_stats[fighter]["slpm_history"].append(this_fight_slpm)
                        if len(fighter_stats[fighter]["slpm_history"]) > 10:
                            fighter_stats[fighter]["slpm_history"] = fighter_stats[fighter]["slpm_history"][-10:]

                    # Track finish outcomes for acceleration
                    is_finish = method_cat in ["ko", "sub"]
                    was_winner = (fighter == r_fighter and row["winner"] == "Red") or (fighter == b_fighter and row["winner"] == "Blue")
                    finish_outcome = 1 if (is_finish and was_winner) else 0

                    fighter_stats[fighter]["finish_rate_history_10"].append(finish_outcome)
                    if len(fighter_stats[fighter]["finish_rate_history_10"]) > 10:
                        fighter_stats[fighter]["finish_rate_history_10"] = fighter_stats[fighter]["finish_rate_history_10"][-10:]

                    # CLUSTER 4: Track losing streak
                    if not was_winner:
                        fighter_stats[fighter]["current_losing_streak"] += 1
                    else:
                        fighter_stats[fighter]["current_losing_streak"] = 0

                # ========== CRITICAL: NOW SYNC POST-FIGHT ELO FOR NEXT FIGHT ==========
                # All opponent classification is complete - safe to update tracked ELO now
                # This ensures next fight sees correct pre-fight ELO for these fighters
                fighter_stats[r_fighter]["elo"] = self.fighter_elos[r_fighter]
                fighter_stats[b_fighter]["elo"] = self.fighter_elos[b_fighter]

        print("\nData leakage fixed successfully!\n")

        # Calculate ELO ratings chronologically (no data leakage)
        df = self.calculate_elo_ratings(df)

        return df

    def build_features(self, df, recompute_elo=True):
        """
        Build all engineered features from base r_*/b_* columns.

        This is the SINGLE SOURCE OF TRUTH for all feature engineering.
        All differentials, polynomials, interactions, and ratios are computed here.

        Args:
            df: DataFrame with base r_*/b_* columns
            recompute_elo: If False, skip ELO differential calculation (used by build_swapped)

        This method is called by:
        1. prepare_core_features() - to build all features from scratch
        2. swap_corners() - after swapping r_/b_ columns and flipping winner
        3. build_swapped() - after swapping for antisymmetrization (recompute_elo=False)

        By recomputing ALL features (including basic differentials) from r_*/b_* bases,
        we achieve mathematically bulletproof corner swapping without fragile
        pattern-matching negation logic.
        """
        # Verify required base columns exist
        if "r_age_at_event" not in df.columns or "b_age_at_event" not in df.columns:
            raise KeyError(
                f"Missing required age columns in build_features()!\n"
                f"   r_age_at_event exists: {'r_age_at_event' in df.columns}\n"
                f"   b_age_at_event exists: {'b_age_at_event' in df.columns}\n"
                f"   This likely means prepare_upcoming_fight() didn't set up the data correctly."
            )

        # ========== STEP 1: CALCULATE ALL BASIC DIFFERENTIALS ==========
        # These must be calculated first as many compound features depend on them

        # Performance/career differentials (from chronologically-corrected stats)
        for stat in ["wins", "losses", "win_loss_ratio", "pro_SLpM", "pro_SApM",
                    "pro_sig_str_acc", "pro_str_def", "pro_td_avg", "pro_td_acc",
                    "pro_td_def", "pro_sub_avg", "ko_rate", "sub_rate", "dec_rate",
                    "recent_form", "win_streak", "loss_streak", "last_5_wins",
                    "days_since_last_fight", "recent_finish_rate", "durability",
                    # Career-based positional/target/control metrics
                    "distance_pct", "clinch_pct", "ground_pct",
                    "head_pct", "body_pct", "leg_pct",
                    "avg_ctrl_sec", "avg_rev",
                    # Opponent quality and momentum
                    "avg_opponent_elo", "distance_pct_momentum", "slpm_momentum", "ctrl_sec_momentum",
                    # Main event experience
                    "main_event_fights",
                    # PHASE 2: Advanced career metrics
                    "total_rounds_fought", "total_fights_fought", "title_fights", "five_round_fights",
                    "last_fight_was_finish", "last_fight_was_win", "last_fight_dominance",
                    "early_finish_rate", "late_finish_rate", "first_round_ko_rate",
                    "fights_last_24_months", "avg_finish_time_last_3", "chin_deterioration"]:
            # Check if columns exist before computing diff
            r_col = f"r_{stat}_corrected"
            b_col = f"b_{stat}_corrected"
            if r_col in df.columns and b_col in df.columns:
                df[f"{stat}_diff_corrected"] = df[r_col] - df[b_col]
            else:
                # If columns don't exist, set diff to 0 (neutral)
                df[f"{stat}_diff_corrected"] = 0.0

        # Physical differentials (not corrected - these are static attributes)
        for stat in ["height", "reach", "weight", "age_at_event", "ape_index"]:
            if f"r_{stat}" in df.columns and f"b_{stat}" in df.columns:
                df[f"{stat}_diff"] = df[f"r_{stat}"] - df[f"b_{stat}"]

        # ELO differentials (calculated from pre-fight ELO ratings)
        # Skip if recompute_elo=False (used by build_swapped - ELO handled separately)
        if recompute_elo and "r_elo_pre_fight" in df.columns and "b_elo_pre_fight" in df.columns:
            df["elo_diff"] = df["r_elo_pre_fight"] - df["b_elo_pre_fight"]

        # opponent_quality_diff: already calculated in fix_data_leakage() with proper chronological data
        # No need to recalculate here

        # ========== STEP 2: BUILD COMPOUND FEATURES FROM DIFFERENTIALS ==========

        # Core derived features
        df["net_striking_advantage"] = (df["r_pro_SLpM_corrected"] - df["r_pro_SApM_corrected"]) - (df["b_pro_SLpM_corrected"] - df["b_pro_SApM_corrected"])
        df["striking_efficiency"] = (df["r_pro_SLpM_corrected"] * df["r_pro_sig_str_acc_corrected"]) - (df["b_pro_SLpM_corrected"] * df["b_pro_sig_str_acc_corrected"])
        df["defensive_striking"] = (df["r_pro_str_def_corrected"] - df["r_pro_SApM_corrected"]) - (df["b_pro_str_def_corrected"] - df["b_pro_SApM_corrected"])
        df["grappling_control"] = (df["r_pro_td_avg_corrected"] * df["r_pro_td_acc_corrected"]) - (df["b_pro_td_avg_corrected"] * df["b_pro_td_acc_corrected"])

        # Experience features
        df["r_total_fights"] = df["r_wins_corrected"] + df["r_losses_corrected"]
        df["b_total_fights"] = df["b_wins_corrected"] + df["b_losses_corrected"]
        df["experience_gap"] = df["r_total_fights"] - df["b_total_fights"]

        # Finish rates
        df["r_finish_rate"] = (df["r_ko_rate_corrected"] + df["r_sub_rate_corrected"]) / 2
        df["b_finish_rate"] = (df["b_ko_rate_corrected"] + df["b_sub_rate_corrected"]) / 2
        df["finish_rate_diff"] = df["r_finish_rate"] - df["b_finish_rate"]

        # Style features
        df["r_striker_score"] = df["r_pro_SLpM_corrected"] - df["r_pro_td_avg_corrected"]
        df["b_striker_score"] = df["b_pro_SLpM_corrected"] - df["b_pro_td_avg_corrected"]
        df["striker_advantage"] = df["r_striker_score"] - df["b_striker_score"]

        df["r_grappler_score"] = df["r_pro_td_avg_corrected"] + df["r_pro_sub_avg_corrected"]
        df["b_grappler_score"] = df["b_pro_td_avg_corrected"] + df["b_pro_sub_avg_corrected"]
        df["grappler_advantage"] = df["r_grappler_score"] - df["b_grappler_score"]

        # Style clash: striker vs grappler matchup
        # Positive = red striker facing blue grappler (striker advantage)
        # Negative = blue striker facing red grappler (grappler advantage)
        # Near 0 = both strikers or both grapplers (no style clash)
        df["striker_vs_grappler"] = (df["r_striker_score"] * df["b_grappler_score"]) - (df["b_striker_score"] * df["r_grappler_score"])

        # Stance matchup features (4 features)
        # Initialize
        df["orthodox_vs_southpaw_advantage"] = 0
        df["orthodox_vs_switch_advantage"] = 0
        df["southpaw_vs_switch_advantage"] = 0
        df["mirror_matchup"] = 0

        # Calculate if stance data is available
        if "r_stance" in df.columns and "b_stance" in df.columns:
            for idx in df.index:
                r_stance = df.at[idx, "r_stance"]
                b_stance = df.at[idx, "b_stance"]

                # Handle missing/NaN stances
                if pd.isna(r_stance):
                    r_stance = "Orthodox"
                if pd.isna(b_stance):
                    b_stance = "Orthodox"

                # Feature 1: Orthodox vs Southpaw (directional - auto-negated on swap via "_advantage")
                if r_stance == "Orthodox" and b_stance == "Southpaw":
                    df.at[idx, "orthodox_vs_southpaw_advantage"] = 1
                elif r_stance == "Southpaw" and b_stance == "Orthodox":
                    df.at[idx, "orthodox_vs_southpaw_advantage"] = -1
                else:
                    df.at[idx, "orthodox_vs_southpaw_advantage"] = 0

                # Feature 2: Orthodox vs Switch (directional - auto-negated on swap via "_advantage")
                if r_stance == "Orthodox" and b_stance == "Switch":
                    df.at[idx, "orthodox_vs_switch_advantage"] = 1
                elif r_stance == "Switch" and b_stance == "Orthodox":
                    df.at[idx, "orthodox_vs_switch_advantage"] = -1
                else:
                    df.at[idx, "orthodox_vs_switch_advantage"] = 0

                # Feature 3: Southpaw vs Switch (directional - auto-negated on swap via "_advantage")
                if r_stance == "Southpaw" and b_stance == "Switch":
                    df.at[idx, "southpaw_vs_switch_advantage"] = 1
                elif r_stance == "Switch" and b_stance == "Southpaw":
                    df.at[idx, "southpaw_vs_switch_advantage"] = -1
                else:
                    df.at[idx, "southpaw_vs_switch_advantage"] = 0

                # Feature 4: Mirror matchup (symmetric - NOT negated on swap)
                if r_stance == b_stance:
                    df.at[idx, "mirror_matchup"] = 1
                else:
                    df.at[idx, "mirror_matchup"] = 0

        # Fighter trajectory is now calculated in fix_data_leakage()

        # Context features
        # Ring rust factor: Performance penalty based on layoff time
        # 0-180 days = no penalty (0)
        # 180-365 days = mild penalty (0.1-0.3)
        # 365-730 days = moderate penalty (0.3-0.6)
        # 730+ days = severe penalty (0.6-1.0)
        df["r_ring_rust"] = 0.0
        df["b_ring_rust"] = 0.0

        for idx, row in df.iterrows():
            for prefix in ["r", "b"]:
                days_off = row.get(f"{prefix}_days_since_last_fight_corrected", 0)
                if pd.notna(days_off) and days_off > 0:
                    if days_off <= 180:
                        rust = 0.0
                    elif days_off <= 365:
                        rust = (days_off - 180) / 185 * 0.3  # 0.0 to 0.3
                    elif days_off <= 730:
                        rust = 0.3 + ((days_off - 365) / 365 * 0.3)  # 0.3 to 0.6
                    else:
                        rust = min(0.6 + ((days_off - 730) / 365 * 0.4), 1.0)  # 0.6 to 1.0 (capped)
                    df.at[idx, f"{prefix}_ring_rust"] = rust

        df["ring_rust_factor"] = df["b_ring_rust"] - df["r_ring_rust"]  # Positive = red advantage

        # Weight class factor (based on fighter's comfort at current weight class)
        # Positive = red advantage (red more comfortable at this weight)
        if "r_weight_comfort" in df.columns and "b_weight_comfort" in df.columns:
            df["weight_class_factor"] = df["r_weight_comfort"] - df["b_weight_comfort"]
        else:
            df["weight_class_factor"] = 0.0  # No advantage if not calculated

        if "is_title_bout" not in df.columns:
            df["is_title_bout"] = 0
        if "total_rounds" not in df.columns:
            df["total_rounds"] = 3

        # NEW: Rounds-based strategy adjustments
        # 5-round fights favor cardio/decision specialists, 3-round favor explosive finishers
        df["rounds_x_cardio"] = df["total_rounds"] * df["dec_rate_diff_corrected"]
        df["rounds_x_finish_rate"] = (5 - df["total_rounds"]) * df["finish_rate_diff"]  # 3-rounders favor finishers
        df["rounds_x_durability"] = df["total_rounds"] * df["durability_diff_corrected"]  # 5-rounders favor durable fighters

        # Momentum features
        df["momentum_swing"] = df["recent_form_diff_corrected"] + (df["win_streak_diff_corrected"] * 0.5)

        # Interaction features
        df["elo_x_form"] = df["elo_diff"] * df["recent_form_diff_corrected"]
        df["reach_x_striking"] = df["reach_diff"] * df["pro_SLpM_diff_corrected"]
        df["experience_x_age"] = df["experience_gap"] * df["age_at_event_diff"]
        df["age_x_striking"] = df["age_at_event_diff"] * df["pro_SLpM_diff_corrected"]
        df["age_x_grappling"] = df["age_at_event_diff"] * df["pro_td_avg_diff_corrected"]
        df["age_x_durability"] = df["age_at_event_diff"] * df["durability_diff_corrected"]
        df["td_x_defense"] = df["pro_td_avg_diff_corrected"] * df["pro_td_def_diff_corrected"]
        df["grappling_x_experience"] = df["grappler_advantage"] * df["experience_gap"]
        df["striking_x_accuracy"] = df["pro_SLpM_diff_corrected"] * df["pro_sig_str_acc_diff_corrected"]
        df["striking_x_defense"] = df["pro_SLpM_diff_corrected"] * df["pro_str_def_diff_corrected"]
        df["form_x_experience"] = df["recent_form_diff_corrected"] * df["experience_gap"]
        df["finish_x_momentum"] = df["finish_rate_diff"] * df["momentum_swing"]

        # height_x_reach: Product of two differentials
        df["height_x_reach"] = df["height_diff"] * df["reach_diff"]

        df["physical_x_striking"] = (df["height_diff"] + df["reach_diff"]) * df["pro_SLpM_diff_corrected"]

        # Elite interaction features
        df["elo_x_win_ratio"] = df["elo_diff"] * df["win_loss_ratio_diff_corrected"]
        df["win_ratio_x_form"] = df["win_loss_ratio_diff_corrected"] * df["recent_form_diff_corrected"]
        df["durability_x_striking"] = df["durability_diff_corrected"] * df["net_striking_advantage"]
        df["elo_x_durability"] = df["elo_diff"] * df["durability_diff_corrected"]
        df["submission_x_grappling"] = df["sub_rate_diff_corrected"] * df["grappler_advantage"]
        df["ko_power_x_striking"] = df["ko_rate_diff_corrected"] * df["striker_advantage"]
        df["momentum_x_win_streak"] = df["momentum_swing"] * df["win_streak_diff_corrected"]

        # Strategic interactions
        df["age_x_win_streak"] = df["age_at_event_diff"] * df["win_streak_diff_corrected"]
        df["elo_x_sub_threat"] = df["elo_diff"] * df["pro_sub_avg_diff_corrected"]
        df["form_x_durability"] = df["recent_form_diff_corrected"] * df["durability_diff_corrected"]
        df["striking_x_grappling_matchup"] = df["striker_advantage"] * df["grappler_advantage"]

        # Compound features
        df["total_finish_threat"] = df["ko_rate_diff_corrected"] + df["sub_rate_diff_corrected"]
        df["elo_x_finish"] = df["elo_diff"] * df["total_finish_threat"]
        df["elite_finisher"] = df["elo_diff"] * df["finish_rate_diff"] * df["recent_form_diff_corrected"]
        df["complete_fighter"] = df["net_striking_advantage"] * df["grappler_advantage"] * df["durability_diff_corrected"]
        df["unstoppable_streak"] = df["win_streak_diff_corrected"] * df["momentum_swing"] * df["recent_form_diff_corrected"]
        df["veteran_advantage"] = df["win_loss_ratio_diff_corrected"] * df["experience_gap"] * (-df["age_at_event_diff"])
        df["momentum_combo"] = df["recent_form_diff_corrected"] * df["win_streak_diff_corrected"]

        df["r_age_prime"] = 1.0 - abs(df["r_age_at_event"] - 29.5) / 10.0
        df["b_age_prime"] = 1.0 - abs(df["b_age_at_event"] - 29.5) / 10.0
        df["age_prime_advantage"] = df["r_age_prime"] - df["b_age_prime"]

        # Layoff freshness vs ring rust (optimal ~90-180 days)
        # Peak performance at ~135 days between fights (4-5 months)
        df["r_layoff_factor"] = 1.0 - abs(df["r_days_since_last_fight_corrected"] - 135) / 200.0  # Peak at 135 days
        df["b_layoff_factor"] = 1.0 - abs(df["b_days_since_last_fight_corrected"] - 135) / 200.0
        df["freshness_advantage"] = df["r_layoff_factor"] - df["b_layoff_factor"]

        # Positional compound interactions
        df["clinch_x_grappling"] = df["clinch_pct_diff_corrected"] * df["grappler_advantage"]
        df["distance_x_striking"] = df["distance_pct_diff_corrected"] * df["net_striking_advantage"]
        df["ground_x_control"] = (
            df["ground_pct_diff_corrected"] *
            (df["avg_ctrl_sec_diff_corrected"] / 300.0).clip(-1, 1)
        )
        df["positional_mastery"] = (
            (abs(df["distance_pct_diff_corrected"]) +
             abs(df["clinch_pct_diff_corrected"]) +
             abs(df["ground_pct_diff_corrected"])) *
            df["grappler_advantage"] *
            df["net_striking_advantage"]
        )
        df["control_dominance"] = (
            (df["avg_ctrl_sec_diff_corrected"] / 300.0).clip(-1, 1) +
            df["avg_rev_diff_corrected"] * 0.1
        )

        # Style clash and matchup features
        # NOTE: style_clash_severity uses abs(), so it's swap-invariant (not negated)
        df["style_clash_severity"] = abs(df["striker_advantage"] * 0.5 + df["grappler_advantage"] * 0.5)
        df["power_vs_technique"] = (df["r_pro_SLpM_corrected"] - df["b_pro_SLpM_corrected"]) * 0.6 + (df["r_pro_sig_str_acc_corrected"] - df["b_pro_sig_str_acc_corrected"]) * 0.4
        df["finish_pressure"] = (df["r_ko_rate_corrected"] - df["b_ko_rate_corrected"]) * 0.5 + (df["r_sub_rate_corrected"] - df["b_sub_rate_corrected"]) * 0.3
        df["offensive_output"] = df["pro_SLpM_diff_corrected"] + df["pro_td_avg_diff_corrected"]
        df["defensive_composite"] = df["pro_str_def_diff_corrected"] + df["pro_td_def_diff_corrected"]
        df["upset_potential"] = -df["elo_diff"] * df["recent_form_diff_corrected"]
        df["skill_momentum"] = df["elo_diff"] * df["momentum_swing"]

        # Additional high-value features (ko and submission specialists)
        df["ko_specialist_gap"] = (
            (df["r_ko_rate_corrected"] * df["r_pro_SLpM_corrected"]) -
            (df["b_ko_rate_corrected"] * df["b_pro_SLpM_corrected"])
        )
        df["submission_specialist_gap"] = (
            (df["r_sub_rate_corrected"] * df["r_pro_sub_avg_corrected"]) -
            (df["b_sub_rate_corrected"] * df["b_pro_sub_avg_corrected"])
        )

        # Elite interaction features
        df["streak_differential"] = df["win_streak_diff_corrected"] * df["loss_streak_diff_corrected"]

        # Elite compound features (desperation, momentum sustainability, adversity)
        # Create intermediate columns for both fighters before taking differential
        df["r_desperation"] = df["r_loss_streak_corrected"] * (df["r_age_at_event"] / 35.0)
        df["b_desperation"] = df["b_loss_streak_corrected"] * (df["b_age_at_event"] / 35.0)
        df["desperation_diff"] = df["r_desperation"] - df["b_desperation"]

        # Momentum sustainability: form divided by layoff penalty
        df["r_momentum_sustainability"] = df["r_recent_form_corrected"] / (df["r_days_since_last_fight_corrected"] / 100.0 + 1.0)
        df["b_momentum_sustainability"] = df["b_recent_form_corrected"] / (df["b_days_since_last_fight_corrected"] / 100.0 + 1.0)
        df["momentum_sustainability_diff"] = df["r_momentum_sustainability"] - df["b_momentum_sustainability"]

        # Adversity experience: win ratio weighted by losses (fighters who maintained good ratio despite many losses)
        df["r_adversity_experience"] = df["r_win_loss_ratio_corrected"] * (df["r_losses_corrected"] + 1)
        df["b_adversity_experience"] = df["b_win_loss_ratio_corrected"] * (df["b_losses_corrected"] + 1)
        df["adversity_experience_diff"] = df["r_adversity_experience"] - df["b_adversity_experience"]

        # PHASE 2: Advanced career tracking features
        # These compound features are calculated from already-computed differentials

        # Career stage features (4 features)
        # Prime years advantage (age 26-33 is optimal)
        df["r_in_prime"] = ((df["r_age_at_event"] >= 26) & (df["r_age_at_event"] <= 33)).astype(float)
        df["b_in_prime"] = ((df["b_age_at_event"] >= 26) & (df["b_age_at_event"] <= 33)).astype(float)
        df["prime_years_advantage"] = df["r_in_prime"] - df["b_in_prime"]

        # Declining phase (penalty for age > 35)
        # Applies penalty for fighters past their prime
        df["r_declining_penalty"] = np.maximum(0, df["r_age_at_event"] - 35) * 0.1
        df["b_declining_penalty"] = np.maximum(0, df["b_age_at_event"] - 35) * 0.1
        df["declining_phase_diff"] = df["r_declining_penalty"] - df["b_declining_penalty"]

        # Rising prospect advantage (age < 26)
        df["r_rising_prospect"] = (df["r_age_at_event"] < 26).astype(float)
        df["b_rising_prospect"] = (df["b_age_at_event"] < 26).astype(float)
        df["rising_prospect_advantage"] = df["r_rising_prospect"] - df["b_rising_prospect"]

        # Age/experience ratio (late bloomer vs prodigy indicator)
        df["r_age_experience_ratio"] = df["r_age_at_event"] / np.maximum(df["r_total_fights"], 1)
        df["b_age_experience_ratio"] = df["b_age_at_event"] / np.maximum(df["b_total_fights"], 1)
        df["age_experience_ratio_diff"] = df["r_age_experience_ratio"] - df["b_age_experience_ratio"]

        # Last fight momentum features
        df["last_fight_finish_momentum"] = (
            df["last_fight_was_finish_diff_corrected"] * df["last_fight_was_win_diff_corrected"]
        )

        df["last_fight_dominance"] = df["last_fight_dominance_diff_corrected"]

        df["last_fight_complete_momentum"] = (
            df["last_fight_was_win_diff_corrected"] +
            df["last_fight_was_finish_diff_corrected"] +
            df["last_fight_dominance_diff_corrected"] * 0.2
        )

        # TIER S: Damage Absorption Trends
        # Chin deterioration (recent damage vs career damage)
        df["chin_deterioration"] = df["chin_deterioration_diff_corrected"]

        # TIER S: Fight Context Features
        # Title fight experience advantage
        df["title_fight_experience"] = df["title_fights_diff_corrected"]

        # Five-round experience advantage
        df["five_round_experience"] = df["five_round_fights_diff_corrected"]

        # Five-round cardio advantage (if 5 rounds, favor better cardio)
        df["five_round_cardio_advantage"] = df["total_rounds"] * df["dec_rate_diff_corrected"]

        # TIER A: Activity & Ring Rust Features
        # Activity level (fights in last 24 months)
        df["activity_level_diff"] = df["fights_last_24_months_diff_corrected"]

        # Layoff severity (non-linear penalty for long layoffs > 365 days)
        df["r_layoff_severity"] = np.maximum(0, (df["r_days_since_last_fight_corrected"] - 365) / 30)
        df["b_layoff_severity"] = np.maximum(0, (df["b_days_since_last_fight_corrected"] - 365) / 30)
        df["layoff_severity_diff"] = df["r_layoff_severity"] - df["b_layoff_severity"]

        # Optimal activity window (90-180 days is ideal)
        df["r_optimal_activity"] = ((df["r_days_since_last_fight_corrected"] >= 90) &
                                     (df["r_days_since_last_fight_corrected"] <= 180)).astype(float)
        df["b_optimal_activity"] = ((df["b_days_since_last_fight_corrected"] >= 90) &
                                     (df["b_days_since_last_fight_corrected"] <= 180)).astype(float)
        df["optimal_activity_advantage"] = df["r_optimal_activity"] - df["b_optimal_activity"]

        # Overactivity penalty (fighting too frequently, < 60 days)
        df["r_overactivity"] = (df["r_days_since_last_fight_corrected"] < 60).astype(float)
        df["b_overactivity"] = (df["b_days_since_last_fight_corrected"] < 60).astype(float)
        df["overactivity_diff"] = df["r_overactivity"] - df["b_overactivity"]

        # TIER A: Finish Timing Specialization
        # Early finisher advantage (rounds 1-2)
        df["early_finisher_advantage"] = df["early_finish_rate_diff_corrected"]

        # Late finisher advantage (rounds 3-5)
        df["late_finisher_advantage"] = df["late_finish_rate_diff_corrected"]

        # First round killer advantage
        df["first_round_killer_advantage"] = df["first_round_ko_rate_diff_corrected"]

        # Finish timing × rounds matchup (early finishers better in 3-rounders)
        df["finish_timing_rounds_matchup"] = (
            df["early_finish_rate_diff_corrected"] * (5 - df["total_rounds"]) +
            df["late_finish_rate_diff_corrected"] * (df["total_rounds"] - 3)
        )

        # TIER A: Advanced Momentum Features
        # Finish momentum acceleration (recent finishes vs career average)
        df["finish_momentum_acceleration"] = (
            (df["r_rolling_finish_rate_3_corrected"] - df["r_finish_rate"]) -
            (df["b_rolling_finish_rate_3_corrected"] - df["b_finish_rate"])
        )

        # TIER B: Grappling Control Efficiency
        # Control time per takedown (quality of control) - avoid division by zero
        df["r_ctrl_per_td"] = df["r_avg_ctrl_sec_corrected"] / np.maximum(df["r_pro_td_avg_corrected"], 0.1)
        df["b_ctrl_per_td"] = df["b_avg_ctrl_sec_corrected"] / np.maximum(df["b_pro_td_avg_corrected"], 0.1)
        df["ctrl_time_per_td_diff"] = df["r_ctrl_per_td"] - df["b_ctrl_per_td"]

        # TIER B: Total output pace (strikes + TDs per minute)
        df["r_total_output_pace"] = (df["r_pro_SLpM_corrected"] + df["r_pro_td_avg_corrected"]) / 15
        df["b_total_output_pace"] = (df["b_pro_SLpM_corrected"] + df["b_pro_td_avg_corrected"]) / 15
        df["total_output_pace_diff"] = df["r_total_output_pace"] - df["b_total_output_pace"]

        # TIER B: Pressure differential (offense - defense)
        df["r_pressure_score"] = (df["r_pro_SLpM_corrected"] - df["r_pro_SApM_corrected"]) + (df["r_pro_td_avg_corrected"] * 0.5)
        df["b_pressure_score"] = (df["b_pro_SLpM_corrected"] - df["b_pro_SApM_corrected"]) + (df["b_pro_td_avg_corrected"] * 0.5)
        df["pressure_differential"] = df["r_pressure_score"] - df["b_pressure_score"]

        # TIER B: Physical Attribute Interactions
        # Reach × Height interaction
        df["r_reach_height_product"] = df["r_reach"] * df["r_height"]
        df["b_reach_height_product"] = df["b_reach"] * df["b_height"]
        df["reach_height_interaction_diff"] = df["r_reach_height_product"] - df["b_reach_height_product"]

        # ========== RESEARCH-BACKED FEATURES (RF/GBDT/SVM Analysis) ==========

        # H2H × ELO: When you've beaten someone before AND have better rating
        df["h2h_x_elo"] = df["h2h_advantage"] * df["elo_diff"]

        # H2H × Recent Form: Previous wins matter more when you're on a hot streak
        df["h2h_x_form"] = df["h2h_advantage"] * df["recent_form_diff_corrected"]

        # Age differential interactions (age_at_event_diff is #2 at 0.015408)
        # Prime age advantage: Being in prime (26-33) vs opponent not in prime
        def calculate_age_prime_score(age):
            if 26 <= age <= 33:
                return 1.0
            elif 23 <= age < 26 or 33 < age <= 36:
                return 0.6
            else:
                return 0.2

        df["r_age_prime_score"] = df["r_age_at_event"].apply(calculate_age_prime_score)
        df["b_age_prime_score"] = df["b_age_at_event"].apply(calculate_age_prime_score)
        df["age_prime_score_diff"] = df["r_age_prime_score"] - df["b_age_prime_score"]

        # Age × Experience: Older fighters with more experience are more dangerous
        df["age_x_experience"] = df["age_at_event_diff"] * (
            (df["r_total_fights"] - df["b_total_fights"]) / 50.0
        ).clip(-1, 1)

        # Win ratio interactions (win_loss_ratio_diff_corrected is #3 at 0.014744)
        # Win ratio × finish rate: High win ratio fighters who finish fights are elite
        df["win_ratio_x_finish"] = df["win_loss_ratio_diff_corrected"] * df["finish_rate_diff"]

        # Win ratio × durability: High win ratio + durability = championship material
        df["win_ratio_x_durability"] = df["win_loss_ratio_diff_corrected"] * df["durability_diff_corrected"]

        # POSITIONAL STRIKING: CLINCH #3, GROUND #5, DISTANCE #6 in model rankings
        # NOW USING CAREER AVERAGES (not fight outcome data!)
        # These represent each fighter's career tendencies for where they land strikes

        # Positional advantage: Dominant in preferred position
        # Higher = better at fighting from distance, clinch, or ground
        df["positional_striking_advantage"] = (
            abs(df["distance_pct_diff_corrected"]) +
            abs(df["clinch_pct_diff_corrected"]) +
            abs(df["ground_pct_diff_corrected"])
        )

        # TARGET AREA DISTRIBUTION: Head/Body/Leg striking (ranked #10)
        # NOW USING CAREER AVERAGES of where strikes land
        # Target diversity: Fighters who can attack all areas are more dangerous
        df["target_distribution_advantage"] = (
            abs(df["head_pct_diff_corrected"]) +
            abs(df["body_pct_diff_corrected"]) +
            abs(df["leg_pct_diff_corrected"])
        )

        # CONTROL & REVERSALS: CTRL #5 in GBDT, REV #8/#9 in RF/GBDT
        # NOW USING CAREER AVERAGES per fight
        # Control dominance: Combines control time + reversals (grappling mastery)
        # Normalize control time (convert seconds to 0-1 scale, typical range 0-300 sec)
        # Already calculated above, but keep for clarity

        # Opponent quality: already calculated in fix_data_leakage(), don't overwrite!
        # df["opponent_quality_diff"] is calculated chronologically with proper historical data

        # Clutch and momentum features
        df["r_clutch_factor"] = df["r_win_loss_ratio_corrected"] * df["r_recent_form_corrected"]
        df["b_clutch_factor"] = df["b_win_loss_ratio_corrected"] * df["b_recent_form_corrected"]
        df["clutch_factor_diff"] = df["r_clutch_factor"] - df["b_clutch_factor"]

        df["r_momentum_quality"] = 0.5
        df["b_momentum_quality"] = 0.5
        df["momentum_quality_diff"] = 0

        # Pressure performance: Performance vs elite opponents + title fight success
        # Win rate vs elite opponents as proxy for pressure performance
        df["r_pressure_performance"] = df["r_win_rate_vs_elite_corrected"].fillna(0.5)
        df["b_pressure_performance"] = df["b_win_rate_vs_elite_corrected"].fillna(0.5)
        df["pressure_performance_diff"] = df["r_pressure_performance"] - df["b_pressure_performance"]

        # Form consistency: Based on variance in recent results
        # Higher win streak or consistent recent form = higher consistency
        # Formula: combine win streak and recent form
        df["r_form_consistency"] = (df["r_recent_form_corrected"].fillna(0.5) + df["r_win_streak_corrected"].fillna(0) * 0.1).clip(upper=1.0)
        df["b_form_consistency"] = (df["b_recent_form_corrected"].fillna(0.5) + df["b_win_streak_corrected"].fillna(0) * 0.1).clip(upper=1.0)
        df["form_consistency_diff"] = df["r_form_consistency"] - df["b_form_consistency"]

        # ========== NEW FEATURE: MAIN EVENT EXPERIENCE ==========
        # Championship fight experience (5-round fights)
        # Simple count differential - big fight experience matters
        df["main_event_experience_diff"] = df["main_event_fights_diff_corrected"]

        # ========== PHASE 1A: POLYNOMIAL FEATURES (NON-LINEAR EFFECTS) ==========

        # TIER 1: Critical Non-Linear Features (Top importances)
        # These capture U-shaped curves, saturation effects, and prime years

        # 1. ELO Squared (elite fighters have exponential advantage)
        df["elo_diff_squared"] = df["elo_diff"] ** 2

        # 2. Age Squared (prime years = 26-33, U-shaped curve)
        df["age_diff_squared"] = df["age_at_event_diff"] ** 2

        # 3. Age Cubic (better captures prime decline after 35)
        df["age_diff_cubic"] = df["age_at_event_diff"] ** 3

        # 4. Win/Loss Ratio Squared (elite records matter more)
        df["win_loss_ratio_diff_squared"] = df["win_loss_ratio_diff_corrected"] ** 2

        # 5. Reach Squared (extreme reach advantages compound)
        df["reach_diff_squared"] = df["reach_diff"] ** 2

        # TIER 2: Momentum Non-Linearity (hot streaks are exponential)

        # 6. Recent Form Squared (momentum snowballs)
        df["recent_form_diff_squared"] = df["recent_form_diff_corrected"] ** 2

        # 7. Win Streak Squared (long streaks = confidence boost)
        df["win_streak_diff_squared"] = df["win_streak_diff_corrected"] ** 2

        # 8. Momentum Swing Squared (extreme momentum matters more)
        df["momentum_swing_squared"] = df["momentum_swing"] ** 2

        # TIER 3: Skill Saturation (diminishing returns at extremes)

        # 9. Striking Volume Squared (high output vs very high output)
        df["pro_SLpM_diff_squared"] = df["pro_SLpM_diff_corrected"] ** 2

        # 10. Striking Accuracy Squared (precision compounds)
        df["pro_sig_str_acc_diff_squared"] = df["pro_sig_str_acc_diff_corrected"] ** 2

        # 11. Takedown Average Squared (TD specialists dominate)
        df["pro_td_avg_diff_squared"] = df["pro_td_avg_diff_corrected"] ** 2

        # 12. Net Striking Squared (elite strikers pull away)
        df["net_striking_advantage_squared"] = df["net_striking_advantage"] ** 2

        # TIER 4: Physical Attributes (size advantages compound)

        # 13. Height Squared (tall fighters control distance exponentially)
        df["height_diff_squared"] = df["height_diff"] ** 2

        # 14. Weight Squared (weight cuts/advantages non-linear)
        df["weight_diff_squared"] = df["weight_diff"] ** 2

        # 15. Ape Index Squared (extreme wingspan = huge advantage)
        df["ape_index_diff_squared"] = df["ape_index_diff"] ** 2

        # TIER 5: Experience Effects (veteran wisdom vs youth)

        # 16. Experience Gap Squared (huge experience gaps matter more)
        df["experience_gap_squared"] = df["experience_gap"] ** 2

        # 17. Days Since Last Fight Squared (ring rust/overtraining curves)
        df["days_since_last_fight_squared"] = df["days_since_last_fight_diff_corrected"] ** 2

        # TIER 6: Finishing Ability (finish rates plateau/accelerate)

        # 18. KO Rate Squared (KO artists have disproportionate impact)
        df["ko_rate_diff_squared"] = df["ko_rate_diff_corrected"] ** 2

        # 19. Sub Rate Squared (submission specialists create fear)
        df["sub_rate_diff_squared"] = df["sub_rate_diff_corrected"] ** 2

        # 20. Finish Rate Squared (combined finishing threat)
        df["finish_rate_diff_squared"] = df["finish_rate_diff"] ** 2

        # TIER 7: Defense/Durability (defense compounds exponentially)

        # 21. Striking Defense Squared (elite defense = hard to hit)
        df["pro_str_def_diff_squared"] = df["pro_str_def_diff_corrected"] ** 2

        # 22. TD Defense Squared (elite wrestlers hard to take down)
        df["pro_td_def_diff_squared"] = df["pro_td_def_diff_corrected"] ** 2

        # 23. Durability Squared (iron chins rare and valuable)
        df["durability_diff_squared"] = df["durability_diff_corrected"] ** 2

        # TIER 8: Positional Mastery (specialist advantages compound)

        # 24. Clinch % Squared (clinch specialists dominate there)
        df["clinch_pct_diff_squared"] = df["clinch_pct_diff_corrected"] ** 2

        # 25. Ground % Squared (ground-and-pound specialists)
        df["ground_pct_diff_squared"] = df["ground_pct_diff_corrected"] ** 2

        # 26. Distance % Squared (distance strikers control range)
        df["distance_pct_diff_squared"] = df["distance_pct_diff_corrected"] ** 2

        # 27. Control Time Squared (control time dominance compounds)
        df["avg_ctrl_sec_diff_squared"] = df["avg_ctrl_sec_diff_corrected"] ** 2

        # TIER 9: Advanced Interaction Squares (amplify synergies)

        # 28. ELO × Form Squared (hot elite fighters unbeatable)
        df["elo_x_form_squared"] = df["elo_x_form"] ** 2

        # 29. Striker Advantage Squared (pure strikers vs pure grapplers)
        df["striker_advantage_squared"] = df["striker_advantage"] ** 2

        # 30. Grappler Advantage Squared (pure grapplers dominate)
        df["grappler_advantage_squared"] = df["grappler_advantage"] ** 2

        # TIER 10: Quality/Momentum Trends (recent changes matter more)

        # 31. SLpM Momentum Squared (improving strikers surge)
        df["slpm_momentum_diff_squared"] = df["slpm_momentum_diff_corrected"] ** 2

        # 32. Control Momentum Squared (grappling improvements)
        df["ctrl_sec_momentum_diff_squared"] = df["ctrl_sec_momentum_diff_corrected"] ** 2

        # TIER 11: Elite Compound Squares (championship factors)

        # 33. Elite Finisher Squared (triple threat advantage)
        df["elite_finisher_squared"] = df["elite_finisher"] ** 2

        # 34. Veteran Advantage Squared (experience in prime)
        df["veteran_advantage_squared"] = df["veteran_advantage"] ** 2

        # 35. Unstoppable Streak Squared (peak momentum)
        df["unstoppable_streak_squared"] = df["unstoppable_streak"] ** 2

        # TIER 12: H2H and Special Cases (history repeats)

        # 36. H2H Advantage Squared (rematches favor previous winner)
        df["h2h_advantage_squared"] = df["h2h_advantage"] ** 2

        # 37. Opponent Quality Squared (strength of schedule non-linear)
        df["avg_opponent_elo_diff_squared"] = df["avg_opponent_elo_diff_corrected"] ** 2

        # 38. Main Event Experience Squared (championship pedigree)
        df["main_event_experience_diff_squared"] = df["main_event_experience_diff"] ** 2

        # TIER 13: Cubic Terms for Complex Curves (age prime is critical)

        # 39. ELO Cubic (extreme rating gaps = certain victory)
        df["elo_diff_cubic"] = df["elo_diff"] ** 3

        # 40. Win/Loss Ratio Cubic (undefeated fighters vs journeymen)
        df["win_loss_ratio_diff_cubic"] = df["win_loss_ratio_diff_corrected"] ** 3

        # 41. Reach Cubic (extreme reach = impossible to close distance)
        df["reach_diff_cubic"] = df["reach_diff"] ** 3

        # 42. Experience Gap Cubic (veterans vs debuters)
        df["experience_gap_cubic"] = df["experience_gap"] ** 3

        # PHASE 2: POLYNOMIAL FEATURES FOR ADVANCED STATS

        # Total rounds/fights fought (experience accumulation compounds)
        df["total_rounds_fought_diff_squared"] = df["total_rounds_fought_diff_corrected"] ** 2
        df["total_fights_fought_diff_squared"] = df["total_fights_fought_diff_corrected"] ** 2

        # Title/5-round experience (championship experience compounds)
        df["title_fights_diff_squared"] = df["title_fights_diff_corrected"] ** 2
        df["five_round_fights_diff_squared"] = df["five_round_fights_diff_corrected"] ** 2

        # Last fight momentum (recent performance compounds)
        df["last_fight_dominance_squared"] = df["last_fight_dominance_diff_corrected"] ** 2
        df["last_fight_complete_momentum_squared"] = df["last_fight_complete_momentum"] ** 2

        # Finish timing rates (specialist advantages compound)
        df["early_finish_rate_diff_squared"] = df["early_finish_rate_diff_corrected"] ** 2
        df["late_finish_rate_diff_squared"] = df["late_finish_rate_diff_corrected"] ** 2
        df["first_round_ko_rate_diff_squared"] = df["first_round_ko_rate_diff_corrected"] ** 2

        # Activity level (fighting frequency matters)
        df["fights_last_24_months_diff_squared"] = df["fights_last_24_months_diff_corrected"] ** 2

        # Prime years advantage (being in prime is critical)
        df["prime_years_advantage_squared"] = df["prime_years_advantage"] ** 2

        # Age experience ratio (late bloomers vs prodigies)
        df["age_experience_ratio_diff_squared"] = df["age_experience_ratio_diff"] ** 2

        # Chin deterioration (damage accumulation is exponential)
        df["chin_deterioration_squared"] = df["chin_deterioration"] ** 2

        # Layoff severity (extreme layoffs have outsized impact)
        df["layoff_severity_diff_squared"] = df["layoff_severity_diff"] ** 2

        # Finish momentum acceleration (improving finishers)
        df["finish_momentum_acceleration_squared"] = df["finish_momentum_acceleration"] ** 2

        # Control efficiency (grappling mastery compounds)
        df["ctrl_time_per_td_diff_squared"] = df["ctrl_time_per_td_diff"] ** 2

        # Total output pace (high pace compounds fatigue)
        df["total_output_pace_diff_squared"] = df["total_output_pace_diff"] ** 2

        # Pressure differential (relentless pressure breaks opponents)
        df["pressure_differential_squared"] = df["pressure_differential"] ** 2

        # ========== PHASE 3: ADVANCED OPPONENT QUALITY, VOLATILITY, AND INTERACTION FEATURES ==========

        # PHASE 3: Calculate differentials for all new features
        for feature in ["avg_opponent_elo_l5", "elo_momentum_vs_competition",
                        "performance_vs_ranked_opponents", "performance_volatility_l10",
                        "finish_rate_acceleration", "slpm_coefficient_of_variation",
                        "performance_decline_velocity", "mileage_adjusted_age",
                        "prime_exit_risk", "career_inflection_point",
                        "title_shot_proximity_score", "tactical_evolution_score",
                        "finish_method_diversity", "aging_power_striker_penalty",
                        "elo_volatility_interaction", "layoff_veteran_interaction",
                        "bayesian_finish_rate", "confidence_weighted_damage_ratio",
                        "distance_from_career_peak", "elite_performance_frequency_l10"]:
            df[f"{feature}_diff"] = df[f"r_{feature}_corrected"] - df[f"b_{feature}_corrected"]

        # Add polynomial features for key PHASE 3 features
        # Use SIGNED squares to preserve directionality: sign(X) * X^2
        # This maintains antisymmetry: if X → -X, then sign(X)*X^2 → -sign(X)*X^2
        df["avg_opponent_elo_l5_diff_squared"] = np.sign(df["avg_opponent_elo_l5_diff"]) * (df["avg_opponent_elo_l5_diff"] ** 2)
        df["elo_momentum_vs_competition_diff_squared"] = np.sign(df["elo_momentum_vs_competition_diff"]) * (df["elo_momentum_vs_competition_diff"] ** 2)
        df["performance_decline_velocity_diff_squared"] = np.sign(df["performance_decline_velocity_diff"]) * (df["performance_decline_velocity_diff"] ** 2)
        df["mileage_adjusted_age_diff_squared"] = np.sign(df["mileage_adjusted_age_diff"]) * (df["mileage_adjusted_age_diff"] ** 2)
        df["layoff_veteran_interaction_diff_squared"] = np.sign(df["layoff_veteran_interaction_diff"]) * (df["layoff_veteran_interaction_diff"] ** 2)
        df["performance_volatility_l10_diff_squared"] = np.sign(df["performance_volatility_l10_diff"]) * (df["performance_volatility_l10_diff"] ** 2)
        df["finish_rate_acceleration_diff_squared"] = np.sign(df["finish_rate_acceleration_diff"]) * (df["finish_rate_acceleration_diff"] ** 2)

        # ========== PHASE 1B: ROLLING STATISTICS DIFFERENTIALS ==========

        # Rolling averages (recent performance trends)
        rolling_features = [
            "rolling_slpm_3", "rolling_slpm_5", "rolling_slpm_10",
            "rolling_sapm_3", "rolling_sapm_5", "rolling_sapm_10",
            "rolling_td_acc_3", "rolling_td_acc_5", "rolling_td_acc_10",
            "rolling_td_def_3", "rolling_td_def_5", "rolling_td_def_10",
            "rolling_ctrl_3", "rolling_ctrl_5", "rolling_ctrl_10",
            "rolling_finish_rate_3", "rolling_finish_rate_5", "rolling_finish_rate_10",
            "rolling_strike_acc_3", "rolling_strike_acc_5", "rolling_strike_acc_10",
            "rolling_damage_3", "rolling_damage_5", "rolling_damage_10",
        ]

        for feature in rolling_features:
            df[f"{feature}_diff"] = df[f"r_{feature}_corrected"] - df[f"b_{feature}_corrected"]

        # Variance/consistency metrics
        variance_features = ["slpm_variance_5", "sapm_variance_5", "performance_consistency_5"]
        for feature in variance_features:
            df[f"{feature}_diff"] = df[f"r_{feature}_corrected"] - df[f"b_{feature}_corrected"]

        # Trend features (improving/declining)
        trend_features = ["slpm_trend_5", "td_acc_trend_5", "finish_rate_trend_5"]
        for feature in trend_features:
            df[f"{feature}_diff"] = df[f"r_{feature}_corrected"] - df[f"b_{feature}_corrected"]

        # ========== PHASE 1C: OPPONENT-ADJUSTED METRICS DIFFERENTIALS ==========

        # Core opponent-adjusted metrics (base differentials)
        opponent_features = [
            "win_rate_vs_elite", "win_rate_vs_strikers", "win_rate_vs_grapplers",
            "win_rate_vs_durable", "win_rate_vs_finishers", "finish_rate_vs_elite",
            "recent_opponent_quality_5", "style_versatility", "step_up_performance"
        ]

        for feature in opponent_features:
            df[f"{feature}_diff"] = df[f"r_{feature}_corrected"] - df[f"b_{feature}_corrected"]

        # Interaction features (opponent-adjusted × other factors)
        # 10. Step-up performance × ELO (elite fighters who perform better vs elite)
        df["step_up_x_elo"] = df["step_up_performance_diff"] * df["elo_diff"]

        # 11. Style versatility × Recent form (well-rounded fighters with momentum)
        df["versatility_x_form"] = df["style_versatility_diff"] * df["recent_form_diff_corrected"]

        # 12. Elite win rate × Opponent quality (proven vs quality opposition)
        df["elite_wins_x_opp_quality"] = df["win_rate_vs_elite_diff"] * df["recent_opponent_quality_5_diff"]

        # 13. Finish rate vs elite × Power metrics (finishers vs best competition)
        df["elite_finish_x_power"] = df["finish_rate_vs_elite_diff"] * df["pro_SLpM_diff_corrected"]

        # 14. Striker wins × Opponent striking (striker killer advantage)
        df["striker_killer_metric"] = df["win_rate_vs_strikers_diff"] * df["pro_SLpM_diff_corrected"]

        # 15. Grappler wins × Opponent grappling (grappler killer advantage)
        df["grappler_killer_metric"] = df["win_rate_vs_grapplers_diff"] * df["pro_td_avg_diff_corrected"]

        # Advanced composite metrics
        # 16. Championship readiness (elite performance + opponent quality + step-up)
        df["championship_readiness"] = (
            df["win_rate_vs_elite_diff"] +
            df["step_up_performance_diff"] +
            (df["recent_opponent_quality_5_diff"] / 100)  # Scale down ELO
        ) / 3

        # 17. Competition level faced (recent opponent quality normalized)
        df["competition_level"] = df["recent_opponent_quality_5_diff"] / 100  # Normalize ELO to similar scale

        # ========== PHASE 1D: STATISTICAL RATIOS ==========

        # TIER 1: Efficiency Ratios (damage/output per resource)
        # 1. Striking output quality (volume × accuracy)
        df["striking_output_quality_diff"] = df["pro_SLpM_diff_corrected"] * df["pro_sig_str_acc_diff_corrected"]

        # 2. Grappling output quality (takedowns × accuracy)
        df["grappling_output_quality_diff"] = df["pro_td_avg_diff_corrected"] * df["pro_td_acc_diff_corrected"]

        # 3. Damage differential ratio (offense/defense in striking)
        # SLpM / SApM shows who deals more than they take
        r_damage_ratio = df["r_pro_SLpM_corrected"] / (df["r_pro_SApM_corrected"] + 0.01)  # Avoid div by 0
        b_damage_ratio = df["b_pro_SLpM_corrected"] / (df["b_pro_SApM_corrected"] + 0.01)
        df["damage_ratio_diff"] = r_damage_ratio - b_damage_ratio

        # 4. Defense-to-offense balance (striking)
        # Higher str_def / lower str_acc = defensive fighter
        r_def_off_balance = (df["r_pro_str_def_corrected"] + 0.01) / (df["r_pro_sig_str_acc_corrected"] + 0.01)
        b_def_off_balance = (df["b_pro_str_def_corrected"] + 0.01) / (df["b_pro_sig_str_acc_corrected"] + 0.01)
        df["defense_offense_balance_diff"] = r_def_off_balance - b_def_off_balance

        # 5. Takedown defense-to-offense balance
        r_td_balance = (df["r_pro_td_def_corrected"] + 0.01) / (df["r_pro_td_acc_corrected"] + 0.01)
        b_td_balance = (df["b_pro_td_def_corrected"] + 0.01) / (df["b_pro_td_acc_corrected"] + 0.01)
        df["td_defense_offense_balance_diff"] = r_td_balance - b_td_balance

        # 6. Finish efficiency (finishes per fight)
        r_finish_rate = df["r_finish_rate"]
        b_finish_rate = df["b_finish_rate"]
        df["finish_efficiency_diff"] = r_finish_rate - b_finish_rate

        # TIER 2: Quality Over Quantity Metrics
        # 7. Precision striking (high accuracy relative to output)
        # Accuracy / Volume = patient, precise striker
        r_precision = (df["r_pro_sig_str_acc_corrected"] + 0.01) / (df["r_pro_SLpM_corrected"] + 0.01)
        b_precision = (df["b_pro_sig_str_acc_corrected"] + 0.01) / (df["b_pro_SLpM_corrected"] + 0.01)
        df["precision_striking_diff"] = r_precision - b_precision

        # 8. Quality grappling (TD accuracy relative to attempts)
        # High TD acc with moderate attempts = efficient grappler
        r_quality_grappling = df["r_pro_td_acc_corrected"] * (df["r_pro_td_avg_corrected"] ** 0.5)
        b_quality_grappling = df["b_pro_td_acc_corrected"] * (df["b_pro_td_avg_corrected"] ** 0.5)
        df["quality_grappling_diff"] = r_quality_grappling - b_quality_grappling

        # 9. Submission threat ratio (sub attempts relative to grappling)
        r_sub_threat = (df["r_pro_sub_avg_corrected"] + 0.01) / (df["r_pro_td_avg_corrected"] + 0.01)
        b_sub_threat = (df["b_pro_sub_avg_corrected"] + 0.01) / (df["b_pro_td_avg_corrected"] + 0.01)
        df["submission_threat_ratio_diff"] = r_sub_threat - b_sub_threat

        # TIER 3: Defensive Efficiency
        # 10. Damage absorption relative to defense (lower is better)
        # High SApM despite high str_def = facing tough competition or getting hit clean
        r_absorption_efficiency = df["r_pro_SApM_corrected"] / (df["r_pro_str_def_corrected"] + 0.01)
        b_absorption_efficiency = df["b_pro_SApM_corrected"] / (df["b_pro_str_def_corrected"] + 0.01)
        df["damage_absorption_efficiency_diff"] = r_absorption_efficiency - b_absorption_efficiency

        # 11. Total defense index (striking + grappling defense combined)
        r_total_defense = (df["r_pro_str_def_corrected"] + df["r_pro_td_def_corrected"]) / 2
        b_total_defense = (df["b_pro_str_def_corrected"] + df["b_pro_td_def_corrected"]) / 2
        df["total_defense_index_diff"] = r_total_defense - b_total_defense

        # 12. Defense versatility (both striking AND grappling defense)
        r_def_versatility = (df["r_pro_str_def_corrected"] * df["r_pro_td_def_corrected"]) ** 0.5
        b_def_versatility = (df["b_pro_str_def_corrected"] * df["b_pro_td_def_corrected"]) ** 0.5
        df["defense_versatility_diff"] = r_def_versatility - b_def_versatility

        # TIER 4: Offensive Versatility
        # 13. Total offense index (striking + grappling output)
        # Normalized to similar scales (SLpM usually 2-6, td_avg usually 1-4)
        r_total_offense = df["r_pro_SLpM_corrected"] + (df["r_pro_td_avg_corrected"] * 1.5)
        b_total_offense = df["b_pro_SLpM_corrected"] + (df["b_pro_td_avg_corrected"] * 1.5)
        df["total_offense_index_diff"] = r_total_offense - b_total_offense

        # 14. Offensive versatility (both striking AND grappling threat)
        # Geometric mean rewards being good at both
        r_off_versatility = (df["r_pro_SLpM_corrected"] * df["r_pro_td_avg_corrected"]) ** 0.5
        b_off_versatility = (df["b_pro_SLpM_corrected"] * df["b_pro_td_avg_corrected"]) ** 0.5
        df["offensive_versatility_diff"] = r_off_versatility - b_off_versatility

        # 15. Strike-to-grapple ratio (striker vs grappler index)
        # >1 = striker, <1 = grappler
        r_striker_index = (df["r_pro_SLpM_corrected"] + 0.1) / (df["r_pro_td_avg_corrected"] + 0.1)
        b_striker_index = (df["b_pro_SLpM_corrected"] + 0.1) / (df["b_pro_td_avg_corrected"] + 0.1)
        df["striker_index_diff"] = r_striker_index - b_striker_index

        # TIER 5: Win Quality Metrics
        # 16. Win/loss ratio squared (magnifies gap between winners and losers)
        df["win_loss_ratio_squared_diff"] = (df["r_win_loss_ratio_corrected"] ** 2) - (df["b_win_loss_ratio_corrected"] ** 2)

        # 17. Experience quality (wins relative to total fights)
        r_experience_quality = df["r_wins_corrected"] / (df["r_wins_corrected"] + df["r_losses_corrected"] + 1)
        b_experience_quality = df["b_wins_corrected"] / (df["b_wins_corrected"] + df["b_losses_corrected"] + 1)
        df["experience_quality_diff"] = r_experience_quality - b_experience_quality

        # 18. Win efficiency (wins per year of career)
        # Assumes fighters with more wins in shorter time = active and successful
        r_win_efficiency = df["r_wins_corrected"] / (df["r_age_at_event"] - 18 + 1)  # Career length approximation
        b_win_efficiency = df["b_wins_corrected"] / (df["b_age_at_event"] - 18 + 1)
        df["win_efficiency_diff"] = r_win_efficiency - b_win_efficiency

        # TIER 6: Recent Form Ratios
        # 19. Recent form to career average ratio (recent performance vs career baseline)
        r_form_ratio = (df["r_recent_form_corrected"] + 5) / (r_experience_quality + 0.5)  # Center recent_form around 5
        b_form_ratio = (df["b_recent_form_corrected"] + 5) / (b_experience_quality + 0.5)
        df["recent_form_ratio_diff"] = r_form_ratio - b_form_ratio

        # 20. Momentum quality (win streak relative to total wins)
        r_momentum_quality = (df["r_win_streak_corrected"] + 1) / (df["r_wins_corrected"] + 1)
        b_momentum_quality = (df["b_win_streak_corrected"] + 1) / (df["b_wins_corrected"] + 1)
        df["momentum_quality_diff"] = r_momentum_quality - b_momentum_quality

        # TIER 7: Physical Efficiency Ratios
        # 21. Reach efficiency (reach relative to height)
        # Ape index already captures this, but ratio form emphasizes it
        r_reach_efficiency = (df["r_reach"] + 1) / (df["r_height"] + 1)
        b_reach_efficiency = (df["b_reach"] + 1) / (df["b_height"] + 1)
        df["reach_efficiency_diff"] = r_reach_efficiency - b_reach_efficiency

        # 22. Size-adjusted striking (SLpM relative to weight class)
        # Heavier fighters often strike less frequently
        # Normalized by dividing by weight (in lbs)
        r_size_adj_striking = df["r_pro_SLpM_corrected"] / ((df["r_weight"] / 100) + 0.01)
        b_size_adj_striking = df["b_pro_SLpM_corrected"] / ((df["b_weight"] / 100) + 0.01)
        df["size_adjusted_striking_diff"] = r_size_adj_striking - b_size_adj_striking

        # 23. Size-adjusted grappling (TD avg relative to weight)
        r_size_adj_grappling = df["r_pro_td_avg_corrected"] / ((df["r_weight"] / 100) + 0.01)
        b_size_adj_grappling = df["b_pro_td_avg_corrected"] / ((df["b_weight"] / 100) + 0.01)
        df["size_adjusted_grappling_diff"] = r_size_adj_grappling - b_size_adj_grappling

        # TIER 8: Advanced Composite Ratios
        # 24. Complete fighter index (offense × defense × finish rate)
        # Geometric mean of all aspects
        r_complete = ((df["r_pro_SLpM_corrected"] + 1) *
                      (df["r_pro_str_def_corrected"] + 0.1) *
                      (df["r_finish_rate"] + 0.1)) ** (1/3)
        b_complete = ((df["b_pro_SLpM_corrected"] + 1) *
                      (df["b_pro_str_def_corrected"] + 0.1) *
                      (df["b_finish_rate"] + 0.1)) ** (1/3)
        df["complete_fighter_index_diff"] = r_complete - b_complete

        # 25. Pressure fighter index (output / defense)
        # High offense, lower defense = aggressive pressure fighter
        r_pressure = (df["r_pro_SLpM_corrected"] + df["r_pro_td_avg_corrected"]) / (df["r_pro_str_def_corrected"] + 0.3)
        b_pressure = (df["b_pro_SLpM_corrected"] + df["b_pro_td_avg_corrected"]) / (df["b_pro_str_def_corrected"] + 0.3)
        df["pressure_fighter_index_diff"] = r_pressure - b_pressure

        # 26. Counter fighter index (defense / offense)
        # High defense, lower offense = counter-striker
        r_counter = (df["r_pro_str_def_corrected"] + 0.1) / (df["r_pro_SLpM_corrected"] + 1)
        b_counter = (df["b_pro_str_def_corrected"] + 0.1) / (df["b_pro_SLpM_corrected"] + 1)
        df["counter_fighter_index_diff"] = r_counter - b_counter

        # 27. Finishing threat composite (KO power + submission threat)
        # Uses finish rate and submission average
        r_finish_threat = (df["r_finish_rate"] + 0.1) * (df["r_pro_sub_avg_corrected"] + 0.1)
        b_finish_threat = (df["b_finish_rate"] + 0.1) * (df["b_pro_sub_avg_corrected"] + 0.1)
        df["finishing_threat_composite_diff"] = r_finish_threat - b_finish_threat

        # TIER 9: Rolling Performance Ratios
        # 28. Recent vs career performance (rolling vs career average)
        r_recent_career_ratio = (df["r_rolling_slpm_5_corrected"] + 0.1) / (df["r_pro_SLpM_corrected"] + 0.1)
        b_recent_career_ratio = (df["b_rolling_slpm_5_corrected"] + 0.1) / (df["b_pro_SLpM_corrected"] + 0.1)
        df["recent_vs_career_striking_diff"] = r_recent_career_ratio - b_recent_career_ratio

        # 29. Consistency ratio (inverse of variance)
        # Lower variance = more consistent
        r_consistency = 1 / (df["r_slpm_variance_5_corrected"] + 0.1)
        b_consistency = 1 / (df["b_slpm_variance_5_corrected"] + 0.1)
        df["striking_consistency_ratio_diff"] = r_consistency - b_consistency

        # 30. Improvement trajectory ratio (trend / career average)
        # Positive trend relative to career baseline = improving fighter
        r_improvement = (df["r_slpm_trend_5_corrected"]) / (df["r_pro_SLpM_corrected"] + 0.1)
        b_improvement = (df["b_slpm_trend_5_corrected"]) / (df["b_pro_SLpM_corrected"] + 0.1)
        df["improvement_trajectory_ratio_diff"] = r_improvement - b_improvement

        return df

    def prepare_core_features(self, df):
        """Prepare only the core 50-60 features"""
        print("\n" + "="*80)
        print("PREPARING CORE FEATURES")
        print("="*80 + "\n")

        # Filter to valid winners only
        if "winner" in df.columns:
            df = df[df["winner"].isin(["Red", "Blue"])].copy()

        # Create method mapping
        method_mapping = {
            "KO/TKO": "KO/TKO", "Submission": "Submission",
            "Decision - Unanimous": "Decision", "Decision - Split": "Decision",
            "Decision - Majority": "Decision", "TKO - Doctor's Stoppage": "KO/TKO",
            "Could Not Continue": "KO/TKO", "DQ": "Decision", "Overturned": "Decision",
        }

        if "method" in df.columns:
            df["method_simple"] = df["method"].map(method_mapping).fillna("Decision")
        else:
            df["method_simple"] = "Decision"

        if "winner" not in df.columns:
            df["winner"] = "Red"

        df["winner_method_simple"] = df["winner"] + "_" + df["method_simple"]

        # Build ALL engineered features from base r_*/b_* columns
        # build_features() is the single source of truth - it calculates ALL differentials
        # and compound features. This ensures corner swapping works correctly.
        df = self.build_features(df)

        # Verify all core features were created
        expected_features = set(self.get_core_feature_names())
        actual_features = set(df.columns)
        missing_features = expected_features - actual_features

        if missing_features:
            print(f"\n⚠️  WARNING: {len(missing_features)} features declared but not created:")
            for feat in sorted(missing_features):
                print(f"   - {feat}")

        print(f"Core features prepared: {len(self.get_core_feature_names())} features")

        return df

    # ===== HYPERPARAMETER OPTIMIZATION WITH OPTUNA =====

    def optimize_hyperparameters(self, X, y, n_trials=100):
        """Optimize hyperparameters using Optuna Bayesian optimization"""
        if not HAS_OPTUNA or not HAS_XGBOOST:
            print("Optuna or XGBoost not available, using default parameters")
            return {
                'n_estimators': 800,
                'max_depth': 7,
                'learning_rate': 0.02,
                'subsample': 0.8,
                'colsample_bytree': 0.8,
                'colsample_bynode': 0.8,
                'reg_alpha': 1.0,
                'reg_lambda': 1.0,
                'min_child_weight': 5,
                'gamma': 0.5,
            }

        print(f"\n{'='*80}")
        print(f"OPTIMIZING HYPERPARAMETERS WITH OPTUNA ({n_trials} trials)")
        print(f"{'='*80}")
        print()

        def objective(trial):
            params = {
                'n_estimators': trial.suggest_int('n_estimators', 300, 2000),  # Wider range for more trees
                'max_depth': trial.suggest_int('max_depth', 3, 12),  # Much deeper trees allowed
                'learning_rate': trial.suggest_float('learning_rate', 0.005, 0.1, log=True),  # Wider learning rate range
                'subsample': trial.suggest_float('subsample', 0.5, 1.0),  # Full range of row sampling
                'colsample_bytree': trial.suggest_float('colsample_bytree', 0.5, 1.0),  # Full range of column sampling
                'colsample_bynode': trial.suggest_float('colsample_bynode', 0.5, 1.0),  # Full range of node sampling
                'reg_alpha': trial.suggest_float('reg_alpha', 0.001, 50, log=True),  # Much wider L1 regularization
                'reg_lambda': trial.suggest_float('reg_lambda', 0.001, 50, log=True),  # Much wider L2 regularization
                'min_child_weight': trial.suggest_int('min_child_weight', 1, 20),  # Wider child weight range
                'gamma': trial.suggest_float('gamma', 0, 10),  # Wider gamma range for split control
                'random_state': 42,
            }

            # Use GPU if available - hyperparameters optimized on GPU work better for GPU training
            if GPU_AVAILABLE['xgboost']:
                params['device'] = 'cuda'
            else:
                params['device'] = 'cpu'
                params['n_jobs'] = -1

            model = XGBClassifier(**params)

            # Time series cross-validation
            cv_scores = []
            tscv = TimeSeriesSplit(n_splits=5)  # 5-fold CV for more robust hyperparameter selection

            for train_idx, val_idx in tscv.split(X):
                X_train_fold = X.iloc[train_idx] if hasattr(X, 'iloc') else X[train_idx]
                X_val_fold = X.iloc[val_idx] if hasattr(X, 'iloc') else X[val_idx]
                y_train_fold = y.iloc[train_idx] if hasattr(y, 'iloc') else y[train_idx]
                y_val_fold = y.iloc[val_idx] if hasattr(y, 'iloc') else y[val_idx]

                # Convert to numpy arrays to avoid dtype issues with XGBoost
                X_train_fold = np.array(X_train_fold) if hasattr(X_train_fold, 'values') else X_train_fold
                X_val_fold = np.array(X_val_fold) if hasattr(X_val_fold, 'values') else X_val_fold
                y_train_fold = np.array(y_train_fold) if hasattr(y_train_fold, 'values') else y_train_fold
                y_val_fold = np.array(y_val_fold) if hasattr(y_val_fold, 'values') else y_val_fold

                model.fit(X_train_fold, y_train_fold)
                score = model.score(X_val_fold, y_val_fold)
                cv_scores.append(score)

            return np.mean(cv_scores)

        # Create study with seeded sampler for deterministic results
        sampler = optuna.samplers.TPESampler(seed=42)
        study = optuna.create_study(direction='maximize', sampler=sampler)
        study.optimize(objective, n_trials=n_trials, show_progress_bar=True)

        print(f"\nBest CV Score: {study.best_value:.4f}")
        print(f"Best Parameters:")
        for key, value in study.best_params.items():
            print(f"  {key}: {value}")

        return study.best_params

    # ===== DYNAMIC ENSEMBLE WEIGHT OPTIMIZATION =====

    def optimize_ensemble_weights(self, estimators, X_train, y_train, X_val, y_val):
        """
        Optimize ensemble weights using grid search on validation set.
        Finds the best combination of weights for each model in the ensemble.
        """
        from itertools import product

        # Convert to numpy arrays
        X_train_np = np.array(X_train)
        y_train_np = np.array(y_train)
        X_val_np = np.array(X_val)
        y_val_np = np.array(y_val)

        # Train all models first
        trained_models = []
        print("Training individual models for weight optimization...")
        for name, model in estimators:
            print(f"  Training {name}...")
            model.fit(X_train_np, y_train_np)
            trained_models.append((name, model))

        # Get predictions from each model on validation set
        model_predictions = []
        for name, model in trained_models:
            if hasattr(model, 'predict_proba'):
                preds = model.predict_proba(X_val_np)[:, 1]
            else:
                preds = model.predict(X_val_np)
            model_predictions.append(preds)

        # Grid search for optimal weights
        # Test combinations of weights from 1-5 for each model
        n_models = len(estimators)
        weight_range = [1, 2, 3, 4, 5]

        best_score = 0
        best_weights = [1] * n_models

        # Generate all weight combinations
        weight_combinations = list(product(weight_range, repeat=n_models))

        print(f"Testing {len(weight_combinations)} weight combinations...")

        for weights in weight_combinations:
            # Compute weighted average of predictions
            weighted_preds = np.zeros(len(X_val_np))
            for i, preds in enumerate(model_predictions):
                weighted_preds += weights[i] * preds
            weighted_preds /= sum(weights)

            # Convert to binary predictions
            binary_preds = (weighted_preds >= 0.5).astype(int)
            score = accuracy_score(y_val_np, binary_preds)

            if score > best_score:
                best_score = score
                best_weights = list(weights)

        print(f"Best validation score with optimized weights: {best_score:.4f}")
        return best_weights

    # ===== RFECV FEATURE SELECTION =====

    def select_features_by_importance(self, X, y):
        """
        RFECV: Recursive Feature Elimination with Cross-Validation
        Automatically finds optimal number of features using cross-validation.
        Trusts RFECV to determine optimal feature count via CV (no artificial caps).
        """
        print("\n" + "="*80)
        print("RFECV FEATURE SELECTION")
        print("="*80 + "\n")

        # Remove constant or near-constant features first
        variances = X.var()
        variance_threshold = variances.quantile(0.01)
        high_variance_features = variances[variances > variance_threshold].index.tolist()
        X_filtered = X[high_variance_features]
        print(f"Preprocessing: Kept {len(high_variance_features)}/{len(X.columns)} features with sufficient variance")

        # Base estimator for RFECV - Use XGBoost for best feature selection accuracy
        if HAS_XGBOOST:
            # XGBoost provides most stable/reliable feature importance for selection
            xgb_params = {
                'n_estimators': 200,  # Increased for more reliable feature ranking (was 100)
                'max_depth': 7,
                'learning_rate': 0.05,
                'subsample': 0.8,
                'colsample_bytree': 0.8,
                'reg_alpha': 1.0,
                'reg_lambda': 1.0,
                'random_state': 42,
            }
            # Add GPU support if available
            if GPU_AVAILABLE['xgboost']:
                xgb_params['device'] = 'cuda'
            else:
                xgb_params['n_jobs'] = -1
            estimator = XGBClassifier(**xgb_params)
        elif HAS_LIGHTGBM:
            lgbm_params = {
                'n_estimators': 200,
                'max_depth': 7,
                'learning_rate': 0.05,
                'subsample': 0.8,
                'colsample_bytree': 0.8,
                'reg_alpha': 1.0,
                'reg_lambda': 1.0,
                'random_state': 42,
                'verbose': -1
            }
            # Add GPU support if available
            if GPU_AVAILABLE['lightgbm']:
                lgbm_params['device'] = 'gpu'
            else:
                lgbm_params['n_jobs'] = -1
            estimator = LGBMClassifier(**lgbm_params)
        else:
            estimator = RandomForestClassifier(
                n_estimators=200,
                max_depth=12,
                random_state=42,
                n_jobs=-1
            )

        # RFECV with TimeSeriesSplit
        min_features = 200  # Aggressive feature selection to reduce overfitting
        n_features = len(high_variance_features)  # Test all high-variance features
        step = 2  # Reduced from 5 to 2 for finer-grained feature selection

        print(f"\nRunning RFECV (testing {min_features}-{n_features} features with 3-fold CV)...")

        rfecv = RFECV(
            estimator=estimator,
            step=step,
            cv=TimeSeriesSplit(n_splits=3),  # 3-fold CV for faster feature selection
            scoring='accuracy',
            min_features_to_select=min_features,
            n_jobs=-1,
            verbose=0  # Suppress verbose output
        )

        # Fit RFECV with animated progress indicator
        import threading
        import time

        # Calculate estimated iterations (for user info)
        total_iterations = (n_features - min_features) // step + 1
        estimated_time = total_iterations * 3 * 1.3  # ~1.3 sec per fold (XGBoost 200 trees, ~209 features, 3-fold CV)
        print(f"Estimated iterations: {total_iterations} feature subsets × 3 folds = {total_iterations * 3} model fits")
        print(f"Estimated time: ~{estimated_time/60:.1f} minutes (XGBoost + 3-fold CV)\n")

        # Start timer
        start_time = time.time()

        # Animated progress indicator
        stop_progress = threading.Event()
        def show_progress():
            spinner = ['|', '/', '-', '\\']
            idx = 0
            dots = 0
            while not stop_progress.is_set():
                elapsed = int(time.time() - start_time)
                mins, secs = divmod(elapsed, 60)

                # Animated bar
                bar_length = 40
                filled = (dots % (bar_length + 1))
                if filled <= bar_length // 2:
                    bar = '=' * filled + '>' + '-' * (bar_length - filled - 1)
                else:
                    bar = '=' * (bar_length - (filled - bar_length // 2)) + '<' + '-' * (filled - bar_length // 2 - 1)

                sys.stdout.write('\r' + ' ' * 80 + '\r')
                sys.stdout.write(f"Running RFECV... [{bar}] {spinner[idx]} ({mins:02d}:{secs:02d})")
                sys.stdout.flush()

                idx = (idx + 1) % len(spinner)
                dots = (dots + 1) % (bar_length * 2)
                time.sleep(0.15)

        progress_thread = threading.Thread(target=show_progress, daemon=True)
        progress_thread.start()

        try:
            rfecv.fit(X_filtered, y)
        finally:
            stop_progress.set()
            progress_thread.join(timeout=0.5)
            elapsed = int(time.time() - start_time)
            mins, secs = divmod(elapsed, 60)
            sys.stdout.write('\r' + ' ' * 80 + '\r')
            bar = '=' * 40
            sys.stdout.write(f"RFECV Complete! [{bar}] ✓ ({mins:02d}:{secs:02d})\n\n")
            sys.stdout.flush()

        # Get RFECV results
        selected_mask = rfecv.support_
        selected_features = X_filtered.columns[selected_mask].tolist()
        n_selected = len(selected_features)
        best_score = rfecv.cv_results_['mean_test_score'].max()

        # Rank selected features by importance
        importances = rfecv.estimator_.feature_importances_
        importance_df = pd.DataFrame({
            'feature': selected_features,
            'importance': importances
        }).sort_values('importance', ascending=False)

        # Results summary
        print("="*80)
        print(f"Selected Features: {n_selected}")
        print(f"Best CV Score:     {best_score:.4f}")
        print("="*80)
        print(f"\nAll {n_selected} Selected Features (Ranked by Importance):")
        print(importance_df[['feature', 'importance']].to_string(index=False))

        # Store for later
        self.feature_importance = importance_df
        self.rfecv = rfecv

        return selected_features

    # ===== TRAINING WITH PROPER VALIDATION =====

    def train(self, df):
        """Train the model with all improvements

        Args:
            df: DataFrame with fight data (should already have data leakage fixed)
        """
        print("\n" + "="*80)
        print("IMPROVED UFC PREDICTOR - TRAINING")
        print("="*80)

        print(f"\nTraining on {len(df)} fights")

        # Prepare core features (builds all features from r_*/b_* bases)
        df = self.prepare_core_features(df)

        # =================================================================
        # ANTISYMMETRIZATION APPROACH (replaces fragile corner swapping)
        # =================================================================
        print(f"\n{'='*80}")
        print("ANTISYMMETRIZATION: Building directional features")
        print(f"{'='*80}")

        # Step 1: Build swapped version (swap corners + recompute features)
        print("\n1. Creating swap-mirror version...")
        df_swap = self.build_swapped(df)
        print(f"   ✓ Swapped version created ({len(df_swap)} fights)")

        # Step 2: Get core feature names
        core_features = self.get_core_feature_names()
        available_features = [f for f in core_features if f in df.columns]
        print(f"\n2. Using {len(available_features)} core features (of {len(core_features)} defined)")

        # Step 3: Antisymmetrize into directional and invariant features
        print("\n3. Antisymmetrizing features...")
        try:
            X_orig = df[available_features].copy().fillna(0)
            X_swap = df_swap[available_features].copy().fillna(0)
        except KeyError as e:
            print(f"\n❌ ERROR: Missing feature column: {e}")
            print(f"\nAvailable in df: {sorted([f for f in available_features if f in df.columns])}")
            print(f"\nMissing in df: {sorted([f for f in available_features if f not in df.columns])}")
            print(f"\nAvailable in df_swap: {sorted([f for f in available_features if f in df_swap.columns])}")
            print(f"\nMissing in df_swap: {sorted([f for f in available_features if f not in df_swap.columns])}")
            raise

        D, I_inv = self.directional_and_invariant(X_orig, X_swap)

        print(f"   ✓ Directional features: {len(D.columns)} (flip on swap)")
        print(f"   ✓ Invariant features: {len(I_inv.columns)} (stay same on swap)")

        # Step 4: Drop flat/constant features from both directional and invariant sets
        print("\n4. Removing flat features...")
        flat_directional = [c for c in D.columns if D[c].nunique(dropna=False) <= 1]
        flat_invariant = [c for c in I_inv.columns if I_inv[c].nunique(dropna=False) <= 1]

        if flat_directional:
            print(f"   ✓ Dropping {len(flat_directional)} flat directional features")
            D = D.drop(columns=flat_directional)
        else:
            print(f"   ✓ No flat directional features found")

        if flat_invariant:
            print(f"   ✓ Dropping {len(flat_invariant)} flat invariant features")
            I_inv = I_inv.drop(columns=flat_invariant)
        else:
            print(f"   ✓ No flat invariant features found")

        # Step 5: Validate invariant features have meaningful variance
        print("\n5. Analyzing invariant features...")
        inv_std = I_inv.std()
        meaningful_invariants = inv_std[inv_std > 0.01]  # Keep features with std > 0.01
        low_variance_invariants = inv_std[inv_std <= 0.01]

        print(f"   ✓ {len(meaningful_invariants)}/{len(I_inv.columns)} invariant features have meaningful variance (std > 0.01)")
        if len(low_variance_invariants) > 0:
            print(f"   ✓ Dropping {len(low_variance_invariants)} low-variance invariant features")
            I_inv = I_inv[meaningful_invariants.index]

        # Step 6: Combine directional AND invariant features
        print(f"\n6. Combining directional and invariant features...")
        X = pd.concat([D, I_inv], axis=1)
        y_winner = (df["winner"] == "Red").astype(int)

        print(f"   ✓ Directional features: {len(D.columns)}")
        print(f"   ✓ Invariant features: {len(I_inv.columns)}")
        print(f"   ✓ Total feature set: {len(X.columns)} features")

        # Diagnostic: Check PHASE 3 feature coverage
        phase3_features = [
            "avg_opponent_elo_l5_diff", "elo_momentum_vs_competition_diff",
            "performance_vs_ranked_opponents_diff", "performance_volatility_l10_diff",
            "finish_rate_acceleration_diff", "slpm_coefficient_of_variation_diff",
            "performance_decline_velocity_diff", "mileage_adjusted_age_diff",
            "prime_exit_risk_diff", "career_inflection_point_diff",
            "title_shot_proximity_score_diff", "tactical_evolution_score_diff",
            "finish_method_diversity_diff", "aging_power_striker_penalty_diff",
            "elo_volatility_interaction_diff", "layoff_veteran_interaction_diff",
            "bayesian_finish_rate_diff", "confidence_weighted_damage_ratio_diff",
            "distance_from_career_peak_diff", "elite_performance_frequency_l10_diff",
            "avg_opponent_elo_l5_diff_squared", "elo_momentum_vs_competition_diff_squared",
            "performance_decline_velocity_diff_squared", "mileage_adjusted_age_diff_squared",
            "layoff_veteran_interaction_diff_squared", "performance_volatility_l10_diff_squared",
            "finish_rate_acceleration_diff_squared"
        ]
        phase3_in_final = [f for f in phase3_features if f in X.columns]
        phase3_missing = [f for f in phase3_features if f not in X.columns]

        print(f"\n   PHASE 3 Feature Status:")
        print(f"   ✓ {len(phase3_in_final)}/27 PHASE 3 features included")
        if phase3_missing:
            print(f"   ⚠ {len(phase3_missing)} PHASE 3 features missing:")
            for f in phase3_missing:
                print(f"      - {f}")
        print(f"\n{'='*80}\n")

        # PROPER VALIDATION: Split into train/val/test
        print("\n" + "="*80)
        print("PROPER VALIDATION SPLIT")
        print("="*80)

        # Sort by date for temporal split
        df_sorted = df.sort_values("event_date").reset_index(drop=True)
        X = X.loc[df_sorted.index]
        y_winner = y_winner.loc[df_sorted.index]

        # 70% train, 15% validation, 15% test
        train_size = int(0.70 * len(df_sorted))
        val_size = int(0.85 * len(df_sorted))

        X_train = X.iloc[:train_size]
        X_val = X.iloc[train_size:val_size]
        X_test = X.iloc[val_size:]

        y_train = y_winner.iloc[:train_size]
        y_val = y_winner.iloc[train_size:val_size]
        y_test = y_winner.iloc[val_size:]

        print(f"\nTrain set: {len(X_train)} fights")
        print(f"Validation set: {len(X_val)} fights")
        print(f"Test set: {len(X_test)} fights (HELD OUT)")

        # =================================================================
        # SYMMETRY ENFORCEMENT: Augment training data with flipped examples
        # =================================================================
        # IMPORTANT: We do NOT use polarity learning because:
        # 1. It can encode biases from unbalanced original data
        # 2. Augmentation teaches the model the correct directionality
        # 3. Model learns from both positive AND negative feature values
        # =================================================================
        print(f"\n{'='*80}")
        print("SYMMETRY ENFORCEMENT: Data Augmentation")
        print(f"{'='*80}")
        print("\n1. Enforcing antisymmetry through data augmentation...")
        print("   Adding mirror examples with proper feature handling:")
        print("   - Directional features: flip sign (D → -D)")
        print("   - Invariant features: keep same (I → I)")
        print("   - Label: flip (y → 1-y)")

        # Identify directional vs invariant columns
        directional_cols = [c for c in X_train.columns if not c.endswith('_inv')]
        invariant_cols = [c for c in X_train.columns if c.endswith('_inv')]

        print(f"   ✓ {len(directional_cols)} directional features will be flipped")
        print(f"   ✓ {len(invariant_cols)} invariant features will stay same")

        # Create flipped versions: negate directional, keep invariant
        X_train_flipped = X_train.copy()
        X_train_flipped[directional_cols] = -X_train[directional_cols]
        # Invariant columns already copied, no need to modify
        y_train_flipped = 1 - y_train

        # Augment ONLY training set (NOT validation - validation must remain independent)
        X_train_aug = pd.concat([X_train, X_train_flipped], axis=0, ignore_index=True)
        y_train_aug = pd.concat([y_train, y_train_flipped], axis=0, ignore_index=True)

        # Shuffle the augmented training data to mix original and flipped examples
        train_indices = np.random.permutation(len(X_train_aug))
        X_train_aug = X_train_aug.iloc[train_indices].reset_index(drop=True)
        y_train_aug = y_train_aug.iloc[train_indices].reset_index(drop=True)

        # Replace training set with augmented version
        original_train_size = len(X_train)
        X_train = X_train_aug
        y_train = y_train_aug

        # Validation and test sets remain original (no augmentation)

        print("\n2. Augmentation complete:")
        print(f"   ✓ Train set: {original_train_size} → {len(X_train)} samples (2x augmentation)")
        print(f"   ✓ Val set: {len(X_val)} samples (no augmentation - independent evaluation)")
        print(f"   ✓ Test set: {len(X_test)} samples (no augmentation - held out)")
        print(f"   ✓ Class balance: {(y_train == 1).sum()}/{len(y_train)} Red wins ({(y_train==1).mean()*100:.1f}%)")
        print(f"{'='*80}\n")

        # Create empty polarity map (no polarity correction needed with augmentation)
        self.polarity_map = {}

        # Feature selection: Use ONLY training set to avoid data leakage
        # RFECV will find optimal number of features automatically via cross-validation
        selected_features = self.select_features_by_importance(X_train, y_train)

        # Check if all selected features exist in X_train
        missing_features = [f for f in selected_features if f not in X_train.columns]
        if missing_features:
            print(f"\n⚠️ WARNING: {len(missing_features)} selected features not in X_train!")
            print(f"  Missing: {missing_features[:5]}")

        # Check for duplicate columns in X_train
        duplicate_cols = X_train.columns[X_train.columns.duplicated()].tolist()
        if duplicate_cols:
            print(f"\n⚠️ WARNING: X_train has {len(duplicate_cols)} duplicate column names!")
            print(f"  Duplicates: {duplicate_cols}")

        # CRITICAL FIX: Remove duplicate columns from ALL dataframes FIRST
        # Keep only the first occurrence of each column
        X_train = X_train.loc[:, ~X_train.columns.duplicated()].copy()
        X_val = X_val.loc[:, ~X_val.columns.duplicated()].copy()
        X_test = X_test.loc[:, ~X_test.columns.duplicated()].copy()

        # Remove duplicates from selected_features list (preserves order)
        selected_features_unique = list(dict.fromkeys(selected_features))

        # Filter to only features that exist in all datasets
        valid_features = [f for f in selected_features_unique if f in X_train.columns and f in X_val.columns and f in X_test.columns]

        # Update datasets with valid features only
        X_train = X_train[valid_features].copy()
        X_val = X_val[valid_features].copy()
        X_test = X_test[valid_features].copy()

        # Update selected_features to the cleaned list
        selected_features = valid_features

        # Create train+val AFTER feature selection (for final training only)
        # Reset indices to avoid any index conflicts during concat
        X_train_reset = X_train.reset_index(drop=True)
        X_val_reset = X_val.reset_index(drop=True)
        y_train_reset = y_train.reset_index(drop=True)
        y_val_reset = y_val.reset_index(drop=True)

        # Augment validation set for final training (more data = better final model)
        # Note: validation was NOT augmented during evaluation, only for final training
        directional_cols_final = [c for c in X_val_reset.columns if not c.endswith('_inv')]
        X_val_flipped_final = X_val_reset.copy()
        X_val_flipped_final[directional_cols_final] = -X_val_reset[directional_cols_final]
        y_val_flipped_final = 1 - y_val_reset

        X_val_aug_final = pd.concat([X_val_reset, X_val_flipped_final], axis=0, ignore_index=True)
        y_val_aug_final = pd.concat([y_val_reset, y_val_flipped_final], axis=0, ignore_index=True)

        # Concatenate along rows (axis=0) - train is already augmented, val now augmented too
        X_train_val = pd.concat([X_train_reset, X_val_aug_final], axis=0, ignore_index=True)
        y_train_val = pd.concat([y_train_reset, y_val_aug_final], axis=0, ignore_index=True)

        # Final verification - ensure ONLY selected features exist
        X_train_val = X_train_val[selected_features].copy()

        available_features = selected_features  # Update for later use

        # Safety check - raise error if mismatch
        if X_train_val.shape[1] != len(selected_features):
            raise ValueError(f"Feature count mismatch! X_train_val has {X_train_val.shape[1]} columns but selected_features has {len(selected_features)} features")

        # Hyperparameter optimization: Use ONLY training set with time-series CV
        if HAS_OPTUNA and HAS_XGBOOST:
            self.best_params = self.optimize_hyperparameters(X_train, y_train, n_trials=25) ### OPTUNA TRIALS ###
        else:
            self.best_params = {
                'n_estimators': 800,
                'max_depth': 7,
                'learning_rate': 0.02,
                'subsample': 0.8,
                'colsample_bytree': 0.8,
                'colsample_bynode': 0.8,
                'reg_alpha': 1.0,
                'reg_lambda': 1.0,
                'min_child_weight': 5,
                'gamma': 0.5,
            }

        # Train final model
        print("\n" + "="*80)
        print("TRAINING MODEL")
        print("="*80 + "\n")

        if HAS_XGBOOST:
            # Use natural class distribution (augmentation handles bias correction)
            scale_weight = (y_train_val == 0).sum() / (y_train_val == 1).sum()

            print(f"Class balancing: scale_pos_weight={scale_weight:.3f} (no adjustment - augmentation balances bias)")

            # Create XGBoost model with GPU support if available
            xgb_params = {
                **self.best_params,
                'scale_pos_weight': scale_weight,
                'random_state': 42,
            }
            # Add GPU support if available
            if GPU_AVAILABLE['xgboost']:
                xgb_params['device'] = 'cuda'
            else:
                xgb_params['device'] = 'cpu'
                xgb_params['n_jobs'] = -1

            xgb_model = XGBClassifier(**xgb_params)

            # ========== PHASE 3: ADVANCED ENSEMBLE WITH DIVERSE MODELS ==========
            print("\n" + "="*80)
            print("PHASE 3: BUILDING ADVANCED ENSEMBLE")
            print("="*80 + "\n")

            estimators = [('xgb', xgb_model)]

            # Add LightGBM if available (different algorithm, often complementary to XGBoost)
            if HAS_LIGHTGBM:
                lgbm_params = {
                    'n_estimators': self.best_params.get('n_estimators', 800),
                    'max_depth': self.best_params.get('max_depth', 7),
                    'learning_rate': self.best_params.get('learning_rate', 0.02),
                    'subsample': self.best_params.get('subsample', 0.8),
                    'colsample_bytree': self.best_params.get('colsample_bytree', 0.8),
                    'reg_alpha': self.best_params.get('reg_alpha', 1.0),
                    'reg_lambda': self.best_params.get('reg_lambda', 1.0),
                    'min_child_samples': self.best_params.get('min_child_weight', 5) * 2,
                    'scale_pos_weight': scale_weight,
                    'random_state': 42,
                    'verbose': -1
                }
                # Add GPU support if available
                if GPU_AVAILABLE['lightgbm']:
                    lgbm_params['device'] = 'gpu'
                else:
                    lgbm_params['n_jobs'] = -1

                lgbm_model = LGBMClassifier(**lgbm_params)
                estimators.append(('lgbm', lgbm_model))

            # Add CatBoost if available (handles categorical features differently, robust to overfitting)
            if HAS_CATBOOST:
                catboost_params = {
                    'iterations': 1000,  # Increased for better convergence and higher weight in ensemble
                    'depth': self.best_params.get('max_depth', 7),
                    'learning_rate': self.best_params.get('learning_rate', 0.02),
                    'bootstrap_type': 'Bernoulli',  # Required for subsample parameter
                    'subsample': self.best_params.get('subsample', 0.8),
                    'l2_leaf_reg': self.best_params.get('reg_lambda', 1.0),
                    'scale_pos_weight': scale_weight,
                    'random_state': 42,
                    'verbose': 0,
                    # Removed aggressive speed optimizations to improve accuracy
                    # border_count default (254) for better split quality
                    # rsm default (1.0) to use all features
                }
                # Add GPU support if available
                if GPU_AVAILABLE['catboost']:
                    catboost_params['task_type'] = 'GPU'
                    catboost_params['devices'] = '0'
                else:
                    catboost_params['thread_count'] = -1

                catboost_model = CatBoostClassifier(**catboost_params)
                estimators.append(('catboost', catboost_model))

            # Add RandomForest for diversity (tree-based but different approach)
            rf_model = RandomForestClassifier(
                n_estimators=600,
                max_depth=12,
                min_samples_split=5,
                min_samples_leaf=2,
                max_features='sqrt',
                class_weight={0: 1.0, 1: scale_weight},
                random_state=42,
                n_jobs=-1
            )
            estimators.append(('rf', rf_model))

            # Add Logistic Regression (linear model for algorithmic diversity)
            # Wrap with StandardScaler since LogReg requires scaled features
            logreg_model = LogisticRegression(
                C=0.1,  # Regularization strength (inverse) - lower = stronger regularization
                penalty='l2',  # Ridge regularization to prevent overfitting
                solver='lbfgs',  # Efficient solver for L2 penalty
                max_iter=1000,
                class_weight={0: 1.0, 1: scale_weight},
                random_state=42,
                n_jobs=-1
            )

            logreg_pipeline = Pipeline([
                ('scaler', StandardScaler()),  # Scale features for Logistic Regression
                ('logreg', logreg_model)
            ])
            estimators.append(('logreg', logreg_pipeline))

            # ========== CHOOSE ENSEMBLE STRATEGY ==========
            # SPEED CONTROL: Set to False for faster training with VotingClassifier
            ENABLE_STACKING = False  # Stacking is slower but often more accurate
            USE_STACKING = ENABLE_STACKING and HAS_STACKING and len(estimators) >= 3

            if USE_STACKING:
                print("\n" + "="*80)
                print("CREATING STACKING ENSEMBLE WITH META-LEARNER")
                print("="*80 + "\n")

                # Stacking ensemble: Base models + meta-learner
                # Meta-learner learns how to best combine base model predictions
                meta_learner = LogisticRegression(
                    C=1.0,              # Regularization strength (inverse) - lower = stronger regularization
                    penalty='l2',       # Ridge regularization
                    max_iter=1000,
                    random_state=42,
                    n_jobs=-1,
                    solver='lbfgs'
                )

                base_model = StackingClassifier(
                    estimators=estimators,
                    final_estimator=meta_learner,
                    cv=3,  # 3-fold CV for generating meta-features (faster than 5-fold)
                    n_jobs=-1,
                    passthrough=False  # Don't pass original features to meta-learner
                )
                print(f"✓ Stacking ensemble created with {len(estimators)} base models + LogisticRegression meta-learner")
                print(f"  Base models: {[name for name, _ in estimators]}")
            else:
                # ========== DYNAMIC WEIGHT OPTIMIZATION (for Voting Ensemble only) ==========
                print(f"\nOptimizing ensemble weights for {len(estimators)} models...")

                # Use validation set to find optimal weights
                optimal_weights = self.optimize_ensemble_weights(
                    estimators, X_train, y_train, X_val, y_val
                )

                print(f"Optimal weights: {dict(zip([name for name, _ in estimators], optimal_weights))}")

                # Voting ensemble with optimized weights
                base_model = VotingClassifier(
                    estimators=estimators,
                    voting='soft',
                    weights=optimal_weights,
                    n_jobs=-1
                )
                print(f"\n✓ Voting ensemble created with {len(estimators)} models: {[name for name, _ in estimators]}")
        else:
            # Fallback: RandomForest
            base_model = RandomForestClassifier(
                n_estimators=600,
                max_depth=10,
                min_samples_split=10,
                random_state=42,
                n_jobs=-1
            )

        # Train final model on train+val
        print("Training final ensemble on train+val...")
        # Convert to numpy to avoid dtype issues
        X_train_val_np = np.array(X_train_val)
        y_train_val_np = np.array(y_train_val)
        base_model.fit(X_train_val_np, y_train_val_np)

        # Calibrate probabilities on train+val
        print("\n" + "="*80)
        print("CALIBRATING PROBABILITIES")
        print("="*80 + "\n")

        self.calibrated_model = CalibratedClassifierCV(
            base_model,
            method='isotonic',
            cv=3  # 3-fold CV (faster than 5-fold)
        )
        self.calibrated_model.fit(X_train_val_np, y_train_val_np)

        self.winner_model = self.calibrated_model

        # DEBUG: Verify feature negation consistency
        directional_by_suffix = [c for c in X_test.columns if not c.endswith('_inv')]
        directional_by_D = [c for c in X_test.columns if c in D.columns]

        print("\n=== FEATURE NEGATION DEBUG ===")
        print(f"Directional by suffix: {len(directional_by_suffix)}")
        print(f"Directional by D.columns: {len(directional_by_D)}")

        if set(directional_by_suffix) != set(directional_by_D):
            print("⚠️ MISMATCH DETECTED!")
            only_suffix = set(directional_by_suffix) - set(directional_by_D)
            only_D = set(directional_by_D) - set(directional_by_suffix)
            if only_suffix:
                print(f"  In suffix but not D: {only_suffix}")
            if only_D:
                print(f"  In D but not suffix: {only_D}")
        else:
            print("✅ Both methods identify same directional features")

        # ========== RED CORNER BIAS DIAGNOSTIC ==========
        print("\n" + "="*80)
        print("RED CORNER BIAS DIAGNOSTIC")
        print("="*80)

        # Test on balanced synthetic data
        n_test = min(1000, len(X_test))  # Use up to 1000 samples
        X_test_balanced = X_test.iloc[:n_test].copy()
        X_test_flipped = X_test_balanced.copy()

        # Flip directional features (match training method: use _inv suffix)
        directional_cols = [c for c in X_test_balanced.columns if not c.endswith('_inv')]
        X_test_flipped[directional_cols] = -X_test_balanced[directional_cols]

        # Predict on original vs flipped
        pred_orig = self.calibrated_model.predict_proba(X_test_balanced)[:, 1]
        pred_flip = self.calibrated_model.predict_proba(X_test_flipped)[:, 1]

        # Check parity: pred_orig + pred_flip should = 1.0
        parity_error = np.abs((pred_orig + pred_flip) - 1.0)
        mean_parity_error = parity_error.mean()
        max_parity_error = parity_error.max()

        # Measure inherent red bias
        mean_pred_orig = pred_orig.mean()
        mean_pred_flip = pred_flip.mean()
        inherent_red_bias = mean_pred_orig - (1 - mean_pred_flip)

        print("Parity Test (should be ~0):")
        print(f"  Mean error: {mean_parity_error:.4f}")
        print(f"  Max error:  {max_parity_error:.4f}")
        print("\nInherent Red Bias:")
        print(f"  Original predictions avg: {mean_pred_orig:.4f} (should be ~0.50)")
        print(f"  Flipped predictions avg:  {mean_pred_flip:.4f} (should be ~0.50)")
        print(f"  Red bias: {inherent_red_bias:+.4f} ({inherent_red_bias*100:+.2f}%)")
        print("\nInterpretation:")
        if abs(inherent_red_bias) < 0.02:
            print("  ✅ GOOD: Bias < 2% - Legitimate (real UFC advantage)")
        elif abs(inherent_red_bias) < 0.05:
            print("  ⚠️  BORDERLINE: Bias 2-5% - Check calibration")
        else:
            print("  ❌ BAD: Bias > 5% - Model learning corner position")
        print("="*80 + "\n")

        # Evaluate on test set (FIRST TIME)
        print("="*80)
        print("FINAL MODEL PERFORMANCE")
        print("="*80)

        # Convert test/train/val X sets to numpy for predictions
        X_test_np = np.array(X_test)
        X_train_np = np.array(X_train)
        X_val_np = np.array(X_val)

        # Get predictions and probabilities
        y_pred_test = self.winner_model.predict(X_test_np)
        y_proba_test = self.winner_model.predict_proba(X_test_np)[:, 1]

        # Calculate all metrics
        train_acc = self.winner_model.score(X_train_np, y_train)
        val_acc = self.winner_model.score(X_val_np, y_val)
        test_acc = accuracy_score(y_test, y_pred_test)
        test_roc_auc = roc_auc_score(y_test, y_proba_test)

        print(f"\nTrain Accuracy:      {train_acc:.4f}")
        print(f"Validation Accuracy: {val_acc:.4f}")
        print(f"Test Accuracy:       {test_acc:.4f} *** HELD OUT ***")
        print(f"\nROC-AUC Score:       {test_roc_auc:.4f} (probability separation)")

        print("\nDetailed Metrics by Corner:")
        print(classification_report(y_test, y_pred_test,
                                   target_names=['Blue Corner', 'Red Corner'],
                                   digits=4))

        # Check overfitting
        train_val_gap = train_acc - val_acc
        if train_val_gap > 0.05:
            print(f"\nWarning: Large train-validation gap ({train_val_gap:.4f}) suggests possible overfitting")
        else:
            print(f"\nTrain-validation gap ({train_val_gap:.4f}) is acceptable")

        # Time series cross-validation for robustness check
        print("\n" + "="*80)
        print("TIME SERIES CROSS-VALIDATION (Robustness Check)")
        print("="*80 + "\n")

        tscv = TimeSeriesSplit(n_splits=5)  # 5-fold CV for more robust evaluation
        cv_scores = []

        for fold, (train_idx, val_idx) in enumerate(tscv.split(X_train_val)):
            X_fold_train = X_train_val.iloc[train_idx]
            X_fold_val = X_train_val.iloc[val_idx]
            y_fold_train = y_train_val.iloc[train_idx]
            y_fold_val = y_train_val.iloc[val_idx]

            # Convert to numpy arrays to avoid dtype issues
            X_fold_train = np.array(X_fold_train)
            X_fold_val = np.array(X_fold_val)

            # Create fold model with GPU support if available
            if HAS_XGBOOST:
                fold_params = {**self.best_params, 'random_state': 42}
                if GPU_AVAILABLE['xgboost']:
                    fold_params['device'] = 'cuda'
                else:
                    fold_params['device'] = 'cpu'
                    fold_params['n_jobs'] = -1
                fold_model = XGBClassifier(**fold_params)
            else:
                fold_model = RandomForestClassifier(n_estimators=600, random_state=42, n_jobs=-1)

            fold_model.fit(X_fold_train, y_fold_train)
            score = fold_model.score(X_fold_val, y_fold_val)
            cv_scores.append(score)
            print(f"  Fold {fold + 1}/5: {score:.4f}")

        print(f"\nCV Mean: {np.mean(cv_scores):.4f} ± {np.std(cv_scores):.4f}")

        # Store training data and feature columns for predictions
        self.df_train = df
        self.feature_columns = selected_features  # Store SELECTED features (post-RFECV)

        # Note: Parity (predict(fight) + predict(swapped) = 1.0) is ENFORCED at prediction time
        # via averaging: red_proba = (p_red_orig + (1 - p_red_flipped)) / 2
        # This is guaranteed regardless of whether the model learns perfect antisymmetry.

        print("\n" + "="*80)
        print("TRAINING COMPLETE")
        print("="*80)
        print(f"\nWinner Model: {test_acc:.4f} accuracy")
        print(f"Features Used: {len(selected_features)}")
        print("Training Method: Antisymmetrization + Data Augmentation (2x)")

        return selected_features

    # ===== PREDICTION METHODS =====

    def get_fighter_latest_stats(self, fighter_name):
        """Get latest stats for fighter from training data"""
        red_fights = self.df_train[
            self.df_train["r_fighter"] == fighter_name
        ].sort_values("event_date", ascending=False)
        blue_fights = self.df_train[
            self.df_train["b_fighter"] == fighter_name
        ].sort_values("event_date", ascending=False)

        if len(red_fights) > 0 and len(blue_fights) > 0:
            red_latest = red_fights.iloc[0]
            blue_latest = blue_fights.iloc[0]
            latest = (
                red_latest
                if red_latest["event_date"] > blue_latest["event_date"]
                else blue_latest
            )
            prefix = (
                "r" if red_latest["event_date"] > blue_latest["event_date"] else "b"
            )
        elif len(red_fights) > 0:
            latest, prefix = red_fights.iloc[0], "r"
        elif len(blue_fights) > 0:
            latest, prefix = blue_fights.iloc[0], "b"
        else:
            return None

        # Get stats - dynamically extract ALL _corrected columns to avoid missing any
        stats = {}

        # Extract all _corrected columns for this fighter
        for col in latest.index:
            if col.startswith(f"{prefix}_") and col.endswith("_corrected"):
                # Remove prefix and _corrected suffix to get stat name
                stat_name = col[len(prefix)+1:-len("_corrected")]
                stats[stat_name] = latest[col]

        # Physical stats (not corrected)
        for stat in ["height", "reach", "weight", "age_at_event", "ape_index", "stance"]:
            col = f"{prefix}_{stat}"
            if col in latest.index:
                stats[stat] = latest[col]
            else:
                stats[stat] = 0 if stat != "stance" else "Orthodox"

        return stats

    def prepare_upcoming_fight(self, fight, feature_columns):
        """Prepare features for an upcoming fight"""
        r_stats = self.get_fighter_latest_stats(fight["red_fighter"])
        b_stats = self.get_fighter_latest_stats(fight["blue_fighter"])

        if not r_stats or not b_stats:
            return None, None, None

        r_total_fights = r_stats["wins"] + r_stats["losses"]
        b_total_fights = b_stats["wins"] + b_stats["losses"]

        fight_features = {
            "r_fighter": fight["red_fighter"],
            "b_fighter": fight["blue_fighter"],
            "total_rounds": fight["total_rounds"],
            "is_title_bout": 1 if fight["total_rounds"] == 5 else 0,
            "r_total_fights": r_total_fights,
            "b_total_fights": b_total_fights,
            "weight_class": fight.get("weight_class", "Unknown"),
        }

        # Physical attributes - include BOTH individual values AND differentials
        # Individual values are needed by build_features() for feature engineering
        for stat in ["height", "reach", "weight", "age_at_event", "ape_index"]:
            fight_features[f"r_{stat}"] = r_stats.get(stat, 0)
            fight_features[f"b_{stat}"] = b_stats.get(stat, 0)
            fight_features[f"{stat}_diff"] = r_stats.get(stat, 0) - b_stats.get(stat, 0)

        # Add ALL _corrected stats dynamically for both red and blue fighters
        # This ensures no KeyErrors when build_features() references any stat
        for prefix, stats in [("r", r_stats), ("b", b_stats)]:
            for stat_name, stat_value in stats.items():
                # Skip physical stats (they're added separately without _corrected suffix)
                if stat_name not in ["height", "reach", "weight", "age_at_event", "ape_index", "stance"]:
                    fight_features[f"{prefix}_{stat_name}_corrected"] = stat_value

        # Compute differentials for common stats (used by some features)
        for stat in r_stats.keys():
            if stat not in ["height", "reach", "weight", "age_at_event", "ape_index", "stance"]:
                if stat in b_stats:
                    fight_features[f"{stat}_diff_corrected"] = r_stats.get(stat, 0) - b_stats.get(stat, 0)

        # Core derived features
        fight_features["net_striking_advantage"] = (
            (r_stats.get("pro_SLpM", 0) - r_stats.get("pro_SApM", 0)) -
            (b_stats.get("pro_SLpM", 0) - b_stats.get("pro_SApM", 0))
        )
        fight_features["striking_efficiency"] = (
            (r_stats.get("pro_SLpM", 0) * r_stats.get("pro_sig_str_acc", 0)) -
            (b_stats.get("pro_SLpM", 0) * b_stats.get("pro_sig_str_acc", 0))
        )
        fight_features["defensive_striking"] = (
            (r_stats.get("pro_str_def", 0) - r_stats.get("pro_SApM", 0)) -
            (b_stats.get("pro_str_def", 0) - b_stats.get("pro_SApM", 0))
        )
        fight_features["grappling_control"] = (
            (r_stats.get("pro_td_avg", 0) * r_stats.get("pro_td_acc", 0)) -
            (b_stats.get("pro_td_avg", 0) * b_stats.get("pro_td_acc", 0))
        )

        fight_features["experience_gap"] = r_total_fights - b_total_fights

        r_finish_rate = (r_stats.get("ko_rate", 0) + r_stats.get("sub_rate", 0)) / 2
        b_finish_rate = (b_stats.get("ko_rate", 0) + b_stats.get("sub_rate", 0)) / 2
        fight_features["finish_rate_diff"] = r_finish_rate - b_finish_rate
        fight_features["r_finish_rate"] = r_finish_rate
        fight_features["b_finish_rate"] = b_finish_rate

        # Style features
        r_striker_score = r_stats.get("pro_SLpM", 0) - r_stats.get("pro_td_avg", 0)
        b_striker_score = b_stats.get("pro_SLpM", 0) - b_stats.get("pro_td_avg", 0)
        fight_features["striker_advantage"] = r_striker_score - b_striker_score
        fight_features["r_striker_score"] = r_striker_score
        fight_features["b_striker_score"] = b_striker_score

        r_grappler_score = r_stats.get("pro_td_avg", 0) + r_stats.get("pro_sub_avg", 0)
        b_grappler_score = b_stats.get("pro_td_avg", 0) + b_stats.get("pro_sub_avg", 0)
        fight_features["grappler_advantage"] = r_grappler_score - b_grappler_score
        fight_features["r_grappler_score"] = r_grappler_score
        fight_features["b_grappler_score"] = b_grappler_score

        # Style clash metric (same formula as training)
        fight_features["striker_vs_grappler"] = (r_striker_score * b_grappler_score) - (b_striker_score * r_grappler_score)

        # Stance matchup features (4 features)
        r_stance = r_stats.get("stance", "Orthodox")
        b_stance = b_stats.get("stance", "Orthodox")

        # Handle missing/NaN stances
        if pd.isna(r_stance) or r_stance == "" or r_stance is None:
            r_stance = "Orthodox"
        if pd.isna(b_stance) or b_stance == "" or b_stance is None:
            b_stance = "Orthodox"

        # Feature 1: Orthodox vs Southpaw (directional)
        if r_stance == "Orthodox" and b_stance == "Southpaw":
            fight_features["orthodox_vs_southpaw_advantage"] = 1
        elif r_stance == "Southpaw" and b_stance == "Orthodox":
            fight_features["orthodox_vs_southpaw_advantage"] = -1
        else:
            fight_features["orthodox_vs_southpaw_advantage"] = 0

        # Feature 2: Orthodox vs Switch (directional)
        if r_stance == "Orthodox" and b_stance == "Switch":
            fight_features["orthodox_vs_switch_advantage"] = 1
        elif r_stance == "Switch" and b_stance == "Orthodox":
            fight_features["orthodox_vs_switch_advantage"] = -1
        else:
            fight_features["orthodox_vs_switch_advantage"] = 0

        # Feature 3: Southpaw vs Switch (directional)
        if r_stance == "Southpaw" and b_stance == "Switch":
            fight_features["southpaw_vs_switch_advantage"] = 1
        elif r_stance == "Switch" and b_stance == "Southpaw":
            fight_features["southpaw_vs_switch_advantage"] = -1
        else:
            fight_features["southpaw_vs_switch_advantage"] = 0

        # Feature 4: Mirror matchup (symmetric)
        if r_stance == b_stance:
            fight_features["mirror_matchup"] = 1
        else:
            fight_features["mirror_matchup"] = 0

        fight_features["r_trajectory_3"] = 0
        fight_features["b_trajectory_3"] = 0
        fight_features["ring_rust_factor"] = 0
        fight_features["weight_class_factor"] = 1.0

        # Momentum features
        fight_features["momentum_swing"] = (
            r_stats.get("recent_form", 0) - b_stats.get("recent_form", 0) +
            (r_stats.get("win_streak", 0) - b_stats.get("win_streak", 0)) * 0.1
        )
        fight_features["style_clash_severity"] = abs(
            fight_features["striker_advantage"] * 0.5 + fight_features["grappler_advantage"] * 0.5
        )
        fight_features["power_vs_technique"] = (
            (r_stats.get("pro_SLpM", 0) - b_stats.get("pro_SLpM", 0)) * 0.6 +
            (r_stats.get("pro_sig_str_acc", 0) - b_stats.get("pro_sig_str_acc", 0)) * 0.4
        )
        fight_features["finish_pressure"] = (
            (r_stats.get("ko_rate", 0) - b_stats.get("ko_rate", 0)) * 0.5 +
            (r_stats.get("sub_rate", 0) - b_stats.get("sub_rate", 0)) * 0.3
        )
        fight_features["upset_potential"] = (
            (b_stats.get("recent_form", 0) - r_stats.get("recent_form", 0)) * 0.4 +
            (b_stats.get("win_streak", 0) - r_stats.get("win_streak", 0)) * 0.3
        )

        # Advanced features
        fight_features["offensive_output"] = (
            (r_stats.get("pro_SLpM", 0) + r_stats.get("pro_td_avg", 0) + r_stats.get("pro_sub_avg", 0)) -
            (b_stats.get("pro_SLpM", 0) + b_stats.get("pro_td_avg", 0) + b_stats.get("pro_sub_avg", 0))
        )
        fight_features["defensive_composite"] = (
            ((r_stats.get("pro_str_def", 0) + r_stats.get("pro_td_def", 0)) / 2) -
            ((b_stats.get("pro_str_def", 0) + b_stats.get("pro_td_def", 0)) / 2)
        )
        fight_features["ko_specialist_gap"] = (
            (r_stats.get("ko_rate", 0) * r_stats.get("pro_SLpM", 0)) -
            (b_stats.get("ko_rate", 0) * b_stats.get("pro_SLpM", 0))
        )
        fight_features["submission_specialist_gap"] = (
            (r_stats.get("sub_rate", 0) * r_stats.get("pro_sub_avg", 0)) -
            (b_stats.get("sub_rate", 0) * b_stats.get("pro_sub_avg", 0))
        )
        fight_features["skill_momentum"] = (
            fight_features["pro_SLpM_diff_corrected"] * fight_features["recent_form_diff_corrected"]
        )

        # Interaction features (will be calculated after ELO is added)
        fight_features["elo_x_form"] = 0  # Placeholder, calculated after ELO
        fight_features["reach_x_striking"] = fight_features["reach_diff"] * fight_features["pro_SLpM_diff_corrected"]
        fight_features["experience_x_age"] = fight_features["experience_gap"] * fight_features["age_at_event_diff"]

        # Advanced interaction features (data-driven from importance analysis)
        fight_features["age_x_striking"] = fight_features["age_at_event_diff"] * fight_features["pro_SLpM_diff_corrected"]
        fight_features["age_x_grappling"] = fight_features["age_at_event_diff"] * fight_features["pro_td_avg_diff_corrected"]
        fight_features["age_x_durability"] = fight_features["age_at_event_diff"] * fight_features["durability_diff_corrected"]
        fight_features["td_x_defense"] = fight_features["pro_td_avg_diff_corrected"] * fight_features["pro_td_def_diff_corrected"]
        fight_features["grappling_x_experience"] = fight_features["grappler_advantage"] * fight_features["experience_gap"]
        fight_features["striking_x_accuracy"] = (
            (r_stats.get("pro_SLpM", 0) - r_stats.get("pro_SApM", 0) - (b_stats.get("pro_SLpM", 0) - b_stats.get("pro_SApM", 0))) *
            fight_features["pro_sig_str_acc_diff_corrected"]
        )
        fight_features["striking_x_defense"] = fight_features["pro_SLpM_diff_corrected"] * fight_features["pro_str_def_diff_corrected"]
        fight_features["form_x_experience"] = fight_features["recent_form_diff_corrected"] * fight_features["experience_gap"]
        fight_features["finish_x_momentum"] = fight_features["finish_rate_diff"] * fight_features["recent_form_diff_corrected"]
        fight_features["height_x_reach"] = fight_features["height_diff"] * fight_features["reach_diff"]
        fight_features["physical_x_striking"] = (fight_features["height_diff"] + fight_features["reach_diff"]) * fight_features["pro_SLpM_diff_corrected"]

        # Elite interaction features (placeholders, some calculated after ELO)
        fight_features["elo_x_win_ratio"] = 0  # Calculated after ELO
        fight_features["win_ratio_x_form"] = fight_features["win_loss_ratio_diff_corrected"] * fight_features["recent_form_diff_corrected"]
        fight_features["durability_x_striking"] = fight_features["durability_diff_corrected"] * (
            (r_stats.get("pro_SLpM", 0) - r_stats.get("pro_SApM", 0)) -
            (b_stats.get("pro_SLpM", 0) - b_stats.get("pro_SApM", 0))
        )
        fight_features["elo_x_durability"] = 0  # Calculated after ELO
        fight_features["submission_x_grappling"] = fight_features["submission_specialist_gap"] * fight_features["grappler_advantage"]
        fight_features["ko_power_x_striking"] = fight_features["ko_rate_diff_corrected"] * (
            (r_stats.get("pro_SLpM", 0) - r_stats.get("pro_SApM", 0)) -
            (b_stats.get("pro_SLpM", 0) - b_stats.get("pro_SApM", 0))
        )
        fight_features["momentum_x_win_streak"] = fight_features["momentum_swing"] * fight_features["win_streak_diff_corrected"]
        fight_features["streak_differential"] = fight_features["win_streak_diff_corrected"] * fight_features["loss_streak_diff_corrected"]

        fight_features["opponent_quality_diff"] = 0
        fight_features["r_clutch_factor"] = r_stats.get("win_loss_ratio", 0) * r_stats.get("recent_form", 0)
        fight_features["b_clutch_factor"] = b_stats.get("win_loss_ratio", 0) * b_stats.get("recent_form", 0)
        fight_features["clutch_factor_diff"] = fight_features["r_clutch_factor"] - fight_features["b_clutch_factor"]
        fight_features["momentum_quality_diff"] = 0
        fight_features["pressure_performance_diff"] = 0
        fight_features["form_consistency_diff"] = 0
        fight_features["h2h_advantage"] = 0
        fight_features["last_5_wins_diff_corrected"] = 0

        # Add ELO features
        if hasattr(self, 'fighter_elos'):
            r_elo = self.fighter_elos.get(fight["red_fighter"], 1500)
            b_elo = self.fighter_elos.get(fight["blue_fighter"], 1500)

            fight_features["r_elo_pre_fight"] = r_elo
            fight_features["b_elo_pre_fight"] = b_elo
            fight_features["elo_diff"] = r_elo - b_elo
        else:
            # Default ELO if not available
            fight_features["r_elo_pre_fight"] = 1500
            fight_features["b_elo_pre_fight"] = 1500
            fight_features["elo_diff"] = 0

        # Calculate elo_x_form and other ELO-dependent features now that ELO is available
        fight_features["elo_x_form"] = fight_features["elo_diff"] * fight_features["recent_form_diff_corrected"]
        fight_features["elo_x_win_ratio"] = fight_features["elo_diff"] * fight_features["win_loss_ratio_diff_corrected"]
        fight_features["elo_x_durability"] = fight_features["elo_diff"] * fight_features["durability_diff_corrected"]

        # NEW: Strategic interaction features
        fight_features["age_x_win_streak"] = fight_features.get("age_at_event_diff", 0) * fight_features.get("win_streak_diff_corrected", 0)
        fight_features["elo_x_sub_threat"] = fight_features["elo_diff"] * fight_features.get("submission_specialist_gap", 0)
        fight_features["form_x_durability"] = fight_features["recent_form_diff_corrected"] * fight_features.get("durability_diff_corrected", 0)
        fight_features["striking_x_grappling_matchup"] = fight_features.get("net_striking_advantage", 0) * fight_features.get("grappler_advantage", 0)
        fight_features["momentum_combo"] = fight_features.get("momentum_x_win_streak", 0) * fight_features["recent_form_diff_corrected"]

        # NEW HIGH-IMPACT: Elite compound features
        fight_features["elite_finisher"] = fight_features["elo_diff"] * fight_features.get("finish_rate_diff", 0) * fight_features["recent_form_diff_corrected"]
        fight_features["veteran_advantage"] = fight_features["win_loss_ratio_diff_corrected"] * fight_features.get("experience_gap", 0) * (-fight_features.get("age_at_event_diff", 0))
        fight_features["complete_fighter"] = fight_features.get("net_striking_advantage", 0) * fight_features.get("grappler_advantage", 0) * fight_features.get("durability_diff_corrected", 0)
        fight_features["total_finish_threat"] = (fight_features.get("ko_rate_diff_corrected", 0) + fight_features.get("sub_rate_diff_corrected", 0)) * fight_features.get("finish_pressure", 0)
        fight_features["unstoppable_streak"] = fight_features.get("win_streak_diff_corrected", 0) * fight_features.get("momentum_swing", 0) * fight_features["recent_form_diff_corrected"]

        # Age prime and freshness factors
        r_age = fight_features.get("r_age_at_event", 30)
        b_age = fight_features.get("b_age_at_event", 30)
        fight_features["age_prime_advantage"] = (1.0 - abs(r_age - 29.5) / 10.0) - (1.0 - abs(b_age - 29.5) / 10.0)

        r_layoff = fight_features.get("r_days_since_last_fight_corrected", 135)
        b_layoff = fight_features.get("b_days_since_last_fight_corrected", 135)
        fight_features["freshness_advantage"] = (1.0 - abs(r_layoff - 135) / 200.0) - (1.0 - abs(b_layoff - 135) / 200.0)

        # Desperation and sustainability
        r_loss_streak = r_stats.get("loss_streak", 0)
        b_loss_streak = b_stats.get("loss_streak", 0)
        fight_features["desperation_diff"] = (r_loss_streak * (r_age / 35.0)) - (b_loss_streak * (b_age / 35.0))

        r_form = r_stats.get("recent_form", 0)
        b_form = b_stats.get("recent_form", 0)
        fight_features["momentum_sustainability_diff"] = (r_form / (r_layoff / 100.0 + 1.0)) - (b_form / (b_layoff / 100.0 + 1.0))

        fight_features["elo_x_finish"] = fight_features["elo_diff"] * fight_features.get("finish_rate_diff", 0)

        r_losses = r_stats.get("losses", 0)
        b_losses = b_stats.get("losses", 0)
        r_win_ratio = r_stats.get("win_loss_ratio", 0)
        b_win_ratio = b_stats.get("win_loss_ratio", 0)
        fight_features["adversity_experience_diff"] = (r_win_ratio * (r_losses + 1)) - (b_win_ratio * (b_losses + 1))

        # RESEARCH-BACKED FEATURES: Enhanced top feature interactions
        # H2H interactions
        fight_features["h2h_x_elo"] = fight_features.get("h2h_advantage", 0) * fight_features.get("elo_diff", 0)
        fight_features["h2h_x_form"] = fight_features.get("h2h_advantage", 0) * fight_features.get("recent_form_diff_corrected", 0)

        # Age prime score (mirroring training calculation)
        def calc_age_prime(age):
            if 26 <= age <= 33:
                return 1.0
            elif 23 <= age < 26 or 33 < age <= 36:
                return 0.6
            else:
                return 0.2

        r_age = r_stats.get("age_at_event", 30)
        b_age = b_stats.get("age_at_event", 30)
        r_age_prime = calc_age_prime(r_age)
        b_age_prime = calc_age_prime(b_age)
        fight_features["age_prime_score_diff"] = r_age_prime - b_age_prime

        # Age × Experience
        experience_diff = (r_total_fights - b_total_fights) / 50.0
        fight_features["age_x_experience"] = fight_features.get("age_at_event_diff", 0) * max(-1, min(1, experience_diff))

        # Win ratio interactions
        fight_features["win_ratio_x_finish"] = fight_features.get("win_loss_ratio_diff_corrected", 0) * fight_features.get("finish_rate_diff", 0)
        fight_features["win_ratio_x_durability"] = fight_features.get("win_loss_ratio_diff_corrected", 0) * fight_features.get("durability_diff_corrected", 0)

        # Career-based positional & target striking (use defaults for upcoming fights)
        # These would be calculated from fighter stats if available
        fight_features["distance_pct_diff_corrected"] = 0
        fight_features["clinch_pct_diff_corrected"] = 0
        fight_features["ground_pct_diff_corrected"] = 0
        fight_features["positional_striking_advantage"] = 0
        fight_features["head_pct_diff_corrected"] = 0
        fight_features["body_pct_diff_corrected"] = 0
        fight_features["leg_pct_diff_corrected"] = 0
        fight_features["target_distribution_advantage"] = 0

        # Career-based control & reversals
        fight_features["avg_ctrl_sec_diff_corrected"] = 0
        fight_features["avg_rev_diff_corrected"] = 0
        fight_features["control_dominance"] = 0

        # Compound positional features (using career averages)
        fight_features["clinch_x_grappling"] = fight_features["clinch_pct_diff_corrected"] * fight_features.get("grappler_advantage", 0)
        fight_features["distance_x_striking"] = fight_features["distance_pct_diff_corrected"] * fight_features.get("net_striking_advantage", 0)
        fight_features["ground_x_control"] = (
            fight_features["ground_pct_diff_corrected"] *
            (fight_features["avg_ctrl_sec_diff_corrected"] / 300.0)
        )
        fight_features["positional_mastery"] = (
            fight_features["positional_striking_advantage"] *
            fight_features.get("grappler_advantage", 0) *
            fight_features.get("net_striking_advantage", 0)
        )

        return fight_features, r_stats, b_stats

    def predict_fight(self, fight_data, feature_columns):
        """Predict winner for a single fight with confidence using antisymmetrization"""

        # Create DataFrame from fight_data
        df_orig = pd.DataFrame([fight_data])

        # Check if features are already computed (validation set) or need to be computed (upcoming fights)
        core_features = self.get_core_feature_names()
        has_precomputed_features = all(f in df_orig.columns for f in core_features[:10])  # Check first 10 features

        if not has_precomputed_features:
            # Case 1: Upcoming fight prediction - compute features from base stats
            df_orig = self.build_features(df_orig, recompute_elo=False)

        # Build swapped version using the proper method
        # This swaps base r_/b_ columns AND recomputes all engineered features
        # Critical for parity: differential features must be recalculated from swapped bases
        df_swap = self.build_swapped(df_orig)

        # Get available core features
        available_features = [f for f in core_features if f in df_orig.columns]

        # Extract features and fill missing values
        X_orig = df_orig[available_features].copy().fillna(0)
        X_swap = df_swap[available_features].copy().fillna(0)

        # Antisymmetrize into directional AND invariant features
        D, I_inv = self.directional_and_invariant(X_orig, X_swap)

        # Combine directional and invariant features (matching training)
        X_combined = pd.concat([D, I_inv], axis=1)

        # Use only selected features (from RFECV during training)
        # Add missing columns with 0 if needed
        X = X_combined.copy()
        for col in feature_columns:
            if col not in X.columns:
                X[col] = 0

        X = X[feature_columns]

        # =================================================================
        # ENFORCE ANTISYMMETRY: Average predictions from both perspectives
        # =================================================================
        # Tree-based models cannot learn perfect antisymmetry, so we enforce it
        # mathematically by averaging predictions from original and flipped features
        #
        # IMPORTANT: Only flip DIRECTIONAL features, keep invariant features same
        #
        # Mathematical guarantee of parity:
        #   p_red_final = (P(Red|D,I) + (1 - P(Red|-D,I))) / 2
        #   p_red_swapped = (P(Red|-D,I) + (1 - P(Red|D,I))) / 2
        #   p_red_final + p_red_swapped = 1.0  [always true by construction]

        # Identify directional vs invariant columns in selected features
        directional_mask = [not col.endswith('_inv') for col in feature_columns]

        # Prediction from original perspective: P(Red | D, I)
        X_np = np.array(X)
        proba_orig = self.winner_model.predict_proba(X_np)[0]
        p_red_orig = proba_orig[1]

        # Prediction from flipped perspective: P(Red | -D, I)
        # Only negate directional features, keep invariant features same
        X_flipped_np = X_np.copy()
        X_flipped_np[:, directional_mask] = -X_flipped_np[:, directional_mask]
        proba_flipped = self.winner_model.predict_proba(X_flipped_np)[0]
        p_red_flipped = proba_flipped[1]

        # Enforce antisymmetry: average the two perspectives
        # This guarantees: predict(fight) + predict(swapped_fight) = 1.0
        red_proba = (p_red_orig + (1 - p_red_flipped)) / 2

        # Reconstruct winner_proba for compatibility
        winner_proba = np.array([1 - red_proba, red_proba])

        # Get winner prediction
        # Use standard 0.5 threshold
        winner_pred = 1 if red_proba >= 0.5 else 0
        winner_name = "Red" if winner_pred == 1 else "Blue"

        # Calculate confidence level
        winner_confidence = winner_proba[winner_pred]
        if winner_confidence >= 0.70:
            confidence_level = "HIGH"
        elif winner_confidence >= 0.60:
            confidence_level = "MEDIUM"
        else:
            confidence_level = "LOW"

        # Betting recommendation (confidence threshold = 65%)
        should_bet = winner_confidence >= 0.65

        return {
            "winner": winner_name,
            "winner_confidence": winner_confidence,
            "confidence_level": confidence_level,
            "red_prob": winner_proba[1],
            "blue_prob": winner_proba[0],
            "should_bet": should_bet,
        }

    def predict_upcoming_fights(self, upcoming_fights, feature_columns):
        """Predict upcoming fights with enhanced confidence metrics"""
        predictions = []
        skipped_fights = []

        for fight in upcoming_fights:
            result = self.prepare_upcoming_fight(fight, feature_columns)
            if not result[0]:
                skipped_fights.append(
                    f"{fight['red_fighter']} vs {fight['blue_fighter']}"
                )
                continue

            fight_features, r_stats, b_stats = result

            pred = self.predict_fight(fight_features, feature_columns)

            predictions.append(
                {
                    "Red Fighter": fight["red_fighter"],
                    "Blue Fighter": fight["blue_fighter"],
                    "Weight Class": fight["weight_class"],
                    "Winner": fight["red_fighter"]
                    if pred["winner"] == "Red"
                    else fight["blue_fighter"],
                    "Win%": pred["winner_confidence"],
                    # "Confidence": pred["confidence_level"],
                    # "Should Bet": "YES" if pred["should_bet"] else "NO",
                }
            )

        return predictions, skipped_fights

    def export_predictions_to_excel(self, predictions, filename="predictions.xlsx"):
        """Export predictions to formatted Excel file with confidence and betting metrics"""
        df = pd.DataFrame(predictions)

        # Ensure percentage columns are in decimal format (0-1 range)
        percentage_columns = ["Win%"]
        for col in percentage_columns:
            if col in df.columns:
                if df[col].max() > 1:
                    df[col] = df[col] / 100
                df[col] = df[col].round(4)

        df.to_excel(filename, index=False, engine="openpyxl")

        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        # Styling
        header_fill = PatternFill(
            start_color="D20A0A", end_color="D20A0A", fill_type="solid"
        )
        header_font = Font(bold=True, size=11, color="FFFFFF")
        left_alignment = Alignment(horizontal="left", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Apply formatting to ALL cells
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.alignment = left_alignment
                cell.border = thin_border

        # Apply header-specific formatting
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Center align specific columns
        for col_name in ["Confidence", "Should Bet"]:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).alignment = center_alignment

        # Format percentage columns
        percentage_col_names = ["Win%"]
        for col_name in percentage_col_names:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.number_format = "0.00%"

        # Highlight high confidence bets (green)
        high_confidence_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        # Low confidence (yellow)
        low_confidence_fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        )

        if "Confidence" in df.columns:
            conf_col_idx = df.columns.get_loc("Confidence") + 1
            for row_idx in range(2, ws.max_row + 1):
                conf_cell = ws.cell(row=row_idx, column=conf_col_idx)
                if conf_cell.value == "HIGH":
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = high_confidence_fill
                elif conf_cell.value == "LOW":
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = low_confidence_fill

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value is None:
                        cell_length = 0
                    elif isinstance(cell.value, (int, float)):
                        cell_length = (
                            10
                            if cell.number_format == "0.00%"
                            else len(str(cell.value))
                        )
                    else:
                        cell_length = len(str(cell.value))

                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass

            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(filename)
        print(f"\nPredictions exported to: {filename}")
        print(f"  - {len(predictions)} fights predicted")
        print(f"  - HIGH confidence fights highlighted in green")
        print(f"  - LOW confidence fights highlighted in yellow")
        return filename


# ===== GUI APPLICATION =====

class UFCPredictorGUI:
    """Complete GUI matching original model with improved backend"""

    def __init__(self, root):
        self.root = root
        self.root.title("UFC Fight Predictor - Winner Only")
        self.root.geometry("1000x800")
        self.root.minsize(700, 550)

        self.data_file_path = tk.StringVar(value=fight_data_path)
        self.output_file_path = tk.StringVar(value="UFC_predictions_winners.xlsx")
        self.predictor = None
        self.create_widgets()

    def create_widgets(self):
        # UFC-themed header
        title_frame = tk.Frame(self.root, bg="#D20A0A")
        title_frame.pack(fill=tk.X)
        tk.Label(
            title_frame,
            text="UFC FIGHT PREDICTOR - WINNER ONLY",
            font=("Arial", 16, "bold"),
            fg="white",
            bg="#D20A0A",
        ).pack(pady=(10, 8))

        # Data file selection
        file_frame = ttk.LabelFrame(self.root, text="Training Data File", padding="5")
        file_frame.pack(fill=tk.X, padx=10, pady=3)
        ttk.Entry(file_frame, textvariable=self.data_file_path, width=70).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(file_frame, text="Browse", command=self.browse_data_file).pack(
            side=tk.LEFT
        )

        # Fights input area
        input_frame = ttk.LabelFrame(
            self.root, text="Enter Fights to Predict", padding="5"
        )
        input_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=3)

        instructions = """Enter fights in CSV format (one per line):
Format: red_fighter,blue_fighter,weight_class,gender,total_rounds

Example:
Max Holloway,Dustin Poirier,Lightweight,Men,5
Ilia Topuria,Charles Oliveira,Lightweight,Men,5
Tatiana Suarez,Amanda Lemos,Women's Strawweight,Women,3"""

        ttk.Label(
            input_frame,
            text=instructions,
            font=("Arial", 8),
            foreground="gray",
            justify=tk.LEFT,
        ).pack(anchor=tk.W, pady=(0, 3))

        self.fights_text = scrolledtext.ScrolledText(
            input_frame, height=8, width=95, wrap=tk.WORD
        )
        self.fights_text.pack(fill=tk.BOTH, expand=True, pady=3)

        # Output file selection
        output_frame = ttk.LabelFrame(self.root, text="Output File", padding="5")
        output_frame.pack(fill=tk.X, padx=10, pady=3)
        ttk.Entry(output_frame, textvariable=self.output_file_path, width=70).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(output_frame, text="Browse", command=self.browse_output_file).pack(
            side=tk.LEFT
        )

        # Buttons
        button_frame = ttk.Frame(self.root, padding="5")
        button_frame.pack(fill=tk.X, padx=10)
        ttk.Button(button_frame, text="Load Sample", command=self.load_sample).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(button_frame, text="Clear", command=self.clear_input).pack(
            side=tk.LEFT, padx=3
        )
        ttk.Button(
            button_frame, text="Generate Predictions", command=self.run_predictions
        ).pack(side=tk.RIGHT, padx=3)

        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(
            self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W
        )
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)

    def browse_data_file(self):
        filename = filedialog.askopenfilename(
            title="Select Fight Data CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if filename:
            self.data_file_path.set(filename)

    def browse_output_file(self):
        filename = filedialog.asksaveasfilename(
            title="Save Predictions As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if filename:
            self.output_file_path.set(filename)

    def load_sample(self):
        sample = """Max Holloway,Dustin Poirier,Lightweight,Men,5
Ilia Topuria,Charles Oliveira,Lightweight,Men,5
Tatiana Suarez,Amanda Lemos,Women's Strawweight,Women,3"""
        self.fights_text.delete("1.0", tk.END)
        self.fights_text.insert("1.0", sample)

    def clear_input(self):
        self.fights_text.delete("1.0", tk.END)

    def parse_fights_input(self, text):
        """Parse CSV fight input"""
        text = text.strip()
        fights = []

        for line in text.split("\n"):
            line = line.strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split(",")]

            if len(parts) != 5:
                raise ValueError(
                    f"Invalid CSV format. Expected 5 fields, got {len(parts)} in line: {line}"
                )

            try:
                fights.append(
                    {
                        "red_fighter": parts[0],
                        "blue_fighter": parts[1],
                        "weight_class": parts[2],
                        "gender": parts[3],
                        "total_rounds": int(parts[4]),
                    }
                )
            except ValueError as e:
                raise ValueError(
                    f"Error parsing line: {line}\nTotal rounds must be a number. {str(e)}"
                )

        if not fights:
            raise ValueError("No valid fight data found.")
        return fights

    def run_predictions(self):
        """Run predictions with model"""
        try:
            self.status_var.set("Loading data and training model...")
            self.root.update()

            fights_text = self.fights_text.get("1.0", tk.END)
            upcoming_fights = self.parse_fights_input(fights_text)

            data_file = self.data_file_path.get()
            if not os.path.exists(data_file):
                raise FileNotFoundError(f"Data file not found: {data_file}")

            df = pd.read_csv(data_file)
            self.status_var.set(
                f"Loaded {len(df)} fights. Training..."
            )
            self.root.update()

            # Create improved predictor
            self.predictor = ImprovedUFCPredictor(use_ensemble=False, debug_mode=False)

            # Train models (output goes to console, not GUI)
            print("\n" + "="*80)
            print("TRAINING UFC PREDICTOR")
            print("="*80 + "\n")

            df = self.predictor.fix_data_leakage(df)
            self.predictor.df_train = df
            feature_columns = self.predictor.train(df)

            self.status_var.set("Generating predictions...")
            self.root.update()

            # Generate predictions
            predictions, skipped_fights = self.predictor.predict_upcoming_fights(
                upcoming_fights, feature_columns
            )

            if not predictions:
                self.status_var.set("No predictions generated")
                messagebox.showerror(
                    "No Predictions",
                    "All fights were skipped due to insufficient fighter data.",
                )
                return

            output_file = self.output_file_path.get()

            # Export to Excel
            with redirect_stdout(io.StringIO()):
                self.predictor.export_predictions_to_excel(predictions, output_file)

            # Clean up temporary files
            cleanup_temp_files()

            # Success message
            success_msg = f"Predictions generated!\n\nSaved to: {output_file}\n\n{len(predictions)} fight(s) predicted"

            if skipped_fights:
                success_msg += f"\n\n{len(skipped_fights)} fight(s) skipped"
                self.status_var.set(
                    f"Success! {len(predictions)} predictions, {len(skipped_fights)} skipped"
                )
                skipped_msg = (
                    success_msg
                    + "\n\nSkipped:\n"
                    + "\n".join(f"- {fight}" for fight in skipped_fights)
                )
                messagebox.showwarning("Predictions Complete", skipped_msg)
            else:
                self.status_var.set(
                    f"Success! All {len(predictions)} predictions saved"
                )
                messagebox.showinfo("Success", success_msg)

        except Exception as e:
            import traceback

            error_details = traceback.format_exc()
            self.status_var.set("Error occurred")
            messagebox.showerror(
                "Error", f"Error:\n\n{str(e)}\n\nDetails:\n{error_details}"
            )


def main():
    """Main function to run the improved UFC Predictor GUI"""
    try:
        root = tk.Tk()
        app = UFCPredictorGUI(root)
        root.mainloop()
    except Exception as e:
        print(f"Error starting UFC Predictor: {e}")
        import traceback
        traceback.print_exc()
    finally:
        cleanup_temp_files()


if __name__ == "__main__":
    # Enable multiprocessing
    if hasattr(sys, "frozen") and sys.frozen:
        import multiprocessing
        multiprocessing.freeze_support()

    main()
    